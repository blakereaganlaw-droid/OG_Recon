#!/usr/bin/env python3
"""
Unit tests for the reconciliation engine.

Covers Section 7 primitives, Section 4 router, Section 5 binder, Section 8
pool dedup, and an end-to-end synthetic run through the forward pipeline
+ independent audit (Section 16 build order steps 1-8, validated on
a synthetic account since real UT source files are not present here).

Run:  python3 -m unittest test_recon -v
"""

import os
import shutil
import tempfile
import unittest
from datetime import date

import recon_engine as E
import recon_audit as A
import run_recon as R


class TestPrimitives(unittest.TestCase):
    def test_N(self):
        self.assertEqual(E.N(None), "")
        self.assertEqual(E.N("  x "), "x")
        self.assertEqual(E.N(5), "5")

    def test_cents(self):
        self.assertEqual(E.cents("123.45"), 12345)
        self.assertEqual(E.cents("$1,234.56"), 123456)
        self.assertEqual(E.cents("(100.00)"), -10000)
        self.assertEqual(E.cents("-50"), -5000)
        self.assertEqual(E.cents(0), 0)
        self.assertEqual(E.cents(10.0), 1000)
        self.assertIsNone(E.cents("abc"))
        self.assertIsNone(E.cents(None))
        self.assertIsNone(E.cents(""))
        # falsy-zero must be preserved, not dropped
        self.assertEqual(E.cents("0.00"), 0)

    def test_parse_date(self):
        self.assertEqual(E.parse_date("2024-03-05"), date(2024, 3, 5))
        self.assertEqual(E.parse_date("03/05/2024"), date(2024, 3, 5))
        self.assertEqual(E.parse_date("3/5/24"), date(2024, 3, 5))
        self.assertEqual(E.parse_date("2024-03-05T00:00:00.000+00:00"), date(2024, 3, 5))
        self.assertEqual(E.parse_date(45356), date(2024, 3, 5))  # Excel serial
        self.assertIsNone(E.parse_date("not a date"))
        self.assertIsNone(E.parse_date(None))

    def test_datetime_before_date(self):
        from datetime import datetime
        self.assertEqual(E.parse_date(datetime(2024, 3, 5, 12, 0)), date(2024, 3, 5))

    def test_znorm(self):
        self.assertEqual(E.znorm("ab-12_34"), "AB1234")
        self.assertEqual(E.znorm(" x.y "), "XY")

    def test_digit_runs(self):
        self.assertEqual(E.digit_runs("ab12345cd678", 5), {"12345"})
        self.assertEqual(E.digit_runs("12 34", 5), set())

    def test_desc_word_tokens(self):
        # >=6-char znorm words carrying a letter; pure numbers and short words
        # (which the numeric/other lanes handle) are excluded.
        t = E.desc_word_tokens("Controlled Disbursing 581_FHB Master Account")
        self.assertIn("CONTROLLED", t)
        self.assertIn("DISBURSING", t)
        self.assertIn("MASTER", t)
        self.assertIn("ACCOUNT", t)
        self.assertNotIn("FHB", t)      # 3 chars
        self.assertNotIn("581", t)      # pure numeric
        # an alphanumeric security/reference code qualifies (has a letter)
        self.assertIn("6698M4UV2", E.desc_word_tokens("CUSIP 6698M4UV2 lot"))

    def test_payer_tokens_excludes_bai2_labels(self):
        toks = E.payer_tokens("SENDING CO NAME ACME WIDGETS COMPANY")
        self.assertIn("ACME", toks)
        self.assertIn("WIDGETS", toks)
        self.assertNotIn("COMPANY", toks)
        self.assertNotIn("NAME", toks)
        self.assertNotIn("SENDING", toks)

    def test_spn_of(self):
        self.assertEqual(E.spn_of("REC-SPN 12345-1"), "SPN12345")
        self.assertEqual(E.spn_of("nothing"), "")

    def test_is_mid(self):
        self.assertTrue(E.is_mid("8012345678"))
        self.assertTrue(E.is_mid("2000123456"))
        self.assertFalse(E.is_mid("6500000097"))  # Heartland excluded
        self.assertFalse(E.is_mid("1234567890"))

    def test_sibling(self):
        self.assertTrue(E.sibling("1234567", "1234568"))
        self.assertTrue(E.sibling("12345670", "12345699"))
        self.assertFalse(E.sibling("1234567", "1234567"))  # equal not sibling
        self.assertFalse(E.sibling("123", "124"))          # too short

    def test_signed_lag(self):
        self.assertEqual(E.signed_lag(date(2024, 3, 10), date(2024, 3, 5)), 5)
        self.assertIsNone(E.signed_lag(None, date(2024, 3, 5)))

    def test_reference_equal(self):
        self.assertTrue(E.reference_equal("ABC-123456", "abc123456"))
        self.assertTrue(E.reference_equal("REF987654", "XREF987654X"))  # containment >=6
        self.assertFalse(E.reference_equal("1234567", "1234568"))       # sibling conflict
        self.assertFalse(E.reference_equal("", "x"))

    def test_reference_tie(self):
        self.assertTrue(E.reference_tie("inv 45678 x", "y45678"))
        self.assertFalse(E.reference_tie("1234567", "1234568"))  # siblings never tie

    def test_date_bands(self):
        self.assertEqual(E.date_band(2), "STRONG")
        self.assertEqual(E.date_band(6), "MODERATE")
        self.assertEqual(E.date_band(12), "WEAK")
        self.assertEqual(E.date_band(20), "SUSPICIOUS")
        self.assertEqual(E.date_band(40), "REJECT")
        self.assertTrue(E.date_ok_merchant(3))
        self.assertFalse(E.date_ok_merchant(10))

    def test_xlsb_integral_float_normalization(self):
        # pyxlsb returns every numeric cell as a float; an integral Oracle id
        # (DEPOSIT_ID 65105) must collapse to int so it does not stringify to
        # "65105.0" and break the MET<->ST bridge join / d:/r: citations.  The
        # engine and the (independent) audit must normalize IDENTICALLY.
        for norm in (E._xlsb_norm, A._xlsb_norm):
            self.assertEqual(norm(65105.0), 65105)
            self.assertIs(type(norm(65105.0)), int)
            self.assertEqual(str(norm(203362.0)), "203362")   # no ".0"
            self.assertEqual(norm(1890.69), 1890.69)          # money untouched
            self.assertEqual(norm("65105"), "65105")          # text untouched
            self.assertIsNone(norm(None))
            self.assertIs(norm(True), True)                    # bool is not a float
            self.assertEqual(norm(1e18), 1e18)                 # beyond guard: untouched


class TestRouter(unittest.TestCase):
    def test_classify(self):
        self.assertEqual(E.classify_file("20240101_FHB_UTC_BSL_UNR.xlsx"), "BSL")
        self.assertEqual(E.classify_file("FHB_UTC_All_Data.xlsx"), "ALL_DATA")
        self.assertEqual(E.classify_file("Account_ST_open.xlsx"), "ST")
        self.assertEqual(E.classify_file("MET_All.xlsx"), "MET")
        self.assertEqual(E.classify_file("Receivables_Receipts.xlsx"), "RECEIPTS")
        self.assertEqual(E.classify_file("Edison_Payments.xlsx"), "EDISON_PAY")
        self.assertIsNone(E.classify_file("random.txt"))

    def test_bsl_excludes_all_data(self):
        # A file that is All_Data must not be misrouted to BSL.
        self.assertEqual(E.classify_file("FHB_UTC_BSL_All_Data.xlsx"), "ALL_DATA")

    def test_infer_account(self):
        self.assertEqual(E.infer_account("20240101_FHB_UTC_BSL.xlsx"), "FHB_UTC")
        self.assertEqual(E.infer_account("Regions_UTM_BSL.xlsx"), "REGIONS_UTM")


class TestBinder(unittest.TestCase):
    def test_content_first(self):
        rows = [
            ("Date", "Amount", "Reference"),
            ("2024-03-05", "100.00", "REF123456"),
            ("2024-03-06", "200.00", "REF223456"),
        ]
        specs = {
            "date": E._rs(True, ["Date"], E.pred_date),
            "amount": E._rs(True, ["Amount"], E.pred_signed_amount),
            "reference": E._rs(False, ["Reference"], E.pred_reference),
        }
        m, hi = E.bind_columns(rows, specs)
        self.assertEqual(hi, 0)
        self.assertEqual(m["date"], 0)
        self.assertEqual(m["amount"], 1)
        self.assertEqual(m["reference"], 2)

    def test_missing_required_raises(self):
        rows = [("Foo", "Bar"), ("x", "y"), ("a", "b")]
        specs = {"amount": E._rs(True, ["Amount"], E.pred_signed_amount)}
        with self.assertRaises(E.InvalidSourceData):
            E.bind_columns(rows, specs)


class TestPoolDedup(unittest.TestCase):
    def test_keep_largest_and_borrow_counterparty(self):
        # Classic total-plus-invoice-splits repeat: largest == sum(rest).
        e_total = E._mk_entry("T1", 10000, date(2024, 3, 5), "REF1", "", "AR", "UNR", True, "RECEIPTS")
        e_split1 = E._mk_entry("T1", 4000, date(2024, 3, 5), "REF1", "ACME CORP", "AR", "UNR", True, "RECEIPTS")
        e_split2 = E._mk_entry("T1", 6000, date(2024, 3, 5), "REF1", "", "AR", "UNR", True, "RECEIPTS")
        out = E._dedup_keep_largest([e_split1, e_total, e_split2], lambda e: e.id)
        self.assertEqual(len(out), 1)
        self.assertEqual(out[0].amount_cents, 10000)          # kept the total
        self.assertEqual(out[0].counterparty, "ACME CORP")     # borrowed

    def test_deposit_correction_detection(self):
        class _B:
            additional_info = ""
            line_info = ""
        b = _B()
        b.line_info = "Line 1 , 2026-06-10 DEPOSIT CORRECTION CREDIT"
        self.assertTrue(E._is_deposit_correction(b))
        b2 = _B()
        b2.additional_info = "CORRECTED DEPOSIT 0123"
        self.assertTrue(E._is_deposit_correction(b2))
        b3 = _B()
        b3.additional_info = "MERCHANT SERVICEDEPOSIT"
        b3.line_info = "Line 2"
        self.assertFalse(E._is_deposit_correction(b3))

    def test_amount_distinctive(self):
        self.assertTrue(E.amount_distinctive(245546969))   # 2,455,469.69
        self.assertTrue(E.amount_distinctive(-154638))     # (1,546.38)
        self.assertFalse(E.amount_distinctive(500000))     # 5,000.00 round
        self.assertFalse(E.amount_distinctive(6500))       # 65.00 too small
        self.assertFalse(E.amount_distinctive(-13000))     # (130.00)

    def test_convera_payables_only(self):
        class _B:
            amount_cents = -59690
            additional_info = "CONVERA TENN DEBITS 260703"
            line_info = "Line 59 , 2026-07-03"
            transaction_type = "Wire"
            reference_raw = ""
            customer_reference = ""
            account_servicer_reference = ""
        b = _B()
        ap = E._mk_entry("1394", -59690, date(2026, 7, 2), "1394", "Stichting EHEDG", "AP", "UNR", True, "ST")
        ext = E._mk_entry("999", -59690, date(2026, 7, 2), "999", "Someone", "EXT", "UNR", True, "MET")
        self.assertTrue(E._type_gate_ok(b, ap))
        self.assertFalse(E._type_gate_ok(b, ext))

    def test_type_incongruent_uncorroborated(self):
        # Owner rule (2026-07-13): type important but NOT dispositive. A Misc
        # BSL vs electronic (ACH) ST amount-only pairing is barred; any
        # reference/payer/MID tie overrides.
        class _B:
            transaction_type = "Miscellaneous"
            reference_raw = "999888"
            additional_info = "DEPOSIT"
            customer_reference = ""
            account_servicer_reference = ""
            line_info = "Line 1"
            recon_reference = "999888"
        b = _B()
        b.ref_digits = E.digit_runs("999888")
        b.payer_tokens = E.payer_tokens("ACME WIDGETS")
        b.mid = ""
        ach = E._mk_entry("X", 5000, date(2026, 6, 1), "111222", "SomeoneElse", "EXT", "UNR", True, "MET", transaction_type="Automated clearing house")
        self.assertTrue(E._type_incongruent_uncorroborated(b, ach))       # barred
        chk = E._mk_entry("Y", 5000, date(2026, 6, 1), "111222", "X", "EXT", "UNR", True, "MET", transaction_type="Check")
        self.assertFalse(E._type_incongruent_uncorroborated(b, chk))      # Check ST: not electronic
        tied = E._mk_entry("Z", 5000, date(2026, 6, 1), "999888", "X", "EXT", "UNR", True, "MET", transaction_type="Automated clearing house")
        self.assertFalse(E._type_incongruent_uncorroborated(b, tied))     # reference tie overrides
        payer = E._mk_entry("W", 5000, date(2026, 6, 1), "000000", "ACME WIDGETS INC", "EXT", "UNR", True, "MET", transaction_type="ACH")
        self.assertFalse(E._type_incongruent_uncorroborated(b, payer))    # payer tie overrides

    def test_ext_stale_candidate_bar(self):
        # Owner rule (2026-07-12): External STs entered 12+ days before the
        # BSL are barred even as Candidates; 8-11 days may still surface;
        # BSL-before-ST unbounded; non-EXT sources untouched.
        class _B:
            pass
        b = _B()
        b.date = date(2026, 6, 12)
        mk = lambda src_, d: E._mk_entry("X1", 1000, d, "R", "C", src_, "UNR", True, "ST")
        self.assertTrue(E._ext_stale_barred(b, mk("EXT", date(2026, 5, 31))))   # 12d stale
        self.assertTrue(E._ext_stale_barred(b, mk("EXT", date(2026, 1, 1))))    # very stale
        self.assertFalse(E._ext_stale_barred(b, mk("EXT", date(2026, 6, 1))))   # 11d stale
        self.assertFalse(E._ext_stale_barred(b, mk("EXT", date(2026, 7, 30))))  # ST after BSL
        self.assertFalse(E._ext_stale_barred(b, mk("AR", date(2025, 1, 1))))    # non-EXT
        e = mk("EXT", None)
        self.assertFalse(E._ext_stale_barred(b, e))                             # no date

    def test_chargeback_mid_gate(self):
        # Owner rule (2026-07-12): negative chargeback/merchant-fee lines
        # pair ONLY with STs carrying the SAME MID — wrong or absent MID is
        # barred even as a Candidate.
        class _B:
            amount_cents = -6000
            additional_info = ("MERCHANT SERVICECHARGEBACK2603178028920588"
                               "ENTRY DESC: CHARGEBACK ID NUMBER: 8028920588")
            line_info = "Line 129 , 2026-03-17"
            transaction_type = "Automated clearing house"
            reference_raw = "8028920588"
            customer_reference = ""
            account_servicer_reference = "8028920588"
        b = _B()
        wrong = E._mk_entry("1027659", -6000, date(2026, 6, 30), "8035758468", "FY26 CC chargeback", "EXT", "UNR", True, "MET")
        right = E._mk_entry("999999", -6000, date(2026, 3, 17), "8028920588", "Chargeback", "EXT", "UNR", True, "MET")
        nomid = E._mk_entry("888888", -6000, date(2026, 3, 17), "REF123456", "Something", "EXT", "UNR", True, "MET")
        self.assertTrue(E._is_card_fee_debit(b))
        self.assertFalse(E._type_gate_ok(b, wrong))   # different MID
        self.assertTrue(E._type_gate_ok(b, right))    # same MID
        self.assertFalse(E._type_gate_ok(b, nomid))   # no MID at all

    def test_distinct_receipts_sharing_label_all_kept(self):
        # Real FHB Master case: two receipts share one transaction-number
        # label but do NOT fit the total-plus-splits signature — both stay,
        # ids disambiguated by amount (base_id preserves the bridge join).
        a = E._mk_entry("SPN070326 ACH HRSA", 1820327, date(2026, 7, 3), "SPN070326 ACH HRSA", "HRSA", "AR", "UNR", True, "ST")
        b = E._mk_entry("SPN070326 ACH HRSA", 10241147, date(2026, 7, 3), "SPN070326 ACH HRSA", "HRSA", "AR", "UNR", True, "ST")
        out = E._dedup_keep_largest([a, b], lambda e: e.id)
        self.assertEqual(len(out), 2)
        self.assertEqual(sorted(e.id for e in out),
                         ["SPN070326 ACH HRSA [102411.47]",
                          "SPN070326 ACH HRSA [18203.27]"])
        self.assertTrue(all(e.base_id == "SPN070326 ACH HRSA" for e in out))

    def test_p9c_partial_reference_group_enriched_review(self):
        # Owner directive (2026-07-14): the ORT/Receivables 1:M reference
        # search always runs.  A bank line whose reference EXACTLY matches a
        # group of >= 2 open Receivables STs that sum SHORT of the BSL (the
        # rest already auto-reconciled) surfaces as an ENRICHED REVIEW naming
        # them — never a Match/Candidate (exact-sum guardrail preserved).
        bsl = E.make_bsl("L1", date(2026, 6, 16), 13275458, "702602081",
                         "702602081", "DEPOSIT 702602081", "Miscellaneous", "")
        pool = [
            E._mk_entry("702602081 [41124.47]", 4112447, date(2026, 6, 10), "702602081", "LSU", "AR", "UNR", True, "ST", spr="SPN113615"),
            E._mk_entry("702602081 [15336.80]", 1533680, date(2026, 6, 10), "702602081", "Emory", "AR", "UNR", True, "ST", spr="SPN113508"),
            E._mk_entry("702602081 [9941.07]", 994107, date(2026, 6, 10), "702602081", "Metis", "AR", "UNR", True, "ST", spr="SPN111531"),
        ]  # open sum 6640234 << BSL 13275458
        p = E.forward_reconcile([bsl], pool, {}, "FHB_TEST", {})[0]
        self.assertEqual(p.kind, E.REVIEW)
        self.assertEqual(p.pass_name, "P9c_ref_1m_review")
        self.assertIn("PARTIAL_REFERENCE_GROUP", p.codes)
        self.assertEqual(len(p.st_entries), 3)          # the 3 open members cited
        self.assertIn("1:M reference tie", p.explanation)

    def test_p9c_requires_two_members_and_distinctive_ref(self):
        # A single reference-tied ST (1:1, not 1:M) does NOT trigger the
        # partial-group review; nor does a short (<6 char) reference — both
        # fall through to the ordinary P10 residual review.
        bsl = E.make_bsl("L1", date(2026, 6, 16), 13275458, "702602081",
                         "702602081", "DEPOSIT", "Miscellaneous", "")
        one = [E._mk_entry("702602081 [41124.47]", 4112447, date(2026, 6, 10), "702602081", "LSU", "AR", "UNR", True, "ST")]
        self.assertEqual(E.forward_reconcile([bsl], one, {}, "T", {})[0].pass_name, "P10_review")
        bsl2 = E.make_bsl("L2", date(2026, 6, 16), 13275458, "2605", "2605", "DEPOSIT", "Miscellaneous", "")
        short = [E._mk_entry("2605 [1.00]", 100, date(2026, 6, 10), "2605", "A", "AR", "UNR", True, "ST"),
                 E._mk_entry("2605 [2.00]", 200, date(2026, 6, 10), "2605", "B", "AR", "UNR", True, "ST")]
        self.assertEqual(E.forward_reconcile([bsl2], short, {}, "T", {})[0].pass_name, "P10_review")

    def test_p9c_rejects_opposite_sign_and_oversized_shared_id(self):
        # A shared ACH originator / company id (not a deposit batch) collides
        # a NEGATIVE debit with many POSITIVE receipts: opposite sign and the
        # group exceeds the bank line -> NOT a partial deposit, no P9c review.
        bsl = E.make_bsl("L1", date(2026, 6, 8), -112500, "8487001827",
                         "8487001827", "ACH COMPANY ID: 8487001827", "Automated clearing house", "")
        pool = [E._mk_entry("A [30000.00]", 3000000, date(2026, 5, 1), "8487001827", "X", "EXT", "UNR", True, "MET"),
                E._mk_entry("B [36059.26]", 3605926, date(2026, 5, 2), "8487001827", "Y", "EXT", "UNR", True, "MET")]
        self.assertNotEqual(E.forward_reconcile([bsl], pool, {}, "T", {})[0].pass_name, "P9c_ref_1m_review")

    def test_payer_family_alias_not_a_contradiction(self):
        # VSHP / TennCare == BlueCare Tennessee: distinct trade names for one
        # payer family are agreement, not contradiction.
        self.assertEqual(E.payer_family("VSHP TN CARE SELECT"), {"BLUECARE"})
        self.assertEqual(E.payer_family("BlueCare Tennessee"), {"BLUECARE"})

        class _B:
            amount_cents = 24673664
            additional_info = "VSHP TN CARE SELCCD_PYMTS  CO NAME: VSHP TN CARE SEL"
            customer_reference = "VSHP TN CARE SEL"
        b = _B()
        bc = E._mk_entry("R1", 10712119, date(2026, 7, 6), "", "BlueCare Tennessee", "AR", "UNR", True, "RECEIPTS")
        self.assertFalse(E.payer_contradiction(b, [bc]))          # same family, no contradiction
        israel = E._mk_entry("R2", 10712119, date(2026, 7, 6), "", "City of Chattanooga", "AR", "UNR", True, "RECEIPTS")
        self.assertTrue(E.payer_contradiction(b, [israel]))        # unrelated payer -> contradiction

    def test_feed_session_label_is_not_a_payer_contradiction(self):
        # Oracle ORT stamps unattributed External lines with a generic
        # "FEED SESSION <n>" batch label in the counterparty column.  That
        # names the load batch, not a payer, so it is SILENCE on payer
        # identity and must NOT contradict a real bank-side payer (owner,
        # 2026-07-16).  A Heartland settlement vs a feed-session ST is not
        # contradicted; a real unrelated payer still is.
        self.assertEqual(E._contra_tokens("TN FEED SESSION 4161 | TN FEED SESSION 4161"), set())
        self.assertEqual(E._contra_tokens("CITY OF CHATTANOOGA"), {"CITY", "CHATTANOOGA"})

        class _B:
            amount_cents = 189069
            additional_info = ("HRTLAND PMT SYS TXNS/FEES 2607136500000097923"
                               "SENDING CO ID: WFBEHPS001 SENDING CO NAME: HRTLAND PMT SYS")
            customer_reference = "HRTLAND PMT SYS"
        b = _B()
        feed = E._mk_entry("1030340", 2619400624, date(2026, 7, 13), "",
                           "TN FEED SESSION 4161 | TN FEED SESSION 4161", "EXT", "UNR", True, "EXT")
        self.assertFalse(E.payer_contradiction(b, [feed]))     # feed-session label = silence
        israel = E._mk_entry("R9", 111, date(2026, 7, 13), "",
                             "City of Chattanooga", "AR", "UNR", True, "RECEIPTS")
        self.assertTrue(E.payer_contradiction(b, [israel]))    # unrelated payer still contradicts

    def test_ref_tied_split_group_outranks_amount_only_coincidence(self):
        # Owner doctrine (2026-07-17, FHB UTIA $40 merchant line): a deposit
        # whose members carry the BSL's reference but sum to it only WITH a
        # closed member (auto-rec split) outranks a coincidental equal-sum
        # OPEN deposit that carries no tie at all.  Previously the
        # uncorroborated coincidence entered the amount-only branch and its
        # payer-contradiction bar skipped PAST the split branch, stranding
        # the line in Review with the ref-tied deposit unmentioned.
        bsl = E.make_bsl("L32", date(2026, 5, 12), 4000, "8042195480", "",
                         "MERCHANT SERVICEDEPOSIT   2605128042195480"
                         "SENDING CO ID: 3084000026 SENDING CO NAME: MERCHANT SERVICE",
                         "Automated clearing house", "")
        open_member = E._mk_entry("987207", 1000, date(2026, 5, 13), "8042195480",
                                  "", "EXT", "UNR", True, "MET",
                                  deposit_id="117748", receipt_id="341093",
                                  transaction_type="Automated clearing house")
        closed_member = E._mk_entry("987229", 3000, date(2026, 5, 13), "8042195480",
                                    "", "EXT", "REC", False, "MET",
                                    deposit_id="117748", receipt_id="341092",
                                    transaction_type="ACH")
        # coincidental open $40 deposit, later date, no tie, contradicting payer
        coincidence = E._mk_entry("999001", 4000, date(2026, 7, 14), "555000111",
                                  "City of Chattanooga", "EXT", "UNR", True, "MET",
                                  deposit_id="128801", receipt_id="900001",
                                  transaction_type="Automated clearing house")
        # extra open same-MID merchant STs (real FHB UTIA condition): they
        # break the P4 phase-1 whole-reference-group sum, so only the
        # deposit-chain (phase 2) view can explain the line.
        extra1 = E._mk_entry("1031250", 2000, date(2026, 5, 14), "8042195480",
                             "", "EXT", "UNR", True, "ST",
                             transaction_type="Automated clearing house")
        extra2 = E._mk_entry("1031251", 1500, date(2026, 5, 15), "8042195480",
                             "", "EXT", "UNR", True, "ST",
                             transaction_type="Automated clearing house")
        pool = [open_member, closed_member, coincidence, extra1, extra2]
        p = E.forward_reconcile([bsl], pool, {}, "FHB_TEST", {})[0]
        self.assertEqual(p.kind, E.CANDIDATE)
        self.assertEqual(p.pass_name, "P4_deposit_group")
        self.assertIn("POSSIBLE_AUTO_REC_SPLIT", p.codes)
        self.assertIn("117748", p.explanation)
        self.assertIn("987229", p.explanation)   # closed member named

    def test_p8c_payer_family_receipt_sum_candidate(self):
        # A VSHP/TennCare ACH covered by two same-day BlueCare receipts whose
        # own references do NOT tie the bank line surfaces as a Candidate via
        # the payer-family tie + exact same-day sum (never a Match).
        bsl = E.make_bsl("L1", date(2026, 7, 6), 24673664, "", "",
                         "VSHP TN CARE SELCCD_PYMTS CO NAME: VSHP TN CARE SEL",
                         "Automated clearing house", "", customer_reference="VSHP TN CARE SEL")
        pool = [
            E._mk_entry("DOC 25903", 10712119, date(2026, 7, 6), "", "BlueCare Tennessee", "AR", "UNR", True, "RECEIPTS"),
            E._mk_entry("DOC 25901", 13961545, date(2026, 7, 6), "", "BlueCare Tennessee", "AR", "UNR", True, "RECEIPTS"),
        ]
        p = E.forward_reconcile([bsl], pool, {}, "FHB_TEST", {})[0]
        self.assertEqual(p.kind, E.CANDIDATE)
        self.assertEqual(p.pass_name, "P8c_payer_family")
        self.assertIn("PAYER_FAMILY_GROUP", p.codes)
        self.assertEqual(len(p.st_entries), 2)
        # a different-day receipt breaks the same-day group -> no P8c candidate
        pool2 = [
            E._mk_entry("DOC 25903", 10712119, date(2026, 7, 6), "", "BlueCare Tennessee", "AR", "UNR", True, "RECEIPTS"),
            E._mk_entry("DOC 25901", 13961545, date(2026, 6, 30), "", "BlueCare Tennessee", "AR", "UNR", True, "RECEIPTS"),
        ]
        self.assertNotEqual(E.forward_reconcile([bsl], pool2, {}, "FHB_TEST", {})[0].pass_name, "P8c_payer_family")


# ---- synthetic end-to-end fixture ------------------------------------

def _write_xlsx(path, sheets):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for title, rows in sheets:
        ws = wb.create_sheet(title=title[:31])
        for r in rows:
            ws.append(list(r))
    wb.save(path)


def _build_fhb_utc_inputs(d, bsl_name="20240101_FHB_UTC_BSL_UNR.xlsx",
                          st_name="20240101_FHB_UTC_Account_ST.xlsx"):
    """Synthetic FHB_UTC account: 2 matchable lines + 1 review line."""
    # BSL: three open bank lines for FHB_UTC.
    bsl = [
        ("Transaction Date", "Amount", "Line Number", "Account Servicer Reference",
         "Additional Information", "Transaction Code"),
        ("2024-03-10", "150.00", "1", "REF100200", "ACH CREDIT VENDOR", "142"),   # exact 1:1
        ("2024-03-12", "300.00", "2", "GRP500600", "DEPOSIT", "174"),             # 1:M group
        ("2024-03-15", "999.99", "3", "NADA000111", "UNKNOWN THING", "142"),      # review
    ]
    _write_xlsx(os.path.join(d, bsl_name), [("Exported", bsl)])
    # ST: open external transactions.
    st = [
        ("Transaction Date", "Amount", "Transaction Number", "Source", "Reference", "Counterparty"),
        ("2024-03-09", "150.00", "ST1", "EXT", "REF100200", "VENDOR INC"),        # matches line 1
        ("2024-03-11", "120.00", "ST2", "EXT", "GRP500600", "PAYER A"),           # part of group
        ("2024-03-11", "180.00", "ST3", "EXT", "GRP500600", "PAYER A"),           # part of group
        ("2024-03-01", "999.99", "ST4", "GL", "NADA000111", "JOURNAL"),           # journal: never matches
    ]
    _write_xlsx(os.path.join(d, st_name), [("Exported", st)])
    return "FHB_UTC"


class TestEndToEnd(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp()
        self.out = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.d, ignore_errors=True)
        shutil.rmtree(self.out, ignore_errors=True)

    def _build_inputs(self):
        return _build_fhb_utc_inputs(self.d)

    def test_forward_and_audit(self):
        account = self._build_inputs()
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["account"], "FHB_UTC")
        self.assertEqual(runlog["bsl_count"], 3)
        summ = runlog["recon_summary"]
        # line 1 -> Match (exact 1:1); line 2 -> Match (1:M group); line 3 -> Review (journal only)
        self.assertEqual(summ["matches"], 2)
        self.assertEqual(summ["reviews"], 1)
        # audit must pass
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))

    def test_conservation_every_bsl_once(self):
        self._build_inputs()
        runlog = E.run(self.d, self.out, present=True)
        s = runlog["recon_summary"]
        self.assertEqual(s["matches"] + s["candidates"] + s["reviews"], runlog["bsl_count"])

    def test_journal_never_matches(self):
        self._build_inputs()
        runlog = E.run(self.d, self.out, present=True)
        # The 999.99 line has only a GL/journal counterpart -> must be Review.
        recon_path = runlog["recon_workbook"]
        tabs, _ = A._read_output_tabs(recon_path)
        review = [r for r in tabs["Review Notes"][3:] if any(A._N(c) for c in r)]
        self.assertTrue(any("999.99" in A._N(r[2]) for r in review))


class TestColumnRobustness(unittest.TestCase):
    """Column movement/renaming/insertion must never change results silently:
    identical output for benign rearrangement, loud named errors for genuine
    ambiguity. Pins the guarantees proven by the mutation sweep."""

    def setUp(self):
        self.a = tempfile.mkdtemp()   # baseline inputs
        self.b = tempfile.mkdtemp()   # mutated inputs
        self.oa = tempfile.mkdtemp()
        self.ob = tempfile.mkdtemp()

    def tearDown(self):
        for d in (self.a, self.b, self.oa, self.ob):
            shutil.rmtree(d, ignore_errors=True)

    def test_leading_date_key_validates_dates(self):
        # An 8-digit account number is not a date; the real stamp wins.
        self.assertEqual(E._leading_date_key("acct_99999999_st_20250101.xlsx"),
                         "20250101")
        self.assertEqual(E._leading_date_key("20240101_FHB_UTC_BSL.xlsx"),
                         "20240101")
        self.assertEqual(E._leading_date_key("undated.xlsx"), "00000000")

    def test_bind_columns_no_header_raises(self):
        rows = [("Foo", "Bar"), ("2024-01-01", "10.00")]
        specs = {"amount": E._rs(True, ["Amount"], E.pred_signed_amount)}
        with self.assertRaises(E.InvalidSourceData):
            E.bind_columns(rows, specs)

    def test_optional_role_blind_tie_unbound(self):
        # Two columns identical on content and header: an optional role must
        # go unbound, never bind leftmost-by-position.
        rows = [("Date", "Amount", "X", "Y"),
                ("2024-01-01", "10.00", "REF1", "REF9"),
                ("2024-01-02", "20.00", "REF2", "REF8")]
        specs = {
            "date": E._rs(True, ["Date"], E.pred_date),
            "amount": E._rs(True, ["Amount"], E.pred_signed_amount),
            "reference": E._rs(False, ["Reference"], E.pred_reference),
        }
        m, _hi = E.bind_columns(rows, specs)
        self.assertNotIn("reference", m)

    def test_required_role_alias_priority_tiebreak(self):
        # Owner, 2026-07-21: the real AP Payments export carries BOTH a "Payee"
        # and a "Supplier or Party" column that hold the same name but differ on
        # a few rows.  Both exact-match the payee alias list and tie on content,
        # so the binder used to fail loud (AmbiguousColumn).  The alias list is
        # ordered by preference, so the earlier alias ("Payee") wins.
        rows = [("Payment Number", "Payee", "Supplier or Party", "Amount"),
                ("1001", "Acme Inc", "Acme Inc", "10.00"),
                ("1002", "Beta LLC", "Beta Corp", "20.00"),   # columns differ here
                ("1003", "Gamma Co", "Gamma Co", "30.00")]
        specs = {
            "payment_number": E._rs(True, ["Payment Number"], E.pred_number),
            "payee": E._rs(True, ["Payee", "Supplier or Party", "Supplier"],
                           E.pred_customer),
            "amount": E._rs(True, ["Amount"], E.pred_signed_amount),
        }
        m, _hi = E.bind_columns(rows, specs)
        self.assertEqual(m["payee"], 1)   # the "Payee" column, not col 2
        # Helper: exact-alias order beats substring; genuine dup header -> None.
        an = [E._norm_header(a) for a in ["Payee", "Supplier or Party"]]
        self.assertEqual(E._alias_priority_winner(
            ("Payee", "Supplier or Party"), [0, 1], an), 0)
        self.assertIsNone(E._alias_priority_winner(
            ("Payee", "Payee"), [0, 1], an))   # identical header -> still ambiguous

    def test_pred_met_description(self):
        self.assertTrue(E.pred_met_description("d:8812345 | r:991 | UT FOUNDATION"))
        self.assertFalse(E.pred_met_description("d:63363 | r:197960"))  # ID-column stub
        self.assertFalse(E.pred_met_description("something | else"))
        self.assertFalse(E.pred_met_description("see attached"))

    def test_mid_master_multi_mid_and_conflict(self):
        gl = "01-1234567-123456-123456"
        p = os.path.join(self.a, "MID_Master.xlsx")
        _write_xlsx(p, [("Sheet1", [("8012345678", "8098765432", gl)])])
        rf = E.RoutedFile("MID_MASTER", p, "MID_Master.xlsx", "all")
        out = E.load_mid_master(rf)
        # Every MID in the row maps; no rightmost-wins column dependence.
        self.assertEqual(out["mid_gl"]["8012345678"], gl)
        self.assertEqual(out["mid_gl"]["8098765432"], gl)
        # Two distinct GL strings in one row: genuinely ambiguous -> loud.
        gl2 = "02-7654321-654321-654321"
        p2 = os.path.join(self.a, "MID_Master_bad.xlsx")
        _write_xlsx(p2, [("Sheet1", [("8012345678", gl, gl2)])])
        with self.assertRaises(E.InvalidSourceData):
            E.load_mid_master(E.RoutedFile("MID_MASTER", p2, "MID_Master_bad.xlsx", "all"))

    def test_multi_file_newest_wins_and_tie_fails_loud(self):
        # Distinct date stamps: newest wins, runs clean.
        _build_fhb_utc_inputs(self.a)
        _build_fhb_utc_inputs(self.a, bsl_name="20230101_FHB_UTC_BSL_UNR.xlsx",
                              st_name="20230101_FHB_UTC_Account_ST.xlsx")
        runlog = E.run(self.a, self.oa, present=True)
        self.assertEqual(runlog["roles_bound"]["BSL"]["file"],
                         "20240101_FHB_UTC_BSL_UNR.xlsx")
        # Same date stamp on two BSL files: one would be silently ignored ->
        # must fail loud instead.
        _build_fhb_utc_inputs(self.b)
        _build_fhb_utc_inputs(self.b, bsl_name="20240101_v2_FHB_UTC_BSL_UNR.xlsx",
                              st_name="20240101_v2_FHB_UTC_Account_ST.xlsx")
        with self.assertRaises(E.InvalidSourceData):
            E.run(self.b, self.ob, present=True)


    def test_column_shuffle_end_to_end_identical(self):
        # The core guarantee: permuted columns + preamble rows + a decoy
        # constant-date column produce cell-for-cell identical output.
        account = _build_fhb_utc_inputs(self.a)
        base = E.run(self.a, self.oa, present=True)

        def _mutate(src, dst, perm, preamble, decoy_at):
            from openpyxl import load_workbook, Workbook
            wb = load_workbook(src)
            rows = [list(r) for r in wb.worksheets[0].iter_rows(values_only=True)]
            out = Workbook(); out.remove(out.active)
            ws = out.create_sheet("Exported")
            for p in preamble:
                ws.append(p)
            for i, r in enumerate(rows):
                new = [r[j] for j in perm]
                if decoy_at is not None:
                    new.insert(decoy_at, "As Of" if i == 0 else "2024-04-01")
                ws.append(new)
            out.save(dst)

        _mutate(os.path.join(self.a, "20240101_FHB_UTC_BSL_UNR.xlsx"),
                os.path.join(self.b, "20240101_FHB_UTC_BSL_UNR.xlsx"),
                [5, 3, 1, 0, 4, 2], [("University of Tennessee",), ()], 2)
        _mutate(os.path.join(self.a, "20240101_FHB_UTC_Account_ST.xlsx"),
                os.path.join(self.b, "20240101_FHB_UTC_Account_ST.xlsx"),
                [5, 4, 3, 2, 1, 0], [], None)
        mut = E.run(self.b, self.ob, present=True)

        self.assertEqual(mut["audit"]["status"], "PASS",
                         msg=str(mut["audit"].get("failures")))
        self.assertEqual(base["recon_summary"], mut["recon_summary"])
        ta, _ = A._read_output_tabs(base["recon_workbook"])
        tb, _ = A._read_output_tabs(mut["recon_workbook"])
        self.assertEqual(ta, tb)

    def test_all_data_never_loaded_and_never_crashes(self):
        # ALL_DATA is recognized but not loaded (forward-only engine): even
        # two undated All_Data workbooks (a tie the newest-wins check would
        # reject for loaded roles) must not abort the forward run.
        _build_fhb_utc_inputs(self.a)
        _write_xlsx(os.path.join(self.a, "FHB_UTC_All_Data.xlsx"),
                    [("Recon History", [("junk",)])])
        _write_xlsx(os.path.join(self.a, "FHB_UTC_v2_All_Data.xlsx"),
                    [("Recon History", [("junk",)])])
        runlog = E.run(self.a, self.oa, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS")
        self.assertIn("not loaded", runlog["all_data"])
        self.assertNotIn("ALL_DATA", runlog["roles_bound"])


    def test_audit_binds_same_columns_as_engine(self):
        # 'Post Date' header plus an inserted date-parseable decoy column:
        # both engine and audit must bind the exact-alias column -> PASS.
        bsl = [
            ("Trade Date", "Post Date", "Amount", "Line Number",
             "Account Servicer Reference", "Additional Information", "Transaction Code"),
            ("2024-03-09", "2024-03-10", "150.00", "1", "REF100200", "ACH CREDIT", "142"),
            ("2024-03-11", "2024-03-12", "300.00", "2", "GRP500600", "DEPOSIT", "174"),
        ]
        st = [
            ("Transaction Date", "Amount", "Transaction Number", "Source", "Reference", "Counterparty"),
            ("2024-03-09", "150.00", "ST1", "EXT", "REF100200", "VENDOR INC"),
            ("2024-03-11", "120.00", "ST2", "EXT", "GRP500600", "PAYER A"),
            ("2024-03-11", "180.00", "ST3", "EXT", "GRP500600", "PAYER A"),
        ]
        _write_xlsx(os.path.join(self.a, "20240101_FHB_UTC_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        _write_xlsx(os.path.join(self.a, "20240101_FHB_UTC_Account_ST.xlsx"),
                    [("Exported", st)])
        runlog = E.run(self.a, self.oa, present=True)
        self.assertEqual(runlog["roles_bound"]["BSL"]["columns"]["date"], 1)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))


class TestPerRunScript(unittest.TestCase):
    """run_recon.py: stage one upload -> immutable run folder -> engine."""

    def setUp(self):
        self.upload = tempfile.mkdtemp()
        self.runs = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.upload, ignore_errors=True)
        shutil.rmtree(self.runs, ignore_errors=True)

    def test_staged_name_prefix_rules(self):
        # Hex upload prefix stripped; plausible YYYYMMDD date prefix kept.
        self.assertEqual(R.staged_name("933782d6-FHB_UTC_BSL.xlsx"),
                         "FHB_UTC_BSL.xlsx")
        self.assertEqual(R.staged_name("20240101-FHB_UTC_BSL.xlsx"),
                         "20240101-FHB_UTC_BSL.xlsx")
        # All-digit but NOT a plausible date (month 20) -> an unlucky random
        # hex prefix; must be stripped or it poisons newest-wins ordering.
        self.assertEqual(R.staged_name("30201090-FHB_UTC_BSL.xlsx"),
                         "FHB_UTC_BSL.xlsx")
        self.assertEqual(R.staged_name("FHB_UTC_BSL.xlsx"), "FHB_UTC_BSL.xlsx")
        self.assertEqual(R.staged_name("933782d6-x.xlsx", strip_prefix=False),
                         "933782d6-x.xlsx")

    def test_perform_run_end_to_end(self):
        _build_fhb_utc_inputs(
            self.upload,
            bsl_name="933782d6-20240101_FHB_UTC_BSL_UNR.xlsx",
            st_name="ab12cd34-20240101_FHB_UTC_Account_ST.xlsx")
        # An unrouted spreadsheet and a non-spreadsheet ride along.
        _write_xlsx(os.path.join(self.upload, "random_notes.xlsx"),
                    [("Sheet1", [("a",)])])
        with open(os.path.join(self.upload, "readme.txt"), "w") as fh:
            fh.write("not a spreadsheet")

        code, report = R.perform_run([self.upload], runs_root=self.runs,
                                     run_id="run_test")
        self.assertEqual(code, 0)
        self.assertEqual(report["account"], "FHB_UTC")
        self.assertEqual(report["audit"], "PASS")

        run_dir = os.path.join(self.runs, "run_test")
        # Upload prefixes stripped in staging.
        self.assertTrue(os.path.exists(os.path.join(
            run_dir, "input", "20240101_FHB_UTC_BSL_UNR.xlsx")))
        self.assertTrue(os.path.exists(report["recon_workbook"]))
        self.assertTrue(os.path.exists(report["runlog"]))

        import json
        with open(os.path.join(run_dir, "manifest.json")) as fh:
            manifest = json.load(fh)
        roles = {e["staged_as"]: e["role"] for e in manifest["files"]}
        self.assertEqual(roles["20240101_FHB_UTC_BSL_UNR.xlsx"], "BSL")
        self.assertEqual(roles["20240101_FHB_UTC_Account_ST.xlsx"], "ST")
        self.assertEqual(roles["random_notes.xlsx"], "UNROUTED")
        self.assertTrue(all(len(e["sha256"]) == 64 for e in manifest["files"]))
        self.assertTrue(any("random_notes.xlsx" in w for w in manifest["warnings"]))
        self.assertEqual(len(manifest["ignored_non_spreadsheets"]), 1)
        cv = manifest["code_versions"]
        self.assertEqual(len(cv["recon_engine.py"]), 64)
        self.assertEqual(len(cv["recon_audit.py"]), 64)

    def test_individual_file_args_and_parent_dir_dedup(self):
        # Documented primary usage: file arguments, including a file passed
        # both explicitly and via its parent directory (one source, no
        # spurious self-collision).
        _build_fhb_utc_inputs(self.upload)
        bsl = os.path.join(self.upload, "20240101_FHB_UTC_BSL_UNR.xlsx")
        st = os.path.join(self.upload, "20240101_FHB_UTC_Account_ST.xlsx")
        code, report = R.perform_run([bsl, st, self.upload, bsl],
                                     runs_root=self.runs, run_id="rfiles")
        self.assertEqual(code, 0)
        self.assertEqual(report["audit"], "PASS")

    def test_audit_fail_exits_2_and_quarantines(self):
        # Force the independent audit to FAIL: perform_run must recover the
        # runlog (written before the gate raises), report FAIL, exit 2 —
        # and the workbooks stay on disk for forensics.
        import recon_audit
        _build_fhb_utc_inputs(self.upload)
        real_audit = recon_audit.audit
        recon_audit.audit = lambda *a, **k: {
            "status": "FAIL", "checks": {}, "failures": ["forced"]}
        try:
            self.assertEqual(
                R.main([self.upload, "--runs-root", self.runs,
                        "--run-id", "rfail"]), 2)
            out = os.path.join(self.runs, "rfail", "outputs")
            self.assertTrue(os.path.exists(
                os.path.join(out, "FHB_UTC_reconciliation.xlsx")))
            # Gate off: same failing audit -> exit 0, run completes.
            code, report = R.perform_run([self.upload], runs_root=self.runs,
                                         run_id="rfail2", present=False)
            self.assertEqual(code, 0)
            self.assertEqual(report["audit"], "FAIL")
        finally:
            recon_audit.audit = real_audit

    def test_no_strip_flag_via_cli(self):
        _build_fhb_utc_inputs(
            self.upload,
            bsl_name="933782d6-20240101_FHB_UTC_BSL_UNR.xlsx",
            st_name="ab12cd34-20240101_FHB_UTC_Account_ST.xlsx")
        self.assertEqual(
            R.main([self.upload, "--runs-root", self.runs, "--run-id", "rk",
                    "--no-strip-upload-prefix"]), 0)
        self.assertTrue(os.path.exists(os.path.join(
            self.runs, "rk", "input",
            "933782d6-20240101_FHB_UTC_BSL_UNR.xlsx")))

    def test_missing_bsl_fails_loud(self):
        _build_fhb_utc_inputs(self.upload)
        os.remove(os.path.join(self.upload, "20240101_FHB_UTC_BSL_UNR.xlsx"))
        with self.assertRaises(R.PerRunError):
            R.perform_run([self.upload], runs_root=self.runs, run_id="r1")

    def test_mixed_accounts_fail_loud(self):
        _build_fhb_utc_inputs(self.upload)
        _build_fhb_utc_inputs(self.upload,
                              bsl_name="20240101_Regions_UTM_BSL_UNR.xlsx",
                              st_name="20240101_Regions_UTM_Account_ST.xlsx")
        with self.assertRaises(R.PerRunError):
            R.perform_run([self.upload], runs_root=self.runs, run_id="r1")

    def test_mixed_account_st_alone_fails_loud(self):
        # A wrong-account ST with the right-account BSL would cross-pollute
        # the pool; the guard must span every routed file, not just BSL.
        _build_fhb_utc_inputs(self.upload,
                              st_name="20240101_Regions_UTM_Account_ST.xlsx")
        with self.assertRaises(R.PerRunError):
            R.perform_run([self.upload], runs_root=self.runs, run_id="r1")

    def test_case_only_collision_fails_loud(self):
        other = tempfile.mkdtemp()
        try:
            _build_fhb_utc_inputs(self.upload)
            _build_fhb_utc_inputs(other,
                                  bsl_name="20240101_fhb_utc_bsl_unr.xlsx",
                                  st_name="20240101_FHB_UTC_ACCOUNT_st.xlsx")
            with self.assertRaises(R.PerRunError):
                R.perform_run([self.upload, other], runs_root=self.runs,
                              run_id="r1")
        finally:
            shutil.rmtree(other, ignore_errors=True)

    def test_failed_preflight_frees_run_id(self):
        # An unusable upload must not leave a half-built folder or burn the id.
        _build_fhb_utc_inputs(self.upload)
        os.remove(os.path.join(self.upload, "20240101_FHB_UTC_BSL_UNR.xlsx"))
        with self.assertRaises(R.PerRunError):
            R.perform_run([self.upload], runs_root=self.runs, run_id="r1")
        self.assertFalse(os.path.exists(os.path.join(self.runs, "r1")))
        # Same id now works with a good upload.
        _build_fhb_utc_inputs(self.upload)
        code, _ = R.perform_run([self.upload], runs_root=self.runs,
                                run_id="r1")
        self.assertEqual(code, 0)

    def test_bad_run_ids_rejected(self):
        _build_fhb_utc_inputs(self.upload)
        for bad in ("..", ".", "a/b", ""):
            with self.assertRaises(R.PerRunError):
                R.perform_run([self.upload], runs_root=self.runs, run_id=bad)

    def test_nonexistent_path_fails_loud(self):
        with self.assertRaises(R.PerRunError):
            R.perform_run([os.path.join(self.upload, "nope.xlsx")],
                          runs_root=self.runs, run_id="r1")

    def test_subfolder_with_spreadsheets_warned(self):
        _build_fhb_utc_inputs(self.upload)
        nested = os.path.join(self.upload, "nested")
        os.makedirs(nested)
        _write_xlsx(os.path.join(nested, "20240201_FHB_UTC_BSL_UNR.xlsx"),
                    [("Exported", [("a",)])])
        code, report = R.perform_run([self.upload], runs_root=self.runs,
                                     run_id="rsub")
        self.assertEqual(code, 0)
        import json
        with open(os.path.join(report["run_dir"], "manifest.json")) as fh:
            manifest = json.load(fh)
        self.assertIn(nested, manifest["ignored_non_spreadsheets"])
        self.assertTrue(any("NOT read" in w for w in manifest["warnings"]))

    def test_collision_fails_loud(self):
        other = tempfile.mkdtemp()
        try:
            _build_fhb_utc_inputs(self.upload)
            _build_fhb_utc_inputs(other)  # same filenames, different folder
            with self.assertRaises(R.PerRunError):
                R.perform_run([self.upload, other], runs_root=self.runs,
                              run_id="r1")
        finally:
            shutil.rmtree(other, ignore_errors=True)

    def test_runs_are_immutable(self):
        _build_fhb_utc_inputs(self.upload)
        code, _ = R.perform_run([self.upload], runs_root=self.runs,
                                run_id="r1")
        self.assertEqual(code, 0)
        with self.assertRaises(R.PerRunError):
            R.perform_run([self.upload], runs_root=self.runs, run_id="r1")

    def test_main_exit_codes(self):
        # Empty upload -> exit 1 (fail loud, no run executed).
        empty = tempfile.mkdtemp()
        try:
            self.assertEqual(
                R.main([empty, "--runs-root", self.runs, "--run-id", "r1"]), 1)
        finally:
            shutil.rmtree(empty, ignore_errors=True)
        # Good upload -> exit 0.
        _build_fhb_utc_inputs(self.upload)
        self.assertEqual(
            R.main([self.upload, "--runs-root", self.runs, "--run-id", "r2"]), 0)


class TestRealDataShapes(unittest.TestCase):
    """Regression tests for the real FHB Master export shapes (2026-07-10
    validation run): router tokens, integer cells, NA placeholders, the MET
    scope join + ST bridge, and the ORT deposit-group pass."""

    def setUp(self):
        self.d = tempfile.mkdtemp()
        self.out = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.d, ignore_errors=True)
        shutil.rmtree(self.out, ignore_errors=True)

    def test_router_real_filenames(self):
        # "_st" must not swallow All_Status / Rosetta_Stone.
        cases = {
            "20260710_Oracle_CM_FHB_Master_BSL_UNR.xlsx": "BSL",
            "20260710_Oracle_CM_FHB_Master_ST_UNR.xlsx": "ST",
            "20260710_Oracle_OTBI_MET_All_Accounts_All_Status.xlsx": "MET",
            "20260710_FHB_Master_BAI2.xlsx": "BAI2",
            "ORT_Misc_All.xlsb": "ORT_MISC",
            "UT_Recon_SPN_Receivables_Rosetta_Stone_Part2_1.xlsx": "RELATIONSHIP_MAP",
            "UT_Recon_Data_Relationship_Map_20260702_1.xlsx": "RELATIONSHIP_MAP",
        }
        for name, want in cases.items():
            self.assertEqual(E.classify_file(name), want, msg=name)
        self.assertEqual(E.infer_account("20260710_FHB_Master_BAI2.xlsx"),
                         "FHB_MASTER")

    def test_integer_reference_cells_bind(self):
        # Oracle exports deliver references/ids as raw ints; parse_date must
        # not swallow small ints as Excel serials and poison the binder.
        self.assertTrue(E.pred_reference(1390))
        self.assertTrue(E.pred_reference("852256"))
        self.assertFalse(E.pred_reference("2026-07-09"))
        from datetime import datetime as _dt
        self.assertFalse(E.pred_reference(_dt(2026, 7, 9)))
        rows = [
            ("Date", "Amount (USD)", "Reference", "Transaction Number", "Source",
             "Payment Reference Number"),
            ("2026-07-10", "10.00", 1390, 17, "External", ""),
            ("2026-07-09", "20.00", 104000120, 1027519, "External", 260238),
        ]
        m, _hi = E.bind_columns(rows, E.ST_ROLES, filename="st.xlsx")
        self.assertEqual(m["transaction_number"], 3)
        self.assertEqual(m["reference"], 2)

    def test_na_placeholder_never_ties(self):
        self.assertEqual(E.clean_ref("NA"), "")
        self.assertEqual(E.clean_ref("n/a"), "")
        self.assertEqual(E.clean_ref("852256"), "852256")
        a = E._mk_entry("T1", 100, date(2026, 7, 1), "NA", "", "EXT", "UNR", True, "ST")
        self.assertEqual(a.reference, "")
        self.assertFalse(E.reference_equal("NA", a.reference))

    def test_account_of_bank_name(self):
        self.assertEqual(E.account_of_bank_name("FHB - Master Account"), "FHB_MASTER")
        self.assertEqual(E.account_of_bank_name("Regions - UTM"), "REGIONS_UTM")
        self.assertIsNone(E.account_of_bank_name("TRUIST BANK - Chattanooga"))

    def test_student_refund_accounts(self):
        # Owner (2026-07-19): Student Refund accounts are distinct depositories
        # and must win over the generic campus token (4-token match first).
        self.assertEqual(E.infer_account("20260719_Oracle_CM_FHB_Student_Refund_UTK_BSL_UNR.xlsx"),
                         "FHB_STUDENT_REFUND_UTK")
        self.assertEqual(E.infer_account("20260719_Oracle_CM_FHB_UTC_Student_Refund_BSL_UNR.xlsx"),
                         "FHB_STUDENT_REFUND_UTC")
        self.assertEqual(E.account_of_bank_name("FHB - Student Refund - UTK"),
                         "FHB_STUDENT_REFUND_UTK")
        self.assertEqual(E.account_of_bank_name("FHB - Student Refund - UTC"),
                         "FHB_STUDENT_REFUND_UTC")
        # a plain UTC file carries neither "student" nor "refund" -> FHB_UTC
        self.assertEqual(E.infer_account("20260716_Oracle_CM_FHB_UTC_BSL_UNR.xlsx"), "FHB_UTC")
        self.assertIsNone(E.account_of_bank_name("FHB - UTFI"))

    def test_check_conflict_bars_payables_candidate(self):
        # Orphan doctrine R8 end-to-end: a CHECK bank line whose only
        # same-amount open Payables ST carries a DIFFERENT check number is a
        # conflict — it must fall to Review, never an amount-only Candidate.
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            ("2026-07-13", "-150.00", "60410073", "CHECK", "60410073",
             "Check", "Line 12 , 2026-07-13", "475"),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        st = [
            ("Date", "Amount (USD)", "Reference", "Transaction Number", "Source", "Transaction Type", "Counterparty"),
            # same amount, DIFFERENT check number
            ("2026-07-13", "-150.00", "60400012", "60400012", "Payables", "Check", "BANNER VENDOR"),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        self.assertEqual(runlog["recon_summary"],
                         {"matches": 0, "candidates": 0, "misdirected": 0, "reviews": 1})

    def _met_rows(self):
        return [
            ("CBE_BANK_ACCOUNT_NAME", "CET_REFERENCE_TEXT", "CET_STATUS",
             "CET_TRANSACTION_DATE", "CET_TRANSACTION_ID", "AMOUNT",
             "TRANSACTION_DATE", "CET_DESCRIPTION", "DEPOSIT_ID", "RECEIPT_ID"),
            # our account: deposit 900 = two open receipts 400 + 600
            ("FHB - Master Account", "DEPREF77", "UNR", "2026-07-01", 501,
             "400.00", "2026-07-01", "d:900 | r:11 | PAYER A", 900, 11),
            ("FHB - Master Account", "DEPREF77", "UNR", "2026-07-01", 502,
             "600.00", "2026-07-01", "d:900 | r:12 | PAYER A", 900, 12),
            # same trx id as the ST export -> must bridge, not duplicate
            ("FHB - Master Account", "REF100200", "UNR", "2026-07-02", 601,
             "150.00", "2026-07-02", "d:901 | r:13 | VENDOR", 901, 13),
            # other account: identical amounts — must never enter the pool
            ("FHB - UTC", "DEPREF77", "UNR", "2026-07-01", 701,
             "1000.00", "2026-07-01", "d:902 | r:14 | PAYER A", 902, 14),
        ]

    def _build(self, bsl_rows):
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl_rows)])
        st = [
            ("Date", "Amount (USD)", "Reference", "Transaction Number", "Source", "Counterparty"),
            ("2026-07-02", "150.00", "REF100200", 601, "External", "VENDOR INC"),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        _write_xlsx(os.path.join(self.d, "20260710_MET_All_Accounts.xlsx"),
                    [("Miscellaneous External Transact", self._met_rows())])

    def test_met_scope_bridge_and_deposit_group(self):
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            # bundled deposit line: no reference, sums d:900 (400+600)
            ("2026-07-03", "1000.00", "NA", "DEPOSIT 011", "NA",
             "Miscellaneous", "Line 1 , 2026-07-03", "174"),
            # exact 1:1 to the bridged ST
            ("2026-07-03", "150.00", "REF100200", "ACH CREDIT", "REF100200",
             "Automated clearing house", "Line 2 , 2026-07-03", "142"),
        ]
        self._build(bsl)
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        # scope: 3 of 4 MET rows are ours; the FHB-UTC row is excluded
        self.assertEqual(runlog["met_scope"],
                         {"rows_total": 4, "rows_in_account": 3, "key": "bank_name"})
        # bridge: trx 601 exists in both ST and MET -> one pool entry.
        # The FHB-UTC row joins as a foreign SHADOW entry (misdirected
        # guardrail 2026-07-18) — in the pool, never matchable.
        self.assertEqual(runlog["met_bridged_to_st"], 1)
        self.assertEqual(runlog["pool_total"], 4)
        self.assertEqual(runlog["met_foreign_account_open_rows"], 1)
        # ACH line -> P3 exact Match; deposit line -> deposit-type consistency
        # confers NOTHING (owner doctrine 2026-07-11): the exact sum alone
        # surfaces as an amount-only Candidate, never a Match.
        self.assertEqual(runlog["recon_summary"],
                         {"matches": 1, "candidates": 1, "misdirected": 0, "reviews": 0})
        self.assertEqual(runlog["forward_pass_counts"].get("P4_deposit_group"), 1)
        tabs_c, _ = A._read_output_tabs(runlog["recon_workbook"])
        cand = [r for r in tabs_c["Candidate Matches"][3:] if any(A._N(c) for c in r)]
        self.assertIn("AMOUNT_ONLY_GROUP", A._N(cand[0][A.COL_EXPL]))
        # the bridged entry carries the ORT coordinates
        tabs, _ = A._read_output_tabs(runlog["recon_workbook"])
        match_rows = [r for r in tabs["Matches"][3:] if any(A._N(c) for c in r)]
        by_amount = {A._N(r[2]): r for r in match_rows}
        self.assertEqual(A._N(by_amount["150.00"][A.COL_DEP]), "901")  # ORT d:
        self.assertEqual(A._N(by_amount["150.00"][A.COL_REC]), "13")   # ORT r:

    def test_met_gl_fallback_scope(self):
        # COA scope key (owner, 2026-07-18): when the MET export lacks the
        # bank-name column, the asset combo's GL cash account scopes the
        # pool — a cross-account row (UTC GL 100310) must never enter.
        met = [
            ("ASSET_CONCATENATED_SEGMENTS", "CET_REFERENCE_TEXT", "CET_STATUS",
             "CET_TRANSACTION_DATE", "CET_TRANSACTION_ID", "AMOUNT",
             "TRANSACTION_DATE", "CET_DESCRIPTION", "DEPOSIT_ID", "RECEIPT_ID"),
            ("01-1100001-000000-100210-000-0000-00-0000", "DEPREF77", "UNR",
             "2026-07-01", 501, "400.00", "2026-07-01", "d:900 | r:11 | PAYER A", 900, 11),
            ("01-1100001-000000-100210-000-0000-00-0000", "DEPREF77", "UNR",
             "2026-07-01", 502, "600.00", "2026-07-01", "d:900 | r:12 | PAYER A", 900, 12),
            ("01-1100001-000000-100210-000-0000-00-0000", "REF100200", "UNR",
             "2026-07-02", 601, "150.00", "2026-07-02", "d:901 | r:13 | VENDOR", 901, 13),
            ("01-1100001-000000-100310-000-0000-40-0000", "DEPREF77", "UNR",
             "2026-07-01", 701, "1000.00", "2026-07-01", "d:902 | r:14 | PAYER A", 902, 14),
        ]
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            ("2026-07-03", "150.00", "REF100200", "ACH CREDIT", "REF100200",
             "Automated clearing house", "Line 2 , 2026-07-03", "142"),
        ]
        self._build(bsl)
        _write_xlsx(os.path.join(self.d, "20260710_MET_All_Accounts.xlsx"),
                    [("Miscellaneous External Transact", met)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        self.assertEqual(runlog["met_scope"],
                         {"rows_total": 4, "rows_in_account": 3, "key": "gl_cash_account"})
        # The UTC-GL row is excluded from matching but joins the misdirected
        # shadow pool (foreign_account set, available=False).
        self.assertEqual(runlog["pool_total"], 4)
        self.assertEqual(runlog["met_foreign_account_open_rows"], 1)

    def test_met_status_outranks_st_export(self):
        # Orphan doctrine R3 (owner, 2026-07-19): if MET says REC for the
        # bridged transaction, the open-looking ST is consumed — never match
        # it, whatever the (stale) ST export says.
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            ("2026-07-03", "150.00", "REF100200", "ACH CREDIT", "REF100200",
             "Automated clearing house", "Line 2 , 2026-07-03", "142"),
        ]
        self._build(bsl)
        rows = self._met_rows()
        # flip the bridged trx (601) to REC in MET while ST export says open
        rows[3] = ("FHB - Master Account", "REF100200", "REC", "2026-07-02", 601,
                   "150.00", "2026-07-02", "d:901 | r:13 | VENDOR", 901, 13)
        _write_xlsx(os.path.join(self.d, "20260710_MET_All_Accounts.xlsx"),
                    [("Miscellaneous External Transact", rows)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        self.assertEqual(runlog["met_status_overrides"], 1)
        self.assertEqual(runlog["recon_summary"]["matches"], 0)
        tabs, _ = A._read_output_tabs(runlog["recon_workbook"])
        rev = [r for r in tabs["Review Notes"][3:] if any(A._N(c) for c in r)]
        blob = " ".join(A._N(r[A.COL_EXPL]) for r in rev)
        self.assertIn("already-reconciled", blob.lower())

    def test_r3_override_requires_amount_equality(self):
        # R3 must NOT close an open ST when the bridged MET row carries a
        # DIFFERENT amount (keep-largest kept a REC total row while the ST
        # export carries only the open split, or a CET-id collision).  The
        # 150.00 ST stays open and matches its bank line.
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            ("2026-07-03", "150.00", "REF100200", "ACH CREDIT", "REF100200",
             "Automated clearing house", "Line 2 , 2026-07-03", "142"),
        ]
        self._build(bsl)
        rows = self._met_rows()
        # bridged trx 601: MET says REC but at a DIFFERENT amount (300.00)
        rows[3] = ("FHB - Master Account", "REF100200", "REC", "2026-07-02", 601,
                   "300.00", "2026-07-02", "d:901 | r:13 | VENDOR", 901, 13)
        _write_xlsx(os.path.join(self.d, "20260710_MET_All_Accounts.xlsx"),
                    [("Miscellaneous External Transact", rows)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        self.assertIsNone(runlog.get("met_status_overrides"))  # amount mismatch: no override
        self.assertEqual(runlog["recon_summary"]["matches"], 1)  # 150.00 ST still open

    def test_dual_fire_twin_never_matches_twice(self):
        # Orphan doctrine 2.3: two identical open STs sharing one MET
        # RECEIPT_ID are a dual-fire pair — exactly one stays available.
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            ("2026-07-03", "150.00", "REF100200", "ACH CREDIT", "REF100200",
             "Automated clearing house", "Line 2 , 2026-07-03", "142"),
        ]
        self._build(bsl)
        rows = self._met_rows()
        # a dual-fire twin of trx 601: different trx id, same receipt r:13
        rows.append(("FHB - Master Account", "REF100200", "UNR", "2026-07-02", 602,
                     "150.00", "2026-07-02", "d:901 | r:13 | VENDOR", 901, 13))
        _write_xlsx(os.path.join(self.d, "20260710_MET_All_Accounts.xlsx"),
                    [("Miscellaneous External Transact", rows)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        self.assertEqual(runlog["dual_fire_twins"], 1)
        # the line still matches — but exactly one of the twins is cited
        self.assertEqual(runlog["recon_summary"]["matches"], 1)

    def test_check_rail_number_conflict(self):
        # Orphan doctrine R8: on check rails the check number is the identity;
        # same amount + different check number is a conflict, never a
        # candidate (the FHB AP $1,100 cascade).
        class _B:
            transaction_type = "Check"
            reference_raw = "0006789599"
            recon_reference = "6789599"
        b = _B()
        other = E._mk_entry("901", 110000, date(2026, 7, 1), "6789601", "", "AP", "UNR", True, "ST")
        same = E._mk_entry("902", 110000, date(2026, 7, 1), "0006789599", "", "AP", "UNR", True, "ST")
        blank = E._mk_entry("903", 110000, date(2026, 7, 1), "", "", "AP", "UNR", True, "ST")
        arrcpt = E._mk_entry("904", 110000, date(2026, 7, 1), "6789601", "", "AR", "UNR", True, "RECEIPTS")
        self.assertTrue(E._check_number_conflict(b, other))   # different check no. (AP)
        self.assertFalse(E._check_number_conflict(b, same))   # same check (zero-padded)
        self.assertFalse(E._check_number_conflict(b, blank))  # silence never conflicts
        self.assertFalse(E._check_number_conflict(b, arrcpt)) # AR receipt is not a check rail
        class _B2:
            transaction_type = "Automated clearing house"
            reference_raw = "0006789599"
            recon_reference = "6789599"
        self.assertFalse(E._check_number_conflict(_B2(), other))  # not a check rail

    def test_cherry_pick_split_review(self):
        # Orphan doctrine signature #6: no open counterpart at exact cents,
        # but an ALL-closed MET deposit sums exactly — enriched Review naming
        # the deposit, never a forced match.
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            ("2026-07-03", "1134.00", "NA", "DEPOSIT 099", "NA",
             "Miscellaneous", "Line 9 , 2026-07-03", "174"),
        ]
        self._build(bsl)
        met = [
            ("CBE_BANK_ACCOUNT_NAME", "CET_REFERENCE_TEXT", "CET_STATUS",
             "CET_TRANSACTION_DATE", "CET_TRANSACTION_ID", "AMOUNT",
             "TRANSACTION_DATE", "CET_DESCRIPTION", "DEPOSIT_ID", "RECEIPT_ID"),
            ("FHB - Master Account", "DEPREF88", "REC", "2026-07-01", 801,
             "400.00", "2026-07-01", "d:910 | r:21 | PAYER X", 910, 21),
            ("FHB - Master Account", "DEPREF88", "REC", "2026-07-01", 802,
             "600.00", "2026-07-01", "d:910 | r:22 | PAYER X", 910, 22),
            ("FHB - Master Account", "DEPREF88", "REC", "2026-07-01", 803,
             "134.00", "2026-07-01", "d:910 | r:23 | PAYER X", 910, 23),
            # bridged open ST for the other line
            ("FHB - Master Account", "REF100200", "UNR", "2026-07-02", 601,
             "150.00", "2026-07-02", "d:901 | r:13 | VENDOR", 901, 13),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_MET_All_Accounts.xlsx"),
                    [("Miscellaneous External Transact", met)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        tabs, _ = A._read_output_tabs(runlog["recon_workbook"])
        rev = [r for r in tabs["Review Notes"][3:] if any(A._N(c) for c in r)]
        target = [r for r in rev if A._N(r[2]).replace(",", "") in ("1134.00", "1 134.00")]
        self.assertEqual(len(target), 1)
        expl = A._N(target[0][A.COL_EXPL])
        self.assertIn("d:910", expl)
        self.assertIn("cherry-picked", expl)
        self.assertIn("POSSIBLE_AUTO_REC_SPLIT", A._N(target[0][A.COL_EXPL]) + " " + expl)

    def test_payments_feed_matches_check(self):
        # PAYMENTS feed (owner, 2026-07-19): an open (Negotiable) AP payment
        # becomes a negative AP pool entry that matches its check bank line;
        # Cleared/reconciled payments are not pooled.
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            ("2026-07-13", "-500.00", "5924", "EPAY 5924", "5924",
             "Check", "Line 1 , 2026-07-13", "475"),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        st = [
            ("Date", "Amount (USD)", "Reference", "Transaction Number", "Source", "Transaction Type"),
            ("2026-07-01", "999.00", "Z1", "Z1", "External", "ACH"),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        pay = [
            ("Payment Number", "Payment Status", "Reconciled", "Payment Amount", "Payee", "Payment Date"),
            ("5924", "Negotiable", "No", "500.00", "ACME VENDOR", "2026-07-13"),
            ("6000", "Cleared", "Yes", "500.00", "OTHER VENDOR", "2026-07-10"),   # reconciled: not pooled
        ]
        _write_xlsx(os.path.join(self.d, "20260719_Oracle_Payables_Payments.xlsx"),
                    [("Exported", pay)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        self.assertEqual(runlog["pool_sizes"].get("PAYMENTS"), 1)  # only the open one
        self.assertEqual(runlog["recon_summary"]["matches"], 1)    # -500 check <-> payment 5924

    def test_reverse_misdirected_from_all_bsl(self):
        # ALL_BSL reverse search (owner, 2026-07-19): a THIS-account open
        # receipt whose bank line landed in ANOTHER account is a read-only,
        # ST-anchored finding (never a workbook placement).
        rows = [
            ("CBA_BANK_ACCOUNT_NAME", "Amount", "CSL_RECON_REFERENCE", "CSL_BOOKING_DATE"),
            ("FHB - UTHSC", "70992.66", "300045836", "2026-07-03"),
            ("FHB - Master Account", "500.00", "999999", "2026-07-01"),  # same account: ignored
        ]
        m, hi = E.bind_columns(rows, E.ALL_BSL_ROLES, filename="all_bsl.xlsx")
        loaded = {"ALL_BSL": {"rows": rows, "map": m, "header_index": hi}}
        e = E._mk_entry("300045836", 7099266, date(2026, 7, 3), "300045836",
                        "City of Memphis", "AR", "UNR", True, "RECEIPTS")
        out = E._reverse_misdirected([e], E.Ledger(), loaded, "FHB_MASTER")
        self.assertEqual(len(out), 1)
        self.assertEqual(out[0]["landed_in"], "FHB_UTHSC")
        self.assertEqual(out[0]["st_id"], "300045836")
        # a receipt with no foreign match yields nothing
        e2 = E._mk_entry("111", 12345, date(2026, 7, 3), "111", "X", "AR", "UNR", True, "RECEIPTS")
        self.assertEqual(E._reverse_misdirected([e2], E.Ledger(), loaded, "FHB_MASTER"), [])

    def test_misdirected_receipt_gets_own_tab(self):
        # Owner HARD GUARDRAIL (2026-07-18, City of Memphis $70,992.66): the
        # bank line lands in THIS account but the receipt remits to ANOTHER
        # bank account, so no ST exists here.  With exact cents + reference
        # tie the pair surfaces on the dedicated Misdirected tab (never a
        # Match/Candidate); amount-only foreign coincidences stay Review.
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            ("2026-07-03", "70992.66", "300045836",
             "CITY OF MEMPHIS PAYABLES 260703300045836", "300045836",
             "Automated clearing house", "Line 26 , 2026-07-03", "142"),
        ]
        _write_xlsx(os.path.join(self.d, "20260718_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        st = [
            ("Date", "Amount (USD)", "Reference", "Transaction Number", "Source", "Counterparty"),
            ("2026-07-02", "150.00", "REF100200", 601, "External", "VENDOR INC"),
        ]
        _write_xlsx(os.path.join(self.d, "20260718_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        receipts = [
            ("Receipt Number", "Status", "Entered Amount", "Customer Name",
             "Receipt Date", "Remittance Bank Account", "Reference"),
            # the misdirected receipt: remits to FHB - UTC, ref ties the BSL
            ("300045836", "Remitted", "70992.66", "City of Memphis",
             "2026-07-03", "FHB - UTC", "300045836"),
            # a same-amount foreign receipt with NO tie: must NOT surface
            ("888777666", "Remitted", "70992.66", "Someone Else",
             "2026-07-03", "FHB - UTHSC", "999111"),
        ]
        _write_xlsx(os.path.join(self.d, "20260718_Oracle_Receivables_Receipts.xlsx"),
                    [("Export to Excel", receipts)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        self.assertEqual(runlog["recon_summary"],
                         {"matches": 0, "candidates": 0, "misdirected": 1, "reviews": 0})
        tabs, _ = A._read_output_tabs(runlog["recon_workbook"])
        mis = [r for r in tabs["Misdirected"][3:] if any(A._N(c) for c in r)]
        self.assertEqual(len(mis), 1)
        expl = A._N(mis[0][A.COL_EXPL])
        self.assertIn("booked to FHB_UTC", expl)
        self.assertIn("300045836", A._N(mis[0][A.COL_ST_NUMS]))

    def test_misdirected_common_name_tie_is_not_placed(self):
        # Distinctive-tie guard (owner HARD GUARDRAIL, 2026-07-21): a foreign
        # receipt whose ONLY tie to the bank line is a COMMON institutional name
        # ("University of Tennessee", boilerplate on nearly every UT line) plus
        # the amount is amount-only in disguise (rule 4 / 8g) — it must NOT be
        # placed on the Misdirected tab.  Real false case: a $10,000 Bill.com
        # receivable rerouted to FHB_UTC via "University of Te" containment.
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Customer Reference", "Transaction Type", "Statement", "Transaction Code"),
            ("2026-07-16", "10000.00", "9961518",
             "University of Te Receivable 260716 996YUENTH151R", "University of Te",
             "Automated clearing house", "Line 74 , 2026-07-16", "142"),
        ]
        _write_xlsx(os.path.join(self.d, "20260718_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        # A realistic population where "University of Tennessee" is boilerplate
        # (on many entries) — that is what makes the name a common, non-
        # distinctive carrier (freq > 3), exactly as in the real Master pool.
        st = [
            ("Date", "Amount (USD)", "Reference", "Transaction Number", "Source", "Counterparty"),
            ("2026-07-02", "150.00", "REF100200", 601, "External", "VENDOR INC"),
            ("2026-07-02", "11.00", "R2", 602, "AR", "University of Tennessee Knoxville"),
            ("2026-07-02", "12.00", "R3", 603, "AR", "University of Tennessee Martin"),
            ("2026-07-02", "13.00", "R4", 604, "AR", "University of Tennessee HSC"),
            ("2026-07-02", "14.00", "R5", 605, "AR", "University of Tennessee Foundation"),
        ]
        _write_xlsx(os.path.join(self.d, "20260718_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        receipts = [
            ("Receipt Number", "Status", "Entered Amount", "Customer Name",
             "Receipt Date", "Remittance Bank Account", "Reference"),
            # foreign receipt: amount matches, but ties ONLY on the common name
            # (payer "University of Tennessee Research Foundation") — no shared
            # numeric reference (its own ref is unrelated).
            ("SPN040226", "Remitted", "10000.00",
             "University of Tennessee Research Foundation", "2026-07-16",
             "FHB - UTC", "8042156011"),
        ]
        _write_xlsx(os.path.join(self.d, "20260718_Oracle_Receivables_Receipts.xlsx"),
                    [("Export to Excel", receipts)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        # The common-name coincidence is NOT a misdirected placement.
        self.assertEqual(runlog["recon_summary"]["misdirected"], 0,
                         msg="common institutional-name tie must not reroute money")

    def test_account_of_gl_segments(self):
        self.assertEqual(E.account_of_gl_segments("01-1100001-000000-100221-000-0000-00-0000"), "FHB_UTIA")
        self.assertEqual(E.account_of_gl_segments("01-1100001-000000-100330-000-0000-00-0000"), "REGIONS_UTM")
        self.assertEqual(E.account_of_gl_segments("01-1100001-000000-100500-000-0000-70-0000"), "FHB_UTHSC")
        self.assertIsNone(E.account_of_gl_segments("01-1100001-000000-100928-000-0000-00-0000"))  # clearing GL: unmapped
        self.assertIsNone(E.account_of_gl_segments("100221"))          # malformed combo
        self.assertIsNone(E.account_of_gl_segments(None))

    def test_deposit_auto_rec_split_is_candidate(self):
        rows = self._met_rows()
        rows[2] = ("FHB - Master Account", "DEPREF77", "REC", "2026-07-01", 502,
                   "600.00", "2026-07-01", "d:900 | r:12 | PAYER A", 900, 12)
        _write_xlsx(os.path.join(self.d, "20260710_MET_All_Accounts.xlsx"),
                    [("Miscellaneous External Transact", rows)])
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            ("2026-07-03", "1000.00", "NA", "DEPOSIT 011", "NA",
             "Miscellaneous", "Line 1 , 2026-07-03", "174"),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["recon_summary"]["candidates"], 1)
        tabs, _ = A._read_output_tabs(runlog["recon_workbook"])
        cand = [r for r in tabs["Candidate Matches"][3:] if any(A._N(c) for c in r)]
        self.assertIn("already-closed", A._N(cand[0][A.COL_EXPL]))


    def test_router_token_hardening(self):
        # Short bare tokens match whole name segments only.
        self.assertIsNone(E.classify_file("Market_Analysis.xlsx"))   # 'ar' in 'Market'
        self.assertIsNone(E.classify_file("Smart_Data.xlsx"))
        self.assertEqual(E.classify_file("ORT_Departments.xlsx"), "DEPT_INFO")
        self.assertEqual(E.classify_file("ORT_Chart_Of_Accounts.xlsx"), "CHART_OF_ACCOUNTS")
        self.assertEqual(E.classify_file("ORT_AR_All.xlsb"), "ORT_AR")
        self.assertEqual(E.classify_file("UT_MID_Master_Consolidated.xlsx"), "MID_MASTER")
        self.assertEqual(E.classify_file("20260710_FHB_Master_BAI2.xlsx"), "BAI2")  # bai+digits

    def test_bai2_enrichment(self):
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            ("2026-07-03", "150.00", "NA", "MERCHANT SERVICEDEPOSIT", "NA",
             "Automated clearing house", "Line 1 , 2026-07-03", "142"),
            ("2026-07-04", "75.00", "NA", "AMBIG", "NA",
             "Automated clearing house", "Line 2 , 2026-07-04", "142"),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        bai = [
            ("Post Date", "Transaction Description", "Amount", "Debit/Credit",
             "Bank Reference", "Customer Reference", "BAI Code", "DETAIL1", "DETAIL2"),
            ("2026-07-03", "ACH CREDIT RECEIVED", "150.00", "Credit",
             "ACH123", "8036121500", "142", "MERCHANT SERVICEDEPOSIT   26", "FULL ADDENDA TEXT"),
            # two BAI2 rows tie on (date, cents) with unrelated details: never guess
            ("2026-07-04", "X", "75.00", "Credit", "B1", "", "142", "ROW ONE", ""),
            ("2026-07-04", "Y", "75.00", "Credit", "B2", "", "142", "ROW TWO", ""),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BAI2.xlsx"),
                    [("CSVEXP", bai)])
        st = [("Date", "Amount (USD)", "Reference", "Transaction Number", "Source", "Counterparty"),
              ("2026-07-02", "150.00", "8036121500", 601, "External", "V")]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["bai2_enrichment"],
                         {"joined": 1, "ambiguous": 1, "no_hit": 0})
        # The BAI2 Customer Reference (a MID) reached the line's addenda:
        # the line classifies MERCHANT and the MID ties the ST reference.
        self.assertEqual(runlog["bsl_by_lane"].get("MERCHANT"), 1)

    def test_mid_directory(self):
        _write_xlsx(os.path.join(self.d, "ORT_Departments.xlsx"),
                    [("Report (3)", [
                        ("Campus Name", "Department Name", "Campus Bank Account",
                         "Dept Bank Account", "Campus Bank Name", "Dept Bank Name", "Credit Card Mid"),
                        ("Knoxville", "Ticket Office", "1", "2",
                         "FHB - Master Account", "FHB - UTIA", "8036121500"),
                        ("Knoxville", "No Mid Dept", "1", "2",
                         "FHB - Master Account", "FHB - Master Account", "N/A"),
                    ])])
        rf = E.RoutedFile("DEPT_INFO", os.path.join(self.d, "ORT_Departments.xlsx"),
                          "ORT_Departments.xlsx", "Report")
        rows, _ = E.read_rows(rf)
        m, hi = E.bind_columns(rows, E.DEPT_INFO_ROLES, filename=rf.filename)
        loaded = {"DEPT_INFO": {"rows": rows, "map": m, "header_index": hi}}
        d = E._mid_directory(loaded, "FHB_MASTER")
        self.assertEqual(d["8036121500"]["department"], "Ticket Office")
        self.assertEqual(d["8036121500"]["home_account"], "FHB_UTIA")
        self.assertEqual(len(d), 1)  # N/A row never enters




    def test_directional_date_doctrine(self):
        # BSL may PRECEDE the ST by any amount (entry lag, no ceiling);
        # BSL trailing the ST beyond the band is stale — never a Match.
        self.assertTrue(E.date_ok_directional(-300))   # ST after BSL: valid
        self.assertTrue(E.date_ok_directional(0))
        self.assertTrue(E.date_ok_directional(7))
        self.assertFalse(E.date_ok_directional(8))     # ST 8+ days before BSL
        self.assertFalse(E.date_ok_directional(300))
        self.assertTrue(E.date_ok_directional(None))   # unknown dates not gated

    def test_type_gate(self):
        b = E.make_bsl("1", date(2026, 7, 1), 1000, "R", "R", "", "Miscellaneous", "399")
        cc = E._mk_entry("S1", 1000, date(2026, 7, 1), "R", "", "EXT", "UNR",
                         True, "ST", transaction_type="Credit Card")
        eft = E._mk_entry("S2", 1000, date(2026, 7, 1), "R", "", "EXT", "UNR",
                          True, "ST", transaction_type="EFT")
        self.assertFalse(E._type_gate_ok(b, cc))
        self.assertTrue(E._type_gate_ok(b, eft))


def _write_csv(path, rows):
    import csv
    with open(path, "w", newline="", encoding="utf-8") as fh:
        csv.writer(fh).writerows(rows)


class TestChartOfAccounts(unittest.TestCase):
    """Tier-1 Chart of Accounts decode (owner COA export, 2026-07-19): a
    loaded, advisory human-label decoder consumed only in the non-gating
    Review / recommended-GL text.  Placements stay byte-identical with and
    without the CoA bundle (campus/entity consistency confers nothing —
    rule 8c)."""

    def setUp(self):
        self.d = tempfile.mkdtemp()
        self.out = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.d, ignore_errors=True)
        shutil.rmtree(self.out, ignore_errors=True)

    # ---- decode helpers ------------------------------------------------
    def test_segment_helpers(self):
        combo = "01-1100001-123456-100210-000-0000-40-0000"
        self.assertEqual(E.segments_of(combo),
                         ["01", "1100001", "123456", "100210", "000", "0000", "40", "0000"])
        self.assertEqual(E.segments_of(""), [])
        self.assertEqual(E.dept_segment_of(combo), "123456")
        self.assertEqual(E.entity_segment_of(combo), "01")
        self.assertIsNone(E.dept_segment_of("01-11"))
        self.assertIsNone(E.entity_segment_of(""))
        # account_of_gl_segments still reads position 4 (behavior unchanged).
        self.assertEqual(E.account_of_gl_segments(combo), "FHB_MASTER")

    def test_coa_decode_and_valid(self):
        combo = "01-1100001-123456-100210-000-0000-00-0000"
        coa = {
            "combo_decode": {combo: {
                "ent": "01", "ent_desc": "UT System",
                "fund": "1100001", "dept": "123456", "dep_desc": "Test Department",
                "account": "100210", "act_desc": "FHB Master Account",
                "intercompany": "00", "itc_desc": "Default Intercompany",
                "act_grp_desc": "TTL Assets"}},
            "entity_desc": {"01": "UT System", "40": "UT Chattanooga"},
            "postable_efdp": {"01-1100001-123456-000"},
        }
        d = E.coa_decode(combo, coa)
        self.assertEqual(d["dep_desc"], "Test Department")
        self.assertEqual(d["ent_desc"], "UT System")
        # CoA absent, or too-short combo -> None (callers no-op).
        self.assertIsNone(E.coa_decode(combo, None))
        self.assertIsNone(E.coa_decode("01-11", coa))
        # unknown combo falls back to per-segment entity/intercompany labels.
        miss = E.coa_decode("40-2200002-999999-100310-000-0000-01-0000", coa)
        self.assertEqual(miss["ent_desc"], "UT Chattanooga")
        self.assertEqual(miss["dep_desc"], "")
        # validity: recognized + postable for the known combo.
        v = E.coa_combo_valid(combo, coa)
        self.assertTrue(v["recognized"] and v["postable"])
        v2 = E.coa_combo_valid("40-2200002-999999-100310-000-0000-01-0000", coa)
        self.assertFalse(v2["recognized"])
        self.assertFalse(v2["postable"])
        self.assertEqual(E.coa_combo_valid(combo, None),
                         {"recognized": False, "postable": False})

    # ---- loader + routing ---------------------------------------------
    def test_router_coa_structural_files(self):
        for name in ("AcctCombos_base.csv", "AcctCombos_6.csv", "Segments.csv",
                     "ComboSets.xlsx", "CombosTech_UTSystem.xlsx",
                     "RelatedValueSets.csv",
                     # the REAL Oracle export names (owner CoA bundle,
                     # 2026-07-20) — previously routed to None and never loaded.
                     "COA_Account_Combinations_COA_Account_Combinations.csv",
                     "COA_Combination_Sets_UTHSC.xlsx",
                     "COA_Combos_Technical_UT_System.xlsx",
                     "COA_Segments_COA_Segments.csv",
                     "COA_Related_Value_Sets_COA_Related_Value_Sets.csv",
                     # a CoA campus name embedding a stray "met" segment must
                     # NOT bind MET (owner, 2026-07-21 — "Vet_Med" mis-typed
                     # "Vet_Met" leaked the UTIA shard into the MET pool and off
                     # the CoA decode).
                     "COA_Combos_Technical_UTIA_Vet_Met_Ag_Research_UT_Extension.xlsx",
                     "COA_Combination_Sets_UTIA_Vet_Med_Ag_Research_UT_Extension.xlsx"):
            self.assertEqual(E.classify_file(name), "CHART_OF_ACCOUNTS", msg=name)
        # A genuine MET export still routes to MET.
        self.assertEqual(E.classify_file("20260720_Oracle_OTBI_MET_FHB_Master.xlsx"), "MET")
        self.assertEqual(E.classify_file("MET_FHB_UTIA_2.csv"), "MET")
        # ORT_Department_* stays DEPT_INFO (bound before CHART_OF_ACCOUNTS).
        self.assertEqual(E.classify_file("ORT_Department_Info.xlsx"), "DEPT_INFO")

    def _write_coa_bundle(self, d):
        _write_csv(os.path.join(d, "AcctCombos_base.csv"), [
            ["ENT_DESC", "FND_DESC", "DEP_DESC", "PGM_DESC", "ACCOUNT_COMBO",
             "ACT_DESC", "ATV_DESC", "ITC_DESC", "ACT_GRP_DESC"],
            ["01-UT System", "1100001-E&G Funds", "123456-Test Department",
             "000-Default Program", "01-1100001-123456-100210-000-0000-00-0000",
             "100210-FHB Master Account", "0000-Default Activity",
             "00-Default Intercompany", "1ZZZZZ-TTL Assets"],
        ])
        _write_csv(os.path.join(d, "Segments.csv"), [
            ["ENABLED_FLAG", "SUMMARY_FLAG", "DESCRIPTION", "SEGMENT_TYPE",
             "START_DATE", "END_DATE", "VALUE"],
            ["Y", "N", "UT System", "Entity", "01/01/2019", "12/31/4712", "01"],
            ["Y", "N", "UT Chattanooga", "Entity", "01/01/2019", "12/31/4712", "40"],
            ["N", "N", "Retired Code", "Entity", "01/01/2019", "12/31/4712", "99"],
        ])
        _write_xlsx(os.path.join(d, "ComboSets.xlsx"), [("Sheet1", [
            ("COA Combination Sets", None),
            ("Run Date:\xa0Jul 18, 2026", None),
            ("Combination Set", "Entity"),
            ("01-1100001-123456-000", "01-UT System"),
            ("1 of 1", None),
        ])])

    def test_load_chart_of_accounts(self):
        self._write_coa_bundle(self.d)
        files = [E.RoutedFile("CHART_OF_ACCOUNTS", os.path.join(self.d, n), n, "Report")
                 for n in ("AcctCombos_base.csv", "Segments.csv", "ComboSets.xlsx")]
        coa = E.load_chart_of_accounts(files)
        self.assertIn("01-1100001-123456-100210-000-0000-00-0000", coa["combo_decode"])
        dec = coa["combo_decode"]["01-1100001-123456-100210-000-0000-00-0000"]
        self.assertEqual(dec["dep_desc"], "Test Department")
        self.assertEqual(dec["ent_desc"], "UT System")
        self.assertEqual(dec["act_desc"], "FHB Master Account")
        self.assertEqual(coa["entity_desc"]["01"], "UT System")
        self.assertNotIn("99", coa["entity_desc"])          # ENABLED_FLAG=N dropped
        self.assertIn("01-1100001-123456-000", coa["postable_efdp"])
        self.assertNotIn("1 of 1", coa["postable_efdp"])    # footer dropped
        # nothing usable -> None (files-optional graceful degradation).
        self.assertIsNone(E.load_chart_of_accounts([]))

    # ---- end-to-end: byte-identical placements, enriched Review --------
    def _met_with_segments(self):
        return [
            ("CBE_BANK_ACCOUNT_NAME", "CET_REFERENCE_TEXT", "CET_STATUS",
             "CET_TRANSACTION_DATE", "CET_TRANSACTION_ID", "AMOUNT",
             "TRANSACTION_DATE", "CET_DESCRIPTION", "DEPOSIT_ID", "RECEIPT_ID",
             "ASSET_CONCATENATED_SEGMENTS"),
            # already-reconciled counterpart at 250.00 carrying an asset combo:
            # falls to P10 Review (closed), where the CoA decodes its dept/entity.
            ("FHB - Master Account", "TIEREF9", "REC", "2026-07-01", 555,
             "250.00", "2026-07-01", "d:970 | r:71 | PAYER A", 970, 71,
             "01-1100001-123456-100210-000-0000-00-0000"),
        ]

    def _build_master(self, with_coa):
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            ("2026-07-03", "250.00", "TIEREF9", "ACH CREDIT", "TIEREF9",
             "Automated clearing house", "Line 1 , 2026-07-03", "142"),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        st = [("Date", "Amount (USD)", "Reference", "Transaction Number", "Source", "Counterparty"),
              ("2026-07-02", "999.00", "OTHER", 601, "External", "V")]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        _write_xlsx(os.path.join(self.d, "20260710_MET_All_Accounts.xlsx"),
                    [("Miscellaneous External Transact", self._met_with_segments())])
        if with_coa:
            self._write_coa_bundle(self.d)

    def _placement_tabs(self, wb_path):
        tabs, _ = A._read_output_tabs(wb_path)
        return {t: [tuple(A._N(c) for c in r) for r in tabs[t][3:] if any(A._N(c) for c in r)]
                for t in ("Matches", "Candidate Matches", "Misdirected")}

    def test_coa_absent_output_and_enrichment(self):
        # Run WITHOUT the CoA bundle.
        self._build_master(with_coa=False)
        rb = E.run(self.d, self.out, present=True)
        self.assertEqual(rb["audit"]["status"], "PASS", msg=str(rb["audit"].get("failures")))
        tabs_b = A._read_output_tabs(rb["recon_workbook"])[0]
        rev_b = [r for r in tabs_b["Review Notes"][3:] if any(A._N(c) for c in r)]
        expl_b = A._N(rev_b[0][A.COL_EXPL])
        place_b = self._placement_tabs(rb["recon_workbook"])
        self.assertNotIn("coa_combo_validity", rb)
        self.assertNotIn("dept Test Department", expl_b)

        # Fresh run WITH the CoA bundle.
        shutil.rmtree(self.out, ignore_errors=True); self.out = tempfile.mkdtemp()
        self._build_master(with_coa=True)
        rc = E.run(self.d, self.out, present=True)
        self.assertEqual(rc["audit"]["status"], "PASS", msg=str(rc["audit"].get("failures")))
        tabs_c = A._read_output_tabs(rc["recon_workbook"])[0]
        rev_c = [r for r in tabs_c["Review Notes"][3:] if any(A._N(c) for c in r)]
        expl_c = A._N(rev_c[0][A.COL_EXPL])
        place_c = self._placement_tabs(rc["recon_workbook"])

        # Placements byte-identical; only the Review explanation is enriched.
        self.assertEqual(place_b, place_c)
        self.assertEqual([r[:A.COL_EXPL] for r in rev_b],
                         [r[:A.COL_EXPL] for r in rev_c])
        self.assertIn("ALREADY_RECONCILED_COUNTERPART", A._N(rev_c[0][A.COL_EXPL]))
        self.assertIn("dept Test Department", expl_c)
        self.assertIn("entity UT System", expl_c)
        # diagnostic counter present, everything recognized + postable.
        self.assertEqual(rc["coa_combo_validity"],
                         {"rows_seen": 1, "unrecognized_combo": 0, "non_postable_efdp": 0})


class TestTieIndexEquivalence(unittest.TestCase):
    """The 6-gram tie index must equal the brute-force full-pool scan —
    list equality INCLUDING order (perf refactor, 2026-07-19)."""

    VOCAB = [
        "HEARTLANDPAY",          # alpha-only >= 6
        "8036830332",            # digits >= 6
        "REF-12345",             # 5-digit run with separator
        "ABC12",                 # short (< 6 znorm)
        "12345678", "12345679",  # sibling pair (must NOT tie)
        "PAYMENT REF 887799 X",  # containment source
        "887799",                # contained token (= 6)
        "88779",                 # 5 chars — no containment tie
        "NA", "", "  - ",        # null-ish
    ]

    def _bsl(self, i, ref, info="", cust=""):
        return E.make_bsl(f"L{i}", date(2026, 7, 10), 1000 + i, ref, ref,
                          info, "Miscellaneous", "174", customer_reference=cust)

    def _entry(self, i, ref, cp=""):
        return E._mk_entry(f"E{i}", 2000 + i, date(2026, 7, 9), ref, cp,
                           "AR", "UNR", True, "ST")

    def test_index_equals_brute_force(self):
        bsls, pool = [], []
        n = 0
        for ref in self.VOCAB:
            bsls.append(self._bsl(n, ref)); n += 1
        # cross-field-only ties: value in info/customer only
        bsls.append(self._bsl(n, "NA", info="pay HEARTLANDPAY today")); n += 1
        bsls.append(self._bsl(n, "NA", cust="887799")); n += 1
        m = 0
        for ref in self.VOCAB:
            pool.append(self._entry(m, ref)); m += 1
        pool.append(self._entry(m, "NA", cp="PAYMENT REF 887799 X")); m += 1
        pool.append(self._entry(m + 1, "OTHER", cp="HEARTLANDPAY"))
        idx = E._build_tie_index(bsls, pool)
        for b in bsls:
            brute = [e for e in pool if E.cross_reference_tie(b, e)]
            self.assertEqual(idx[b.line_key], brute, msg=b.recon_reference)
        # sanity: the fixture actually exercises ties and non-ties
        total = sum(len(v) for v in idx.values())
        self.assertGreater(total, 4)
        self.assertLess(total, len(bsls) * len(pool))

    def test_fast_path_equivalence(self):
        vals = self.VOCAB + ["  X-887799 ", "8877990"]
        for a in vals:
            for b in vals:
                self.assertEqual(
                    E.reference_equal(a, b),
                    E._reference_equal_z(E.znorm(a), E.znorm(b)), msg=(a, b))
                self.assertEqual(
                    E.sibling(a, b),
                    E._sibling_z(E.znorm(a), E.znorm(b)), msg=(a, b))

    def test_bucket_liveness_within_pass(self):
        # Doctrine 6: an ST consumed by an earlier BSL in the same pass must
        # be invisible to later BSLs even through the amount buckets.
        b1 = E.make_bsl("L1", date(2026, 7, 10), 5000, "TIEREF77", "TIEREF77",
                        "", "Miscellaneous", "174")
        b2 = E.make_bsl("L2", date(2026, 7, 10), 5000, "TIEREF77", "TIEREF77",
                        "", "Miscellaneous", "174")
        e = E._mk_entry("ST1", 5000, date(2026, 7, 9), "TIEREF77", "", "AR",
                        "UNR", True, "ST")
        placements = E.forward_reconcile([b1, b2], [e], {}, "FHB_MASTER", {})
        cited = [p for p in placements if p.st_entries]
        self.assertEqual(len(cited), 1)


class TestRoutingHardening(unittest.TestCase):
    """File-name protection (critical review, 2026-07-19): reconciled
    forensic exports must never be misread as open exports; paginated MET
    shards union; the audit re-parses the SAME file the engine reconciled."""

    def setUp(self):
        self.d = tempfile.mkdtemp()
        self.out = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.d, ignore_errors=True)
        shutil.rmtree(self.out, ignore_errors=True)

    def test_reconciled_exports_never_bind_open_roles(self):
        cases = {
            "20260719_Oracle_CM_Reconciled_FHB_UTC_Student_Refund_BSL.xlsx": "RECONCILED",
            "Reconciled_BSL_FHB_Master.xlsx": "RECONCILED",
            "20260719_Oracle_CM_Reconciled_System_Transactions_FHB_Master.xlsx": "RECONCILED",
            "Reconciliation_Report_Reconciliation_History.xlsx": "RECONCILED",
            "20260719_Oracle_CM_FHB_Master_BSL_UNR.xlsx": "BSL",
            "20260719_Oracle_CM_FHB_Master_ST_UNR.xlsx": "ST",
        }
        for name, want in cases.items():
            self.assertEqual(E.classify_file(name), want, msg=name)

    def test_met_pagination_union(self):
        # Two same-date MET page shards are pages of ONE export: unioned,
        # identical binding required.
        hdr = ("CET Transaction ID", "Amount", "Transaction Date", "CET Status",
               "CBE Bank Account Name")
        _write_xlsx(os.path.join(self.d, "20260710_Oracle_OTBI_MET_FHB_UTC.xlsx"),
                    [("Sheet1", [hdr,
                     ("MET1", "100.00", "2026-07-01", "UNR", "FHB - UTC")])])
        _write_xlsx(os.path.join(self.d, "20260710_Oracle_OTBI_MET_FHB_UTC_2.xlsx"),
                    [("Sheet1", [hdr,
                     ("MET2", "200.00", "2026-07-02", "UNR", "FHB - UTC")])])
        _write_xlsx(os.path.join(self.d, "20260710_FHB_UTC_BSL_UNR.xlsx"),
                    [("Exported", [("Date", "Amount (USD)", "Reference",
                                    "Additional Information",
                                    "Account Servicer Reference",
                                    "Transaction Type", "Statement",
                                    "Transaction Code"),
                                   ("2026-07-01", "10.00", "R1", "", "R1",
                                    "Miscellaneous", "L1", "174")])])
        by_role = E.route_folder(self.d)
        loaded = E.load_and_bind(by_role, {})
        met = loaded["MET"]
        ids = {E.N(r[met["map"]["trx_id"]]) for r in met["rows"][met["header_index"] + 1:]}
        self.assertEqual(ids, {"MET1", "MET2"})

    def test_audit_reparse_ignores_newer_all_bsl_and_reconciled(self):
        bsl_rows = [("Date", "Amount (USD)", "Reference", "Additional Information",
                     "Account Servicer Reference", "Transaction Type", "Statement",
                     "Transaction Code"),
                    ("2026-07-01", "10.00", "R1", "", "R1", "Miscellaneous",
                     "Line 1 , 2026-07-01", "174")]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl_rows)])
        # NEWER-dated all-accounts export + a reconciled forensic export:
        # both carry 'bsl' in the name; neither may win the C1 re-parse.
        big = [bsl_rows[0]] + [
            ("2026-07-0%d" % (i % 9 + 1), "99.00", "X", "", "X",
             "Miscellaneous", f"Line {i} x", "174") for i in range(5)]
        _write_xlsx(os.path.join(self.d, "20260720_Oracle_OTBI_All_BSL_UNR.xlsx"),
                    [("Exported", big)])
        _write_xlsx(os.path.join(self.d, "20260721_Reconciled_BSL_FHB_Master.xlsx"),
                    [("Exported", big)])
        bag = A._reparse_source_bsls(self.d)
        self.assertEqual(len(bag), 1)          # the single account line only

    def test_audit_c2_resums_match_amounts(self):
        # A Match whose cited ST amounts do not sum to the bank line must
        # fail C2 — re-summed from the workbook itself, no pool needed.
        bsl = [("Date", "Amount (USD)", "Reference", "Additional Information",
                "Account Servicer Reference", "Transaction Type", "Statement",
                "Transaction Code"),
               ("2026-07-01", "150.00", "REF88001", "", "REF88001",
                "Miscellaneous", "Line 1 , 2026-07-01", "174")]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        st = [("Date", "Amount (USD)", "Reference", "Transaction Number",
               "Source", "Counterparty"),
              ("2026-06-30", "150.00", "REF88001", 601, "Receivables", "V")]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS")
        self.assertEqual(runlog["recon_summary"]["matches"], 1)
        # Tamper the ST Amounts cell -> audit must now FAIL C2.
        from openpyxl import load_workbook
        wb = load_workbook(runlog["recon_workbook"])
        ws = wb["Matches"]
        ws.cell(row=4, column=11, value="140.00")
        wb.save(runlog["recon_workbook"])
        res = A.audit(self.d, runlog["recon_workbook"], "FHB_MASTER")
        self.assertEqual(res["checks"]["C2"], "FAIL")
        self.assertTrue(any("do not sum" in f for f in res["failures"]))

    def test_ignored_files_announced_immediately(self):
        # Owner hard requirement (2026-07-19): the router announces every
        # name-based skip the moment it happens — reason + rename fix — and
        # run() records it in the runlog.
        import io, contextlib
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", [("Date", "Amount (USD)", "Reference",
                                    "Additional Information",
                                    "Account Servicer Reference",
                                    "Transaction Type", "Statement",
                                    "Transaction Code"),
                                   ("2026-07-01", "10.00", "R1", "", "R1",
                                    "Miscellaneous", "L1", "174")])])
        _write_xlsx(os.path.join(self.d, "mystery_export.xlsx"),
                    [("Sheet1", [("a",)])])          # no router rule
        with open(os.path.join(self.d, "notes.txt"), "w") as fh:
            fh.write("not a BAI2 file")              # .txt without BAI token
        with open(os.path.join(self.d, "legacy.xls"), "w") as fh:
            fh.write("old excel")                    # unreadable legacy format
        err = io.StringIO()
        skipped = []
        with contextlib.redirect_stderr(err):
            by_role = E.route_folder(self.d, skipped)
        text = err.getvalue()
        self.assertEqual(len(skipped), 3)
        for frag in ("mystery_export.xlsx", "notes.txt", "legacy.xls",
                     "IGNORED (file name)", "FIX:"):
            self.assertIn(frag, text)
        self.assertNotIn("BSL_UNR", text)            # routed files: no warning
        # every skip carries a reason and a suggestion
        for s in skipped:
            self.assertTrue(s["reason"] and s["suggestion"], msg=s)
        # run() records the same list in the runlog
        with contextlib.redirect_stderr(io.StringIO()):
            runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(len(runlog["files_ignored_by_name"]), 3)

    def test_recon_history_advisory_audit(self):
        # R2 (orphan doctrine): a staged Reconciled_* export yields advisory
        # orphan findings — placements byte-identical with or without it.
        bsl = [("Date", "Amount (USD)", "Reference", "Additional Information",
                "Account Servicer Reference", "Transaction Type", "Statement",
                "Transaction Code"),
               ("2026-07-01", "150.00", "NA", "MYSTERY PAYMENT", "NA",
                "Miscellaneous", "Line 1 , 2026-07-01", "174")]
        st = [("Date", "Amount (USD)", "Reference", "Transaction Number",
               "Source", "Counterparty"),
              ("2026-06-30", "700.00", "OPENREF9", 601, "Receivables", "V")]
        rec = [("Reconciled Group", "Transaction Source", "Reference", "Date",
                "Amount (USD)", "Transaction Type", "Transaction ID",
                "Journal", "Journal Batch", "Journal Line", "Batch Name",
                "Payment File Reference", "Receipt Batch Number",
                "Cleared Date", "Automatically Reconciled", "Created By",
                "Creation Date"),
               # ESSADMIN group at the open ST's cents (R2 neighborhood)
               ("Group 1", "Statement", "OPENREF9", "2026-06-29", "700.00",
                "ACH", "", "", "", "", "", "", "", "", "Y", "ESSADMIN", "2026-06-29"),
               ("Group 1", "External", "OPENREF9", "2026-06-29", "700.00",
                "", "900001", "", "", "", "", "", "", "", "Y", "ESSADMIN", "2026-06-29"),
               # duplicate feed: same signature, two DISTINCT transaction ids
               ("Group 2", "External", "DUPREF88", "2026-06-20", "55.00",
                "", "800001", "", "", "", "", "", "", "", "Y", "OIC_SYSTEM_USER", "2026-06-20"),
               ("Group 3", "External", "DUPREF88", "2026-06-21", "55.00",
                "", "800002", "", "", "", "", "", "", "", "Y", "OIC_SYSTEM_USER", "2026-06-21"),
               # group at the Review line's cents (history neighborhood)
               ("Group 4", "Statement", "X1", "2026-06-15", "150.00",
                "MSC", "", "", "", "", "", "", "", "", "N", "WKITTS2", "2026-06-15"),
               ("Group 4", "Receivables", "X1", "2026-06-15", "150.00",
                "", "700001", "", "", "", "", "", "", "", "N", "WKITTS2", "2026-06-15")]
        for with_rec in (False, True):
            d, out = tempfile.mkdtemp(), tempfile.mkdtemp()
            _write_xlsx(os.path.join(d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                        [("Exported", bsl)])
            _write_xlsx(os.path.join(d, "20260710_FHB_Master_ST_UNR.xlsx"),
                        [("Exported", st)])
            if with_rec:
                _write_xlsx(os.path.join(d, "20260710_Oracle_CM_Reconciled_FHB_Master_BSL.xlsx"),
                            [("Exported", rec)])
            runlog = E.run(d, out, present=True)
            tabs, _ = A._read_output_tabs(runlog["recon_workbook"])
            dump = {t: [[A._N(c) for c in r] for r in rows]
                    for t, rows in tabs.items()}
            if not with_rec:
                base_dump = dump
                self.assertNotIn("recon_history_orphans", runlog)
                self.assertFalse(os.path.exists(os.path.join(
                    out, "FHB_MASTER_orphan_findings.md")))
            else:
                self.assertEqual(dump, base_dump)     # placements untouched
                rep = runlog["recon_history_orphans"]
                self.assertEqual(rep["dup_feed_signatures"], 1)
                self.assertEqual(rep["open_st_automation_neighborhood"], 1)
                self.assertEqual(rep["review_lines_with_history_neighborhood"], 1)
                text = open(rep["report_path"]).read()
                self.assertIn("800001", text)
                self.assertIn("ESSADMIN", text)
                self.assertIn("Group 4", text)
            shutil.rmtree(d, ignore_errors=True)

    def test_recon_history_option_c_orphan_suppression(self):
        # R2 option C: an open pool ST whose transaction id is a leg of an
        # ESSADMIN/OIC REC group at matching cents is an orphan — suppressed
        # from matching (available=False), so the bank line it would have
        # matched falls to Review instead of a Match.
        bsl = [("Date", "Amount (USD)", "Reference", "Additional Information",
                "Account Servicer Reference", "Transaction Type", "Statement",
                "Transaction Code"),
               ("2026-07-01", "700.00", "OPENREF9", "", "OPENREF9",
                "Miscellaneous", "Line 1 , 2026-07-01", "174")]
        st = [("Date", "Amount (USD)", "Reference", "Transaction Number",
               "Source", "Counterparty"),
              ("2026-06-30", "700.00", "OPENREF9", "900042", "Receivables", "V")]
        rec = [("Reconciled Group", "Transaction Source", "Reference", "Date",
                "Amount (USD)", "Transaction Type", "Transaction ID",
                "Journal", "Journal Batch", "Journal Line", "Batch Name",
                "Payment File Reference", "Receipt Batch Number",
                "Cleared Date", "Automatically Reconciled", "Created By",
                "Creation Date"),
               ("Group 9", "Statement", "OPENREF9", "2026-06-29", "700.00",
                "ACH", "", "", "", "", "", "", "", "", "Y", "ESSADMIN", "2026-06-29"),
               # the SAME transaction id (900042) is a reconciled leg -> orphan
               ("Group 9", "Receivables", "OPENREF9", "2026-06-29", "700.00",
                "", "900042", "", "", "", "", "", "", "", "Y", "ESSADMIN", "2026-06-29")]

        def run_it(with_rec):
            d, out = tempfile.mkdtemp(), tempfile.mkdtemp()
            _write_xlsx(os.path.join(d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                        [("Exported", bsl)])
            _write_xlsx(os.path.join(d, "20260710_FHB_Master_ST_UNR.xlsx"),
                        [("Exported", st)])
            if with_rec:
                _write_xlsx(os.path.join(d, "20260710_Oracle_CM_Reconciled_FHB_Master_ST.xlsx"),
                            [("Exported", rec)])
            rl = E.run(d, out, present=True)
            shutil.rmtree(d, ignore_errors=True)
            return rl

        without = run_it(False)
        # without the reconciled export the open ST 900042 is a live
        # reference-tied counterpart -> a Match.
        self.assertEqual(without["recon_summary"]["matches"], 1)
        with_rec = run_it(True)
        self.assertEqual(with_rec["recon_history_consumed"], 1)
        # the orphan is suppressed -> the line can no longer match it.
        self.assertEqual(with_rec["recon_summary"]["matches"], 0)
        self.assertEqual(with_rec["recon_summary"]["reviews"], 1)
        self.assertEqual(with_rec["audit"]["status"], "PASS")

    def test_recon_history_option_c_human_group_not_suppressed(self):
        # A HUMAN-reconciled group (not ESSADMIN/OIC) must NOT flip
        # availability — deliberate reconciliations are not second-guessed.
        st = [("Date", "Amount (USD)", "Reference", "Transaction Number",
               "Source"), ("2026-06-30", "700.00", "R", "900042", "Receivables")]
        rec = [("Reconciled Group", "Transaction Source", "Reference", "Date",
                "Amount (USD)", "Transaction Type", "Transaction ID",
                "Automatically Reconciled", "Created By"),
               ("Group 9", "Receivables", "R", "2026-06-29", "700.00", "",
                "900042", "N", "WKITTS2")]
        rf = E.RoutedFile("RECONCILED", os.path.join(self.d, "r.xlsx"),
                          "20260710_Oracle_CM_Reconciled_FHB_Master_ST.xlsx", "multi")
        _write_xlsx(rf.path, [("Exported", rec)])
        rh = E.load_recon_history([rf], "FHB_MASTER")
        self.assertIn("900042", rh["by_txn_id"])
        self.assertNotIn("ESSADMIN", rh["by_txn_id"]["900042"]["creators"])
        # a pool entry at that id would NOT be flipped (creator not automation)
        self.assertFalse(rh["by_txn_id"]["900042"]["creators"] & E._AUTOMATION_CREATORS)

    def test_recon_history_wrong_account_skipped(self):
        files = [E.RoutedFile("RECONCILED", "/nonexistent/x.xlsx",
                              "20260710_Oracle_CM_Reconciled_FHB_UTC_BSL.xlsx",
                              "multi")]
        rh = E.load_recon_history(files, "FHB_MASTER")
        self.assertTrue(rh is None or not rh.get("coarse"))

    def test_otbi_recon_report_parse_and_identity_findings(self):
        # The Oracle OTBI Recon Report rendering (three sheets: Bank
        # Statement Lines / AR Matched / MISC Receipts, actor in "Rec By",
        # reconciled leg per row) parses into the R2 indexes, and the
        # same-transaction-identity findings surface even when the MISC
        # actor is blank (the rendering omits Rec By on that sheet) — WITHOUT
        # flipping availability (option C is gated on a confirmed automation
        # actor, so placements stay byte-identical).
        bsl = [("Date", "Amount (USD)", "Reference", "Additional Information",
                "Account Servicer Reference", "Transaction Type", "Statement",
                "Transaction Code"),
               ("2026-07-01", "460.00", "REF123456", "DEPOSIT",
                "REF123456", "Miscellaneous", "Line 1 , 2026-07-01", "174")]
        st = [("Date", "Amount (USD)", "Reference", "Transaction Number",
               "Source", "Counterparty"),
              # matches the bank line by reference -> a Match citing 1031272
              ("2026-06-30", "460.00", "REF123456", "1031272", "External", "V"),
              # a second open ST that is a reconciled leg but matches nothing
              ("2026-06-30", "999.00", "OTHERREF7", "1031289", "External", "W")]
        # OTBI recon report: three sheets, MISC sheet has NO Rec By column.
        bsl_sheet = [("Rec Grp", "Rec By", "Amnt", "Trx Type", "Rec Status",
                      "Bank Accnt"),
                     ("G1", "OIC_SYSTEM_USER", "460.00", "MSC", "REC", "")]
        ar_sheet = [("Rcpt Num", "Amnt", "Rec Grp", "Rec By", "Strc Pay Ref",
                     "Bank Accnt")]
        misc_sheet = [("Trx Id", "Rec Num", "Rec Amnt", "Rec Grp", "Trx Type",
                       "Rec Ref", "Bank Accnt"),
                      ("1031272", "1031272", "460.00", "G9", "MSC",
                       "REF123456", ""),
                      ("1031289", "1031289", "999.00", "G9", "MSC",
                       "OTHERREF7", "")]

        def run_it(with_rec):
            d, out = tempfile.mkdtemp(), tempfile.mkdtemp()
            _write_xlsx(os.path.join(d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                        [("Exported", bsl)])
            _write_xlsx(os.path.join(d, "20260710_FHB_Master_ST_UNR.xlsx"),
                        [("Exported", st)])
            if with_rec:
                _write_xlsx(
                    os.path.join(d, "20260710_Oracle_OTBI_Recon_Report.xlsx"),
                    [("Bank Statement Lines", bsl_sheet),
                     ("AR Matched", ar_sheet),
                     ("MISC Receipts", misc_sheet)])
            rl = E.run(d, out, present=True)
            tabs, _ = A._read_output_tabs(rl["recon_workbook"])
            dump = {t: [[A._N(c) for c in r] for r in rows]
                    for t, rows in tabs.items()}
            shutil.rmtree(d, ignore_errors=True)
            return rl, dump, out

        base_rl, base_dump, _ = run_it(False)
        self.assertEqual(base_rl["recon_summary"]["matches"], 1)
        rl, dump, out = run_it(True)
        # advisory only: placements byte-identical, no availability flip.
        self.assertEqual(dump, base_dump)
        self.assertNotIn("recon_history_consumed", rl)
        self.assertEqual(rl["audit"]["status"], "PASS")
        rep = rl["recon_history_orphans"]
        # the Match cites 1031272 (a reconciled leg) -> flagged.
        self.assertEqual(rep["placements_citing_reconciled_id"], 1)
        # the unmatched open ST 1031289 is a same-identity orphan.
        self.assertGreaterEqual(rep["open_st_reconciled_identity"], 1)
        with open(rep["report_path"]) as fh:
            text = fh.read()
        self.assertIn("1031272", text)
        self.assertIn("already-reconciled", text)

    def test_multi_bai2_index_union_and_dedup(self):
        # Two BAI2 files with overlapping windows: the index unions their
        # coverage but dedups the SAME transaction (same bank reference),
        # keeping the richer addenda — a duplicate must not become a second
        # candidate that makes enrichment decline.
        raw = "\n".join([
            "01,000000000,000000000,260718,0759,1,,,2/",
            "02,084000026,084000026,1,260710,,USD,2/",
            "16,142,5100,Z,26191005187395,001,",
            "88,MERCHANT SERVICEDEPOSIT",
            "88,Customer ID: 8035701948",
            "99,0,1,2/",
        ])
        with open(os.path.join(self.d, "20260718_FHB_Master_BAI2.txt"), "w") as fh:
            fh.write(raw)
        # spreadsheet BAI2 covering an EARLIER window + the same 07-10 record
        _write_xlsx(os.path.join(self.d, "20260715_FHB_Master_BAI2.xlsx"),
                    [("first", [("Post Date", "Amount", "Bank Reference",
                                 "Customer Reference", "BAI Code"),
                                ("2026-07-10", "51.00", "26191005187395", "001", "142"),
                                ("2026-06-05", "77.00", "26156000000001", "002", "142")])])
        files = [E.RoutedFile("BAI2", os.path.join(self.d, n), n, "first")
                 for n in sorted(os.listdir(self.d), reverse=True)
                 if "BAI2" in n]
        loaded = E.load_and_bind({"BAI2": files}, {})
        self.assertEqual(len(loaded["BAI2"]["files"]), 2)
        idx = E._bai2_index(loaded)
        # overlapping record deduped to ONE candidate, richer (txt) details
        cands = idx[(date(2026, 7, 10), 5100)]
        self.assertEqual(len(cands), 1)
        self.assertIn("Customer ID: 8035701948", cands[0]["details"])
        # June record from the older file present (window union)
        self.assertIn((date(2026, 6, 5), 7700), idx)

    def test_single_bai2_shape_unchanged(self):
        _write_xlsx(os.path.join(self.d, "20260715_FHB_Master_BAI2.xlsx"),
                    [("first", [("Post Date", "Amount", "Bank Reference",
                                 "Customer Reference", "BAI Code"),
                                ("2026-07-10", "51.00", "26191005187395", "001", "142")])])
        files = [E.RoutedFile("BAI2", os.path.join(self.d, n), n, "first")
                 for n in sorted(os.listdir(self.d)) if "BAI2" in n]
        loaded = E.load_and_bind({"BAI2": files}, {})
        self.assertNotIn("files", loaded["BAI2"])     # legacy single-dict shape
        idx = E._bai2_index(loaded)
        self.assertIn((date(2026, 7, 10), 5100), idx)

    def test_bsl_pagination_union_end_to_end(self):
        # Two same-date BSL page shards: engine unions them, conservation
        # spans both pages, and the audit's C1 re-parse unions the same
        # shards (a single-file read would falsely fail C1).
        hdr = ("Date", "Amount (USD)", "Reference", "Additional Information",
               "Account Servicer Reference", "Transaction Type", "Statement",
               "Transaction Code")
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", [hdr,
                     ("2026-07-01", "10.00", "R1", "", "R1", "Miscellaneous",
                      "Line 1 , 2026-07-01", "174")])])
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR_2.xlsx"),
                    [("Exported", [hdr,
                     ("2026-07-02", "20.00", "R2", "", "R2", "Miscellaneous",
                      "Line 2 , 2026-07-02", "174")])])
        st = [("Date", "Amount (USD)", "Reference", "Transaction Number",
               "Source", "Counterparty"),
              ("2026-06-30", "999.00", "X", 601, "Receivables", "V")]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["bsl_count"], 2)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        self.assertIn("union_of", runlog["roles_bound"]["BSL"])

    def test_pagination_duplicate_upload_fails_loud(self):
        # An identical data row across two same-date shards is a duplicate
        # upload, not pagination — unioning would double-count money.
        hdr = ("Date", "Amount (USD)", "Reference", "Additional Information",
               "Account Servicer Reference", "Transaction Type", "Statement",
               "Transaction Code")
        row = ("2026-07-01", "10.00", "R1", "", "R1", "Miscellaneous",
               "Line 1 , 2026-07-01", "174")
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", [hdr, row])])
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR_2.xlsx"),
                    [("Exported", [hdr, row])])
        with self.assertRaises(E.InvalidSourceData) as cm:
            E.run(self.d, self.out, present=True)
        self.assertIn("duplicate", str(cm.exception).lower())

    def test_non_paginatable_role_still_fails_loud_on_tie(self):
        # DEPT_INFO is not paginatable: two same-date files stay a conflict.
        hdr = ("Department Name", "Campus Name", "Credit Card Mid")
        _write_xlsx(os.path.join(self.d, "20260710_ORT_Department_Info.xlsx"),
                    [("Report", [hdr, ("Dept A", "UTK", "8011111111")])])
        _write_xlsx(os.path.join(self.d, "20260710_ORT_Department_Info_2.xlsx"),
                    [("Report", [hdr, ("Dept B", "UTC", "8022222222")])])
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", [("Date", "Amount (USD)", "Reference",
                                    "Additional Information",
                                    "Account Servicer Reference",
                                    "Transaction Type", "Statement",
                                    "Transaction Code"),
                                   ("2026-07-01", "10.00", "R1", "", "R1",
                                    "Miscellaneous", "L1", "174")])])
        with self.assertRaises(E.InvalidSourceData) as cm:
            E.run(self.d, self.out, present=True)
        self.assertIn("tie", str(cm.exception).lower())

    def test_c11_fabricated_st_id_fails(self):
        # A Match citing an ST id absent from every source export fails C11.
        bsl = [("Date", "Amount (USD)", "Reference", "Additional Information",
                "Account Servicer Reference", "Transaction Type", "Statement",
                "Transaction Code"),
               ("2026-07-01", "150.00", "REF88001", "", "REF88001",
                "Miscellaneous", "Line 1 , 2026-07-01", "174")]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        st = [("Date", "Amount (USD)", "Reference", "Transaction Number",
               "Source", "Counterparty"),
              ("2026-06-30", "150.00", "REF88001", 601, "Receivables", "V")]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS")
        self.assertEqual(runlog["audit"]["checks"]["C11"], "PASS")
        # Tamper the cited ST number to a fabricated id -> C11 FAIL.
        from openpyxl import load_workbook
        wb = load_workbook(runlog["recon_workbook"])
        ws = wb["Matches"]
        ws.cell(row=4, column=10, value="999999999")
        wb.save(runlog["recon_workbook"])
        res = A.audit(self.d, runlog["recon_workbook"], "FHB_MASTER")
        self.assertEqual(res["checks"]["C11"], "FAIL")
        self.assertTrue(any("C11" in f and "999999999" in f for f in res["failures"]))

    def test_split_ids_semantics(self):
        # '; '-joined cell: one id embeds ', ' — must NOT be over-split.
        cell = "Correct Parsing Rule - Reference: a, b; 601"
        self.assertEqual(A._split_ids(cell),
                         ["Correct Parsing Rule - Reference: a, b", "601"])
        # ', '-joined cell: plain multi-id split.
        self.assertEqual(A._split_ids("601, 602"), ["601", "602"])
        # Disambiguation suffix KEPT by default (it IS the identity, rule 8)
        self.assertEqual(A._split_ids("1 [0.01], 1 [10.00]"),
                         ["1 [0.01]", "1 [10.00]"])
        # ... and stripped for the C11 source-membership lookup.
        self.assertEqual(A._split_ids("1 [0.01], 1 [10.00]", strip_disambig=True),
                         ["1", "1"])
        self.assertEqual(A._split_ids("1024656 [-728.95]", strip_disambig=True),
                         ["1024656"])

    def test_reparse_valid_st_ids_sources(self):
        # Superset spans ST + receipts (incl. DOC synth) + payments + MET,
        # with .xlsb-style '.0' collapse; Edison/reconciled files excluded.
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", [("Date", "Amount (USD)", "Reference",
                                    "Transaction Number", "Source"),
                                   ("2026-06-30", "1.00", "R", "601", "Receivables")])])
        _write_xlsx(os.path.join(self.d, "20260710_Oracle_Receivables_Receipts.xlsx"),
                    [("Export to Excel", [("Receipt Number", "Document Number",
                                           "Receipt Amount"),
                                          ("364912", "9001", "5.00"),
                                          ("", "77001", "6.00")])])
        _write_xlsx(os.path.join(self.d, "20260710_Oracle_Payables_Payments.xlsx"),
                    [("first", [("Payment Number", "Payment Amount", "Payment Status"),
                                ("60410073", "7.00", "Negotiable")])])
        _write_xlsx(os.path.join(self.d, "20260710_Oracle_OTBI_MET_FHB.xlsx"),
                    [("Sheet1", [("CET_TRANSACTION_ID", "AMOUNT"),
                                 ("65105.0", "8.00")])])
        _write_xlsx(os.path.join(self.d, "Edison_Payments.xlsx"),
                    [("sheet1", [("Reference", "Amount"), ("0007000001", "9.00")])])
        ids, digits = A._reparse_valid_st_ids(self.d)
        for want in ("601", "364912", "DOC 77001", "60410073", "65105"):
            self.assertIn(want, ids, msg=sorted(ids))
        self.assertNotIn("0007000001", ids)      # Edison excluded
        self.assertIn("60410073", digits)

    def test_chatt_account_tokens(self):
        self.assertEqual(E.infer_account("20260715_FHB_UT_Chatt_Student_Refund_BAI2.csv"),
                         "FHB_STUDENT_REFUND_UTC")
        self.assertEqual(E.account_of_bank_name("FHB - UT Chatt"), "FHB_UTC")


class TestGuardrails2026(unittest.TestCase):
    """Hard-guardrail regressions (critical review, 2026-07-19): false-match
    and false-candidate holes closed at the pass level."""

    def _fwd(self, bsls, pool):
        runlog = {}
        return E.forward_reconcile(bsls, pool, {}, "FHB_MASTER", runlog)

    def _one(self, placements, line_key):
        return next(p for p in placements if p.bsl.line_key == line_key)

    def test_sibling_reference_is_conflict_not_candidate(self):
        # Rule 7: equal-length numeric refs differing in the last 1-2 digits
        # are CONFLICTS — barred even from the amount-only Candidate lane.
        b = E.make_bsl("L1", date(2026, 7, 10), 12345, "12345678", "12345678",
                       "", "Miscellaneous", "174")
        e = E._mk_entry("ST1", 12345, date(2026, 7, 9), "12345679", "", "AR",
                        "UNR", True, "ST")
        p = self._one(self._fwd([b], [e]), "L1")
        self.assertEqual(p.kind, E.REVIEW, msg=p.explanation)

    def test_distinctive_match_demoted_by_pool_twin(self):
        b = E.make_bsl("L1", date(2026, 7, 10), 245546969, "NA", "NA",
                       "", "Miscellaneous", "174")
        open_st = E._mk_entry("ST1", 245546969, date(2026, 7, 9), "REFA", "",
                              "AR", "UNR", True, "ST")
        # control: unique across the whole pool -> distinctive Match
        p = self._one(self._fwd([b], [open_st]), "L1")
        self.assertEqual(p.kind, E.MATCH)
        self.assertIn("DISTINCTIVE_AMOUNT", p.codes)
        # a CLOSED twin at the same signed cents kills pool-side uniqueness
        closed_twin = E._mk_entry("ST2", 245546969, date(2026, 6, 1), "REFB",
                                  "", "AR", "REC", False, "ST")
        p = self._one(self._fwd([b], [open_st, closed_twin]), "L1")
        self.assertEqual(p.kind, E.CANDIDATE, msg=p.explanation)
        self.assertNotIn("DISTINCTIVE_AMOUNT", p.codes)
        # a foreign-shadow twin kills it too
        shadow = E._mk_entry("ST3", 245546969, date(2026, 7, 8), "REFC", "",
                             "EXT", "UNR", True, "MET")
        shadow.foreign_account = "FHB_UTHSC"
        shadow.available = False
        p = self._one(self._fwd([b], [open_st, shadow]), "L1")
        self.assertEqual(p.kind, E.CANDIDATE, msg=p.explanation)

    def test_pm_weak_digit_tie_rejected(self):
        # A 5-digit-run coincidence in the addenda must NOT reroute money to
        # a foreign account; a znorm >= 6 reference tie still does.
        weak = E.make_bsl("L1", date(2026, 7, 10), 7099266, "NA", "NA",
                          "ZIP 38105 REMIT", "Automated clearing house", "142")
        f = E._mk_entry("300045836", 7099266, date(2026, 7, 9), "38105",
                        "City of Memphis", "AR", "UNR", True, "RECEIPTS")
        f.foreign_account = "FHB_MASTER_X"
        f.available = False
        p = self._one(self._fwd([weak], [f]), "L1")
        self.assertEqual(p.kind, E.REVIEW, msg=p.explanation)
        strong = E.make_bsl("L2", date(2026, 7, 10), 7099266, "300045836",
                            "300045836", "", "Automated clearing house", "142")
        f2 = E._mk_entry("300045836", 7099266, date(2026, 7, 9), "300045836",
                         "City of Memphis", "AR", "UNR", True, "RECEIPTS")
        f2.foreign_account = "FHB_MASTER_X"
        f2.available = False
        p = self._one(self._fwd([strong], [f2]), "L2")
        self.assertEqual(p.kind, E.MISDIRECTED, msg=p.explanation)

    def test_pm_never_cites_one_foreign_entry_twice(self):
        mk = lambda k: E.make_bsl(k, date(2026, 7, 10), 7099266, "300045836",
                                  "300045836", "", "Automated clearing house", "142")
        f = E._mk_entry("300045836", 7099266, date(2026, 7, 9), "300045836",
                        "City of Memphis", "AR", "UNR", True, "RECEIPTS")
        f.foreign_account = "FHB_MASTER_X"
        f.available = False
        placements = self._fwd([mk("L1"), mk("L2")], [f])
        kinds = sorted(p.kind for p in placements)
        self.assertEqual(kinds, sorted([E.MISDIRECTED, E.REVIEW]))

    def test_p7_spn_screens(self):
        # Uncorroborated SPN singleton no longer produces a P7 Candidate; the
        # line falls through to the screened later lanes.
        b = E.make_bsl("L1", date(2026, 7, 10), 55555, "NA", "NA",
                       "", "Miscellaneous", "174")
        e = E._mk_entry("SPN-1", 55555, date(2026, 7, 9), "OTHERREF", "",
                        "AR", "UNR", True, "RECEIPTS")
        e.spn = "260709650"
        p = self._one(self._fwd([b], [e]), "L1")
        self.assertNotEqual(p.pass_name, "P7_spn", msg=p.explanation)
        # Corroborated but 12+ days stale External member: barred from P7.
        b2 = E.make_bsl("L2", date(2026, 7, 30), 55555, "260709650",
                        "260709650", "", "Miscellaneous", "174")
        e2 = E._mk_entry("SPN-2", 55555, date(2026, 7, 1), "260709650", "",
                         "EXT", "UNR", True, "MET")
        e2.spn = "260709650"
        p = self._one(self._fwd([b2], [e2]), "L2")
        self.assertNotEqual((p.pass_name, p.kind), ("P7_spn", E.MATCH),
                            msg=p.explanation)

    def test_p4_competing_equal_sum_group_is_candidate(self):
        # Whole tied set sums exactly AND a proper sub-cluster also sums
        # exactly (the rest is a reversal pair netting zero): ambiguous ->
        # AMBIGUOUS_GROUP Candidate, never a Match.
        b = E.make_bsl("L1", date(2026, 7, 10), 10000, "GRPREF99", "GRPREF99",
                       "", "Miscellaneous", "174")
        def rc(i, cents, cp):
            e = E._mk_entry(f"R{i}", cents, date(2026, 7, 9), "GRPREF99", cp,
                            "AR", "UNR", True, "RECEIPTS")
            return e
        pool = [rc(1, 6000, "Payer X"), rc(2, 4000, "Payer X"),
                rc(3, -4000, "Payer Y"), rc(4, 4000, "Payer Y")]
        p = self._one(self._fwd([b], pool), "L1")
        self.assertEqual(p.kind, E.CANDIDATE, msg=p.explanation)
        self.assertIn("AMBIGUOUS_GROUP", p.codes)

    def test_description_number_creates_reference_tie(self):
        # MET/ORT free-text descriptions (owner, 2026-07-20): a >=6-digit number
        # embedded in the description ("...LGIP 44711210") is now a searchable
        # reference — a bank line carrying that number ties to the entry, where
        # before the description was invisible to the cross-reference screen.
        b = E.make_bsl("L1", date(2026, 7, 10), 500000, "44711210", "44711210",
                       "LGIP TRANSFER 44711210", "Automated clearing house", "142")
        e = E._mk_entry("ST9", 500000, date(2026, 7, 9), "OTHERREF", "",
                        "EXT", "UNR", True, "MET",
                        description="495-FHB - Master Account-LGIP 44711210")
        self.assertIn("44711210", e.desc_refs)
        p = self._one(self._fwd([b], [e]), "L1")
        self.assertEqual(p.kind, E.MATCH, msg=p.explanation)
        self.assertEqual([x.id for x in p.st_entries], ["ST9"])
        # A common description WORD must NOT tie (only distinctive numbers do).
        b2 = E.make_bsl("L2", date(2026, 7, 10), 700000, "NA", "NA",
                        "FHB MASTER ACCOUNT DEPOSIT", "Miscellaneous", "174")
        e2 = E._mk_entry("ST8", 700000, date(2026, 7, 9), "ZZZ", "",
                         "EXT", "UNR", True, "MET",
                         description="Controlled Disbursing 581_FHB Master Account")
        self.assertEqual(e2.desc_refs, ())   # "581" is <6 digits; words excluded
        p2 = self._one(self._fwd([b2], [e2]), "L2")
        self.assertNotEqual(p2.kind, E.MATCH, msg=p2.explanation)

    def test_merchant_mid_ambiguity_does_not_cannibalize_deposit(self):
        # Merchant grouping-key guard (rule 7 + doctrine 8c; owner: minimize
        # false candidates).  A MID is a GROUPING key shared across all of a
        # merchant's receipts, not a 1:1 identity.  Real UTSO case: a $150
        # merchant line found FOUR equal open $150 receipts all carrying the
        # same MID; P3 coin-flipped one and CONSUMED it — but that receipt was
        # a member of the ORT deposit summing to a SEPARATE $460 line, which
        # was then stranded.  P3 must defer the whole ambiguity to the deposit
        # lane, which sums the whole deposit.
        MID = "8037859173"
        line460 = E.make_bsl("L460", date(2026, 7, 13), 46000, MID, MID,
                             f"MERCHANT SERVICE MERCH DEP {MID}",
                             "Automated clearing house", "142")
        line150 = E.make_bsl("L150", date(2026, 7, 13), 15000, MID, MID,
                             f"MERCHANT SERVICE MERCH DEP {MID}",
                             "Automated clearing house", "142")
        self.assertEqual(line460.lane, E.LANE_MERCHANT)
        self.assertEqual(line150.lane, E.LANE_MERCHANT)

        def rc(i, cents, dep):
            return E._mk_entry(i, cents, date(2026, 7, 13), MID, "Touchnet",
                               "EXT", "UNR", True, "MET", deposit_id=dep,
                               receipt_id=f"r{i}")
        # deposit A ($460) = $310 + $150; deposit B ($150) = single $150.
        r_a1 = rc("1000", 31000, "500001")
        r_a2 = rc("1001", 15000, "500001")   # sorts first among the $150s
        r_b1 = rc("1002", 15000, "500002")
        # two extra loose $150 receipts (no deposit) to force P3 ambiguity.
        r_x1 = rc("1003", 15000, "")
        r_x2 = rc("1004", 15000, "")
        pool = [r_a1, r_a2, r_b1, r_x1, r_x2]

        placements = self._fwd([line460, line150], pool)
        p460 = self._one(placements, "L460")
        p150 = self._one(placements, "L150")

        # The $460 line matches its deposit — impossible if P3 had eaten r1001.
        self.assertEqual(p460.kind, E.MATCH, msg=p460.explanation)
        self.assertEqual({e.id for e in p460.st_entries}, {"1000", "1001"})
        self.assertEqual(p460.pass_name, "P4_deposit_group", msg=p460.explanation)
        # The $150 line is resolved through the deposit lane, never a P3
        # coin-flip that cites one arbitrary same-MID receipt.
        self.assertNotEqual(p150.pass_name, "P3_exact_1to1", msg=p150.explanation)
        for p in placements:
            self.assertNotIn("MULTIPLE_EQUAL_CANDIDATES", p.codes or [])

    def test_p3_transparency_defers_two_equal_same_mid_settlements(self):
        # Transparency touch, P3 arm (owner, 2026-07-21): the P3 date filter can
        # narrow a merchant tie-set to one open same-MID receipt by dropping an
        # out-of-window twin.  When >= 2 DISTINCT available same-MID settlements
        # equal the line, that narrowing is a silent date-window pick — defer so
        # the merchant lane names the ambiguity instead of 1:1-ing one.
        MID = "8037859173"
        line = E.make_bsl("L1", date(2026, 7, 17), 15000, MID, MID,
                          f"MERCHANT SERVICE {MID}", "Automated clearing house", "142")

        def rc(i, dt):
            return E._mk_entry(i, 15000, dt, MID, "Touchnet", "EXT", "UNR",
                               True, "MET", receipt_id=f"r{i}")
        a = rc("2001", date(2026, 7, 15))   # 2d before -> passes P3 date gate
        b = rc("2002", date(2026, 7, 7))    # 10d before -> P3 drops it (>8d), not stale-barred (<12d)
        p = self._one(self._fwd([line], [a, b]), "L1")
        self.assertNotEqual(p.pass_name, "P3_exact_1to1", msg=p.explanation)
        self.assertIn("MULTIPLE_EQUAL_CANDIDATES", p.codes or [])

    def test_p3_lone_same_mid_receipt_still_matches(self):
        # Byte-safety: a lone available same-MID receipt equal to the line is a
        # legitimate 1:1 merchant settlement and still Matches in P3 (only >= 2
        # DISTINCT available same-MID settlements defer — real Master Line 109's
        # single open -$15 chargeback must keep its exact-MID match).
        MID = "8037859173"
        line = E.make_bsl("L1", date(2026, 7, 13), 15000, MID, MID,
                          f"MERCHANT SERVICE {MID}", "Automated clearing house", "142")
        e = E._mk_entry("3000", 15000, date(2026, 7, 12), MID, "Touchnet",
                        "EXT", "UNR", True, "MET", deposit_id="700001",
                        receipt_id="r3000")
        p = self._one(self._fwd([line], [e]), "L1")
        self.assertEqual(p.kind, E.MATCH, msg=p.explanation)
        self.assertEqual(p.pass_name, "P3_exact_1to1", msg=p.explanation)

    def test_p6_transparency_two_equal_same_mid_receipts(self):
        # Transparency touch (owner, 2026-07-21): two same-MID card receipts
        # each equal the settlement, one INSIDE the 1-4d window and one outside.
        # P4 phase 2 groups by deposit id; these loose receipts carry none, so
        # the merchant lane (P6) is first to see them — and its window match
        # would silently pick the in-window receipt.  Instead the line surfaces
        # as an ambiguous Candidate naming both (shared MID is a grouping key).
        MID = "2000002247"
        line = E.make_bsl("L75", date(2026, 7, 17), 7500, MID, MID,
                          f"MERCHANT BANKCD {MID}", "Automated clearing house", "142")

        def rc(i, dt):
            return E._mk_entry(i, 7500, dt, MID, "Heartland", "EXT", "UNR",
                               True, "MET", receipt_id=f"r{i}")
        a = rc("4000", date(2026, 7, 15))   # 2d before -> in the 1-4d window
        b = rc("4001", date(2026, 7, 7))    # 10d before -> out of window
        p = self._one(self._fwd([line], [a, b]), "L75")
        self.assertEqual(p.kind, E.CANDIDATE, msg=p.explanation)
        self.assertEqual(p.pass_name, "P6_merchant", msg=p.explanation)
        self.assertIn("MULTIPLE_EQUAL_CANDIDATES", p.codes or [])
        self.assertIn("4000", p.explanation)
        self.assertIn("4001", p.explanation)

    def test_p6_single_equal_same_mid_receipt_not_flagged(self):
        # Byte-safety of the transparency touch: ONE same-MID receipt equal to
        # the settlement in the card window still Matches (no false ambiguity).
        MID = "2000002247"
        line = E.make_bsl("L75", date(2026, 7, 17), 7500, MID, MID,
                          f"MERCHANT BANKCD {MID}", "Automated clearing house", "142")
        e = E._mk_entry("4100", 7500, date(2026, 7, 15), MID, "Heartland",
                        "EXT", "UNR", True, "MET", receipt_id="r4100")
        p = self._one(self._fwd([line], [e]), "L75")
        self.assertEqual(p.kind, E.MATCH, msg=p.explanation)

    def test_ort_misc_reference_crossref_index(self):
        # ORT raw-activity reference cross-reference (owner "cross-reference
        # everything", 2026-07-20): the ORT_Misc / ORT_AR REFERENCE_TEXT column,
        # keyed by Parked Receipt ID, supplies per-receipt bank reference
        # numbers the MET export may lack.  Numbers only, >= 6 digits.
        rows = [["Parked Receipt ID", "REFERENCE_TEXT"],
                ["500123", "8042156011"],
                ["500123", "8042156011"],       # dup item row -> deduped
                ["500124", "600255"],
                ["500125", "ABC"],              # non-numeric -> excluded
                ["500126", "12345"]]            # < 6 digits -> excluded
        loaded = {"ORT_MISC": {"rows": rows, "header_index": 0,
                               "map": {"parked_receipt_id": 0, "reference_text": 1}}}
        idx = E._ort_misc_ref_index(loaded)
        self.assertEqual(idx.get("500123"), {"8042156011"})
        self.assertEqual(idx.get("500124"), {"600255"})
        self.assertNotIn("500125", idx)
        self.assertNotIn("500126", idx)
        # Absent report -> empty (pure no-op).
        self.assertEqual(E._ort_misc_ref_index({}), {})


class TestCMConfig(unittest.TestCase):
    """CM Configuration exports (owner, 2026-07-19): the five CFG_* loaders,
    the Oracle-LIKE simulator, and the raw BAI2 .txt reader."""

    def setUp(self):
        self.d = tempfile.mkdtemp()
        self.out = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.d, ignore_errors=True)
        shutil.rmtree(self.out, ignore_errors=True)

    def test_cfg_router(self):
        cases = {
            "CM_Configurations_Transaction_Creation_Rules.xlsx": "CFG_TCR",
            "CM_Configurations_Parse_Rules.xlsx": "CFG_PARSE",
            "CM_Configurations_Matching_Rules.xlsx": "CFG_MATCHING",
            "CM_Configurations_Tolerance_Rules.xlsx": "CFG_TOLERANCE",
            "CM_Configurations_Recon_Rulesets.xlsx": "CFG_RULESETS",
            "FHB_UTHSC_BAI2.txt": "BAI2",
        }
        for name, want in cases.items():
            self.assertEqual(E.classify_file(name), want, msg=name)

    def test_cfg_tcr_loader(self):
        # Paginated export shape: preamble rows, header at row 6, footer.
        rows = [("",) * 16] * 4 + [
            ("Transaction Creation Rules",) + ("",) * 15,
            ("Run Date: 7/19/2026",) + ("",) * 15,
            ("", "BNK ACCNTNME", "ENBLD", "SQC NUM", "RLE NME", "DSCRPT",
             "TRX CDE", "TRX TPE", "SRCH FLD", "SRCH STRNG", "CASH", "OFFSET",
             "ACCNT FLG", "LST UPDTE", "UPDTEBY", ""),
            ("", "FHB - Master Account", "Y", "1", "EFT 495_Test", "d",
             "495", "EFT", "ADDENDA", "%LGIP%", "01-1100098-000000-100210-000-0000-00-0000",
             "01-1100001-000000-100930-000-0000-00-0000", "Y",
             "2026-04-02 17:03:10", "AOWENS43", ""),
            ("",) * 16,                            # blank spacer
            ("", "BNK ACCNTNME", "ENBLD", "SQC NUM", "RLE NME", "DSCRPT",
             "TRX CDE", "TRX TPE", "SRCH FLD", "SRCH STRNG", "CASH", "OFFSET",
             "ACCNT FLG", "LST UPDTE", "UPDTEBY", ""),  # repeated page header
            ("", "FHB - UTHSC", "N", "2", "ACH 142_Test", "d",
             "142", "ACH", "", "", "01-1100098-000000-100500-000-0000-00-0000",
             "", "Y", "2026-01-01 00:00:00", "X", ""),
            ("1 of 2",) + ("",) * 15,              # footer
        ]
        p = os.path.join(self.d, "CM_Configurations_Transaction_Creation_Rules.xlsx")
        _write_xlsx(p, [("Sheet1", rows)])
        rf = E.RoutedFile("CFG_TCR", p, os.path.basename(p), "first")
        rules = E.load_cm_config("CFG_TCR", rf)
        self.assertEqual(len(rules), 2)            # spacer/header/footer dropped
        self.assertEqual(rules[0]["bank_account"], "FHB - Master Account")
        self.assertEqual(rules[0]["trx_code"], "495")
        self.assertEqual(rules[0]["search_string"], "%LGIP%")
        self.assertEqual(rules[1]["enabled"], "N")

    def test_like_match(self):
        self.assertTrue(E.like_match("%LGIP Account No.: 447112-10%",
                                     "PRIOR TEXT LGIP Account No.: 447112-10 TRAILER"))
        self.assertTrue(E.like_match("RETURN SETTLE%", "RETURN SETTLEMENT X"))
        self.assertFalse(E.like_match("RETURN SETTLE%", "X RETURN SETTLEMENT"))
        self.assertTrue(E.like_match("%PAY%REVERSAL%", "ACH PAY 33 REVERSAL 9"))
        self.assertTrue(E.like_match("A_C", "AbC"))          # _ = one char
        self.assertFalse(E.like_match("A_C", "AbbC"))
        self.assertFalse(E.like_match("", "anything"))       # blank never fires

    def test_raw_bai2_txt_reader(self):
        raw = "\n".join([
            "01,000000000,000000000,260719,0759,2659562,,,2/",
            "02,084000026,084000026,1,260718,,USD,2/",
            "03,90603,USD,010,0,,,015,0,,,/",
            "88,060,0,,,063,0,,,072,0,,,/",                      # 03 continuation: ignored
            "16,142,5100,Z,25202003732875,08035701948,",
            "88,MERCHANT SERVICEDEPOSIT 2507218035701948",
            "88,Customer ID: 8035701948",
            "88,Trace Number: 084000023732875",
            "16,475,250000,Z,99887766,00001234,",
            "49,0,2/",
            "98,0,1,2/",
            "99,0,1,2/",
        ])
        p = os.path.join(self.d, "FHB_UTHSC_BAI2.txt")
        with open(p, "w") as fh:
            fh.write(raw)
        rows = E._read_bai2_txt(p)
        self.assertEqual(len(rows), 3)                    # header + 2 details
        hdr = rows[0]
        self.assertEqual(hdr[:6], ("Post Date", "Transaction Description",
                                   "Amount", "Bank Reference",
                                   "Customer Reference", "BAI Code"))
        # credit 142: positive, group as-of date, addenda in DETAIL columns
        self.assertEqual(rows[1][0], "2026-07-18")
        self.assertEqual(E.cents(rows[1][2]), 5100)
        self.assertEqual(rows[1][5], "142")
        self.assertIn("Customer ID: 8035701948", rows[1])
        # the 03-record continuation must NOT leak into detail addenda
        self.assertFalse(any("060,0" in str(c) for c in rows[1]))
        # debit 475: negated
        self.assertEqual(E.cents(rows[2][2]), -250000)
        # binds the existing BAI2 role vocabulary
        m, hi = E.bind_columns(rows, E.BAI2_ROLES, filename="FHB_UTHSC_BAI2.txt")
        self.assertEqual(hi, 0)
        self.assertEqual(m["bai_code"], 5)

    def test_bai2_content_sniff_routes_unnamed_txt(self):
        # A raw BAI2 whose FILENAME carries no BAI token (real bank export
        # "BAIEXP_07202026_071541.txt") is recognized by its 01/02/16 record
        # structure and routed as BAI2 — never silently dropped over a name.
        raw = "\n".join([
            "01,000000000,000000000,260720,1916,2663008,,,2/",
            "02,084000026,084000026,1,260718,,USD,2/",
            "16,142,5100,Z,25202003732875,08035701948,",
            "88,MERCHANT SERVICEDEPOSIT",
            "49,0,2/", "98,0,1,2/", "99,0,1,2/",
        ])
        p = os.path.join(self.d, "BAIEXP_07202026_071541.txt")
        with open(p, "w") as fh:
            fh.write(raw)
        self.assertIsNone(E.classify_file("BAIEXP_07202026_071541.txt"))
        self.assertTrue(E._looks_like_bai2(p))
        # a non-BAI2 .txt must NOT sniff as BAI2
        q = os.path.join(self.d, "notes.txt")
        with open(q, "w") as fh:
            fh.write("just some notes\nnothing structured\n")
        self.assertFalse(E._looks_like_bai2(q))
        # route_folder (needs a BSL present) routes the content-BAI2 as BAI2
        # and leaves the plain .txt skipped.
        _write_xlsx(os.path.join(self.d, "20260720_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", [
                        ("Date", "Amount (USD)", "Reference", "Additional Information",
                         "Account Servicer Reference", "Transaction Type",
                         "Statement", "Transaction Code"),
                        ("2026-07-20", "1.00", "NA", "X", "NA", "Miscellaneous",
                         "Line 1 , 2026-07-20", "174")])])
        skipped = []
        by_role = E.route_folder(self.d, skipped)
        self.assertIn("BAI2", by_role)
        self.assertTrue(any(f.filename == "BAIEXP_07202026_071541.txt"
                            for f in by_role["BAI2"]))
        self.assertTrue(any(s["file"] == "notes.txt" for s in skipped))

    def test_bai2_sheet_content_sniff_routes_unnamed_xlsx(self):
        # A BAI2 SPREADSHEET whose filename carries no BAI token
        # ("20260720_FHB_Master.xlsx", sheet 'CSVEXP_...') is recognized by its
        # 'BAI Code' + date columns and routed as BAI2.
        _write_xlsx(os.path.join(self.d, "20260720_FHB_Master.xlsx"),
                    [("CSVEXP_07202026", [
                        ("Post Date", "Bank ID", "Account Number", "Account Name",
                         "Transaction Description", "Amount", "Bank Reference",
                         "Customer Reference", "BAI Code", "Currency",
                         "Debit/Credit", "DETAIL1"),
                        ("2026-07-20", "084000026", "15", "UT General Acct",
                         "ACH CREDIT", "500000", "ACH260720", "50571", "142",
                         "USD", "Credit", "ASAP GRANT")])])
        self.assertIsNone(E.classify_file("20260720_FHB_Master.xlsx"))
        self.assertTrue(E._looks_like_bai2_sheet(
            os.path.join(self.d, "20260720_FHB_Master.xlsx")))
        # a non-BAI2 spreadsheet must NOT sniff
        _write_xlsx(os.path.join(self.d, "random_export.xlsx"),
                    [("s", [("Foo", "Bar"), ("1", "2")])])
        self.assertFalse(E._looks_like_bai2_sheet(
            os.path.join(self.d, "random_export.xlsx")))
        _write_xlsx(os.path.join(self.d, "20260720_Oracle_CM_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", [
                        ("Date", "Amount (USD)", "Reference", "Additional Information",
                         "Account Servicer Reference", "Transaction Type",
                         "Statement", "Transaction Code"),
                        ("2026-07-20", "1.00", "NA", "X", "NA", "Miscellaneous",
                         "Line 1 , 2026-07-20", "174")])])
        skipped = []
        by_role = E.route_folder(self.d, skipped)
        self.assertIn("BAI2", by_role)
        self.assertTrue(any(f.filename == "20260720_FHB_Master.xlsx"
                            for f in by_role["BAI2"]))
        self.assertTrue(any(s["file"] == "random_export.xlsx" for s in skipped))

    def test_cfg_tcr_orphan_rows_load(self):
        # The real export carries rules with a BLANK bank-account cell
        # (detached rules, 3 still enabled) — they must load, not be
        # silently dropped as spacer rows (doctrine rule 3).
        rows = [("",) * 16] * 6 + [
            ("", "BNK ACCNTNME", "ENBLD", "SQC NUM", "RLE NME", "DSCRPT",
             "TRX CDE", "TRX TPE", "SRCH FLD", "SRCH STRNG", "CASH", "OFFSET",
             "ACCNT FLG", "LST UPDTE", "UPDTEBY", ""),
            ("", "FHB - UTHSC", "Y", "1", "normal rule", "d", "142", "ACH",
             "", "", "01-x", "", "Y", "", "A", ""),
            ("", "", "Y", "", "DIR 555 orphan rule", "d", "555", "MSC",
             "", "", "", "", "Y", "", "A", ""),
        ]
        p = os.path.join(self.d, "CM_Configurations_Transaction_Creation_Rules.xlsx")
        _write_xlsx(p, [("Sheet1", rows)])
        rules = E.load_cm_config("CFG_TCR", E.RoutedFile("CFG_TCR", p, os.path.basename(p), "first"))
        self.assertEqual(len(rules), 2)
        orphan = [r for r in rules if not r["bank_account"]]
        self.assertEqual(len(orphan), 1)
        self.assertEqual(orphan[0]["name"], "DIR 555 orphan rule")

    def test_recommend_gl_offset_only_and_foreign_excluded(self):
        # Adversarial-review fixes: the CoA fallback recommends ONLY the
        # OFFSET combo (the ECT posting side) and never cites a
        # foreign-account shadow entry (rule 8g).
        coa = {"combo_decode": {}, "entity_desc": {"01": "UT System"},
               "postable_efdp": set()}
        loaded = {"CHART_OF_ACCOUNTS": coa}
        b = E.make_bsl("1", date(2026, 7, 1), 25000, "R", "R", "", "Miscellaneous", "174")
        # foreign shadow entry with offset combo: must NOT be recommended
        foreign = E._mk_entry("F1", 25000, date(2026, 7, 1), "X", "", "EXT",
                              "UNR", True, "MET",
                              offset_segments="70-1100001-000000-461000-000-0000-00-0000")
        foreign.foreign_account = "FHB_UTHSC"
        foreign.available = False
        self.assertEqual(E.recommend_gl_string(b, loaded, [foreign]), "")
        # asset-only entry (no offset combo): nothing to recommend
        asset_only = E._mk_entry("A1", 25000, date(2026, 7, 1), "X", "", "EXT",
                                 "UNR", True, "MET",
                                 asset_segments="01-1100001-000000-100210-000-0000-00-0000")
        self.assertEqual(E.recommend_gl_string(b, loaded, [asset_only]), "")
        # in-account entry with offset combo: recommended, entity decoded
        ok = E._mk_entry("K1", 25000, date(2026, 7, 1), "X", "", "EXT",
                         "UNR", True, "MET",
                         offset_segments="01-1100001-011413-546500-260-0000-00-0000")
        got = E.recommend_gl_string(b, loaded, [asset_only, foreign, ok])
        self.assertIn("01-1100001-011413-546500-260-0000-00-0000", got)
        self.assertIn("UT System", got)

    def test_config_audit_end_to_end(self):
        # CFG files present: placements byte-identical, artifacts written,
        # creation-failure vs uncovered classification correct.
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            # claimed by the enabled TCR, but NO pool entry -> creation failure
            ("2026-07-16", "290.50", "NA", "USDA TREAS 310 MISC PAY", "626001636",
             "Automated clearing house", "Line 1 , 2026-07-16", "142"),
            # recurring uncovered signature (x2)
            ("2026-07-15", "-100.00", "NA", "STATE-TNRECEIPTSTNRECEIPTS 1", "NA",
             "Automated clearing house", "Line 2 , 2026-07-15", "451"),
            ("2026-07-16", "-200.00", "NA", "STATE-TNRECEIPTSTNRECEIPTS 2", "NA",
             "Automated clearing house", "Line 3 , 2026-07-16", "451"),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        st = [("Date", "Amount (USD)", "Reference", "Transaction Number", "Source", "Counterparty"),
              ("2026-07-02", "999.00", "OTHER", 601, "External", "V")]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        tcr = [("",) * 16] * 6 + [
            ("", "BNK ACCNTNME", "ENBLD", "SQC NUM", "RLE NME", "DSCRPT",
             "TRX CDE", "TRX TPE", "SRCH FLD", "SRCH STRNG", "CASH", "OFFSET",
             "ACCNT FLG", "LST UPDTE", "UPDTEBY", ""),
            ("", "FHB - Master Account", "Y", "334", "142-FHB - Master Account-626001636",
             "d", "142", "ACH", "ACCOUNT_SERV_REFERENCE", "626001636",
             "01-1100098-000000-100210-000-0000-00-0000", "", "Y", "", "A", ""),
        ]
        _write_xlsx(os.path.join(self.d, "CM_Configurations_Transaction_Creation_Rules.xlsx"),
                    [("Sheet1", tcr)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        ca = runlog["config_audit"]
        self.assertEqual(ca["tcr"]["creation_failures"], 1)
        self.assertEqual(ca["tcr"]["uncovered_review_lines"], 2)
        self.assertEqual(ca["tcr"]["uncovered_recurring_signatures"], 1)
        self.assertTrue(os.path.exists(ca["report_path"]))
        self.assertTrue(os.path.exists(ca["json_path"]))
        text = open(ca["report_path"]).read()
        self.assertIn("creation FAILURES", text)
        self.assertIn("626001636", text)
        self.assertIn("STATE-TNRECEIPTS", text)
        self.assertIn("SIMULATED", text)

    def test_edison_annotation(self):
        # Edison (State of TN) exports annotate stranded State lines with the
        # payment reference/invoice — text only, placements untouched.
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            # ref-tied Edison payment (payment ref digits in the ASR)
            ("2026-07-08", "586.66", "NA", "STATE-TN PAYMNTS", "0007041551",
             "Automated clearing house", "Line 1 , 2026-07-08", "142"),
            # amount matches TWO payments, no ref tie -> never guessed
            ("2026-07-09", "400.00", "NA", "STATE-TN PAYMNTS", "NA",
             "Automated clearing house", "Line 2 , 2026-07-09", "142"),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        st = [("Date", "Amount (USD)", "Reference", "Transaction Number", "Source", "Counterparty"),
              ("2026-07-02", "999.00", "OTHER", 601, "External", "V")]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        _write_xlsx(os.path.join(self.d, "Edison_Payments.xlsx"), [("sheet1", [
            ("Reference", "Invoice Number", "Payment Date", "Amount", "Currency"),
            ("0007041551", "12052025", "2026-07-07", "586.66", "USD"),
            ("0007000001", "INV-A", "2026-07-07", "400.00", "USD"),
            ("0007000002", "INV-B", "2026-07-08", "400.00", "USD"),
        ])])
        _write_xlsx(os.path.join(self.d, "Edison_Invoices.xlsx"), [("sheet1", [
            ("Invoice Number", "Invoice Date", "Gross Amt", "Currency",
             "Approval Status", "Due Date", "Voucher"),
            ("12052025", "2026-06-30", "586.66", "USD", "Approved",
             "2026-07-15", "00169979"),
        ])])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        self.assertEqual(runlog["recon_summary"]["reviews"], 2)
        tabs, _ = A._read_output_tabs(runlog["recon_workbook"])
        rev = {A._N(r[2]): A._N(r[A.COL_EXPL])
               for r in tabs["Review Notes"][3:] if any(A._N(c) for c in r)}
        self.assertIn("Edison: State of TN payment 0007041551", rev["586.66"])
        self.assertIn("invoice '12052025'", rev["586.66"])
        self.assertIn("Approved", rev["586.66"])
        self.assertIn("voucher 00169979", rev["586.66"])
        # ambiguous amount-only: no Edison note at all
        self.assertNotIn("Edison", rev["400.00"])

    def test_gms_aging_annotation(self):
        # GMS Sponsored AR Aging (owner, 2026-07-20): annotate stranded
        # sponsored-AR lines with sponsor/award/aging context.  Advisory only
        # (placements untouched); GMS rows never place.  The synthetic report
        # carries a realistic multi-row preamble (incl. an "Invoice Number:"
        # filter-label collision row) ABOVE the real header past row 12 — so
        # this also pins the widened header scan.
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            # (1) invoice-number tie: the 9-digit invoice is in the addenda
            ("2026-07-08", "5121.79", "NA",
             "ACH CREDIT SPONSORED INV 110012367 FEDERAL DRAWDOWN", "NA",
             "Automated clearing house", "Line 1 , 2026-07-08", "142"),
            # (2) amount + sponsor-name tie (no invoice number in the text)
            ("2026-07-09", "35205.00", "NA",
             "AAA FOUNDATION FOR TRAFFIC SAFETY REMITTANCE", "NA",
             "Automated clearing house", "Line 2 , 2026-07-09", "142"),
            # (3) amount collides with TWO sponsors, no tie -> never guessed
            ("2026-07-10", "4000.00", "NA", "MISC DEPOSIT", "NA",
             "Automated clearing house", "Line 3 , 2026-07-10", "142"),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        st = [("Date", "Amount (USD)", "Reference", "Transaction Number", "Source", "Counterparty"),
              ("2026-07-02", "999.00", "OTHER", 601, "External", "V")]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        H = ("Bill to Customer Number", "Bill to Customer Name", "Account Number",
             "Account Name", "Legal Entity", "Award Number", "Award Type",
             "Grants Administrator", "Invoice Number", "Owning Org", "Award PI",
             "Payment Terms", "Invoice Date", "Invoiced Amount", "Amount Applied Pay",
             "Amount Applied CM", "Amount Outstanding", "Days Outstanding",
             "0 to 30", "31 to 60", "61 to 90", "91 to 120", "Over 120")
        blank = tuple("" for _ in H)
        preamble = [
            ("RPT1 - Sponsored AR Aging Report",) + blank[1:],
            blank, blank, blank, blank, blank,
            ("Business Unit:", "", "Legal Entity:", "", "Customer:") + blank[5:],
            blank,
            # the collision row: an "Invoice Number:" / "Award Number:" filter label
            ("Grants Administrator:", "", "Award Number:", "", "Invoice Number:") + blank[5:],
            ("Aging Bucket:", "", "As of Date:", "07/20/2026") + blank[4:],
            blank, blank,          # header now lands at row 13 (0-indexed)
        ]
        data = [
            ("432853", "1890 Universities Foundation", "308366", "1890 Univ Fo",
             "UT Extension", "2101045", "Federal", "Stacy Keisling", "110012367",
             "180008-Center of Farm Mgmt", "Tori Griffin", "IMMEDIATE",
             "05/18/2026", "5121.79", "0.0", "0.0", "5121.79", "63",
             "0.0", "0.0", "5121.79", "0.0", "0.0"),
            ("146302", "AAA Foundation for Traffic Safety", "202157", "AAA Fdn",
             "UT Knoxville", "2004896", "Private", "Eva Vickers", "110012912",
             "100430-Ctr Transport", "Christopher Cherry", "IMMEDIATE",
             "06/03/2026", "35205.00", "0.0", "0.0", "35205.00", "47",
             "0.0", "35205.00", "0.0", "0.0", "0.0"),
            # two DIFFERENT sponsors at $4000 -> amount-only ambiguity
            ("100069", "American Diabetes Association", "200072", "ADA",
             "UT Knoxville", "2100737", "Private", "Eva Vickers", "110014190",
             "100227-Kinesiology", "Lyndsey H", "IMMEDIATE", "06/30/2026",
             "4000.00", "0.0", "0.0", "4000.00", "20", "4000.00", "0.0", "0.0", "0.0", "0.0"),
            ("100104", "AMETEK Advanced", "200107", "AMETEK", "UT Knoxville",
             "2101345", "Private", "Eva Vickers", "110012052", "100426-Mech",
             "Anming Hu", "IMMEDIATE", "05/04/2026", "4000.00", "0.0", "0.0",
             "4000.00", "77", "0.0", "0.0", "4000.00", "0.0", "0.0"),
        ]
        _write_xlsx(
            os.path.join(self.d, "RPT_GMS_001__Sponsored_AR_Aging_Report_RPT1.xlsx"),
            [("Sheet1", preamble + [H] + data)])
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        # GMS is advisory only: all three lines remain Reviews (no placement).
        self.assertEqual(runlog["recon_summary"]["reviews"], 3)
        self.assertEqual(runlog["recon_summary"]["matches"], 0)
        self.assertEqual(runlog["sponsored_projects"]["invoices"], 4)
        tabs, _ = A._read_output_tabs(runlog["recon_workbook"])
        rev = {A._N(r[2]): A._N(r[A.COL_EXPL])
               for r in tabs["Review Notes"][3:] if any(A._N(c) for c in r)}
        # (1) invoice-number tie names the invoice + sponsor + award
        self.assertIn("sponsored invoice 110012367", rev["5,121.79"])
        self.assertIn("1890 Universities Foundation", rev["5,121.79"])
        self.assertIn("award 2101045 Federal", rev["5,121.79"])
        # (2) amount + sponsor-name tie
        self.assertIn("sponsored invoice 110012912", rev["35,205.00"])
        self.assertIn("AAA Foundation for Traffic Safety", rev["35,205.00"])
        # (3) amount collides with two sponsors, no tie -> never guessed
        self.assertNotIn("GMS:", rev["4,000.00"])

    def test_gms_aging_absent_is_noop(self):
        # With no sponsored report present, the annotation is a pure no-op.
        b = E.make_bsl("L1", date(2026, 7, 10), 5121, "110012367", "110012367",
                       "INV 110012367", "Automated clearing house", "142")
        self.assertEqual(E._sponsored_note(b, None), "")
        self.assertIsNone(E._sponsored_index({}))

    def test_award_conversion_sponsor_ref_annotation(self):
        # Active Oracle Award Conversion Report (owner, 2026-07-21): a federal
        # deposit cites the SPONSOR's OWN contract number, not our internal SPN.
        # The report's `Sponsor Number` column resolves that contract number to
        # our SPN/award.  Pipe-delimited .txt; advisory annotation only.
        self.assertEqual(
            E.classify_file("Active_Oracle_Award_Conversion_Report.txt"),
            "GMS_AWARD_CONVERSION")
        bsl = [
            ("Date", "Amount (USD)", "Reference", "Additional Information",
             "Account Servicer Reference", "Transaction Type", "Statement", "Transaction Code"),
            # the addenda carries the sponsor's contract number, NOT an SPN
            ("2026-07-16", "4500.00", "NA",
             "USDA FOREST SERVICE PMT AP25PPQFO000C140 DRAWDOWN", "NA",
             "Automated clearing house", "Line 5 , 2026-07-16", "142"),
        ]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_BSL_UNR.xlsx"),
                    [("Exported", bsl)])
        st = [("Date", "Amount (USD)", "Reference", "Transaction Number", "Source", "Counterparty"),
              ("2026-07-02", "999.00", "OTHER", 601, "External", "V")]
        _write_xlsx(os.path.join(self.d, "20260710_FHB_Master_ST_UNR.xlsx"),
                    [("Exported", st)])
        # pipe-delimited report (sniffed by the reader); one active award whose
        # Sponsor Number is the contract number the bank line cites.
        conv = [
            ["Project Number", "Award Number", "Award Name", "Sponsor Number",
             "ERA Award Number", "Award PI"],
            ["SPN112933", "2100490", "USDA APHIS TN F25 Field Crop",
             "AP25PPQFO000C140", "A012345", "Jerome Grant"],
        ]
        with open(os.path.join(self.d, "Active_Oracle_Award_Conversion_Report.txt"),
                  "w", encoding="utf-8") as fh:
            fh.write("\n".join("|".join(r) for r in conv) + "\n")
        runlog = E.run(self.d, self.out, present=True)
        self.assertEqual(runlog["audit"]["status"], "PASS",
                         msg=str(runlog["audit"].get("failures")))
        tabs, _ = A._read_output_tabs(runlog["recon_workbook"])
        rev = {A._N(r[2]): A._N(r[A.COL_EXPL])
               for r in tabs["Review Notes"][3:] if any(A._N(c) for c in r)}
        note = rev.get("4,500.00", "")
        self.assertIn("SPN SPN112933", note)
        self.assertIn("award 2100490", note)
        self.assertIn("sponsor ref AP25PPQFO000C140", note)

    def test_run_recon_stages_bai2_txt(self):
        # A native BAI2 .txt must be staged by the per-run wrapper; any other
        # .txt stays ignored (preserves ignored_non_spreadsheets semantics).
        self.assertTrue(R._stageable("20260718_FHB_Master_BAI2.txt"))
        self.assertFalse(R._stageable("readme.txt"))
        self.assertTrue(R._stageable("20260710_FHB_Master_BSL_UNR.xlsx"))

    def test_student_refund_uthsc_utm_utso_accounts(self):
        # Config export exposed three more Student Refund depositories; the
        # generic campus token must not swallow them (misdirected scope, 8g).
        self.assertEqual(E.account_of_bank_name("FHB - Student Refund - UTHSC"),
                         "FHB_STUDENT_REFUND_UTHSC")
        self.assertEqual(E.account_of_bank_name("FHB - Student Refund - UTM"),
                         "FHB_STUDENT_REFUND_UTM")
        self.assertEqual(E.account_of_bank_name("FHB - Student Refund - UTSO"),
                         "FHB_STUDENT_REFUND_UTSO")
        self.assertEqual(E.account_of_bank_name("FHB - UTHSC"), "FHB_UTHSC")
        self.assertEqual(E.account_of_bank_name("FHB - Accounts Payable"), "FHB_AP")


if __name__ == "__main__":
    unittest.main(verbosity=2)
