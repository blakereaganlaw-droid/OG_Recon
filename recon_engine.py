#!/usr/bin/env python3
"""
UT Cash Management Reconciliation Engine
========================================

A production-grade, deterministic reconciliation engine for University of
Tennessee bank accounts in Oracle Cash Management (DASH).

Forward matching engine: match open bank statement lines (BSL) to open
system transactions (ST), classifying each line Match / Candidate / Review.
Every available source feeds the candidate pool — ST exports, Receivables
receipts, and the MET/ORT chain (d:/r: deposit-receipt bridge).

The BACKWARD engine (re-audit reconciled groups -> recommend unwinds) was
split into the separate Unreconcile2 project so this engine stays fast:
an ALL_DATA workbook is recognized but never loaded here.

Temperament (binding, see BUILD SPEC sections 0 & 15):
  * Determinism first — same inputs, identical outputs, no float, no clock.
  * Fail loud, never silent — a missing required column/file/relationship
    raises a named exception (InvalidSourceData / MissingRequiredFile /
    AmbiguousColumn) naming the file, the role, and the candidates.
  * Every input row is accounted for exactly once.
  * The pipeline is a fixed sequence of stages; a later stage never overrides
    an earlier one; availability is re-derived inside each loop.

Dependencies: Python 3.10+ standard library + openpyxl + decimal only.
No pandas (it float-coerces zero-padded references and merchant IDs).

This module is self-contained (portable to Grok / ChatGPT / Perplexity as
either executable code or an operating procedure). The independent audit lives
in recon_audit.py and imports nothing from here.
"""

from __future__ import annotations

import argparse
import functools
import json
import os
import re
import sys
from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

# openpyxl is imported lazily inside the workbook writer / loader helpers so
# that the primitives and pure logic can be imported and unit-tested without it.


# ======================================================================
# Named exceptions (Section 0.2)
# ======================================================================

class ReconError(Exception):
    """Base class for every engine error."""


class InvalidSourceData(ReconError):
    def __init__(self, file, role, detail):
        self.file, self.role, self.detail = file, role, detail
        super().__init__(f"InvalidSourceData(file={file!r}, role={role!r}, detail={detail})")


class MissingRequiredFile(ReconError):
    def __init__(self, role, detail=""):
        self.role, self.detail = role, detail
        super().__init__(f"MissingRequiredFile(role={role!r}, detail={detail})")


class AmbiguousColumn(ReconError):
    def __init__(self, file, role, candidates):
        self.file, self.role, self.candidates = file, role, candidates
        super().__init__(
            f"AmbiguousColumn(file={file!r}, role={role!r}, candidates={candidates})"
        )


# ======================================================================
# Section 7 — Normalization primitives (implement first; unit-tested)
# ======================================================================

_MID_RE = re.compile(r"^(80\d{8}|2000\d{6})$")
_HEARTLAND_MID = "6500000097"
_SPN_RE = re.compile(r"SPN\s*\d+", re.IGNORECASE)
_DIGIT_RUN_RE = re.compile(r"\d+")
_NON_ALNUM_RE = re.compile(r"[^A-Za-z0-9]+")
_ALPHA_TOKEN_RE = re.compile(r"[A-Za-z]{4,}")

# Generic stopwords removed from payer token sets.  Explicitly includes the
# BAI2 field labels (never tokenize them as payer text) and channel words.
PAYER_STOPWORDS = {
    # BAI2 / addenda structural labels
    "COMPANY", "CUSTOMER", "ENTRY", "DESCRIPTION", "DESC", "NAME", "ID",
    "SENDING", "REF", "REFERENCE", "ADDENDA", "INFO", "INFORMATION", "NUMBER",
    "ACCOUNT", "TRACE", "BATCH", "INDIVIDUAL", "IDENTIFICATION",
    # 2026-07-12 (adversarial verification): 'Class Code: CCD' in BAI2
    # addenda made CLASS overlap 'Restorative Class 6' and fabricated a
    # payer tie on two UTHSC candidates — structural labels, all of them.
    "CLASS", "CODE", "CUST", "RECEIVED",
    # channel / generic words
    "ACH", "WIRE", "DEPOSIT", "DEPOSITS", "PAYMENT", "PAYMENTS", "MERCHANT",
    "SERVICE", "SERVICES", "STATE", "TENNESSEE", "TENN", "UNIVERSITY",
    "CREDIT", "DEBIT", "CHECK", "MISC", "MISCELLANEOUS", "TRANSFER", "ZBA",
    "SETTLEMENT", "BANKCARD", "CARD", "FUNDS", "TRANSACTION", "TRAN", "EFT",
    "THE", "AND", "FOR", "FROM", "WITH", "INC", "LLC", "CORP", "CO",
}


def N(v) -> str:
    """Null-safe string: '' if None else str(v).strip()."""
    return "" if v is None else str(v).strip()


def cents(v):
    """Signed integer cents via Decimal.  Handles '$', thousands commas, and
    (parentheses) = negative.  Returns None if unparseable (never raises)."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return int(Decimal(v) * 100)
    if isinstance(v, float):
        # Route through str to avoid binary float noise; spec forbids float
        # math but Excel hands us floats, so normalize deterministically.
        s = repr(v)
    else:
        s = str(v).strip()
    if s == "":
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1].strip()
    s = s.replace("$", "").replace(",", "").replace(" ", "")
    if s.startswith("-"):
        neg = True
        s = s[1:]
    elif s.startswith("+"):
        s = s[1:]
    if s == "":
        return None
    try:
        d = (Decimal(s) * 100).quantize(Decimal("1"))
    except (InvalidOperation, ValueError):
        return None
    c = int(d)
    return -c if neg else c


_DATE_FORMATS = ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%d-%b-%Y", "%d-%b-%y")
# Excel serial epoch (1899-12-30 accounts for the Lotus 1900 leap-year bug).
_EXCEL_EPOCH = date(1899, 12, 30)


def parse_date(v):
    """Return a datetime.date, or None if unparseable.

    Accepts date, datetime (checked FIRST because datetime subclasses date),
    ISO strings incl. '...000+00:00', US formats, and Excel serial numbers.
    """
    if v is None:
        return None
    # datetime BEFORE date (Section 6 guard).
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        # Excel serial date.  Reject implausible small/large numbers.
        n = int(v)
        if 1 <= n <= 60000:
            try:
                return _EXCEL_EPOCH.fromordinal(_EXCEL_EPOCH.toordinal() + n)
            except (OverflowError, ValueError):
                return None
        return None
    s = str(v).strip()
    if s == "":
        return None
    # ISO with time / timezone, e.g. 2024-03-05T00:00:00.000+00:00
    m = re.match(r"^(\d{4}-\d{2}-\d{2})[T ]", s)
    if m:
        try:
            return date.fromisoformat(m.group(1))
        except ValueError:
            pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Bare numeric string could be an Excel serial exported as text.
    if s.isdigit():
        n = int(s)
        if 1 <= n <= 60000:
            return _EXCEL_EPOCH.fromordinal(_EXCEL_EPOCH.toordinal() + n)
    return None


@functools.lru_cache(maxsize=1 << 18)
def znorm(s) -> str:
    """Uppercase, strip non-alphanumerics (for reference equality)."""
    return _NON_ALNUM_RE.sub("", N(s)).upper()


def digit_runs(s, n=5) -> set:
    """Set of maximal digit substrings of length >= n (for reference-tie)."""
    return {r for r in _DIGIT_RUN_RE.findall(N(s)) if len(r) >= n}


def payer_tokens(s) -> set:
    """Set of alpha tokens length>=4, minus the generic stopword set."""
    return {t.upper() for t in _ALPHA_TOKEN_RE.findall(N(s))
            if t.upper() not in PAYER_STOPWORDS}


def spn_of(s) -> str:
    """The SPN token (regex (?i)SPN\\s*\\d+) found in a receipt/transaction
    number, else ''.  Normalized to 'SPN' + digits (no internal space)."""
    m = _SPN_RE.search(N(s))
    if not m:
        return ""
    return "SPN" + re.sub(r"\D", "", m.group(0))


def is_mid(s) -> bool:
    """True if the znorm digits match the MID pattern and != Heartland id."""
    z = znorm(s)
    if z == _HEARTLAND_MID:
        return False
    return bool(_MID_RE.match(z))


def sibling(a, b) -> bool:
    """True if a,b are equal-length pure-numeric of length>=7 differing only
    in the last 1-2 digits (a conflict, never a tie)."""
    za, zb = znorm(a), znorm(b)
    if not (za.isdigit() and zb.isdigit()):
        return False
    if len(za) != len(zb) or len(za) < 7:
        return False
    if za == zb:
        return False
    # Prefixes must match up to the last two digits.
    if za[:-2] != zb[:-2]:
        return False
    # And they differ somewhere in the final 1-2 positions (guaranteed since
    # za != zb and prefixes to -2 are equal).
    return True


def signed_lag(bsl_date, st_date):
    """(BSL date - ST date).days, or None if either is missing."""
    if bsl_date is None or st_date is None:
        return None
    return (bsl_date - st_date).days


# ---- Reference equality / tie (Section 7 trailer) --------------------

_NULL_REF_TOKENS = {"NA", "NONE", "NULL", "UNKNOWN"}


def clean_ref(v):
    """Null out placeholder reference tokens: a literal 'NA' must never tie
    another literal 'NA' (37 of 283 real FHB Master BSL references are 'NA')."""
    s = N(v)
    if znorm(s) in _NULL_REF_TOKENS or not re.search(r"[A-Za-z0-9]", s):
        return ""
    return s


def _blankish(v):
    """A placeholder cell ('NA', '.', '-') is blank for content sampling:
    real Oracle exports pad id columns with them, and counting them dilutes
    the true column's content score below junk text columns."""
    s = N(v)
    return not s or znorm(s) in _NULL_REF_TOKENS or not re.search(r"[A-Za-z0-9]", s)


def _sibling_z(za, zb) -> bool:
    """`sibling` over PRE-COMPUTED znorms (perf fast path; identical result —
    sibling is a pure function of the two znorms)."""
    if not (za.isdigit() and zb.isdigit()):
        return False
    if len(za) != len(zb) or len(za) < 7:
        return False
    if za == zb:
        return False
    return za[:-2] == zb[:-2]


def _reference_equal_z(za, zb) -> bool:
    """`reference_equal` over PRE-COMPUTED znorms (perf fast path)."""
    if not za or not zb:
        return False
    hit = (za == zb) or (len(za) >= 6 and za in zb) or (len(zb) >= 6 and zb in za)
    if not hit:
        return False
    return not _sibling_z(za, zb)


def reference_equal(a, b) -> bool:
    """znorm equality OR full containment of one znorm token inside the other
    for length >= 6.  A sibling pair is a conflict, never equal."""
    return _reference_equal_z(znorm(a), znorm(b))


def reference_tie(a, b) -> bool:
    """Weaker corroboration: a shared digit_run of length >= 5 (truncation
    allowed).  A sibling pair never ties."""
    if sibling(a, b):
        return False
    ra, rb = digit_runs(a, 5), digit_runs(b, 5)
    if ra & rb:
        return True
    # Truncation allowed: one run is a prefix of the other (length >= 5).
    for x in ra:
        for y in rb:
            if len(x) >= 5 and len(y) >= 5 and (x.startswith(y) or y.startswith(x)):
                return True
    return False


# ======================================================================
# Section 4 — File router.  Identify every file by NAME, then bind by header.
# ======================================================================

@dataclass
class RouterRule:
    role: str
    contains: list          # all tokens must appear (case-insensitive)
    excludes: list          # none may appear
    any_of: list            # at least one must appear (empty = no constraint)
    sheet: str
    required: bool


# First matching rule wins (Section 4.1).  `any_of` handles "several filename
# shapes"; `excludes` handles the "not All_Data" style negative constraints.
ROUTER_TABLE = [
    # ALL_BSL (owner, 2026-07-19): the all-accounts open bank-statement-line
    # export (OTBI) — feeds the misdirected search's foreign-BSL index, NOT
    # this account's BSL reconciliation.  Must precede BSL/MET (it embeds both
    # 'bsl' and 'oracle_otbi'); the single-account BSL never carries an 'all'
    # segment, so it still binds BSL.
    # RECONCILED forensic exports (owner uploads, 2026-07-19): reconciled
    # BSL/ST/receipt/history reports.  Recognized so they can NEVER be
    # misread as the open exports (a "Reconciled_..._Student_Refund_BSL"
    # file used to bind the hard-required BSL role and would have poisoned
    # the whole run with already-reconciled lines) — but never loaded by
    # the forward engine (they fuel the config audit's offline analysis
    # and Unreconcile2).  Must precede ALL_BSL/BSL/ST/MET.
    RouterRule("RECONCILED", [], [], ["reconciled", "reconciliation_report"], "multi", False),
    RouterRule("ALL_BSL", ["all", "bsl"], ["all_data", "enriched", "reconciled"], [], "first", False),
    # 'enriched' excluded: an Enriched_..._BSL_... workbook is the parsed
    # enrichment source (ENRICHED below), not a competing BSL export.
    RouterRule("BSL", ["bsl"], ["all_data", "enriched", "reconciled"], [], "first", True),
    RouterRule("ALL_DATA", ["all_data"], [], [], "multi", False),
    # 'oracle_otbi' also routes here: OTBI is the MET report's source system
    # and real exports sometimes omit the MET token
    # (20260711_Oracle_OTBI_Regions_UTIA_All_Status.xlsx).
    RouterRule("MET", [], [], ["met", "oracle_otbi"], "first", False),
    # "_st" alone is too greedy: it matches All_Status / Rosetta_Stone /
    # _Statement.  Require a separator (or end) after the token.
    RouterRule("ST", [], [], ["_st_", "_st.", "account_st"], "first", False),
    # 'ar_matched' excluded: "AR_Matched_Invoice_Receipts_AR_Deposit_Receipts_
    # All_NonMisc" embeds 'receipts_all' but is the receipt-APPLICATION feed
    # (ACRA/ABA), not a receipts export — it routes to AR_MATCHED below.
    RouterRule("RECEIPTS", [], ["ar_matched"], ["receivables_receipts", "receipts_all", "oracle_receipts"], "Export to Excel", False),
    # PAYMENTS (owner, 2026-07-19): the AP payment feed — the Payables
    # analogue of RECEIPTS.  Requires BOTH tokens so Edison_Payments (no
    # 'payables') and account files never match.
    RouterRule("PAYMENTS", ["payables", "payments"], [], [], "first", False),
    RouterRule("DEPT_INFO", [], [], ["ort_department", "department_info"], "Report", False),
    # CHART_OF_ACCOUNTS (owner COA export, 2026-07-19): the combination-universe
    # reference bundle — the seven AcctCombos shards, Segments.csv, ComboSets /
    # CombosTech.  A multi-file role; load_chart_of_accounts dispatches each file
    # by name and skips the rest (RelatedValueSets, the ORT_Activity routing
    # table).  Placed AFTER DEPT_INFO so ORT_Department_* still bind DEPT_INFO.
    RouterRule("CHART_OF_ACCOUNTS", [], [],
               ["chart_of_accounts", "gl_departments", "acctcombos",
                "combosets", "combostech", "segments", "relatedvaluesets"],
               "Report", False),
    RouterRule("ORT_AR", ["ort", "_ar"], [], [], "Report", False),
    RouterRule("ORT_MISC", ["ort", "misc"], [], [], "Report", False),
    RouterRule("BAI2", ["bai"], [], [], "first", False),
    RouterRule("EDISON_PAY", ["edison_payments"], [], [], "first", False),
    RouterRule("EDISON_INV", ["edison_invoices"], [], [], "first", False),
    RouterRule("MID_MASTER", ["mid_master"], [], [], "all", False),
    RouterRule("ENRICHED", [], [], ["enriched", "crossref"], "first", False),
    RouterRule("APPLIED_UNAPPLIED", ["applied", "unapplied"], [], [], "first", False),
    RouterRule("CONTRACTS_INV", ["contracts_to_receivable_invoices"], [], [], "first", False),
    RouterRule("GMS_AGING", [], [], ["gms_001", "sponsored_aging"], "first", False),
    RouterRule("AR_INVOICES", ["ar_invoices"], [], [], "first", False),
    RouterRule("AR_MATCHED", [], [], ["ar_matched", "deposit_receipts"], "first", False),
    RouterRule("AR_UNAPPLIED_SUMMARY", [], [], ["ar_063", "unapplied_receipts_summary"], "first", False),
    RouterRule("GMS_SPONSOR_MAP", ["rpt_gms_0"], [], [], "first", False),
    RouterRule("CFG_MATCHING", ["matching_rules"], [], [], "first", False),
    RouterRule("CFG_PARSE", ["parse_rules"], [], [], "first", False),
    RouterRule("CFG_TOLERANCE", ["tolerance_rules"], [], [], "first", False),
    RouterRule("CFG_RULESETS", ["recon_rulesets"], [], [], "first", False),
    RouterRule("CFG_TCR", ["transaction_creation_rules"], [], [], "first", False),
    RouterRule("RELATIONSHIP_MAP", [], [], ["relationship_map", "rosetta"], "reference", False),
]

# Roles the pipeline treats as hard requirements at the top level (Section 4.1
# "Required?" column reads "yes"/conditional).  BSL is the sole unconditional
# requirement.  ALL_DATA is recognized but not loaded — it fuels the backward
# engine, which lives in the separate Unreconcile2 project.
HARD_REQUIRED_ROLES = {"BSL"}

# Account tokens recognized in filenames (Section 1.2 + 4.1).
_ACCOUNT_TOKENS = [
    # Student Refund accounts (owner, 2026-07-19) — distinct depository
    # statements ("FHB - Student Refund - UTK/UTC").  Listed FIRST so the
    # 4-token match wins over the generic campus token (a plain FHB UTC file
    # carries neither "student" nor "refund", so it still binds FHB_UTC).
    ("FHB_STUDENT_REFUND_UTK", ["fhb", "student", "refund", "utk"]),
    ("FHB_STUDENT_REFUND_UTC", ["fhb", "student", "refund", "utc"]),
    # Long-form campus spelling ("FHB_UT_Chatt_Student_Refund_BAI2") —
    # without it the file carries account None and the mixed-account
    # preflight guard is blind to it (critical review, 2026-07-19).
    ("FHB_STUDENT_REFUND_UTC", ["fhb", "student", "refund", "chatt"]),
    # The TCR config export (owner, 2026-07-19) names Student Refund
    # depositories for UTHSC/UTM/UTSO too — without these entries the
    # generic campus token would swallow them ("FHB - Student Refund -
    # UTHSC" -> FHB_UTHSC), poisoning the misdirected shadow scope (8g).
    ("FHB_STUDENT_REFUND_UTHSC", ["fhb", "student", "refund", "uthsc"]),
    ("FHB_STUDENT_REFUND_UTM", ["fhb", "student", "refund", "utm"]),
    ("FHB_STUDENT_REFUND_UTSO", ["fhb", "student", "refund", "utso"]),
    ("FHB_MASTER", ["fhb", "master"]),
    ("FHB_UTHSC", ["fhb", "uthsc"]),
    ("FHB_UTIA", ["fhb", "utia"]),
    ("FHB_UTC", ["fhb", "utc"]),
    ("FHB_UTC", ["fhb", "chatt"]),   # "FHB - UT Chatt" long form
    ("FHB_UTM", ["fhb", "utm"]),
    ("FHB_UTSO", ["fhb", "utso"]),
    ("FHB_AP", ["fhb", "ap"]),
    # Long-form bank name ("FHB - Accounts Payable"): "ap" only matches a
    # whole name segment, so the spelled-out form needs its own entry.
    ("FHB_AP", ["fhb", "accounts", "payable"]),
    ("REGIONS_UTIA", ["regions", "utia"]),
    ("REGIONS_UTIPS", ["regions", "utips"]),
    ("REGIONS_UTM", ["regions", "utm"]),
    ("REGIONS_MASTER", ["regions", "master"]),
    ("REGIONS_UTSI", ["regions", "utsi"]),
    ("REGIONS_UTHSC", ["regions", "uthsc"]),
]

_YYYYMMDD_RE = re.compile(r"(\d{8})")


def _name_segments(low):
    return [s for s in re.split(r"[^a-z0-9]+", low) if s]


def _token_hit(low, segments, tok):
    """Filename-token semantics: a token carrying a separator ('_st_',
    'all_data') or 5+ chars matches as a substring; a short bare token
    ('ar', 'st', 'met', 'bai') matches only a WHOLE name segment, with a
    numeric suffix allowed ('bai' matches 'bai2') — so 'ar' can never fire
    inside 'chart' or 'departments'."""
    if len(tok) >= 5 or any(ch in tok for ch in "_.-"):
        return tok in low
    for seg in segments:
        if seg == tok or (seg.startswith(tok) and seg[len(tok):].isdigit()):
            return True
    return False


def _tokens_present(fname_lower, tokens):
    segments = _name_segments(fname_lower)
    return all(_token_hit(fname_lower, segments, t) for t in tokens)


def classify_file(filename) -> str | None:
    """Return the internal role key for a filename, or None if unmatched.
    First matching router rule wins (Section 4.1)."""
    low = filename.lower()
    segments = _name_segments(low)
    for rule in ROUTER_TABLE:
        if rule.contains and not all(_token_hit(low, segments, t) for t in rule.contains):
            continue
        if rule.excludes and any(_token_hit(low, segments, t) for t in rule.excludes):
            continue
        if rule.any_of and not any(_token_hit(low, segments, t) for t in rule.any_of):
            continue
        if not rule.contains and not rule.any_of:
            continue  # a rule with no positive tokens matches nothing
        return rule.role
    return None


def infer_account(filename) -> str | None:
    """Normalize the account token from a BSL/ALL_DATA filename to short form."""
    low = filename.lower()
    for short, tokens in _ACCOUNT_TOKENS:
        if _tokens_present(low, tokens):
            return short
    return None


def account_of_bank_name(name):
    """Map a long bank-account name ("FHB - Master Account") to the short
    engine account (FHB_MASTER) — the scope join of ORT doc section 4.6.
    Returns None for names belonging to no configured account."""
    low = N(name).lower()
    for short, tokens in _ACCOUNT_TOKENS:
        if _tokens_present(low, tokens):
            return short
    return None


# GL cash-account map (owner COA export, 2026-07-18): the Chart of Accounts
# assigns each depository bank account a natural-account code, and the MET
# export stamps it in ASSET_CONCATENATED_SEGMENTS
# (ENTITY-FUND-DEPT-ACCOUNT-...).  This is an INDEPENDENT second scope key:
# real exports carry cross-account postings the bank-name join cannot see
# (38 rows on the UTHSC GL 100500 inside the Master MET; 34 rows on the
# Master GL 100210 inside the UTIA MET).  Only codes whose engine account
# exists in _ACCOUNT_TOKENS are mapped; clearing/payroll/CBORD GLs are
# deliberately absent (they are not depository statements).
_GL_CASH_ACCOUNTS = {
    "100210": "FHB_MASTER",
    "100221": "FHB_UTIA",
    "100226": "FHB_UTSO",
    "100310": "FHB_UTC",
    "100330": "REGIONS_UTM",
    "100335": "REGIONS_MASTER",
    "100350": "REGIONS_UTIPS",
    "100360": "REGIONS_UTIA",
    "100384": "REGIONS_UTSI",
    "100390": "REGIONS_UTHSC",
    "100500": "FHB_UTHSC",
}


def segments_of(segments):
    """Split a concatenated GL combination into its dash-delimited segments.
    The SINGLE split point for combo strings — positional layout
    [ENTITY, FUND, DEPT, ACCOUNT, PROGRAM, ACTIVITY, INTERCOMPANY, FUTURE]
    (widths 2/7/6/6/3/4/2/4) — so every decoder reuses one parse and the
    "Account is position 4, not last" gotcha lives in exactly one place.
    Returns [] for an empty/blank string."""
    s = N(segments)
    if not s:
        return []
    return [p.strip() for p in s.split("-")]


def account_of_gl_segments(segments):
    """Map a concatenated GL combination ("01-1100001-000000-100221-...") to
    the short engine account via its natural-account segment (position 4 of
    the ENTITY-FUND-DEPT-ACCOUNT-... combo).  Returns None for malformed
    combos and for GLs belonging to no configured depository account."""
    parts = segments_of(segments)
    if len(parts) < 4:
        return None
    return _GL_CASH_ACCOUNTS.get(parts[3])


def dept_segment_of(segments):
    """Department segment (position 3) of a GL combo, or None if absent."""
    parts = segments_of(segments)
    return parts[2] if len(parts) > 2 else None


def entity_segment_of(segments):
    """Entity segment (position 1) of a GL combo, or None if absent."""
    parts = segments_of(segments)
    return parts[0] if parts else None


def _coa_label(v):
    """The human label carried in an AcctCombos "<code>-<label>" *_DESC cell
    ("01-UT System" -> "UT System").  A cell with no dash is returned as-is."""
    s = N(v)
    return s.split("-", 1)[1].strip() if "-" in s else s


def coa_decode(segments, coa):
    """Decode a concatenated GL combination into human labels through the
    loaded Chart of Accounts (or None).  ADVISORY ONLY — campus/entity
    "consistency" confers no matching evidence (rule 8c); this feeds Review
    and recommended-GL text, never a placement.  Returns a dict
    {ent, ent_desc, fund, dept, dep_desc, account, act_desc, intercompany,
    itc_desc, act_grp_desc} or None when the CoA is absent or the string
    yields <4 segments (callers no-op)."""
    if not coa:
        return None
    parts = segments_of(segments)
    if len(parts) < 4:
        return None
    combo = "-".join(parts)
    decoded = coa.get("combo_decode", {}).get(combo)
    if decoded is not None:
        return decoded
    # Miss: fall back to per-segment labels.  The Entity value set decodes
    # BOTH the Entity segment (pos1) and the Intercompany segment (pos7),
    # which reuse the same value set; Account/Dept stay raw codes.
    entity_desc = coa.get("entity_desc", {})
    itc = parts[6] if len(parts) > 6 else ""
    return {
        "ent": parts[0], "ent_desc": entity_desc.get(parts[0], ""),
        "fund": parts[1], "dept": parts[2], "dep_desc": "",
        "account": parts[3], "act_desc": "",
        "intercompany": itc, "itc_desc": entity_desc.get(itc, ""),
        "act_grp_desc": "",
    }


def coa_combo_valid(segments, coa):
    """{'recognized': the full 8-seg combo is in the AcctCombos universe,
    'postable': its E-F-D-P subkey (segments 1,2,3,5) is in the ComboSet
    whitelist}.  Feeds ONLY the coa_combo_validity diagnostic counter — it
    never drops or downgrades a row."""
    if not coa:
        return {"recognized": False, "postable": False}
    parts = segments_of(segments)
    if len(parts) < 5:
        return {"recognized": False, "postable": False}
    combo = "-".join(parts)
    efdp = "-".join((parts[0], parts[1], parts[2], parts[4]))
    return {"recognized": combo in coa.get("combo_decode", {}),
            "postable": efdp in coa.get("postable_efdp", set())}


def _leading_date_key(filename):
    """Newest plausible YYYYMMDD stamp in the filename ('00000000' if none).
    An 8-digit run that does not parse as a calendar date (an account or
    merchant number like 99999999) is ignored, never mistaken for a date."""
    best = "00000000"
    for m in _YYYYMMDD_RE.finditer(filename):
        s = m.group(1)
        y, mo, d = int(s[:4]), int(s[4:6]), int(s[6:8])
        if 1990 <= y <= 2099 and 1 <= mo <= 12 and 1 <= d <= 31 and s > best:
            best = s
    return best


@dataclass
class RoutedFile:
    role: str
    path: str
    filename: str
    sheet_hint: str


# Roles the router recognizes but the FORWARD engine never reads — named so
# the owner is told, per file, that the engine will not use it.
_RECOGNIZED_NOT_LOADED = {
    "ALL_DATA": "lifecycle workbook — Unreconcile2's fuel",
    "RECONCILED": "reconciled forensic export — offline config-audit / Unreconcile2 fuel",
    "CONTRACTS_INV": "contracts-to-invoices reference — not consumed by the forward passes",
    "GMS_AGING": "sponsored AR aging — not consumed by the forward passes",
    "AR_INVOICES": "AR invoice listing — not consumed by the forward passes",
    "AR_UNAPPLIED_SUMMARY": "unapplied summary — not consumed by the forward passes",
    "GMS_SPONSOR_MAP": "sponsor map — not consumed by the forward passes",
    "RELATIONSHIP_MAP": "relationship/rosetta doc — reference only",
}


def _skip_notice(name, reason, suggestion=""):
    """Owner hard requirement (2026-07-19): the MOMENT the router decides to
    ignore a file because of its NAME, say so — on stderr, with the reason
    and a rename suggestion.  Nothing is ever ignored silently."""
    msg = f"IGNORED (file name): {name} — {reason}"
    if suggestion:
        msg += f"  FIX: {suggestion}"
    print(msg, file=sys.stderr)
    return {"file": name, "reason": reason, "suggestion": suggestion}


def route_folder(input_dir, skipped=None) -> dict:
    """Scan input_dir, classify each file, resolve multi-file-per-role by
    newest-YYYYMMDD (union+dedup handled downstream).  Returns
    {role: [RoutedFile, ...]} plus asserts hard-required roles present.
    EVERY skipped file is announced on stderr the moment it is skipped
    (and appended to `skipped` when a list is passed) — name-based
    ignoring is never silent."""
    if not os.path.isdir(input_dir):
        raise MissingRequiredFile("INPUT_DIR", f"{input_dir} is not a directory")
    by_role = {}
    if skipped is None:
        skipped = []
    rule_by_role = {r.role: r for r in ROUTER_TABLE}
    for name in sorted(os.listdir(input_dir)):
        path = os.path.join(input_dir, name)
        if not os.path.isfile(path):
            continue
        low = name.lower()
        if low.endswith(".xls"):
            skipped.append(_skip_notice(
                name, "legacy .xls format is not readable",
                "re-export as .xlsx or .csv"))
            continue
        if not low.endswith((".xlsx", ".xlsm", ".xlsb", ".csv", ".txt")):
            skipped.append(_skip_notice(
                name, "not a supported data format (.xlsx/.xlsm/.xlsb/.csv, "
                      "or .txt for raw BAI2)",
                "convert to a spreadsheet export, or rename with a _BAI2 "
                "token if it is a raw bank transmission"))
            continue
        role = classify_file(name)
        if role is None:
            skipped.append(_skip_notice(
                name, "no router rule matches this name",
                "rename with the export's role token (see FILE_NAMING.md: "
                "BSL / _ST_ / MET / Receipts / Payables_Payments / BAI2 ...)"))
            continue
        # Raw bank files (owner, 2026-07-19): a .txt is accepted ONLY as a
        # native BAI2 transmission ("FHB_UTHSC_BAI2.txt") — the raw reader
        # parses type-16/88 records with the FULL untruncated addenda the
        # Oracle BSL feed cuts off.  Any other .txt stays unrouted.
        if low.endswith(".txt") and role != "BAI2":
            skipped.append(_skip_notice(
                name, f".txt is accepted only for raw BAI2 transmissions "
                      f"(this name routes as {role})",
                "export as .xlsx/.csv, or add a _BAI2 token if it is a raw "
                "bank file"))
            continue
        rf = RoutedFile(role, path, name, rule_by_role[role].sheet)
        by_role.setdefault(role, []).append(rf)
        if role in _RECOGNIZED_NOT_LOADED:
            print(f"NOTE (file name): {name} recognized as {role} — "
                  f"{_RECOGNIZED_NOT_LOADED[role]}; the forward engine will "
                  "not read it.", file=sys.stderr)

    # Newest-wins ordering within a role (leading YYYYMMDD desc); ties keep all
    # for union+dedup by the loaders.
    for role, files in by_role.items():
        files.sort(key=lambda f: _leading_date_key(f.filename), reverse=True)

    for role in HARD_REQUIRED_ROLES:
        if role not in by_role:
            raise MissingRequiredFile(role, f"no file in {input_dir} matched role {role}")
    return by_role


# ======================================================================
# Section 5 — Column binding (content first, header as tiebreak)
# ======================================================================

@dataclass
class RoleSpec:
    required: bool
    header_aliases: list
    predicate: object  # fn(cell) -> bool


@functools.lru_cache(maxsize=1 << 17)
def _norm_header(h) -> str:
    return re.sub(r"[^A-Z0-9]", "", N(h).upper())


# ---- content predicates (Section 5.2) --------------------------------

def pred_date(v):
    return parse_date(v) is not None


def pred_signed_amount(v):
    return cents(v) is not None


def pred_reference(v):
    s = N(v)
    if not s:
        return False
    # Reject only genuinely date-shaped values.  A bare integer is a
    # plausible reference: real Oracle exports deliver references and
    # transaction numbers as ints (1390, 852256), which parse_date would
    # otherwise swallow as Excel serials and poison the content score.
    if isinstance(v, (datetime, date)):
        return False
    if isinstance(v, str) and not s.isdigit() and parse_date(s) is not None:
        return False
    return bool(re.search(r"[A-Za-z0-9]", s))


_STATUS_VOCAB = {
    "REC", "UNR", "VOID", "APP", "CLEARED", "CONFIRMED", "REMITTED",
    "UNAPPLIED", "REVERSED", "OP", "CL", "OPEN", "CLOSED", "APPLIED", "MATCHED",
}


def pred_status(v):
    return _norm_header(v) in _STATUS_VOCAB


def pred_customer(v):
    s = N(v)
    if not s or not re.search(r"[A-Za-z]", s):
        return False
    digits = sum(c.isdigit() for c in s)
    return digits / max(1, len(s)) < 0.4


_TXN_TYPE_VOCAB = {
    "ACH", "CHK", "MSC", "EFT", "BKA", "BKF", "ZBA", "CREDITCARD", "CHECK",
    "MISCELLANEOUS", "WIRE", "DEPOSIT", "CREDIT", "DEBIT",
}


def pred_txn_type(v):
    return _norm_header(v) in _TXN_TYPE_VOCAB


_GL_RE = re.compile(r"^\d{2}-\d{7}-\d{6}-\d{6}")


def pred_gl_string(v):
    return bool(_GL_RE.match(N(v)))


def pred_mid(v):
    return is_mid(v)


def pred_number(v):
    s = N(v)
    return bool(s) and bool(re.search(r"\d", s))


_MET_DESC_RE = re.compile(r"(?i)\b[dr]\s*:\s*\d+")


def pred_met_description(v):
    """MET CET_DESCRIPTION carries 'd:N | r:N | payer/purpose...' — at least
    two pipe segments.  The composite ID column ('d:N | r:N', one pipe) and
    generic text columns must not satisfy this, so content scoring cannot
    prefer a stub column over the real description (real FHB Master lesson:
    the ID column scored 1.0 and hid every payer string)."""
    return N(v).count("|") >= 2


def pred_any(v):  # non-empty
    return N(v) != ""


def _cols_identical(data_rows, cols):
    """True when every listed column carries value-for-value identical content
    over the sampled data rows (normalized text).  A bank export that repeats
    a column verbatim (e.g. BAI2 with two 'Amount' columns) is not a real
    ambiguity: either binding yields identical output, so the leftmost is
    chosen deterministically instead of failing loud."""
    ref = None
    for c in cols:
        vals = tuple(N(r[c]) if c < len(r) else "" for r in data_rows)
        if ref is None:
            ref = vals
        elif vals != ref:
            return False
    return True


def _signed_twin(data_rows, cols):
    """Disambiguate a signed/unsigned amount twin: some bank exports carry the
    same amount twice — once signed, once as magnitude with the sign in a
    separate Debit/Credit column.  When every sampled row agrees in absolute
    cents across the tied columns and exactly one column carries a negative
    value, that signed column is the real amount role.  Returns its index or
    None when the pattern doesn't hold."""
    has_negative = {c: False for c in cols}
    for r in data_rows:
        mags = set()
        for c in cols:
            v = r[c] if c < len(r) else None
            if _blankish(v):
                return None
            try:
                cts = cents(v)
            except Exception:
                return None
            if cts is None:
                return None
            mags.add(abs(cts))
            if cts < 0:
                has_negative[c] = True
        if len(mags) != 1:
            return None
    signed = [c for c in cols if has_negative[c]]
    return signed[0] if len(signed) == 1 else None


def bind_columns(rows, role_specs, filename="<rows>", header_scan=12, sample=50):
    """Bind each role to a column index by scanning content first, header as
    tiebreak (Section 5.1).  Returns (mapping {role: col_index}, header_index).

    rows: list of tuples already read from the sheet.
    """
    if not rows:
        # Every required role fails.
        for role, spec in role_specs.items():
            if spec.required:
                raise InvalidSourceData(filename, role, "empty sheet")
        return {}, 0

    all_aliases = []
    for spec in role_specs.values():
        all_aliases.extend(_norm_header(a) for a in spec.header_aliases)
    max_arity = len(role_specs)

    # 1. Locate the header row.  Never assume row 0: a deep report preamble
    # (banner + total rows) bound at row 0 silently emits phantom data rows.
    header_index = None
    best_hits = -1
    scan = min(header_scan, len(rows))
    need_hits = min(2, len({a for a in all_aliases if a})) or 1
    for i in range(scan):
        row = rows[i]
        nonempty = sum(1 for c in row if N(c) != "")
        hits = 0
        for c in row:
            hc = _norm_header(c)
            if hc and any(hc == a or (a and a in hc) for a in all_aliases):
                hits += 1
        # Prefer a row with enough columns AND enough alias hits.
        if nonempty >= min(max_arity, 2) and hits >= need_hits and hits > best_hits:
            best_hits = hits
            header_index = i
    if header_index is None:
        if any(s.required for s in role_specs.values()):
            raise InvalidSourceData(
                filename, "HEADER",
                f"no header row found in the first {scan} rows "
                f"(need >= {need_hits} recognized column headers); "
                "refusing to assume row 0")
        header_index = 0
    header = rows[header_index]
    ncols = max(len(r) for r in rows)
    data_rows = rows[header_index + 1: header_index + 1 + sample]

    # Content sampling spans the WHOLE sheet, not the first `sample` rows:
    # a sparsely-populated column (e.g. Structured Payment Reference blank on
    # every recent Journal row) must be scored on the values it actually
    # carries, or reference-shaped neighbor columns blind-tie it out of
    # binding (real FHB Master lesson, 2026-07-11: an unbound SPR column
    # silently gutted the 4x4 cross-reference screen).  Up to `sample`
    # non-blank values are collected per column in one pass.
    col_samples = {c: [] for c in range(ncols)}
    unfilled = set(col_samples)
    for r in rows[header_index + 1:]:
        if not unfilled:
            break
        for c in list(unfilled):
            if c < len(r) and not _blankish(r[c]):
                bucket = col_samples[c]
                bucket.append(r[c])
                if len(bucket) >= sample:
                    unfilled.discard(c)

    # 2/3. Score every column for every role; content decides, header breaks ties.
    mapping = {}
    for role, spec in role_specs.items():
        alias_norms = [_norm_header(a) for a in spec.header_aliases]
        scored = []
        for col in range(ncols):
            hcell = header[col] if col < len(header) else ""
            hnorm = _norm_header(hcell)
            if hnorm and hnorm in alias_norms:
                header_score = 3
            elif hnorm and any(a and a in hnorm for a in alias_norms):
                header_score = 2
            else:
                header_score = 0
            sampled = col_samples[col]
            if sampled:
                content_score = sum(1 for c in sampled if spec.predicate(c)) / len(sampled)
            else:
                content_score = 0.0
            scored.append((content_score, header_score, col))
        scored.sort(key=lambda t: (t[0], t[1]), reverse=True)
        best = scored[0]
        # Ambiguity: tie on both content and header for a required role.
        if spec.required and len(scored) > 1:
            second = scored[1]
            if (best[0], best[1]) == (second[0], second[1]) and best[0] > 0:
                tied = [t[2] for t in scored if (t[0], t[1]) == (best[0], best[1])]
                if _cols_identical(data_rows, tied):
                    best = (best[0], best[1], min(tied))
                else:
                    signed = _signed_twin(data_rows, tied)
                    if signed is not None:
                        best = (best[0], best[1], signed)
                    else:
                        raise AmbiguousColumn(filename, role, tied)
        if spec.required and best[0] == 0:
            raise InvalidSourceData(
                filename, role,
                f"no column satisfied content predicate (header={_norm_header(header[best[2]]) if best[2] < len(header) else ''!r})")
        if not spec.required:
            # An optional role never binds by position alone: leave it unbound
            # on zero evidence or a blind tie (same content AND header score
            # on two columns) rather than guess the leftmost column.
            if best[0] == 0 and best[1] == 0:
                continue
            if len(scored) > 1 and (best[0], best[1]) == (scored[1][0], scored[1][1]):
                tied = [t[2] for t in scored if (t[0], t[1]) == (best[0], best[1])]
                if _cols_identical(data_rows, tied):
                    best = (best[0], best[1], min(tied))
                else:
                    continue
        mapping[role] = best[2]
    return mapping, header_index


# ---- Role specs per file (Section 5.3) -------------------------------

def _rs(required, aliases, pred):
    return RoleSpec(required, aliases, pred)


BSL_ROLES = {
    "date": _rs(True, ["Transaction Date", "Booking Date", "Value Date", "Date", "Post Date"], pred_date),
    "amount": _rs(True, ["Amount", "Transaction Amount", "Signed Amount"], pred_signed_amount),
    "line_key": _rs(True, ["Line Number", "Statement Line", "Statement", "Line", "Bank Statement Line", "Sequence"], pred_number),
    "reference": _rs(False, ["Reference", "Bank Reference", "Recon Reference"], pred_reference),
    "account_servicer_reference": _rs(False, ["Account Servicer Reference", "Servicer Reference"], pred_reference),
    "customer_reference": _rs(False, ["Customer Reference"], pred_reference),
    "additional_info": _rs(False, ["Additional Information", "Additional Info", "Addenda"], pred_any),
    "transaction_type": _rs(False, ["Transaction Type", "Type"], pred_any),
    "transaction_code": _rs(False, ["Transaction Code", "Bank Transaction Code", "Code"], pred_any),
}

ST_ROLES = {
    "date": _rs(True, ["Transaction Date", "Date", "Accounting Date", "GL Date"], pred_date),
    "amount": _rs(True, ["Amount", "Transaction Amount", "Signed Amount"], pred_signed_amount),
    "transaction_number": _rs(True, ["Transaction Number", "Transaction", "Trx Number", "Transaction Num", "Number"], pred_reference),
    "source": _rs(True, ["Source", "Transaction Source", "Origin"], pred_any),
    "reference": _rs(False, ["Reference", "Recon Match Reference", "Match Reference"], pred_reference),
    "structured_payment_reference": _rs(False, ["Structured Payment Reference", "Strc Pay Ref"], pred_reference),
    "counterparty": _rs(False, ["Counterparty", "Customer", "Payer", "Customer Name"], pred_customer),
    "transaction_type": _rs(False, ["Transaction Type", "Type"], pred_any),
}

RECEIPTS_ROLES = {
    "receipt_number": _rs(True, ["Receipt Number", "Receipt Num"], pred_reference),
    "document_number": _rs(False, ["Document Number", "Doc Number"], pred_reference),
    # 'Status' is the lifecycle (Cleared/Remitted/Reversed); the export's
    # separate 'State' column (Applied/Unapplied) must not tie it.
    "status": _rs(True, ["Status", "Receipt Status"], pred_status),
    "amount": _rs(True, ["Receipt Amount", "Entered Amount", "Amount"], pred_signed_amount),
    "customer_name": _rs(True, ["Customer Name", "Customer", "Payer"], pred_customer),
    "receipt_date": _rs(False, ["Receipt Date", "Date"], pred_date),
    "deposit_date": _rs(False, ["Deposit Date", "Cleared Date"], pred_date),
    "batch_number": _rs(False, ["Batch Number", "Receipt Batch"], pred_reference),
    "reference": _rs(False, ["Reference", "Remittance Reference"], pred_reference),
    "unapplied_amount": _rs(False, ["Unapplied Amount"], pred_signed_amount),
    # Misdirected-SPN detector (owner hard guardrail, 2026-07-18): the
    # receipts export is system-wide; this column names which bank account
    # each receipt actually remits to.  Predicate demands a value that maps
    # to a configured account so 'Remittance Bank' / '...Currency' twins
    # can never bind here.
    "remittance_bank_account": _rs(False, ["Remittance Bank Account"],
                                   lambda v: account_of_bank_name(v) is not None),
}

# AP payment feed (owner, 2026-07-19): the Payables analogue of RECEIPTS.
# `Payment Number` is the AP identity (= ST Transaction Number on the AP
# rows); `Payment Amount` is a POSITIVE disbursement magnitude the pool
# negates to the bank's signed cents; `Payment Status` / `Reconciled` give
# the open/closed lifecycle.
PAYMENTS_ROLES = {
    "payment_number": _rs(True, ["Payment Number", "Payment Num", "Check Number"], pred_reference),
    "amount": _rs(True, ["Payment Amount", "Amount"], pred_signed_amount),
    "status": _rs(True, ["Payment Status", "Status"], pred_any),
    "payee": _rs(True, ["Payee", "Supplier or Party", "Payee Name", "Supplier"], pred_customer),
    "reconciled": _rs(False, ["Reconciled"], pred_any),
    "payment_date": _rs(False, ["Payment Date", "Date"], pred_date),
    "payment_document": _rs(False, ["Payment Document"], pred_any),
}

# All-accounts open bank-statement-line export (owner, 2026-07-19): the
# reverse-direction evidence for the misdirected search — an open BSL in
# ANOTHER account that a THIS-account open ST/receipt actually funds.  The
# OTBI shape uses CSL_/CBA_ headers.
ALL_BSL_ROLES = {
    "bank_account_name": _rs(True, ["CBA Bank Account Name", "Bank Account Name", "Bank Account"], pred_any),
    "amount": _rs(True, ["Amount", "Signed Amount"], pred_signed_amount),
    "date": _rs(False, ["CSL Booking Date", "Booking Date", "Statement Date", "Date"], pred_date),
    "reference": _rs(False, ["CSL Recon Reference", "Recon Reference", "Reference"], pred_reference),
    "customer_reference": _rs(False, ["CSL Customer Reference", "Customer Reference"], pred_reference),
    "addenda": _rs(False, ["CSL Addenda Txt", "Addenda", "Additional Information"], pred_any),
}

MET_ROLES = {
    "trx_id": _rs(True, ["CET Transaction ID", "Transaction Number", "Trx Id", "Transaction Id"], pred_reference),
    "amount": _rs(True, ["Amount", "Transaction Amount"], pred_signed_amount),
    "transaction_date": _rs(True, ["Transaction Date", "Date"], pred_date),
    "status": _rs(True, ["CET Status", "Status", "State"], pred_any),
    # Optional: the real OTBI export carries native DEPOSIT_ID / RECEIPT_ID
    # columns (authoritative; the d:/r: description parse is the fallback).
    "description": _rs(False, ["CET Description", "Description", "Desc"], pred_met_description),
    "deposit_id": _rs(False, ["Deposit ID"], pred_number),
    "receipt_id": _rs(False, ["Receipt ID"], pred_number),
    "reference_text": _rs(False, ["CET Reference Text", "Reference Text"], pred_reference),
    "bank_account_name": _rs(False, ["CBE Bank Account Name", "Bank Account Name", "Bank Account"], pred_any),
    "transaction_type": _rs(False, ["CET Transaction Type", "Transaction Type"], pred_any),
    "cleared_date": _rs(False, ["Cleared Date"], pred_date),
    "offset": _rs(False, ["Offset Concatenated Segments", "Offset", "GL"], pred_any),
    # COA scope key (owner, 2026-07-18): the asset combo's natural account
    # identifies the depository bank account independently of the bank name.
    "asset_segments": _rs(False, ["Asset Concatenated Segments"], pred_any),
}




BAI2_ROLES = {
    "post_date": _rs(True, ["Post Date", "Date"], pred_date),
    "amount": _rs(True, ["Amount"], pred_signed_amount),
    "description": _rs(False, ["Transaction Description"], pred_any),
    "bank_reference": _rs(False, ["Bank Reference"], pred_reference),
    "customer_reference": _rs(False, ["Customer Reference"], pred_reference),
    "bai_code": _rs(False, ["BAI Code"], pred_number),
}

DEPT_INFO_ROLES = {
    "department": _rs(True, ["Department Name"], pred_customer),
    "campus": _rs(False, ["Campus Name"], pred_customer),
    "dept_bank_name": _rs(False, ["Dept Bank Name"], pred_any),
    "campus_bank_name": _rs(False, ["Campus Bank Name"], pred_any),
    "mid": _rs(False, ["Credit Card Mid"], pred_mid),
}

# Edison (State of Tennessee) exports (owner, 2026-07-19): the STATE'S OWN
# record of payments to UT and UT's invoices into Edison.  ANNOTATION ONLY —
# the C6 State pass stays retired; State lines reconcile through the normal
# lanes, and Edison names WHAT a stranded State line is (payment reference,
# invoice, date) for the reviewer's manual ECT.  Never a pool source.
EDISON_PAY_ROLES = {
    "reference": _rs(True, ["Reference"], pred_reference),
    "invoice": _rs(False, ["Invoice Number"], pred_any),
    "date": _rs(False, ["Payment Date", "Date"], pred_date),
    "amount": _rs(True, ["Amount"], pred_signed_amount),
}

EDISON_INV_ROLES = {
    "invoice": _rs(True, ["Invoice Number"], pred_any),
    "date": _rs(False, ["Invoice Date"], pred_date),
    "amount": _rs(False, ["Gross Amt", "Gross Amount"], pred_signed_amount),
    "status": _rs(False, ["Approval Status", "Status"], pred_any),
    "voucher": _rs(False, ["Voucher"], pred_reference),
}



ENRICHED_ROLES = {
    # Pre-parsed enriched BSL workbook (all rows, full year): parsed trace
    # ids / MIDs / reconciliation keys the Oracle feed does not carry.
    # Joined back onto the open BSLs by Statement line, then (date, cents).
    "date": _rs(True, ["Date"], pred_date),
    "amount": _rs(True, ["Amount (USD)", "Amount"], pred_signed_amount),
    "line_key": _rs(False, ["Statement"], pred_any),
    "additional_info": _rs(False, ["Additional Information"], pred_any),
    "parsed_trace": _rs(False, ["Parsed_Trace_ID", "Parsed Trace ID"], pred_reference),
    "parsed_mid": _rs(False, ["Parsed_MID", "Parsed MID"], pred_mid),
    "recon_key": _rs(False, ["Reconciliation_Key", "Reconciliation Key"], pred_any),
    "category": _rs(False, ["Transaction_Category", "Transaction Category"], pred_any),
}

AR_MATCHED_ROLES = {
    # ACRA/ABA receipt-application feed (AR_Matched_Invoice_Receipts...):
    # per-receipt bank-deposit dates and application context.
    "receipt_number": _rs(True, ["ACRA Receipt Number", "Receipt Number"], pred_reference),
    "amount": _rs(True, ["ACRA Amount", "Amount"], pred_signed_amount),
    "receipt_date": _rs(False, ["ACRA Receipt Date", "Receipt Date"], pred_date),
    "deposit_date": _rs(False, ["ACRA Deposit Date", "Deposit Date"], pred_date),
    "status": _rs(False, ["ACRA Status", "Status"], pred_any),
    "bank_account_name": _rs(False, ["CBA Bank Account Name", "Bank Account Name"], pred_any),
    "comments": _rs(False, ["ACRA Comments", "Comments"], pred_any),
}

APPLIED_UNAPPLIED_ROLES = {
    "trx_number": _rs(True, ["Trx Number", "Transaction Number"], pred_reference),
    "receipt_number": _rs(True, ["Receipt Number"], pred_reference),
    "contract_number": _rs(False, ["Contract Number"], pred_reference),
    "customer_name": _rs(False, ["Customer Name", "Customer"], pred_customer),
    "accounted_applied_amount": _rs(False, ["Accounted Applied Amount", "Applied Amount"], pred_signed_amount),
    "accounted_unapplied_amount": _rs(False, ["Accounted Unapplied Amount", "Unapplied Amount"], pred_signed_amount),
}


# ======================================================================
# Workbook / CSV loading (Section 3)
# ======================================================================

def _read_sheet_rows(path, sheet_hint):
    """Read rows from an xlsx sheet (data_only, read_only=False per Section 3).
    Returns (rows, sheet_title).  sheet_hint 'first' or a title substring."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = _pick_sheet(wb, sheet_hint, os.path.basename(path))
        ws.reset_dimensions()  # never trust stated dimensions (Oracle BI)
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        title = ws.title
        wb.close()
    except Exception as e:
        raise InvalidSourceData(os.path.basename(path), "WORKBOOK", f"load failed: {e}")
    if len(rows) <= 1:
        # Pathological export the streaming reader cannot see — take the
        # slow full-parse path before concluding the sheet is empty.
        try:
            wb = load_workbook(path, read_only=False, data_only=True)
        except Exception as e:
            raise InvalidSourceData(os.path.basename(path), "WORKBOOK", f"load failed: {e}")
        ws = _pick_sheet(wb, sheet_hint, os.path.basename(path))
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        title = ws.title
        wb.close()
    return rows, title


def _match_sheet(wb, title_substr, filename):
    """Resolve a sheet by title substring: an exact title wins; several
    non-exact matches are ambiguous (fail loud, never take the first tab);
    zero matches returns None (caller decides the spec-mandated fallback)."""
    low = title_substr.lower().strip()
    matches = [ws for ws in wb.worksheets if low in ws.title.lower()]
    for ws in matches:
        if ws.title.lower().strip() == low:
            return ws
    if len(matches) > 1:
        raise InvalidSourceData(
            filename, title_substr,
            f"multiple sheets match {title_substr!r}: "
            f"{[ws.title for ws in matches]} — rename or pass the exact title")
    return matches[0] if matches else None


def _pick_sheet(wb, sheet_hint, filename="<workbook>"):
    if sheet_hint in ("first", "multi", "all", "reference", "stream", None):
        return wb.worksheets[0]
    ws = _match_sheet(wb, sheet_hint, filename)
    # No match: spec §4.1 mandates the first-sheet fallback for named hints.
    return ws if ws is not None else wb.worksheets[0]


def _read_csv_rows(path):
    import csv
    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        rows = list(csv.reader(fh))
    if rows:
        # Strip leading UTF-8 BOM from the first header cell.
        first = rows[0]
        if first and first[0].startswith("﻿"):
            first[0] = first[0].lstrip("﻿")
    return [tuple(r) for r in rows]


def _bai2_date(yymmdd):
    """BAI2 YYMMDD -> ISO date string ('' when malformed).  BAI2 years are
    two-digit; the format postdates 2000, so 20YY is always correct here."""
    s = N(yymmdd)
    if len(s) == 6 and s.isdigit():
        return f"20{s[0:2]}-{s[2:4]}-{s[4:6]}"
    return ""


def _read_bai2_txt(path):
    """Parse a NATIVE BAI2 transmission (.txt) into spreadsheet-shaped rows
    binding the existing BAI2_ROLES: header row + one row per type-16 detail
    record, with the 88-continuation addenda preserved UNTRUNCATED in
    DETAIL1..DETAILn columns (the whole point — the Oracle BSL feed cuts the
    addenda; the raw file carries Customer ID / Trace Number / TRN1 / check
    numbers in full).  Record types: 01 file header, 02 group header (as-of
    date, YYMMDD), 03 account identifier, 16 detail
    (,type-code,amount-in-minor-units,funds-type,bank-ref,customer-ref,text),
    88 continuation of the PREVIOUS record, 49/98/99 trailers.  Sign follows
    the BAI type-code convention: 100-399 credit (+), 400-699 debit (-).
    Money stays integer cents throughout (no float)."""
    details = []          # [date, desc, signed_amount_str, bank_ref, cust_ref, code, [addenda...]]
    group_date = ""
    cur = None            # the open type-16 record collecting 88 lines
    max_addenda = 0
    with open(path, encoding="utf-8", errors="replace") as fh:
        for raw in fh:
            line = raw.rstrip("\r\n").rstrip()
            if line.endswith("/"):
                line = line[:-1].rstrip()
            if not line:
                continue
            rec, _, rest = line.partition(",")
            if rec == "02":
                f = rest.split(",")
                # 02,receiver,originator,group-status,as-of-date,...
                group_date = _bai2_date(f[3]) if len(f) > 3 else group_date
                cur = None
            elif rec == "16":
                f = rest.split(",")
                code = N(f[0]) if f else ""
                minor = N(f[1]) if len(f) > 1 else ""
                sign = ""
                if code.isdigit() and 400 <= int(code) <= 699:
                    sign = "-"
                # integer minor units -> decimal string (never float)
                digits = minor if minor.isdigit() else ""
                if digits:
                    cents_ = int(digits)
                    amt = f"{sign}{cents_ // 100}.{cents_ % 100:02d}"
                else:
                    amt = ""
                bank_ref = N(f[3]) if len(f) > 3 else ""
                cust_ref = N(f[4]) if len(f) > 4 else ""
                text = ",".join(f[5:]).strip() if len(f) > 5 else ""
                cur = [group_date, text, amt, bank_ref, cust_ref, code, []]
                details.append(cur)
            elif rec == "88":
                if cur is not None:
                    cur[6].append(rest.strip())
                    max_addenda = max(max_addenda, len(cur[6]))
            elif rec in ("03", "49", "98", "99", "01"):
                if rec != "03":
                    pass
                cur = None  # continuations after non-16 records are not detail addenda
    header = ["Post Date", "Transaction Description", "Amount",
              "Bank Reference", "Customer Reference", "BAI Code"] + \
             [f"DETAIL{i + 1}" for i in range(max_addenda)]
    rows = [tuple(header)]
    for d in details:
        rows.append(tuple(d[:6] + d[6] + [""] * (max_addenda - len(d[6]))))
    return rows


def _xlsb_norm(v):
    """pyxlsb returns EVERY numeric cell as a float, so an integral Oracle id
    (DEPOSIT_ID/RECEIPT_ID/TRANSACTION_ID = 65105) arrives as 65105.0 and would
    stringify to "65105.0" — breaking the MET<->ST bridge join and the d:/r:
    deposit-group citations (audit C4 fails: "not a real MET deposit").  Collapse
    integral floats back to int so an .xlsb read matches the .csv/.xlsx text form
    byte-for-byte downstream.  Money (1890.69) and any true fractional serial are
    left untouched; the 1e15 guard keeps the collapse inside float's exact-integer
    range; bool is not a float so it is unaffected."""
    if isinstance(v, float) and -1e15 < v < 1e15 and v.is_integer():
        return int(v)
    return v


def _read_xlsb_rows(path, sheet_hint):
    """Binary workbook reader (pyxlsb).  Dates arrive as Excel serials, which
    parse_date already accepts."""
    try:
        from pyxlsb import open_workbook
    except ImportError:
        raise InvalidSourceData(
            os.path.basename(path), "WORKBOOK",
            ".xlsb requires the pyxlsb package (pip install pyxlsb) — or "
            "re-export the file as .xlsx")
    with open_workbook(path) as wb:
        names = wb.sheets
        target = names[0]
        if sheet_hint not in ("first", "multi", "all", "reference", "stream", None):
            low = sheet_hint.lower().strip()
            matches = [n for n in names if low in n.lower()]
            exact = [n for n in matches if n.lower().strip() == low]
            if exact:
                target = exact[0]
            elif len(matches) > 1:
                raise InvalidSourceData(
                    os.path.basename(path), sheet_hint,
                    f"multiple sheets match {sheet_hint!r}: {matches}")
            elif matches:
                target = matches[0]
        with wb.get_sheet(target) as ws:
            rows = [tuple(_xlsb_norm(c.v) for c in row) for row in ws.rows()]
    return rows, target


def read_rows(routed_file: RoutedFile):
    """Dispatch to xlsx/xlsb/csv loader based on extension.  Wraps failures
    in InvalidSourceData (Section 15.5)."""
    ext = os.path.splitext(routed_file.path)[1].lower()
    try:
        if ext == ".csv":
            return _read_csv_rows(routed_file.path), "csv"
        if ext == ".txt":
            # Only BAI2 .txt files are routed (route_folder); parse natively.
            return _read_bai2_txt(routed_file.path), "bai2-raw"
        if ext == ".xlsb":
            return _read_xlsb_rows(routed_file.path, routed_file.sheet_hint)
        return _read_sheet_rows(routed_file.path, routed_file.sheet_hint)
    except InvalidSourceData:
        raise
    except Exception as e:
        raise InvalidSourceData(routed_file.filename, routed_file.role, f"read failed: {e}")

# ======================================================================
# Section 8 — Candidate ST pool (all-status, deduped)
# ======================================================================

# Status vocabularies for availability (Section 8.2).
OPEN_RECEIPT_STATUSES = {"UNR", "REMITTED", "CONFIRMED", "UNAPPLIED", "OPEN", "OP"}
CLOSED_RECEIPT_STATUSES = {"CLEARED", "APP", "APPLIED", "REVERSED", "VOID", "REC", "CLOSED", "CL"}
# AP payment lifecycle (owner, 2026-07-19): "Negotiable" = issued/outstanding
# (open money that has not cleared the bank); "Cleared" (Reconciled=Yes) is
# already reconciled; "Voided" is cancelled — neither is open.
OPEN_PAYMENT_STATUSES = {"NEGOTIABLE", "ISSUED", "UNR", "OPEN"}
JOURNAL_SOURCES = {"GL", "JOURNAL", "JE", "MANUAL"}


@dataclass
class PoolEntry:
    id: str
    amount_cents: int
    date: object            # datetime.date or None
    reference: str
    znref: str
    digits: set
    payer_tokens: set
    counterparty: str
    source: str             # AR / EXT / AP / PAY / GL ...
    status: str
    available: bool
    spn: str
    is_mid: bool
    origin: str = ""        # which file/role produced it (for run log)
    deposit_id: str = ""    # ORT d:
    receipt_id: str = ""    # ORT r:
    transaction_type: str = ""  # Credit Card / Check / EFT (type gate)
    spr: str = ""           # Structured Payment Reference (screened separately)
    base_id: str = ""       # undisambiguated id when distinct rows share one label
    foreign_account: str = ""  # set when the entry belongs to ANOTHER bank
    #                            account (misdirected-SPN shadow pool; owner
    #                            hard guardrail 2026-07-18) — never eligible
    #                            for normal Matches/Candidates
    asset_segments: str = ""   # raw MET ASSET_CONCATENATED_SEGMENTS combo,
    #                            decoded through the CoA for Review/GL text only
    #                            (advisory; campus consistency confers nothing)
    offset_segments: str = ""  # raw MET OFFSET_CONCATENATED_SEGMENTS combo —
    #                            the ECT posting side (§6 account-string
    #                            sourcing); the CoA-recommended GL comes from
    #                            HERE, never from the cash-side asset combo


def _mk_entry(id, amount_cents, dt, reference, counterparty, source, status,
              available, origin, deposit_id="", receipt_id="", transaction_type="",
              spr="", asset_segments="", offset_segments=""):
    ref = clean_ref(reference)
    return PoolEntry(
        id=N(id),
        amount_cents=amount_cents,
        date=dt,
        reference=ref,
        znref=znorm(ref),
        digits=digit_runs(ref),
        payer_tokens=payer_tokens(counterparty) | payer_tokens(reference),
        counterparty=N(counterparty),
        source=N(source).upper(),
        status=_norm_header(status),
        available=available,
        spn=spn_of(id) or spn_of(reference),
        is_mid=is_mid(reference),
        origin=origin,
        deposit_id=N(deposit_id),
        receipt_id=N(receipt_id),
        transaction_type=N(transaction_type),
        spr=clean_ref(spr),
        asset_segments=N(asset_segments),
        offset_segments=N(offset_segments),
    )


def _dedup_keep_largest(entries, keyfn):
    """Dedup by keyfn keeping the largest-magnitude amount (the total, not a
    split).  When the kept total lacks a counterparty, borrow it from a dropped
    split (Section 8.1).

    The keep-largest doctrine targets total-plus-invoice-split repeats, whose
    signature is largest == sum(rest) in signed cents (a lone verbatim repeat
    also satisfies it).  Same-key rows that do NOT sum that way are distinct
    receipts sharing a transaction-number label (real FHB Master case: two
    'SPN070326 ACH HRSA' receipts whose SUM is the bank line) — all are kept,
    with the display id disambiguated by amount so the ledger, conservation
    assert, and audit C3 treat them as separate consumables."""
    groups = {}
    for e in entries:
        groups.setdefault(keyfn(e), []).append(e)
    out = []
    for key, members in groups.items():
        members_sorted = sorted(
            members,
            key=lambda e: (abs(e.amount_cents), e.id),
            reverse=True,
        )
        keep = members_sorted[0]
        rest = members_sorted[1:]
        if rest and keep.amount_cents != sum(m.amount_cents for m in rest):
            # Distinct receipts sharing a label — keep every one.
            for m in members_sorted:
                m.base_id = m.id
                m.id = f"{m.id} [{m.amount_cents / 100:.2f}]"
                out.append(m)
            continue
        if not keep.counterparty:
            for m in rest:
                if m.counterparty:
                    keep.counterparty = m.counterparty
                    keep.payer_tokens = keep.payer_tokens | payer_tokens(m.counterparty)
                    break
        out.append(keep)
    return out


def build_pool(loaded: dict, account: str, runlog: dict) -> list:
    """Build the union pool from loaded data sources (Section 8).

    `loaded` maps role -> {'rows', 'map', 'header_index'} for bound files.
    Returns a list of PoolEntry.  Records pool sizes in runlog.
    """
    pool = []
    counts = {}

    # 1. Open non-Receivables STs from ST (exclude Journal from eligibility but
    #    keep in the pool so P10 can explain journal-only lines; dedup
    #    keep-largest by transaction number).
    st = loaded.get("ST")
    if st:
        rows, m, hi = st["rows"], st["map"], st["header_index"]
        raw = []
        for r in rows[hi + 1:]:
            amt = cents(_cell(r, m.get("amount")))
            if amt is None:
                continue
            source = N(_cell(r, m.get("source"))).upper()
            txn = N(_cell(r, m.get("transaction_number")))
            if not txn:
                continue
            src_norm = _classify_source(source)
            e = _mk_entry(
                id=txn,
                amount_cents=amt,
                dt=parse_date(_cell(r, m.get("date"))),
                reference=_cell(r, m.get("reference")),
                spr=_cell(r, m.get("structured_payment_reference")),
                counterparty=_cell(r, m.get("counterparty")),
                source=src_norm,
                status="UNR",
                available=(src_norm not in JOURNAL_SOURCES),  # journals kept but ineligible
                origin="ST",
                transaction_type=_cell(r, m.get("transaction_type")),
            )
            # ST AR rows always enter the pool; when a RECEIPTS export also
            # exists it MERGES onto them below (authoritative for lifecycle
            # status) instead of replacing them — the ST export often carries
            # reference/SPR/counterparty detail the receipts export lacks
            # (real FHB Master case: a 7-receipt reference group whose ties
            # lived only in the ST export).
            raw.append(e)
        deduped = _dedup_keep_largest(raw, lambda e: e.id)
        pool.extend(deduped)
        counts["ST"] = len(deduped)

    # 2. All-status Receivables receipts from RECEIPTS (authoritative).
    rc = loaded.get("RECEIPTS")
    if rc:
        rows, m, hi = rc["rows"], rc["map"], rc["header_index"]
        raw = []
        for r in rows[hi + 1:]:
            amt = cents(_cell(r, m.get("amount")))
            recno = N(_cell(r, m.get("receipt_number")))
            if _blankish(recno):
                # 'NA'-numbered receipts are real rows sharing a placeholder —
                # falling through to keep-largest dedup on the id 'NA' would
                # collapse them all into one (2026-07-12 finding: two BlueCare
                # receipts summing to a $246,736.64 bank line vanished).
                doc = N(_cell(r, m.get("document_number")))
                recno = f"DOC {doc}" if doc and not _blankish(doc) else ""
            if amt is None or not recno:
                continue
            status = _norm_header(_cell(r, m.get("status")))
            available = status in OPEN_RECEIPT_STATUSES
            e = _mk_entry(
                id=recno,
                amount_cents=amt,
                dt=parse_date(_cell(r, m.get("receipt_date")) or _cell(r, m.get("deposit_date"))),
                reference=_cell(r, m.get("reference")) or recno,
                counterparty=_cell(r, m.get("customer_name")),
                source="AR",
                status=status or "UNR",
                available=available,
                origin="RECEIPTS",
            )
            # Misdirected-SPN detector (owner hard guardrail, 2026-07-18):
            # remember which bank account this receipt remits to.  The
            # decision (foreign vs ours) happens at merge time below — a
            # receipt that merges onto one of THIS account's STs is ours by
            # definition, whatever the remittance column says.
            if m.get("remittance_bank_account") is not None and account and account != "UNKNOWN":
                rem = account_of_bank_name(_cell(r, m.get("remittance_bank_account")))
                if rem is not None and rem != account:
                    e.foreign_account = rem
            raw.append(e)
        deduped = _dedup_keep_largest(raw, lambda e: e.id)
        by_id = {}
        for p in pool:
            if p.source == "AR":
                by_id.setdefault(p.base_id or p.id, []).append(p)
        merged_rc = appended = foreign_rc = 0
        for e in deduped:
            group = by_id.get(e.base_id or e.id) or []
            # Merge by equal signed cents only — same label, different amount
            # is a distinct receipt (mirrors the MET bridge rule).
            equal = [p for p in group if p.amount_cents == e.amount_cents]
            prev = equal[0] if len(equal) == 1 else None
            if prev is not None:
                # An ST exists in THIS account's export — not misdirected.
                prev.status = e.status or prev.status
                prev.available = e.available
                if not prev.counterparty and e.counterparty:
                    prev.counterparty = e.counterparty
                    prev.payer_tokens = prev.payer_tokens | payer_tokens(e.counterparty)
                if (not prev.reference or prev.reference == prev.id) and \
                        e.reference and e.reference != e.id:
                    prev.reference = e.reference
                    prev.znref = znorm(e.reference)
                    prev.digits = digit_runs(e.reference)
                if prev.date is None:
                    prev.date = e.date
                merged_rc += 1
            elif e.foreign_account:
                # Shadow pool (owner hard guardrail, 2026-07-18): the receipt
                # remits to ANOTHER bank account and has no ST here.  Kept in
                # the pool but never available to normal passes — only the
                # dedicated misdirected search may cite it.
                e.available = False
                pool.append(e)
                foreign_rc += 1
            else:
                pool.append(e)
                appended += 1
        counts["RECEIPTS"] = appended
        runlog["receipts_merged_to_st"] = merged_rc
        if foreign_rc:
            counts["RECEIPTS_FOREIGN"] = foreign_rc
            runlog["receipts_foreign_account"] = foreign_rc

    # 2b. Open Payables payments from PAYMENTS (owner, 2026-07-19): the AP
    #     analogue of RECEIPTS.  Only OUTSTANDING payments are open money —
    #     "Negotiable" (Reconciled=No, not Voided); "Cleared"/"Voided" are
    #     reconciled or cancelled.  Payment Amount is a positive disbursement
    #     magnitude; negate it to the bank line's signed cents.  Merge onto the
    #     ST export's AP rows by payment number (the check identity) at equal
    #     signed cents; otherwise append a new open AP entry.
    pm = loaded.get("PAYMENTS")
    if pm:
        rows, m, hi = pm["rows"], pm["map"], pm["header_index"]
        raw = []
        for r in rows[hi + 1:]:
            mag = cents(_cell(r, m.get("amount")))
            pno = N(_cell(r, m.get("payment_number")))
            if mag is None or not pno:
                continue
            status = _norm_header(_cell(r, m.get("status")))
            reconciled = _norm_header(_cell(r, m.get("reconciled")))
            voided = "VOID" in status
            is_open = (not voided) and (status in OPEN_PAYMENT_STATUSES
                                        or (reconciled == "NO" and status not in CLOSED_RECEIPT_STATUSES))
            if not is_open:
                continue  # reconciled/cleared/voided payments are not open money
            e = _mk_entry(
                id=pno,
                amount_cents=-mag,           # disbursement: the bank sees it negative
                dt=parse_date(_cell(r, m.get("payment_date"))),
                reference=pno,               # the payment/check number is the identity
                counterparty=_cell(r, m.get("payee")),
                source="AP",
                status=status or "UNR",
                available=True,
                origin="PAYMENTS",
                transaction_type="Check",
            )
            raw.append(e)
        deduped = _dedup_keep_largest(raw, lambda e: e.id)
        by_id = {}
        for p in pool:
            if p.source == "AP":
                by_id.setdefault(p.base_id or p.id, []).append(p)
        merged_pm = appended_pm = 0
        for e in deduped:
            group = by_id.get(e.base_id or e.id) or []
            equal = [p for p in group if p.amount_cents == e.amount_cents]
            prev = equal[0] if len(equal) == 1 else None
            if prev is not None:
                # ST export already carries this payment — payment feed is
                # authoritative for AP payee/date, never re-opens a consumed ST.
                if not prev.counterparty and e.counterparty:
                    prev.counterparty = e.counterparty
                    prev.payer_tokens = prev.payer_tokens | payer_tokens(e.counterparty)
                if prev.date is None:
                    prev.date = e.date
                merged_pm += 1
            else:
                pool.append(e)
                appended_pm += 1
        counts["PAYMENTS"] = appended_pm
        runlog["payments_merged_to_st"] = merged_pm

    # 3. ORT receipts from MET for the ECT chain (index by d: and reference).
    met = loaded.get("MET")
    if met:
        rows, m, hi = met["rows"], met["map"], met["header_index"]
        raw = []
        met_total = 0
        coa = loaded.get("CHART_OF_ACCOUNTS")
        coa_validity = {"rows_seen": 0, "unrecognized_combo": 0,
                        "non_postable_efdp": 0}
        have_account = bool(account) and account != "UNKNOWN"
        scoped = m.get("bank_account_name") is not None and have_account
        # COA fallback scope (owner, 2026-07-18): when the export lacks a
        # bank-name column, the asset combo's GL cash account is the scope
        # key — otherwise an all-accounts export would leak every account
        # into the pool.  When BOTH keys are bound, bank name stays
        # authoritative (it names the statement the row belongs to) and
        # GL disagreements are surfaced in the runlog, never silently used.
        gl_scoped = (not scoped and have_account
                     and m.get("asset_segments") is not None)
        gl_conflicts = 0
        foreign_raw = []
        for r in rows[hi + 1:]:
            met_total += 1
            # Scope join (ORT doc section 4.6): the MET export spans EVERY
            # account; keep only rows whose long bank-account name maps to
            # the account being reconciled.  Cross-account rows never enter
            # the pool as matchable entries — but rows belonging to a
            # DIFFERENT configured account feed the misdirected shadow pool
            # (owner hard guardrail, 2026-07-18) instead of being dropped.
            foreign_acc = ""
            if scoped:
                row_acc = account_of_bank_name(_cell(r, m.get("bank_account_name")))
                if row_acc != account:
                    if row_acc is None:
                        continue
                    foreign_acc = row_acc
                elif m.get("asset_segments") is not None:
                    gl_acc = account_of_gl_segments(_cell(r, m.get("asset_segments")))
                    if gl_acc is not None and gl_acc != account:
                        gl_conflicts += 1
            elif gl_scoped:
                gl_acc = account_of_gl_segments(_cell(r, m.get("asset_segments")))
                if gl_acc != account:
                    if gl_acc is None:
                        continue
                    foreign_acc = gl_acc
            amt = cents(_cell(r, m.get("amount")))
            trx = N(_cell(r, m.get("trx_id")))
            if amt is None or not trx:
                continue
            asset_seg = N(_cell(r, m.get("asset_segments")))
            offset_seg = N(_cell(r, m.get("offset")))
            desc = N(_cell(r, m.get("description")))
            dep_desc, rec_desc, _payer = parse_met_description(desc)
            # Native DEPOSIT_ID / RECEIPT_ID columns are authoritative; the
            # d:/r: description parse is the fallback (they agree on 160,692
            # of 160,692 overlapping real rows).
            dep_id = N(_cell(r, m.get("deposit_id"))) or dep_desc
            rec_id = N(_cell(r, m.get("receipt_id"))) or rec_desc
            ref_text = clean_ref(_cell(r, m.get("reference_text")))
            cleared = parse_date(_cell(r, m.get("cleared_date")))
            status_raw = _norm_header(_cell(r, m.get("status")))
            # Status is authoritative when present; the cleared-date=>closed
            # inference (Section 8.3) applies only when status is absent —
            # the real OTBI export stamps CET_CLEARED_DATE on UNR rows too.
            if status_raw:
                closed = status_raw in CLOSED_RECEIPT_STATUSES
            else:
                closed = cleared is not None
            e = _mk_entry(
                id=trx,
                amount_cents=amt,
                dt=parse_date(_cell(r, m.get("transaction_date"))),
                reference=ref_text or _met_reference(desc) or trx,
                counterparty=_payer,
                source="EXT",
                status=status_raw or ("REC" if closed else "UNR"),
                available=not closed,
                origin="MET",
                deposit_id=dep_id,
                receipt_id=rec_id,
                transaction_type=_cell(r, m.get("transaction_type")),
                asset_segments=asset_seg,
                offset_segments=offset_seg,
            )
            if foreign_acc:
                # Shadow entry: only the misdirected search may cite it, and
                # only while it is still OPEN on its own account.
                if e.available:
                    e.foreign_account = foreign_acc
                    e.available = False
                    foreign_raw.append(e)
                continue
            # CoA combo-validity diagnostic (owner, 2026-07-19): decode each
            # in-account MET combo against the CoA universe.  Counter only —
            # like met_gl_conflicts it never drops or downgrades a row.
            if coa is not None and asset_seg:
                coa_validity["rows_seen"] += 1
                v = coa_combo_valid(asset_seg, coa)
                if not v["recognized"]:
                    coa_validity["unrecognized_combo"] += 1
                if not v["postable"]:
                    coa_validity["non_postable_efdp"] += 1
            raw.append(e)
        deduped = _dedup_keep_largest(raw, lambda e: e.id)
        counts["MET"] = len(deduped)
        if foreign_raw:
            foreign_ded = _dedup_keep_largest(foreign_raw, lambda e: e.id)
            pool.extend(foreign_ded)
            counts["MET_FOREIGN"] = len(foreign_ded)
            runlog["met_foreign_account_open_rows"] = len(foreign_ded)
        if scoped:
            runlog["met_scope"] = {"rows_total": met_total, "rows_in_account": len(raw),
                                   "key": "bank_name"}
            if gl_conflicts:
                # Rows whose bank name says this account but whose GL cash
                # account belongs to another depository (cross-account
                # postings) — kept (bank name is authoritative), surfaced
                # for the reconciler.
                runlog["met_gl_conflicts"] = gl_conflicts
        elif gl_scoped:
            runlog["met_scope"] = {"rows_total": met_total, "rows_in_account": len(raw),
                                   "key": "gl_cash_account"}
        if coa is not None and coa_validity["rows_seen"]:
            runlog["coa_combo_validity"] = coa_validity
        # MET <-> ST bridge (CET_TRANSACTION_ID == ST Transaction Number, a
        # 1:1 bijection): the same transaction must never sit in the pool
        # twice, or reference groups double-count.  The ST export row is
        # canonical; borrow the bridge fields the MET copy carries.  MET-only
        # transactions join the pool as their own entries.
        by_id = {}
        for p in pool:
            by_id.setdefault(p.base_id or p.id, []).append(p)
        merged = 0
        dropped_mismatch = 0
        met_status_overrides = 0
        for e in deduped:
            group = by_id.get(e.base_id or e.id)
            if not group:
                pool.append(e)
                continue
            prev = None
            if len(group) == 1:
                prev = group[0]
            else:
                # Disambiguated same-label STs: the MET copy of the same
                # transaction carries the same signed cents — merge into that
                # member; never guess among unequal ones.
                equal = [p for p in group if p.amount_cents == e.amount_cents]
                if len(equal) == 1:
                    prev = equal[0]
                else:
                    dropped_mismatch += 1
                    continue
            merged += 1
            if not prev.deposit_id:
                prev.deposit_id = e.deposit_id
            if not prev.receipt_id:
                prev.receipt_id = e.receipt_id
            if not prev.asset_segments and e.asset_segments:
                prev.asset_segments = e.asset_segments  # CoA decode (advisory)
            if not prev.offset_segments and e.offset_segments:
                prev.offset_segments = e.offset_segments  # ECT posting side
            if not prev.counterparty and e.counterparty:
                prev.counterparty = e.counterparty
                prev.payer_tokens = prev.payer_tokens | payer_tokens(e.counterparty)
            if (not prev.reference or prev.reference == prev.id) and \
                    e.reference and e.reference != e.id:
                prev.reference = e.reference
                prev.znref = znorm(e.reference)
                prev.digits = digit_runs(e.reference)
                prev.is_mid = prev.is_mid or e.is_mid
            # Orphan doctrine R3 (owner, 2026-07-19): MET status OUTRANKS the
            # ST export.  If MET says the transaction's money is already
            # consumed (REC / cleared), the open-looking ST is an orphan or a
            # stale-export artifact — never consume it, whatever the ST
            # export's recon status says.  Guarded by exact signed-cent
            # equality: a singleton same-label merge with a different amount
            # (keep-largest kept a REC total row while the ST export carries
            # only an open split; or a CET id collides with an unrelated
            # receipt number) must NOT flip availability — only the true
            # same-transaction bridge does.
            if prev.available and not e.available and \
                    prev.amount_cents == e.amount_cents:
                prev.available = False
                prev.status = e.status or prev.status
                met_status_overrides += 1
        runlog["met_bridged_to_st"] = merged
        if met_status_overrides:
            runlog["met_status_overrides"] = met_status_overrides
        if dropped_mismatch:
            runlog["met_bridge_amount_mismatch_dropped"] = dropped_mismatch

    # AR receipt-application enrichment (AR_MATCHED): join by receipt number
    # (scoped to this account's bank name) to backfill missing dates with the
    # BANK deposit date and missing counterparties with the receipt comments.
    am = loaded.get("AR_MATCHED")
    if am:
        rows, m, hi = am["rows"], am["map"], am["header_index"]
        by_recno = {}
        for r in rows[hi + 1:]:
            bank = account_of_bank_name(_cell(r, m.get("bank_account_name")))
            if bank and account and account != "UNKNOWN" and bank != account:
                continue
            recno = clean_ref(_cell(r, m.get("receipt_number")))
            if not recno:
                continue
            by_recno.setdefault(recno, r)
        enriched = 0
        for e in pool:
            if e.source != "AR":
                continue
            r = by_recno.get(clean_ref(e.base_id or e.id)) or by_recno.get(clean_ref(e.reference))
            if r is None:
                continue
            touched = False
            if e.date is None:
                d = parse_date(_cell(r, m.get("deposit_date"))) or                     parse_date(_cell(r, m.get("receipt_date")))
                if d is not None:
                    e.date = d
                    touched = True
            if not e.counterparty:
                c = N(_cell(r, m.get("comments")))
                if c:
                    e.counterparty = c
                    e.payer_tokens = e.payer_tokens | payer_tokens(c)
                    touched = True
            if touched:
                enriched += 1
        runlog["ar_matched_enriched"] = enriched
        runlog["ar_matched_rows"] = len(by_recno)

    # Orphan doctrine 2.3 (owner, 2026-07-19): one ORT receipt occasionally
    # spawns two identical open external STs (dual-fire; 132 confirmed on FHB
    # Master).  The twin is not new money — prior reconciliation consumed one
    # copy's line, or will.  Keep exactly ONE available (deterministic order),
    # mark the twin(s) unavailable so nothing ever matches them.
    by_fire = {}
    for e in pool:
        if e.available and e.receipt_id and not e.foreign_account:
            by_fire.setdefault((_norm_id(e.receipt_id), e.amount_cents), []).append(e)
    dual_fire_twins = 0
    for key, group in by_fire.items():
        if len(group) > 1 and len({e.id for e in group}) > 1:
            for twin in _sorted(group)[1:]:
                twin.available = False
                dual_fire_twins += 1
    if dual_fire_twins:
        runlog["dual_fire_twins"] = dual_fire_twins

    runlog["pool_sizes"] = counts
    runlog["pool_total"] = len(pool)
    # Status breakdown for the run log.
    by_status = {}
    for e in pool:
        by_status[e.status] = by_status.get(e.status, 0) + 1
    runlog["pool_by_status"] = by_status
    return pool


def _classify_source(source):
    """Normalize a raw source string to one of AR/EXT/AP/PAY/GL (Section 8)."""
    s = _norm_header(source)
    if not s:
        return "EXT"
    if "AR" in s or "RECEIV" in s:
        return "AR"
    if "AP" in s or "PAYABLE" in s:
        return "AP"
    if "PAY" in s:
        return "PAY"
    if "JOURNAL" in s or s == "GL" or "MANUAL" in s or s == "JE":
        return "GL"
    if "EXT" in s or "ECT" in s or "ORT" in s:
        return "EXT"
    return "EXT"


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


# ---- MET description parser (Section 5.4) ----------------------------

def parse_met_description(desc):
    """Split CET_DESCRIPTION on '|'.  seg0->d: deposit id; seg1->r: receipt id;
    seg2+ joined -> payer/purpose.  Returns (deposit_id, receipt_id, payer)."""
    parts = [p.strip() for p in N(desc).split("|")]
    dep = rec = ""
    if len(parts) >= 1:
        m = re.search(r"d:\s*(\d+)", parts[0], re.IGNORECASE)
        if m:
            dep = m.group(1)
    if len(parts) >= 2:
        m = re.search(r"r:\s*(\d+)", parts[1], re.IGNORECASE)
        if m:
            rec = m.group(1)
    payer = " | ".join(parts[2:]).strip() if len(parts) > 2 else ""
    # Fallback: search whole string for d:/r: if the split missed them.
    if not dep:
        m = re.search(r"d:\s*(\d+)", N(desc), re.IGNORECASE)
        if m:
            dep = m.group(1)
    if not rec:
        m = re.search(r"r:\s*(\d+)", N(desc), re.IGNORECASE)
        if m:
            rec = m.group(1)
    return dep, rec, payer


def _met_reference(desc):
    """Best-effort reference extracted from the payer/purpose segment."""
    _, _, payer = parse_met_description(desc)
    return payer


# ======================================================================
# BSL model + lane classification (Section 9, P1)
# ======================================================================

LANE_STATE = "STATE"
LANE_MERCHANT = "MERCHANT"
LANE_GENERAL = "GENERAL"

# FHB transaction codes whose whole Account Servicer Reference is the match key.
FHB_WHOLE_REF_CODES = {"142", "174", "175", "165", "244", "495", "451", "699",
                       "475", "357", "631", "661"}

_MERCHANT_KEYWORDS = ("MERCHANT SERVICE", "BANKCARD", "TOUCHNET", "CYBERSOURCE",
                      "PAYMENTECH")
_STATE_KEYWORDS = ("STATE-TN", "STATE OF TENN")


@dataclass
class BSL:
    line_key: str
    date: object
    amount_cents: int
    recon_reference: str
    reference_raw: str
    additional_info: str
    transaction_type: str
    transaction_code: str
    customer_reference: str = ""
    account_servicer_reference: str = ""
    lane: str = LANE_GENERAL
    ref_digits: set = field(default_factory=set)
    payer_tokens: set = field(default_factory=set)
    mid: str = ""
    line_info: str = ""


def build_recon_reference(reference, account_servicer_reference, transaction_code):
    """RECON_REFERENCE is the whole Account Servicer Reference for the FHB
    whole-ref transaction codes (Parse Rules (X~)); fall back to reference."""
    asr = clean_ref(account_servicer_reference)
    code = N(transaction_code)
    reference = clean_ref(reference)
    if code in FHB_WHOLE_REF_CODES and asr:
        return asr
    if asr:
        return asr
    return N(reference)


def extract_ach_payer(additional_info):
    """Parse ACH payer from addenda 'SENDING CO NAME: <X> ENTRY DESC' and from
    'Company Name: <X>' in an enriched addenda blob.  Never tokenize BAI2
    labels as payer text."""
    s = N(additional_info)
    m = re.search(r"SENDING\s+CO\s+NAME:\s*(.+?)\s+ENTRY\s+DESC", s, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.search(r"Company\s+Name:\s*(.+?)(?:$|\s{2,}|\|)", s, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return ""


def _mid_from_text(additional_info):
    """Extract a Customer ID MID from addenda if present."""
    for tok in re.findall(r"\d{10}", N(additional_info)):
        if is_mid(tok):
            return znorm(tok)
    return ""


def classify_lane(bsl: BSL) -> str:
    """P1 lane selector: STATE before MERCHANT before GENERAL."""
    info = bsl.additional_info.upper()
    ref = bsl.recon_reference.upper()
    # STATE
    if info.startswith("STATE-TN") or any(k in info for k in _STATE_KEYWORDS):
        return LANE_STATE
    # MERCHANT
    if bsl.mid or is_mid(bsl.recon_reference):
        return LANE_MERCHANT
    if any(k in info for k in _MERCHANT_KEYWORDS) or any(k in ref for k in _MERCHANT_KEYWORDS):
        return LANE_MERCHANT
    return LANE_GENERAL


def make_bsl(line_key, dt, amount_cents, reference, account_servicer_reference,
             additional_info, transaction_type, transaction_code,
             customer_reference=""):
    recon_ref = build_recon_reference(reference, account_servicer_reference, transaction_code)
    ach_payer = extract_ach_payer(additional_info)
    mid = _mid_from_text(additional_info) or (znorm(recon_ref) if is_mid(recon_ref) else "")
    bsl = BSL(
        line_key=N(line_key),
        date=dt,
        amount_cents=amount_cents,
        recon_reference=recon_ref,
        reference_raw=clean_ref(reference),
        additional_info=N(additional_info),
        customer_reference=clean_ref(customer_reference),
        account_servicer_reference=clean_ref(account_servicer_reference),
        transaction_type=N(transaction_type),
        transaction_code=N(transaction_code),
        ref_digits=digit_runs(recon_ref),
        payer_tokens=payer_tokens(recon_ref) | payer_tokens(ach_payer) | payer_tokens(additional_info),
        mid=mid,
    )
    bsl.lane = classify_lane(bsl)
    # Human-readable line info (no formulas; static).
    info = N(additional_info) or N(reference) or N(transaction_type)
    bsl.line_info = f"{bsl.line_key} {info}".strip()[:200]
    return bsl


# ======================================================================
# Date doctrine (Section 11)
# ======================================================================

DATE_CEILING = 15


def date_band(lag):
    """Non-State band label for a signed lag (BSL - ST)."""
    if lag is None:
        return "UNKNOWN"
    a = abs(lag)
    if a <= 3:
        return "STRONG"
    if a <= 7:
        return "MODERATE"
    if a <= 15:
        return "WEAK"
    if a <= 30:
        return "SUSPICIOUS"
    return "REJECT"


def date_ok_directional(lag):
    """Owner doctrine (2026-07-11, final): the date gate applies ONLY when
    the ST was entered 8 or more days before the BSL (lag = BSL - ST >= 8).
    An ST entered after the BSL — by any amount, even months — is valid;
    unknown dates are not gated."""
    return lag is None or lag < 8


def date_ok_merchant(lag):
    """Card window: ST precedes BSL by 1-4 days (lag in 1..4)."""
    return lag is not None and 1 <= lag <= 4


# ======================================================================
# Consumption ledger (Section 2.7, 8.4)
# ======================================================================

class Ledger:
    """One consumption ledger spanning Matches and Candidates.  An ST named in
    a Candidate is barred from later Match promotion."""

    def __init__(self):
        self._consumed = set()       # matched (hard-consumed)
        self._candidate_barred = set()

    def is_available(self, entry: PoolEntry):
        return entry.available and entry.id not in self._consumed and entry.id not in self._candidate_barred

    def consume_match(self, ids):
        for i in ids:
            self._consumed.add(i)

    def consume_candidate(self, ids):
        for i in ids:
            self._candidate_barred.add(i)

    def consumed(self):
        return set(self._consumed)

    def barred(self):
        return set(self._candidate_barred)


# ======================================================================
# Classification result model
# ======================================================================

MATCH = "Match"
CANDIDATE = "Candidate"
REVIEW = "Review"
# Misdirected (owner HARD GUARDRAIL, 2026-07-18): the bank deposit landed in
# THIS account but the system transaction/receipt was booked to a DIFFERENT
# bank account (any source — AR/AP/EXT; any bank — FHB/Regions; any campus).
# Real case: City of Memphis $70,992.66 hit FHB UTHSC while receipt 300045836
# remits to FHB Master — "300045836 ST does not exist" on UTHSC.  These pairs
# get their own workbook tab; they can never auto-reconcile without a reroute.
MISDIRECTED = "Misdirected"

CONF_HIGH = "High"
CONF_MEDIUM = "Medium"
CONF_LOW = "Low"


@dataclass
class Placement:
    bsl: BSL
    kind: str               # Match / Candidate / Review
    confidence: str
    st_entries: list        # list[PoolEntry]
    codes: list             # exception / cause codes
    explanation: str
    pass_name: str
    deposit_ids: list = field(default_factory=list)
    receipt_ids: list = field(default_factory=list)


# ======================================================================
# Section 9 — FORWARD pipeline (fixed order; later never overrides earlier)
# ======================================================================

def _bsl_znorms(bsl):
    """Non-empty znorms of the BSL's 5 identifier fields (perf fast path —
    an empty znorm can never satisfy reference_equal)."""
    out = []
    for v in (bsl.reference_raw, bsl.additional_info, bsl.customer_reference,
              bsl.account_servicer_reference, bsl.recon_reference):
        if v:
            z = znorm(v)
            if z:
                out.append(z)
    return tuple(out)


def _entry_znorms(e):
    """Non-empty znorms of the ST entry's 4 identifier fields."""
    out = []
    for v in (e.reference, e.id, e.spr, e.counterparty):
        if v:
            z = znorm(v)
            if z:
                out.append(z)
    return tuple(out)


def _cross_reference_tie_z(bz, ez):
    """cross_reference_tie over precomputed znorm tuples (identical result:
    a truthy raw with an empty znorm can never hit reference_equal, so
    restricting to non-empty znorms loses nothing)."""
    for zb in bz:
        for zs in ez:
            if _reference_equal_z(zb, zs):
                return True
    return False


def cross_reference_tie(bsl, e):
    """Owner-mandated multi-cross-reference screen (2026-07-11): EVERY BSL
    identifier field — Reference, Additional Information (full text, BAI2-
    enriched), Customer Reference, Account Servicer Reference — is compared
    against EVERY ST identifier field — Reference, Transaction Number,
    Structured Payment Reference, Counterparty.  Full cell contents; no
    field skipped, no truncation."""
    return _cross_reference_tie_z(_bsl_znorms(bsl), _entry_znorms(e))


def _grams6(z):
    """All character 6-grams of a znorm string (index keys)."""
    return [z[i:i + 6] for i in range(len(z) - 5)]


def _build_tie_index(bsls, pool):
    """{bsl.line_key -> [pool entries with cross_reference_tie, POOL ORDER]}.

    Candidate generation is a provable superset of reference_equal hits
    (equality or containment >= 6 always shares a 6-gram with the indexed
    side; below-6 hits require exact equality, kept in a separate exact
    dict; cross-length-class hits below 6 are impossible), then every
    candidate pair is verified with the EXACT tie predicate — so the result
    equals the brute-force full-pool scan, including order.  Built ONCE per
    run: cross_reference_tie depends only on fields immutable after
    build_pool.  Availability is deliberately NOT indexed (doctrine rule 6:
    it is derived through the ledger at every access)."""
    grams = {}      # 6-gram -> set of bsl indices
    exact = {}      # znorm shorter than 6 -> set of bsl indices
    bz_list = []
    for bi, b in enumerate(bsls):
        bz = _bsl_znorms(b)
        bz_list.append(bz)
        for z in bz:
            if len(z) >= 6:
                for g in set(_grams6(z)):
                    grams.setdefault(g, set()).add(bi)
            else:
                exact.setdefault(z, set()).add(bi)
    tie_index = {b.line_key: [] for b in bsls}
    for e in pool:
        ez = _entry_znorms(e)
        if not ez:
            continue
        cand = set()
        for z in ez:
            if len(z) >= 6:
                for g in _grams6(z):
                    s = grams.get(g)
                    if s:
                        cand |= s
            s = exact.get(z)
            if s:
                cand |= s
        for bi in cand:
            if _cross_reference_tie_z(bz_list[bi], ez):
                tie_index[bsls[bi].line_key].append(e)
    return tie_index


def cross_digit_tie(bsl, e):
    """Weak digit-run corroboration over the same 4x4 field grid as
    cross_reference_tie: a shared >=5-digit run anywhere between the BSL's
    identifier fields and the ST's.  Supports a DATE_GATE Candidate, never a
    Match."""
    bsl_vals = (bsl.reference_raw, bsl.additional_info, bsl.customer_reference,
                bsl.account_servicer_reference, bsl.recon_reference)
    st_vals = (e.reference, e.id, e.spr, e.counterparty)
    for b in bsl_vals:
        if not b:
            continue
        for s in st_vals:
            if s and reference_tie(b, s):
                return True
    return False


_PAYER_LABEL_RE = re.compile(
    r"(?:CO(?:MPANY)?\s+NAME|CUSTOMER\s+NAME|INDIVIDUAL\s+NAME)"
    r"\s*:?\s*([A-Za-z0-9 &.,'\-]{3,40})", re.I)
_CONTRA_STOP = {"THE", "AND", "FOR", "FROM", "WITH", "INC", "LLC", "CORP",
                "CO", "OF"}

# Oracle ORT bank-feed load stamps unattributed External lines with a generic
# "FEED SESSION <n>" batch label in the counterparty column (owner, 2026-07-16:
# "review all cells and all characters, especially the description text").  That
# label names the load BATCH, not a payer/beneficiary/originator — so it is
# SILENCE on payer identity, never a contradicting name.  Strip it before
# tokenizing so a real bank-side payer (e.g. a Heartland settlement) is not
# falsely contradicted by a feed-session artifact (owner doctrine: "silence
# never contradicts").
_FEED_SESSION_RE = re.compile(r"(?i)\bFEED\s+SESSION\b")


def _contra_tokens(text):
    cleaned = _FEED_SESSION_RE.sub(" ", N(text))
    return {w.upper() for w in _ALPHA_TOKEN_RE.findall(cleaned)
            if len(w) >= 3 and w.upper() not in _CONTRA_STOP}


def _bsl_payer_name_tokens(bsl):
    """Tokens of the payer NAMES the bank line actually discloses: the
    Customer Reference field plus 'CO NAME:' / 'CUSTOMER NAME:' labeled
    segments of the (BAI2-enriched) addenda.  Deliberately NOT the filtered
    payer_tokens set — common words like STATE/PAYMENTS are stopworded there
    for tie purposes, which would fabricate contradictions between identical
    payers ('STATE-TN PAYMNTS' vs 'State of TN Payments')."""
    toks = set()
    cr = N(bsl.customer_reference)
    if cr and not _blankish(cr):
        toks |= _contra_tokens(cr)
    for m in _PAYER_LABEL_RE.finditer(N(bsl.additional_info)):
        toks |= _contra_tokens(m.group(1))
    return toks


def _token_overlap(a, b):
    """Shared token, tolerating bank-side truncation (prefix, len >= 4)."""
    if a & b:
        return True
    for x in a:
        for y in b:
            if len(x) >= 4 and len(y) >= 4 and (x.startswith(y) or y.startswith(x)):
                return True
    return False


# Payer-family aliases (owner, 2026-07-14): distinct trade names for the SAME
# payer are NOT a contradiction and DO corroborate.  VSHP (Volunteer State
# Health Plan) / TN CARE SELECT / TennCare is BlueCare Tennessee (BlueCross
# BlueShield of TN)'s managed-care product.  Each entry maps a canonical family
# to the distinctive znorm tokens that denote it; keep tokens specific enough
# not to collide with unrelated payers.
_PAYER_FAMILY_ALIASES = [
    ("BLUECARE", {"BLUECARE", "BCBST", "VSHP", "VOLUNTEERSTATEHEALTH",
                  "TENNCARE", "TNCARE", "TNCARESELECT"}),
]


def payer_family(text) -> set:
    """The canonical payer-family id(s) a piece of payer text belongs to."""
    z = znorm(text)
    return {canon for canon, toks in _PAYER_FAMILY_ALIASES
            if any(t in z for t in toks)}


def _bsl_payer_families(bsl) -> set:
    return payer_family(bsl.additional_info) | payer_family(bsl.customer_reference)


def payer_contradiction(bsl, entries):
    """Owner directive (2026-07-11, answer-key review): an amount-only pairing
    whose payer texts actively disagree is wrong — 'City of Chattanooga has
    nothing to do with Israel.'  When the bank line names a payer (labeled
    addenda name / Customer Reference) AND the ST side carries counterparty
    text (ST export or MET CET_DESCRIPTION) and the two share no token, the
    pairing is contradicted.  Consulted only by zero-corroboration
    placements — a reference tie always outranks payer text; silence on
    either side never contradicts.

    Credits only: on a debit the bank line names the payment CHANNEL
    (Convera, card processor) while the ST counterparty names the
    BENEFICIARY — different roles, not comparable, never a contradiction."""
    if bsl.amount_cents <= 0:
        return False
    bt = _bsl_payer_name_tokens(bsl)
    if not bt:
        return False
    st = set()
    for e in entries:
        st |= _contra_tokens(e.counterparty)
    if not st:
        return False
    if _token_overlap(bt, st):
        return False
    # A shared payer FAMILY (e.g. VSHP / TennCare == BlueCare) is agreement,
    # not contradiction, even when the surface tokens differ.
    bfam = _bsl_payer_families(bsl)
    if bfam and any(bfam & payer_family(e.counterparty) for e in entries):
        return False
    return True


def _is_deposit_correction(bsl):
    """Deposit-correction bank lines are manual fixes (owner, 2026-07-11):
    they rarely have an ST and need a manual ECT.  They may surface as
    flagged Candidates — never as Matches from amount-sum passes."""
    txt = _norm_header(bsl.additional_info) + _norm_header(bsl.line_info)
    return "CORRECTION" in txt or "CORRECTED" in txt


def amount_distinctive(c):
    """Owner doctrine (2026-07-11): amount alone can suffice ONLY when the
    exact signed-cents value is statistically unlikely to collide — a rare
    combination of digits.  Deterministic proxy: non-zero cents AND at least
    $1,000 in magnitude (e.g. 2,455,469.69 qualifies; 500.00, 60.00, or any
    round-dollar amount never does)."""
    return c % 100 != 0 and abs(c) >= 100000


def _pool_amount_twin_free(amount, pool_amount_counts, cited):
    """Hard guardrail (adversarial review, 2026-07-19): the distinctive-amount
    exception claims the amount is UNIQUE on the pool side — so every pool
    entry at that signed-cents value must be among the cited entries.  A
    closed, foreign-shadow, or out-of-band open twin at the same amount means
    the digit combination is not rare (the money may already be consumed, or
    belong to another account) and amount-only evidence stays a Candidate."""
    return pool_amount_counts.get(amount, 0) == \
        sum(1 for e in cited if e.amount_cents == amount)


def _is_convera(bsl):
    return "CONVERA" in _norm_header(bsl.additional_info) + _norm_header(bsl.line_info)


STALE_CANDIDATE_CEILING = 12


def _ext_stale_barred(bsl, e):
    """Owner rule (2026-07-12): an External-source ST entered 12 or more
    days BEFORE the BSL statement date is almost certainly not the
    counterpart — it may not appear even as a Candidate.  (A Match is
    already barred at 8+ days stale; 8-11 days stale may still surface as a
    flagged Candidate.  BSL-before-ST stays unbounded-valid, and non-EXT
    sources are untouched.)"""
    if e.source != "EXT":
        return False
    lag = signed_lag(bsl.date, e.date)
    return lag is not None and lag >= STALE_CANDIDATE_CEILING


@functools.lru_cache(maxsize=1 << 15)
def _mid_tokens(text):
    """Every MID-shaped 10-digit run in the text (80xxxxxxxx / 2000xxxxxx,
    Heartland excluded), frozen for set algebra."""
    return frozenset(run for run in re.findall(r"\d{10}", N(text)) if is_mid(run))


def _is_card_fee_debit(bsl):
    """Owner rule (2026-07-12): a NEGATIVE bank line for a card chargeback or
    merchant fee ('chargeback(s)', 'merchant fee(s)' and derivatives) belongs
    to exactly one merchant — the MID is the critical matching string."""
    if bsl.amount_cents >= 0:
        return False
    txt = _norm_header(bsl.additional_info) + _norm_header(bsl.line_info) +         _norm_header(bsl.transaction_type)
    return "CHARGEBACK" in txt or "MERCHANTFEE" in txt


def _bsl_card_mids(bsl):
    return (_mid_tokens(bsl.reference_raw) | _mid_tokens(bsl.additional_info)
            | _mid_tokens(bsl.customer_reference)
            | _mid_tokens(bsl.account_servicer_reference)
            | _mid_tokens(bsl.line_info))


def _electronic_type(t):
    z = _norm_header(t)
    return ("ACH" in z or "AUTOMATEDCLEARINGHOUSE" in z or "EFT" in z
            or "ELECTRONICFUNDS" in z)


def _misc_bsl(bsl):
    return "MISC" in _norm_header(bsl.transaction_type)


def _has_any_tie(bsl, e):
    """Any corroboration that overrides transaction-type incongruence: a
    reference tie (the 4x4 screen or a digit-run), a payer-token overlap, or
    a shared MID."""
    if cross_reference_tie(bsl, e) or cross_digit_tie(bsl, e):
        return True
    if e.payer_tokens & bsl.payer_tokens:
        return True
    if _bsl_card_mids(bsl) & (_mid_tokens(e.reference) | _mid_tokens(e.id)
                              | _mid_tokens(e.counterparty)):
        return True
    return False


def _check_number_conflict(bsl, e):
    """Orphan doctrine R8 (owner, 2026-07-19): on check rails the check
    number IS the identity.  Same amount + a DIFFERENT check number is a
    conflict, never a partial match — this is exactly how the FHB AP $1,100
    auto-rec cascade propagated.  Applies only when BOTH sides carry a
    numeric check-shaped reference (>= 4 digits); a blank reference is
    silence and never conflicts.  Compare canonically (leading zeros
    stripped; text reads only — float coercion corrupts 0006789599)."""
    if "CHECK" not in _norm_header(bsl.transaction_type):
        return False
    # Check-vs-check only: a check number conflicts with another CHECK/Payables
    # entry, not with an AR receipt that merely carries a numeric reference —
    # barring a legitimate receipt candidate here would be over-broad.
    if "CHECK" not in _norm_header(e.transaction_type) and e.source != "AP":
        return False
    bref = znorm(bsl.reference_raw) or znorm(bsl.recon_reference)
    eref = znorm(e.reference)
    if not (bref.isdigit() and len(bref) >= 4 and eref.isdigit() and len(eref) >= 4):
        return False
    return bref.lstrip("0") != eref.lstrip("0")


def _type_incongruent_uncorroborated(bsl, e):
    """Owner rule (2026-07-13): transaction type is important but not
    dispositive.  Be suspicious of an electronic (ACH/EFT) ST paired to a
    Miscellaneous bank line on amount ALONE — with no reference, payer, or
    MID tie it is almost certainly a coincidence and may not surface even as
    a Candidate.  Any tie overrides (type is not dispositive)."""
    return (_misc_bsl(bsl) and _electronic_type(e.transaction_type)
            and not _has_any_tie(bsl, e))


def _type_gate_ok(bsl, e):
    """Deterministic type gate (ORT doc section 5.2): a Credit Card ST never
    pairs with a Check or Miscellaneous bank line.  EFT is never rejected on
    the label alone.  Convera lines (owner, 2026-07-11) are international
    wires and ALWAYS Payables — they never pair with a non-Payables ST."""
    if _is_convera(bsl) and e.source != "AP":
        return False
    # Chargeback / merchant-fee debits (owner, 2026-07-12): when the bank
    # line carries a MID, ONLY an ST carrying the SAME MID may pair with it —
    # even as a Candidate.  A chargeback against MID 8028920588 has nothing
    # to do with an ST for MID 8035758468.
    if _is_card_fee_debit(bsl):
        bmids = _bsl_card_mids(bsl)
        if bmids:
            emids = (_mid_tokens(e.reference) | _mid_tokens(e.id)
                     | _mid_tokens(e.spr) | _mid_tokens(e.counterparty))
            if not (bmids & emids):
                return False
    et = _norm_header(e.transaction_type)
    if "CREDITCARD" in et:
        bt = _norm_header(bsl.transaction_type)
        if "CHECK" in bt or "MISC" in bt:
            return False
    return True


def _amount_matches(bsl, entry):
    return entry.amount_cents == bsl.amount_cents


def forward_reconcile(bsls, pool, loaded, account, runlog):
    """Run P0-P10 over the account's open BSLs.  Returns list[Placement].
    P0 (load/validate) has already run by the time we get here."""
    ledger = Ledger()
    placements = {}          # line_key -> Placement (each BSL placed once)
    pass_counts = {}

    def place(bsl, kind, confidence, entries, codes, explanation, pass_name):
        if bsl.line_key in placements:
            return  # a later pass never overrides an earlier one
        # HARD GUARDRAIL (owner doctrine, 2026-07-11): a Match or Candidate
        # without STs, or whose cited STs do not sum exactly to the BSL in
        # signed cents, must not exist.  Engine invariant — fail loud.
        if kind in (MATCH, CANDIDATE, MISDIRECTED):
            if not entries:
                raise ReconError(
                    f"engine bug: {kind} for BSL {bsl.line_key} cites no ST "
                    f"({pass_name})")
            cited = sum(e.amount_cents for e in entries)
            if cited != bsl.amount_cents:
                raise ReconError(
                    f"engine bug: {kind} for BSL {bsl.line_key} cites STs "
                    f"summing {cited} != BSL {bsl.amount_cents} ({pass_name})")
        ids = [e.id for e in entries]
        if kind == MATCH:
            ledger.consume_match(ids)
        elif kind in (CANDIDATE, MISDIRECTED):
            ledger.consume_candidate(ids)
        p = Placement(
            bsl=bsl, kind=kind, confidence=confidence, st_entries=list(entries),
            codes=list(codes), explanation=explanation, pass_name=pass_name,
            deposit_ids=[e.deposit_id for e in entries if e.deposit_id],
            receipt_ids=[e.receipt_id for e in entries if e.receipt_id],
        )
        placements[bsl.line_key] = p
        pass_counts[pass_name] = pass_counts.get(pass_name, 0) + 1

    def unplaced():
        return [b for b in bsls if b.line_key not in placements]

    bsl_amount_counts = {}
    for b in bsls:
        bsl_amount_counts[b.amount_cents] = bsl_amount_counts.get(b.amount_cents, 0) + 1
    # Whole-pool amount census (hard guardrail, 2026-07-19): the distinctive-
    # amount lanes must prove the amount is rare across the ENTIRE pool —
    # a closed, foreign-shadow, or out-of-band twin at the same signed cents
    # means the digit combination is NOT unique and amount-only evidence
    # stays insufficient (rule 4).
    pool_amount_counts = {}
    for e in pool:
        pool_amount_counts[e.amount_cents] = pool_amount_counts.get(e.amount_cents, 0) + 1
    # Cross-reference tie relation, computed ONCE (perf refactor 2026-07-19):
    # provably equal to the per-BSL full-pool scans it replaces (see
    # _build_tie_index).  Indexed over the FULL pool — P4's reference-group
    # lane needs closed/shadow members for its auto-rec-split evidence.
    tie_index = _build_tie_index(bsls, pool)
    # Static-field buckets (perf, 2026-07-19): order-preserving partitions of
    # the pool keyed ONLY on fields immutable after build_pool.  Availability
    # is deliberately NOT indexed — every consumer still filters through
    # ledger.is_available at access (doctrine rule 6).
    by_amount = {}
    ar_pool = []
    mid_pool = {}
    znref_pool = {}
    pool_pos = {}
    for i, e in enumerate(pool):
        pool_pos[id(e)] = i
        by_amount.setdefault(e.amount_cents, []).append(e)
        if e.source == "AR":
            ar_pool.append(e)
        if e.is_mid:
            mid_pool.setdefault(e.znref, []).append(e)
        for z in {z for z in (e.znref, znorm(e.id)) if len(z) >= 6}:
            znref_pool.setdefault(z, []).append(e)

    # ---- P2 DATA_FEED_ERROR sweep -----------------------------------
    feed_errors = _p2_data_feed_errors(loaded, account)
    runlog.setdefault("p2_data_feed_errors", len(feed_errors))

    # ---- P3 Exact reference 1:1 (Amount + Reference) ----------------
    stale_1to1 = {}          # line_key -> out-of-band ties, placed after P4
    for bsl in unplaced():
        cands = []
        for e in tie_index.get(bsl.line_key, ()):
            if not ledger.is_available(e):
                continue
            if not _amount_matches(bsl, e):
                continue
            # Guardrails: journal never matches; MID-into-non-merchant
            # conflict; deterministic type gate (Credit Card ST never pairs
            # with a Check/Miscellaneous line).
            if e.source in JOURNAL_SOURCES:
                continue
            if e.is_mid and bsl.lane != LANE_MERCHANT:
                continue
            if not _type_gate_ok(bsl, e):
                continue
            cands.append(e)
        ties = cands
        cands = [e for e in ties
                 if date_ok_directional(signed_lag(bsl.date, e.date))]
        if len(cands) == 1:
            e = cands[0]
            lag = signed_lag(bsl.date, e.date)
            conf = CONF_HIGH if abs(lag) <= DATE_CEILING else CONF_MEDIUM
            note = "" if abs(lag) <= DATE_CEILING else \
                f" (receipt-entry lag {-lag}d — BSL precedes ST; no ceiling)"
            place(bsl, MATCH, conf, [e],
                  [], f"Exact amount + reference tie (lag {lag}d, band {date_band(lag)})"
                  f"{note}; source {e.source}.",
                  "P3_exact_1to1")
        elif len(cands) > 1:
            ordered = _sorted(cands)
            place(bsl, CANDIDATE, CONF_MEDIUM, [ordered[0]],
                  ["MULTIPLE_EQUAL_CANDIDATES"],
                  f"{len(cands)} open entries match amount+reference; citing "
                  f"{ordered[0].id}, alternates: "
                  + ", ".join(e.id for e in ordered[1:]) + ".",
                  "P3_exact_1to1")
        elif ties:
            # Directional doctrine: the only ties here TRAIL the ST beyond
            # the band (stale-ST) — Candidate, never Match.  DEFERRED until
            # after P4: a stale 1:1 coincidence must not pre-empt a live ORT
            # deposit-chain group for the same BSL (real FHB Master case:
            # a 77d-stale reference tie shadowed deposit d:128369's exact
            # 4-receipt sum).
            stale_1to1[bsl.line_key] = ties

    # ---- P4 1:M ECT / ORT reference group (the workhorse) -----------
    for bsl in unplaced():
        group, closed_members, competing = _p4_reference_group(
            bsl, tie_index.get(bsl.line_key, ()), ledger)
        if group is None:
            continue
        total = sum(e.amount_cents for e in group)
        plausible = all(date_ok_directional(signed_lag(bsl.date, e.date))
                        for e in group)
        if total == bsl.amount_cents and not plausible and not closed_members \
                and not competing:
            if any(_ext_stale_barred(bsl, e) for e in group):
                continue  # 12+ days stale External member: not even a Candidate
            place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(group),
                  ["DATE_CONFLICT"],
                  f"1:M reference group of {len(group)} receipt(s) sums to BSL "
                  "but the BSL trails a member ST beyond the band (stale-ST).",
                  "P4_ref_group")
            continue
        if total == bsl.amount_cents and plausible:
            if closed_members:
                if any(_ext_stale_barred(bsl, e) for e in group):
                    place(bsl, REVIEW, CONF_LOW, [],
                          ["POSSIBLE_AUTO_REC_SPLIT"],
                          "Reference group sums to BSL only with member(s) "
                          "12+ days stale (External): "
                          + ", ".join(sorted(e.id for e in group))
                          + " — not citable as a Candidate; run Unreconcile2.",
                          "P4_ref_group")
                    continue
                place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(group),
                      ["POSSIBLE_AUTO_REC_SPLIT"],
                      "Reference group sums exactly but includes already-closed member(s): "
                      + ", ".join(sorted(m.id for m in closed_members))
                      + " (auto-rec-stranded; run Unreconcile2 to re-audit the group).",
                      "P4_ref_group")
            elif competing:
                place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(group),
                      ["AMBIGUOUS_GROUP"],
                      "Reference group sums exactly but a competing equal-sum group exists.",
                      "P4_ref_group")
            else:
                place(bsl, MATCH, CONF_HIGH, _sorted(group),
                      [], f"1:M reference group of {len(group)} deduped receipt(s) sums to BSL; reference tie holds.",
                      "P4_ref_group")

    # ---- P4 phase 1b: Receivables counterparty group -----------------
    # Reverse-engineered from the reference reconciliation (2026-07-11):
    # combined federal remittances land as ONE bank line covering several
    # Receivables receipts that share a counterparty and each individually
    # cross-tie the BSL (e.g. two 'SPN070326 ACH HRSA' receipts under one
    # HHS TREAS deposit).  The whole tie-set can never sum — a truncated
    # shared reference prefix fans the 4x4 screen out to thousands of
    # entries — so the sum is taken per (AR, counterparty) group.
    for bsl in unplaced():
        groups = {}
        for e in tie_index.get(bsl.line_key, ()):
            if e.source != "AR" or not e.counterparty:
                continue
            if not ledger.is_available(e):
                continue
            if not _type_gate_ok(bsl, e):
                continue
            groups.setdefault(znorm(e.counterparty), []).append(e)
        exact = [(cp, g) for cp, g in groups.items()
                 if len(g) >= 2 and sum(x.amount_cents for x in g) == bsl.amount_cents]
        if len(exact) == 1:
            cp, group = exact[0]
            place(bsl, MATCH, CONF_HIGH, _sorted(group), [],
                  f"Receivables counterparty group '{group[0].counterparty}' — "
                  f"{len(group)} receipt(s) sum exactly to BSL; every member "
                  "cross-reference-tied.",
                  "P4_cp_group")
        elif len(exact) > 1:
            ordered = sorted(exact)
            cp, group = ordered[0]
            place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(group),
                  ["MULTIPLE_EQUAL_CANDIDATES"],
                  f"{len(exact)} equal-sum tied counterparty groups; citing "
                  f"'{group[0].counterparty}'.",
                  "P4_cp_group")

    # ---- P4 phase 2: ORT deposit group (d: chain) --------------------
    # Primary ORT path (relationship doc §3): index MET rows by DEPOSIT_ID;
    # the deposit's deduped open receipts sum to one BSL.  An exact sum is
    # necessary, never sufficient — corroborate by (a) reference tie,
    # or (b) payer tie (deposit-type consistency confers NOTHING —
    # owner, 2026-07-11).  All-status context: closed members completing
    # the sum mean an auto-rec split -> Candidate naming them.
    deposit_index = {}
    for e in pool:
        if e.deposit_id and e.source not in JOURNAL_SOURCES:
            deposit_index.setdefault(e.deposit_id, []).append(e)
    # Dedup each deposit by receipt_id (dual-fire) and by id, deterministically.
    for dep, members in deposit_index.items():
        seen_id, seen_rid, uniq = set(), set(), []
        for e in _sorted(members):
            if e.id in seen_id:
                continue
            if e.receipt_id:
                if e.receipt_id in seen_rid:
                    continue
                seen_rid.add(e.receipt_id)
            seen_id.add(e.id)
            uniq.append(e)
        deposit_index[dep] = uniq

    def _deposit_corroboration(bsl, members):
        for e in members:
            if reference_equal(bsl.recon_reference, e.reference) or                reference_equal(bsl.recon_reference, e.id) or                reference_tie(bsl.recon_reference, e.reference):
                return "reference tie"
        # Merchant-lane (MID) bank lines corroborate ONLY through a reference/
        # MID tie (owner, 2026-07-11): a card deposit that doesn't carry the
        # line's MID belongs to some other department — payer text and
        # deposit-type consistency prove nothing in the merchant lane.
        if bsl.lane == LANE_MERCHANT:
            return ""
        for e in members:
            if e.payer_tokens & bsl.payer_tokens:
                return "payer tie"
        # Deposit-type consistency confers NOTHING (owner, 2026-07-11
        # false-match review): "if that's the only criteria, there is no
        # candidate."  A deposit-typed line with only an exact sum falls to
        # the amount-only path like any other line.
        return ""

    for bsl in unplaced():
        exact_open, split_groups = [], []
        for dep, members in deposit_index.items():
            gated = [e for e in members if _type_gate_ok(bsl, e)]
            if not gated:
                continue
            open_m = [e for e in gated if ledger.is_available(e)]
            if not open_m:
                continue
            if sum(e.amount_cents for e in open_m) == bsl.amount_cents:
                exact_open.append((dep, open_m))
            elif sum(e.amount_cents for e in gated) == bsl.amount_cents:
                closed = [e for e in gated if not e.available]
                if closed:
                    split_groups.append((dep, gated, closed))
        correction = _is_deposit_correction(bsl)
        # Directional date gate: every member may precede the BSL by at
        # most 8 days OR be entered after it by any amount (entry lag has
        # no ceiling); a deposit the BSL trails beyond the band is stale.
        dated = [(d, g) for d, g in exact_open
                 if all(date_ok_directional(signed_lag(bsl.date, e.date))
                        for e in g)]
        if not dated and exact_open:
            # 12+ days stale External members bar the whole group from
            # candidacy (owner rule, 2026-07-12).
            exact_open = [(d, g) for d, g in exact_open
                          if not any(_ext_stale_barred(bsl, e) for e in g)]
        if not dated and exact_open:
            corro = [(d, g) for d, g in exact_open if _deposit_corroboration(bsl, g)]
            if len(corro) == 1:
                dep, group = corro[0]
                why = _deposit_corroboration(bsl, group)
                place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(group),
                      ["DATE_CONFLICT"],
                      f"ORT deposit d:{dep} sums to BSL with {why}, but the "
                      "BSL trails the deposit beyond the band (stale-ST).",
                      "P4_deposit_group")
                continue
            if corro:
                dep, group = sorted(corro)[0]
                why = _deposit_corroboration(bsl, group)
                place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(group),
                      ["DATE_CONFLICT", "MULTIPLE_EQUAL_CANDIDATES"],
                      f"{len(corro)} corroborated out-of-band deposits sum to "
                      f"BSL; citing d:{dep} ({why}).",
                      "P4_deposit_group")
                continue
            # Owner call (2026-07-11): stale amount-only deposit sums surface
            # as LOW-confidence Candidates (double-flagged) rather than
            # vanishing into Review — the reviewer decides.  A payer
            # contradiction bars even the Candidate.
            uncontra = [(d, g) for d, g in exact_open
                        if not payer_contradiction(bsl, g)
                        and not all(_type_incongruent_uncorroborated(bsl, e) for e in g)]
            if not uncontra:
                continue
            dep, group = sorted(uncontra)[0]
            stale_codes = ["AMOUNT_ONLY_GROUP", "DATE_CONFLICT"]
            if correction:
                stale_codes.append("MANUAL_ECT")
            place(bsl, CANDIDATE, CONF_LOW, _sorted(group),
                  stale_codes,
                  f"ORT deposit d:{dep} sums to BSL but carries no reference/"
                  "payer corroboration and no member date in band"
                  + (f"; {len(uncontra)} equal-sum deposit(s)."
                     if len(uncontra) > 1 else "."),
                  "P4_deposit_group")
            continue
        exact_open = dated
        # Owner doctrine (2026-07-17, FHB UTIA $40 merchant line): reference
        # ties outrank amount-only evidence.  When a split group's members
        # carry the BSL's reference (the deposit chain's own id — for a
        # merchant line, the MID, "the critical matching string") and NO
        # open exact-sum deposit carries any corroboration of its own, the
        # ref-tied auto-rec split is the true chain.  Previously the
        # uncorroborated equal-sum coincidence entered the amount-only
        # branch and its payer-contradiction bar `continue`d PAST the
        # split-group branch, stranding the line in Review while the
        # reference-tied deposit went unmentioned.
        ref_split = [(d, g, c) for d, g, c in split_groups
                     if any(reference_equal(bsl.recon_reference, e.reference)
                            or reference_equal(bsl.recon_reference, e.id)
                            or reference_tie(bsl.recon_reference, e.reference)
                            for e in g)]
        if ref_split and not any(_deposit_corroboration(bsl, g)
                                 for _, g in exact_open):
            exact_open = []
            split_groups = ref_split
        if len(exact_open) == 1:
            dep, group = exact_open[0]
            why = _deposit_corroboration(bsl, group)
            if why and correction:
                # Owner (2026-07-11): corrections are manual fixes — surface
                # the exact-sum group as a Candidate flagged for a manual ECT,
                # never a Match.
                place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(group),
                      ["MANUAL_ECT"],
                      f"Deposit correction — manual ECT required. ORT deposit "
                      f"d:{dep} sums to BSL ({why}), cited for reference only.",
                      "P4_deposit_group")
            elif why:
                place(bsl, MATCH, CONF_HIGH, _sorted(group), [],
                      f"ORT deposit d:{dep} — {len(group)} open receipt(s) sum to BSL; {why}.",
                      "P4_deposit_group")
            else:
                if payer_contradiction(bsl, group):
                    continue
                if all(_type_incongruent_uncorroborated(bsl, e) for e in group):
                    continue  # Misc BSL vs all-electronic group, amount-only
                # Type-incongruence (8e) is checked BEFORE the distinctive
                # lane (adversarial review, 2026-07-19): doctrine names BOTH
                # amount-only lanes (P9b + P4 deposit groups) — a distinctive
                # amount is still amount-only evidence and never overrides an
                # electronic-vs-Miscellaneous incongruence with zero ties.
                if amount_distinctive(bsl.amount_cents) and \
                        bsl_amount_counts.get(bsl.amount_cents) == 1 and \
                        _pool_amount_twin_free(bsl.amount_cents,
                                               pool_amount_counts, group) and \
                        not correction:
                    place(bsl, MATCH, CONF_MEDIUM, _sorted(group),
                          ["DISTINCTIVE_AMOUNT"],
                          f"Distinctive-amount ORT deposit group: d:{dep} is "
                          "the ONLY deposit summing to this non-round amount, "
                          "and no other bank line carries it — a rare digit "
                          "combination is valid match evidence (owner "
                          "doctrine).",
                          "P4_deposit_group")
                    continue
                codes = ["AMOUNT_ONLY_GROUP"]
                note = ""
                if correction:
                    codes.append("MANUAL_ECT")
                    note = " Deposit correction — manual ECT required; cited for reference only."
                place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(group),
                      codes,
                      f"ORT deposit d:{dep} sums to BSL but carries no reference/payer corroboration." + note,
                      "P4_deposit_group")
        elif len(exact_open) > 1:
            corroborated = [(d, g) for d, g in exact_open if _deposit_corroboration(bsl, g)]
            if len(corroborated) == 1:
                dep, group = corroborated[0]
                why = _deposit_corroboration(bsl, group)
                if correction:
                    place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(group),
                          ["MANUAL_ECT"],
                          f"Deposit correction — manual ECT required. ORT "
                          f"deposit d:{dep} sums to BSL ({why}), cited for "
                          "reference only.",
                          "P4_deposit_group")
                else:
                    place(bsl, MATCH, CONF_HIGH, _sorted(group), [],
                          f"ORT deposit d:{dep} uniquely corroborated among {len(exact_open)} equal-sum deposits.",
                          "P4_deposit_group")
            else:
                uncontra = [(d, g) for d, g in exact_open
                            if not payer_contradiction(bsl, g)
                            and not all(_type_incongruent_uncorroborated(bsl, e) for e in g)]
                if not uncontra:
                    continue
                dep, group = sorted(uncontra)[0]
                multi_codes = ["MULTIPLE_EQUAL_CANDIDATES"]
                if correction:
                    multi_codes.append("MANUAL_ECT")
                place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(group),
                      multi_codes,
                      f"{len(uncontra)} deposits sum to BSL: "
                      + ", ".join(f"d:{d}" for d, _ in sorted(uncontra)) + ".",
                      "P4_deposit_group")
        elif split_groups:
            dep, group, closed = sorted(split_groups)[0]
            if any(_ext_stale_barred(bsl, e) for e in group):
                place(bsl, REVIEW, CONF_LOW, [],
                      ["POSSIBLE_AUTO_REC_SPLIT"],
                      f"ORT deposit d:{dep} sums to BSL only with member(s) "
                      "12+ days stale (External), closed: "
                      + ", ".join(sorted(m.id for m in closed))
                      + " — not citable as a Candidate; run Unreconcile2.",
                      "P4_deposit_group")
            else:
                place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(group),
                      ["POSSIBLE_AUTO_REC_SPLIT"],
                      f"ORT deposit d:{dep} sums to BSL only with already-closed member(s): "
                      + ", ".join(sorted(m.id for m in closed))
                      + " (auto-rec split; run Unreconcile2).",
                      "P4_deposit_group")

    runlog["p5_state"] = "retired (owner doctrine 2026-07-11: Edison pass eliminated)"

    # ---- P6 Merchant / MID ------------------------------------------
    _p6_merchant(unplaced, place, mid_pool, ledger, loaded, runlog, account)

    # ---- P7 Receivables SPN group -----------------------------------
    _p7_spn(unplaced, place, pool, ledger, loaded, runlog)

    # ---- P8 Named-payer rules (Amount + Payer) ----------------------
    _p8_named_payer(unplaced, place, by_amount, ledger, runlog)

    # ---- P8c Payer-family receipt-sum group (owner 2026-07-14) -------
    # A managed-care remittance lands as ONE positive ACH covering several
    # Receivables receipts from the same payer FAMILY, deposited the SAME day,
    # whose per-receipt references do not tie the bank line (blank / parked
    # SPN).  An exact same-day same-family receipt sum, corroborated by a
    # payer-FAMILY tie (VSHP / TennCare == BlueCare Tennessee), clears the
    # amount-alone bar.  It is a CANDIDATE, never a Match: the addenda's
    # structured references may cite other receipts (reference ambiguity ->
    # route to Candidate).  Whole-group sum by (family, same day) only; never
    # a blind subset-sum.
    for bsl in unplaced():
        if bsl.amount_cents <= 0:
            continue
        fams = _bsl_payer_families(bsl)
        if not fams:
            continue
        groups = {}
        for e in ar_pool:
            if not ledger.is_available(e):
                continue
            if not _type_gate_ok(bsl, e):
                continue
            if signed_lag(bsl.date, e.date) != 0:      # same deposit day only
                continue
            shared = fams & payer_family(e.counterparty)
            if not shared:
                continue
            groups.setdefault(frozenset(shared), []).append(e)
        for fam, g in sorted(groups.items(), key=lambda kv: sorted(kv[0])):
            seen, uniq = set(), []
            for e in _sorted(g):
                if e.id in seen:
                    continue
                seen.add(e.id)
                uniq.append(e)
            if uniq and sum(e.amount_cents for e in uniq) == bsl.amount_cents:
                place(bsl, CANDIDATE, CONF_MEDIUM, uniq,
                      ["PAYER_FAMILY_GROUP"],
                      f"Payer-family receipt group ({'/'.join(sorted(fam))}): "
                      f"{len(uniq)} same-day Receivables receipt(s) from "
                      f"{uniq[0].counterparty!r} sum exactly to the BSL. Payer-family "
                      "tie corroborates (VSHP / TennCare == BlueCare Tennessee). "
                      "Candidate, not Match — the addenda's structured payment "
                      "references may cite other receipts of the same payer.",
                      "P8c_payer_family")
                break
    runlog["p8c_payer_family"] = "ran"

    # ---- P8b Deferred stale 1:1 (out-of-band amount+reference ties) --
    # Held back from P3 so a stale coincidence never shadows a live ORT
    # chain / merchant / SPN group; whatever survives the group passes is
    # placed here with the same wording P3 used.
    for bsl in unplaced():
        ties = stale_1to1.get(bsl.line_key)
        if not ties:
            continue
        live = [e for e in ties
                if ledger.is_available(e) and not _ext_stale_barred(bsl, e)]
        if len(live) == 1:
            e = live[0]
            lag = signed_lag(bsl.date, e.date)
            place(bsl, CANDIDATE, CONF_MEDIUM, [e],
                  ["DATE_CONFLICT"],
                  f"Exact amount + reference tie to {e.id}, but the BSL "
                  f"trails the ST by {lag}d — stale-ST pairing, never a Match.",
                  "P8b_stale_1to1")
        elif live:
            ordered = _sorted(live)
            place(bsl, CANDIDATE, CONF_MEDIUM, [ordered[0]],
                  ["MULTIPLE_EQUAL_CANDIDATES", "DATE_CONFLICT"],
                  f"{len(live)} out-of-band amount+reference ties; citing "
                  f"{ordered[0].id}, alternates: "
                  + ", ".join(e.id for e in ordered[1:]) + ".",
                  "P8b_stale_1to1")

    # ---- P9 Payables debit ------------------------------------------
    for bsl in unplaced():
        if bsl.amount_cents >= 0:
            continue
        cands = [e for e in by_amount.get(bsl.amount_cents, ())
                 if e.source == "AP" and ledger.is_available(e)]
        tied = [e for e in cands if cross_reference_tie(bsl, e)]
        if len(tied) == 1:
            e = tied[0]
            place(bsl, MATCH, CONF_HIGH, [e],
                  [], "Negative BSL matched to single reference-tied open Payables ST.",
                  "P9_payables")
        elif cands:
            # Orphan doctrine R8 (owner, 2026-07-19): on check rails the check
            # number IS the identity — a same-amount AP check with a DIFFERENT
            # check number is a conflict, never an amount-only candidate (the
            # FHB AP $1,100 cascade).  Payer contradiction bars the rest.
            cands = [e for e in cands
                     if not payer_contradiction(bsl, [e])
                     and not _check_number_conflict(bsl, e)]
            if not cands:
                continue
            ordered = _sorted(cands)
            place(bsl, CANDIDATE, CONF_LOW, [ordered[0]],
                  ["MISSING_REFERENCE"],
                  "Payables amount match without a clean reference tie; citing "
                  f"{ordered[0].id}"
                  + (", alternates: " + ", ".join(e.id for e in ordered[1:])
                     if len(ordered) > 1 else "") + ".",
                  "P9_payables")

    # ---- P9b Amount-only singles -> LOW candidates -------------------
    # Reverse-engineered from the reference reconciliation (2026-07-11): an
    # unplaced BSL with open exact-cents ST(s) surfaces as a LOW Candidate
    # naming them — the hard guardrail still bars a Match without
    # corroboration, and a payer contradiction bars even the Candidate.
    for bsl in unplaced():
        elig = []
        for e in by_amount.get(bsl.amount_cents, ()):
            if not ledger.is_available(e):
                continue
            if e.source in JOURNAL_SOURCES:
                continue
            if e.is_mid and bsl.lane != LANE_MERCHANT:
                continue
            if not _type_gate_ok(bsl, e):
                continue
            if payer_contradiction(bsl, [e]):
                continue
            if _type_incongruent_uncorroborated(bsl, e):
                continue  # Misc BSL vs electronic ST, amount-only (owner 2026-07-13)
            if _check_number_conflict(bsl, e):
                continue  # check rails: different check number = conflict (R8)
            if sibling(bsl.recon_reference, e.reference):
                continue  # sibling references are CONFLICTS (rule 7) — like
                #           R8's check rails, a conflict bars even the
                #           amount-only Candidate (hard guardrail 2026-07-19)
            elig.append(e)
        if not elig:
            continue
        inband = [e for e in elig
                  if date_ok_directional(signed_lag(bsl.date, e.date))]
        if inband:
            ordered = _sorted(inband)
            if len(inband) == 1 and amount_distinctive(bsl.amount_cents) and \
                    bsl_amount_counts.get(bsl.amount_cents) == 1 and \
                    _pool_amount_twin_free(bsl.amount_cents,
                                           pool_amount_counts, [ordered[0]]) and \
                    not _is_deposit_correction(bsl):
                place(bsl, MATCH, CONF_MEDIUM, [ordered[0]],
                      ["DISTINCTIVE_AMOUNT"],
                      f"Distinctive-amount 1:1: {ordered[0].id} is the ONLY "
                      "open ST at this non-round amount and no other bank "
                      "line carries it — a rare digit combination is valid "
                      "match evidence (owner doctrine).",
                      "P9b_amount_only")
                continue
            codes = ["AMOUNT_ONLY"]
            note = ""
            if _is_deposit_correction(bsl):
                codes.append("MANUAL_ECT")
                note = " Deposit correction — manual ECT required."
            place(bsl, CANDIDATE, CONF_LOW, [ordered[0]],
                  codes,
                  f"AMOUNT_ONLY: {len(inband)} open ST(s) at exact cents but "
                  f"zero cross-reference corroboration; citing {ordered[0].id}"
                  + (", alternates: " + ", ".join(e.id for e in ordered[1:])
                     if len(ordered) > 1 else "")
                  + ". Hard guardrail bars Match without corroboration." + note,
                  "P9b_amount_only")
            continue
        # Only stale exact-cents STs remain: surface as a DATE_GATE Candidate
        # when at least a digit-run tie corroborates; a zero-evidence stale
        # coincidence stays Review.
        evid = [e for e in elig
                if cross_digit_tie(bsl, e) and not _ext_stale_barred(bsl, e)]
        if evid:
            ordered = _sorted(evid)
            lag = signed_lag(bsl.date, ordered[0].date)
            place(bsl, CANDIDATE, CONF_LOW, [ordered[0]],
                  ["DATE_GATE"],
                  f"DATE_GATE: only exact-cents ST(s) precede the BSL by 8+ "
                  f"days (lag {lag}d); digit-run evidence only; citing "
                  f"{ordered[0].id}. Revised gate bars Match.",
                  "P9b_amount_only")
    runlog["p9b_amount_only"] = "ran"

    # ---- PM Misdirected search (owner HARD GUARDRAIL, 2026-07-18) ----
    # The bank deposit landed in THIS account but the system transaction /
    # receipt was booked to a DIFFERENT bank account (any source, any bank,
    # any campus — real case: City of Memphis $70,992.66 in FHB UTHSC while
    # receipt 300045836 remits to FHB Master).  Search the foreign shadow
    # pool with FULL corroboration only: exact signed cents AND a reference
    # tie (never amount alone).  These placements get their own workbook tab
    # — they can never auto-reconcile without a reroute/ECT.
    foreign = [e for e in pool if e.foreign_account]
    if foreign:
        for bsl in unplaced():
            hits = []
            for e in foreign:
                if e.id in ledger.barred():
                    continue  # already cited by an earlier line's placement —
                    #           one foreign entry never explains two bank lines
                if not _amount_matches(bsl, e):
                    continue
                if e.source in JOURNAL_SOURCES:
                    continue
                if not _type_gate_ok(bsl, e):
                    continue
                # Hard guardrail (adversarial review, 2026-07-19): a workbook
                # placement to a FOREIGN account requires the strong tie —
                # znorm equality/containment >= 6 chars (cross_reference_tie),
                # the same standard as P9c and _reverse_misdirected.  The
                # 5-digit-run grid tie (cross_digit_tie) can fire on a ZIP,
                # date fragment, or truncated trace in the addenda and is NOT
                # acceptable evidence for rerouting money across accounts.
                if not cross_reference_tie(bsl, e):
                    continue
                # No payer_contradiction screen here: it is consulted only by
                # zero-corroboration placements (owner doctrine — reference
                # ties outrank payer text), and every PM hit carries a
                # reference tie by construction ("STATE OF NE" paying for
                # "University of Nebraska" is the same transaction).
                hits.append(e)
            if not hits:
                continue
            ordered = _sorted(hits)
            chosen = ordered[0]
            place(bsl, MISDIRECTED, CONF_MEDIUM, [chosen],
                  ["MISDIRECTED"],
                  f"Misdirected: this bank line landed in {account} but the "
                  f"matching system entry {chosen.id} (exact cents + reference "
                  f"tie) is booked to {chosen.foreign_account}. It can never "
                  "auto-reconcile here — reroute the receipt/ST or raise an ECT."
                  + (" Alternates: " + ", ".join(
                        f"{e.id} ({e.foreign_account})" for e in ordered[1:4])
                     if len(ordered) > 1 else ""),
                  "PM_misdirected")
    runlog["pm_misdirected"] = "ran"

    # ---- P9c ORT / Receivables 1:M reference search (HARD GUARDRAIL,
    #      owner 2026-07-14) --------------------------------------------
    # The ORT chain (the bridge/chain for External STs) and the Receivables
    # ST reference columns MUST always be searched for 1:M ties.  When a bank
    # line's reference ties to a group of open non-journal, non-merchant STs
    # that does NOT sum to the BSL (exact and exact-with-closed groups were
    # already placed by P4), the remaining member(s) are already
    # auto-reconciled (stranded) and absent from the UNR export.  The
    # exact-sum invariant (owner 2026-07-11) bars a Match/Candidate here, so
    # surface an ENRICHED REVIEW that NAMES the reference-tied open members,
    # the partial sum, and the shortfall — an auto-rec split to re-audit in
    # Unreconcile2 — rather than a bare 'no counterpart' note.  Placed before
    # P10 so the residual pass only sees lines with no reference tie at all.
    for bsl in unplaced():
        # STRICT deposit-batch keys: the bank line's own reference fields
        # (NOT the broad addenda/counterparty grid, which is only safe under
        # P4's exact-sum filter), znorm-exact, >= 6 chars to avoid short
        # date-like coincidences.
        bkeys = {z for z in (znorm(bsl.reference_raw), znorm(bsl.recon_reference),
                             znorm(bsl.account_servicer_reference)) if len(z) >= 6}
        if not bkeys:
            continue
        members = []
        cand, seen = [], set()
        for k in sorted(bkeys):
            for e in znref_pool.get(k, ()):
                if id(e) not in seen:
                    seen.add(id(e))
                    cand.append(e)
        cand.sort(key=lambda e: pool_pos[id(e)])
        for e in cand:
            if not ledger.is_available(e):
                continue
            if e.source in JOURNAL_SOURCES:
                continue
            if e.is_mid and bsl.lane != LANE_MERCHANT:
                continue
            if not _type_gate_ok(bsl, e):
                continue
            if _ext_stale_barred(bsl, e):
                continue  # 12+ day stale External member: not citable (owner 8d)
            # (znref_pool membership == the original EXACT-equality key join:
            # each entry is indexed under znorm(reference) and znorm(id) >= 6.)
            members.append(e)
        # A genuine 1:M deposit batch has >= 2 open members sharing the exact
        # reference; a lone coincidental tie is left to the ordinary review.
        if len(members) < 2:
            continue
        # deterministic de-dup by id and by receipt_id (dual-fire guard)
        seen_id, seen_rid, uniq = set(), set(), []
        for e in _sorted(members):
            if e.id in seen_id:
                continue
            if e.receipt_id and e.receipt_id in seen_rid:
                continue
            seen_id.add(e.id)
            if e.receipt_id:
                seen_rid.add(e.receipt_id)
            uniq.append(e)
        if len(uniq) < 2:
            continue
        total = sum(e.amount_cents for e in uniq)
        if total == bsl.amount_cents:
            continue  # exact group — P4 owns it (defensive; should not reach here)
        # Directional plausibility: the open members of a partial deposit are a
        # SUBSET of the bank total, so they must share the BSL's sign and sum to
        # STRICTLY LESS than it in magnitude.  A group that is opposite-sign or
        # exceeds the bank line is not a partial deposit — it is a coincidental
        # collision on a shared ACH originator / company ID (e.g. a $(1,125)
        # debit vs 18 positive receipts summing $66k on a shared company id),
        # never a genuine 1:M deposit.
        if bsl.amount_cents == 0:
            continue
        same_sign = (total > 0) == (bsl.amount_cents > 0) and total != 0
        if not same_sign or abs(total) >= abs(bsl.amount_cents):
            continue
        short = bsl.amount_cents - total
        src = uniq[0].source
        ref_shown = sorted(bkeys & set().union(
            *[{znorm(e.reference), znorm(e.id)} for e in uniq]))
        place(bsl, REVIEW, CONF_LOW, uniq,
              ["PARTIAL_REFERENCE_GROUP", "POSSIBLE_AUTO_REC_SPLIT"],
              f"ORT/Receivables 1:M reference tie (ref {', '.join(ref_shown)}): "
              f"{len(uniq)} open {src} ST(s) carry the BSL reference and sum to "
              f"{_usd(total)} of {_usd(bsl.amount_cents)} (short {_usd(short)} — the "
              "remaining member(s) were already auto-reconciled and are absent from "
              "the UNR export). The shared-reference tie is dispositive that these "
              "belong to this deposit; an auto-rec split — run Unreconcile2 to re-audit "
              "and free the stranded member(s). Not citable as a Match/Candidate: the "
              "open subset does not sum to the bank line (exact-sum guardrail).",
              "P9c_ref_1m_review")
    runlog["p9c_ref_1m_review"] = "ran"


    # ---- P10 Residual -> Review -------------------------------------
    coa = loaded.get("CHART_OF_ACCOUNTS")
    # P10 reuses the run-level by_amount bucket (same construction: full
    # pool, pool order) for _p10_review_cause and the recommended-GL fallback.
    edison_idx, edison_inv = _edison_index(loaded)
    for bsl in unplaced():
        exact_entries = by_amount.get(bsl.amount_cents, [])
        codes, expl = _p10_review_cause(bsl, pool, feed_errors, deposit_index,
                                        coa, exact_entries)
        if _is_deposit_correction(bsl):
            codes = ["MANUAL_ECT"] + [c for c in codes if c != "MANUAL_ECT"]
            expl = ("Deposit correction — rarely has an ST; manual ECT "
                    "required. " + expl)
        gl = recommend_gl_string(bsl, loaded, exact_entries)
        if gl:
            expl += f" Recommended GL: {gl}."
        expl += _edison_note(bsl, edison_idx, edison_inv)
        place(bsl, REVIEW, CONF_LOW, [], codes, expl, "P10_review")

    # ---- Reverse misdirected search (owner, 2026-07-19) -------------
    # The mirror of PM: a THIS-account open ST/receipt whose bank line
    # actually landed in ANOTHER account.  ALL_BSL supplies every open bank
    # line system-wide; we scan the account's still-open, unconsumed pool
    # entries for an exact signed-cent + reference tie to a foreign open BSL.
    # Read-only: these are ST-anchored findings (no THIS-account BSL to
    # place), so they go to the run log for the reconciler — never a workbook
    # placement, and never affecting BSL conservation.
    runlog["reverse_misdirected"] = _reverse_misdirected(
        pool, ledger, loaded, account)

    runlog["forward_pass_counts"] = pass_counts
    ordered = [placements[b.line_key] for b in bsls]

    # Consumption & conservation asserts (Section 9 trailer).
    _assert_conservation(bsls, ordered)
    runlog["consumed_st_ids"] = len(ledger.consumed())
    runlog["barred_st_ids"] = len(ledger.barred())
    return ordered


def _sorted(entries):
    """Total order for determinism (amount, then date, then id)."""
    return sorted(entries, key=lambda e: (e.amount_cents,
                                          e.date.toordinal() if e.date else 0,
                                          e.id))


def _norm_id(v):
    """Normalize a numeric id cell: ints/floats from Excel/xlsb ('64017.0')
    become bare digit strings."""
    s = N(v)
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _reverse_misdirected(pool, ledger, loaded, account):
    """Mirror of PM (owner, 2026-07-19): find THIS-account open, unconsumed
    ST/receipts whose bank line actually landed in ANOTHER account, using the
    all-accounts open-BSL export (ALL_BSL).  Match on exact signed cents AND
    znorm-EXACT reference equality (>= 6 chars; containment/addenda-embedding
    deliberately NOT accepted — a runlog claim about another account's line
    keeps the strictest tie).  Read-only and ST-anchored — there is no
    THIS-account BSL to place, so findings go to the run log for the
    reconciler, never a workbook row (BSL conservation is untouched)."""
    ab = loaded.get("ALL_BSL")
    if not ab or not account or account == "UNKNOWN":
        return []
    rows, m, hi = ab["rows"], ab["map"], ab["header_index"]
    by_amt = {}
    for r in rows[hi + 1:]:
        acc = account_of_bank_name(_cell(r, m.get("bank_account_name")))
        if acc is None or acc == account:
            continue  # only OTHER configured accounts are "foreign"
        amt = cents(_cell(r, m.get("amount")))
        if amt is None:
            continue
        fref = clean_ref(_cell(r, m.get("reference")))
        by_amt.setdefault(amt, []).append((acc, fref,
                                           parse_date(_cell(r, m.get("date")))))
    if not by_amt:
        return []
    findings, seen = [], set()
    for e in pool:
        if not ledger.is_available(e) or e.foreign_account or e.source in JOURNAL_SOURCES:
            continue
        # Require a UNIQUE per-transaction reference: a MID / merchant id is a
        # shared originator carried by many transactions, so exact amount + MID
        # is a coincidence, not proof the same money landed elsewhere.
        if not (e.znref and len(e.znref) >= 6) or e.is_mid or is_mid(e.reference):
            continue
        for acc, fref, fdate in by_amt.get(e.amount_cents, ()):
            if e.znref == znorm(fref):
                key = (e.id, acc)
                if key in seen:
                    continue
                seen.add(key)
                findings.append({
                    "st_id": e.id, "source": e.source,
                    "amount": e.amount_cents / 100, "landed_in": acc,
                    "foreign_bsl_reference": fref,
                    "date": fdate.isoformat() if fdate else None,
                })
                break
    findings.sort(key=lambda f: (f["landed_in"], f["amount"], f["st_id"]))
    return findings


def _p2_data_feed_errors(loaded, account=None):
    """Open ORT parked receipt ids absent from MET => DATA_FEED_ERROR.
    Prefers the ORT export's bound Parked Receipt ID column (scoped to the
    account's bank name); falls back to the r:-token text scan."""
    met = loaded.get("MET")
    if not met:
        return []
    met_receipt_ids = set()
    rows, m, hi = met["rows"], met["map"], met["header_index"]
    for r in rows[hi + 1:]:
        rid = _norm_id(_cell(r, m.get("receipt_id")))
        if rid:
            met_receipt_ids.add(rid)
        desc = N(_cell(r, m.get("description")))
        _, rec_id, _ = parse_met_description(desc)
        if rec_id:
            met_receipt_ids.add(rec_id)
    errors = []
    for role in ("ORT_AR", "ORT_MISC"):
        src = loaded.get(role)
        if not src:
            continue
        rows, m, hi = src["rows"], src["map"], src["header_index"]
        prid = m.get("parked_receipt_id")
        bank = m.get("bank_name")
        if prid is not None:
            for r in rows[hi + 1:]:
                if bank is not None and account and account != "UNKNOWN" and \
                        account_of_bank_name(_cell(r, bank)) != account:
                    continue
                rid = _norm_id(_cell(r, prid))
                if rid and rid not in met_receipt_ids:
                    errors.append(rid)
            continue
        for r in rows[hi + 1:]:
            # fallback heuristic: any r:<id> token in a description-like cell
            joined = " ".join(N(c) for c in r)
            for mm in re.finditer(r"r:\s*(\d+)", joined, re.IGNORECASE):
                rid = mm.group(1)
                if rid not in met_receipt_ids:
                    errors.append(rid)
    return sorted(set(errors))


def _p4_reference_group(bsl, tied_entries, ledger):
    """Assemble the all-status deduped reference group for a BSL via the
    owner-mandated 4x4 cross-reference screen (`tied_entries` = the
    precomputed cross-tie slice for this BSL, pool order, ALL statuses).
    Returns (group_entries or None, closed_members, competing_bool)."""
    members = []
    for e in tied_entries:
        if e.source in JOURNAL_SOURCES:
            continue
        if not _type_gate_ok(bsl, e):
            continue
        members.append(e)
    if not members:
        return None, [], False
    # Dedup by id keeping largest (already deduped in pool, but ORT+RECEIPTS
    # may overlap on reference; keep unique ids).
    seen, uniq = set(), []
    for e in members:
        if e.id in seen:
            continue
        seen.add(e.id)
        uniq.append(e)
    # Dual-fire guard (ORT doc section 6.3): one ORT receipt id spawning two
    # identical STs — dedup by receipt_id before summing, deterministic keep.
    seen_rid, deduped = set(), []
    for e in _sorted(uniq):
        if e.receipt_id:
            if e.receipt_id in seen_rid:
                continue
            seen_rid.add(e.receipt_id)
        deduped.append(e)
    uniq = deduped
    open_members = [e for e in uniq if ledger.is_available(e)]
    closed_members = [e for e in uniq if not e.available]
    # A group needs at least one open member to be actionable.
    if not open_members:
        return None, [], False
    total_open = sum(e.amount_cents for e in open_members)
    if total_open == bsl.amount_cents:
        # Competing equal-sum detection (hard guardrail, 2026-07-19 — this
        # branch was specified but never implemented): when a PROPER
        # sub-cluster of the tied set (clustered by deposit id, else payer,
        # else the entry itself) ALSO sums exactly to the BSL, the remainder
        # nets to zero — reversal-pair noise inside the group — and which
        # assembly is the true counterpart is ambiguous.  Ambiguity is a
        # Candidate, never a Match (rule 4: evidence must pick ONE story).
        clusters = {}
        for e in open_members:
            key = ("d", e.deposit_id) if e.deposit_id else (
                ("c", znorm(e.counterparty)) if N(e.counterparty) else ("i", e.id))
            clusters.setdefault(key, []).append(e)
        competing = len(clusters) > 1 and any(
            sum(m.amount_cents for m in ms) == bsl.amount_cents
            for ms in clusters.values())
        return open_members, [], competing
    # Try open + closed sum (auto-rec-stranded case).
    total_all = sum(e.amount_cents for e in uniq)
    if total_all == bsl.amount_cents and closed_members:
        return uniq, closed_members, False
    return None, [], False








def _p6_merchant(unplaced, place, mid_pool, ledger, loaded, runlog, account=None):
    """Merchant / MID (Section 9 P6)."""
    mid_dir = _mid_directory(loaded, account)

    for bsl in unplaced():
        if bsl.lane != LANE_MERCHANT:
            continue
        mid = bsl.mid or (znorm(bsl.recon_reference) if is_mid(bsl.recon_reference) else "")
        if not mid:
            # merchant text but no MID: defer to reference/chain lane, not straight to Review.
            continue
        group = [e for e in mid_pool.get(mid, ())
                 if ledger.is_available(e)]
        # card window: ST precedes BSL by 1-4 days.
        in_window = [e for e in group if date_ok_merchant(signed_lag(bsl.date, e.date))]
        stale = [e for e in group if (signed_lag(bsl.date, e.date) or 0) > 30]
        total = sum(e.amount_cents for e in in_window)
        if in_window and total == bsl.amount_cents:
            place(bsl, MATCH, CONF_HIGH, _sorted(in_window),
                  [], f"Same-MID card group ({len(in_window)}) in 1-4d window sums to settlement.",
                  "P6_merchant")
        elif stale and not in_window:
            place(bsl, REVIEW, CONF_LOW, [], ["MID_GUARDRAIL"],
                  "Stale (>30d) same-MID group; likely auto-rec artifact."
                  + _mid_note(bsl, mid_dir, account), "P6_merchant")
        elif group:
            # 12+ days stale External members may not be cited in a
            # Candidate (owner rule, 2026-07-12).
            eligible = [e for e in group if not _ext_stale_barred(bsl, e)]
            whole = sum(e.amount_cents for e in eligible)
            singles = [e for e in eligible if e.amount_cents == bsl.amount_cents]
            if eligible and whole == bsl.amount_cents:
                place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(eligible),
                      ["GROUPING_CONFLICT"],
                      "Same-MID receipts sum to settlement but not inside the 1-4d window.",
                      "P6_merchant")
            elif singles:
                best = _sorted(singles)[0]
                place(bsl, CANDIDATE, CONF_MEDIUM, [best],
                      ["GROUPING_CONFLICT"],
                      f"Same-MID receipt {best.id} equals the settlement out of window; "
                      f"{len(group)} same-MID receipt(s) present.", "P6_merchant")
            else:
                place(bsl, REVIEW, CONF_LOW, [], ["GROUPING_CONFLICT"],
                      f"Same-MID receipts present ({len(group)}) but no subset sums "
                      "to the settlement.", "P6_merchant")
        # else: no card receipt -> defer to P7/P10.
    runlog["p6_merchant"] = "ran"


def _p7_spn(unplaced, place, pool, ledger, loaded, runlog):
    """Receivables SPN group (Section 9 P7).  Hard-guardrailed (adversarial
    review, 2026-07-19): P7 runs BEFORE the screened amount-only lanes, so it
    applies the same screens they do — type gate, stale ceiling, directional
    date plausibility — and its zero-corroboration singleton path is GONE
    (an uncorroborated single receipt falls through to P9b, which screens
    payer contradiction and type incongruence before any Candidate)."""
    # SPN partition is BSL-independent: build once (perf, 2026-07-19); the
    # per-BSL screens (type gate, stale ceiling) and availability are applied
    # inside the loop, so behavior is unchanged.
    spn_groups = {}
    for e in pool:
        if e.spn:
            spn_groups.setdefault(_spn_root(e.spn), []).append(e)
    for bsl in unplaced():
        if bsl.lane == LANE_MERCHANT:
            continue
        by_spn = {}
        for root, entries in spn_groups.items():
            live = [e for e in entries
                    if ledger.is_available(e)
                    and _type_gate_ok(bsl, e)
                    and not _ext_stale_barred(bsl, e)]
            if live:
                by_spn[root] = live
        placed = False
        for root, members in sorted(by_spn.items()):
            # corroborator: shared SPN root AND a reference tie to the BSL —
            # bare membership never suffices (rule 4).
            corroborated = reference_tie(bsl.recon_reference, root) or \
                any(reference_tie(bsl.recon_reference, e.reference) or
                    reference_equal(bsl.recon_reference, e.reference) for e in members)
            if not corroborated:
                continue
            total = sum(e.amount_cents for e in members)
            if total != bsl.amount_cents:
                continue
            plausible = all(date_ok_directional(signed_lag(bsl.date, e.date))
                            for e in members)
            if not plausible:
                # Same shape as P4's stale-group handling: corroborated but
                # out of band is a flagged Candidate, never a Match.
                place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(members),
                      ["DATE_CONFLICT"],
                      f"SPN group {root} of {len(members)} receipt(s) sums to "
                      "BSL with a reference tie, but a member date is out of "
                      "band.",
                      "P7_spn")
                placed = True
                break
            place(bsl, MATCH, CONF_HIGH, _sorted(members), [],
                  f"SPN group {root} of {len(members)} receipt(s) sums to BSL"
                  "; corroborated.",
                  "P7_spn")
            placed = True
            break
        if placed:
            continue
    runlog["p7_spn"] = "ran"


def _spn_root(spn):
    """Strip only a trailing -N sequence, never a space-separated reference."""
    return re.sub(r"-\d+$", "", N(spn))


# Named-payer rules (Section 9 P8).  keyword -> counterparty token set.
NAMED_PAYER_RULES = [
    ("TVA", {"TVA", "TENNESSEEVALLEYAUTHORITY"}),
    ("ASAP", {"ENERGY", "ASAP", "DEPTOFENERGY"}),
    ("NSF", {"NSF", "NATIONALSCIENCEFOUNDATION"}),
    ("UT-BATTELLE", {"BATTELLE", "UTBATTELLE"}),
    ("JEFFERSON SCIENCE", {"JEFFERSON"}),
    ("PRINCETON PLASMA", {"PRINCETON"}),
    ("FISH", {"FISH", "WILDLIFE"}),
    ("AMERICAN HEART", {"AMERICANHEART"}),
    ("STATE OF TENNESSEE", {"STATEOFTENNESSEE"}),
]


def _p8_named_payer(unplaced, place, by_amount, ledger, runlog):
    for bsl in unplaced():
        info_z = znorm(bsl.additional_info) + znorm(bsl.recon_reference)
        for keyword, cp_tokens in NAMED_PAYER_RULES:
            if znorm(keyword) not in info_z:
                continue
            cands = []
            for e in by_amount.get(bsl.amount_cents, ()):
                if not ledger.is_available(e):
                    continue
                if not _type_gate_ok(bsl, e):
                    continue
                lag = signed_lag(bsl.date, e.date)
                if not date_ok_directional(lag):
                    continue
                cp_z = znorm(e.counterparty)
                if any(tok in cp_z for tok in cp_tokens):
                    cands.append(e)
            if len(cands) == 1:
                place(bsl, MATCH, CONF_HIGH, cands, [],
                      f"Named-payer rule '{keyword}': amount + counterparty tie.",
                      "P8_named_payer")
                break
            elif len(cands) > 1:
                place(bsl, CANDIDATE, CONF_MEDIUM, _sorted(cands),
                      ["MULTIPLE_EQUAL_CANDIDATES"],
                      f"Named-payer rule '{keyword}': multiple counterparty matches.",
                      "P8_named_payer")
                break
    runlog["p8_named_payer"] = "ran"


def _mid_directory(loaded, account):
    """MID -> {department, campus, home_account} from DEPT_INFO (+ any MIDs
    the MID master names).  Used to annotate merchant lines; never a gate."""
    out = {}
    di = loaded.get("DEPT_INFO")
    if di:
        rows, m, hi = di["rows"], di["map"], di["header_index"]
        for r in rows[hi + 1:]:
            mid = znorm(clean_ref(_cell(r, m.get("mid"))))
            if not mid or not is_mid(mid):
                continue
            home = account_of_bank_name(_cell(r, m.get("dept_bank_name"))) or \
                   account_of_bank_name(_cell(r, m.get("campus_bank_name")))
            out.setdefault(mid, {
                "department": N(_cell(r, m.get("department"))),
                "campus": N(_cell(r, m.get("campus"))),
                "home_account": home,
            })
    return out


def _mid_note(bsl, mid_dir, account):
    info = mid_dir.get(bsl.mid or znorm(bsl.recon_reference) or "")
    if not info:
        return ""
    note = f" MID belongs to {info['department']} ({info['campus']})."
    if info["home_account"] and account and account != "UNKNOWN" and \
            info["home_account"] != account:
        note += f" ADVISORY: MID's home account is {info['home_account']}, not {account}."
    return note


# ---- Edison (State of TN) annotation (owner, 2026-07-19) --------------

def _edison_index(loaded):
    """{signed cents -> [(payment reference, invoice, iso date)]} from the
    Edison payments export, plus {invoice -> (status, voucher)} from the
    invoice export.  Returns (None, {}) when EDISON_PAY is absent."""
    ep = loaded.get("EDISON_PAY")
    if not ep:
        return None, {}
    rows, m, hi = ep["rows"], ep["map"], ep["header_index"]
    idx = {}
    for r in rows[hi + 1:]:
        amt = cents(_cell(r, m.get("amount")))
        ref = N(_cell(r, m.get("reference")))
        if amt is None or not ref:
            continue
        dt = parse_date(_cell(r, m.get("date")))
        idx.setdefault(amt, []).append(
            (ref, N(_cell(r, m.get("invoice"))), dt.isoformat() if dt else ""))
    inv_info = {}
    ei = loaded.get("EDISON_INV")
    if ei:
        rows, m, hi = ei["rows"], ei["map"], ei["header_index"]
        for r in rows[hi + 1:]:
            inv = N(_cell(r, m.get("invoice")))
            if inv:
                inv_info.setdefault(inv, (N(_cell(r, m.get("status"))),
                                          N(_cell(r, m.get("voucher")))))
    return idx, inv_info


def _edison_note(bsl, edison_idx, inv_info):
    """Name the Edison payment a stranded State line corresponds to —
    ANNOTATION ONLY (the C6 State pass is retired; Edison records are the
    payer's, never pool entries, so they can never place anything).  A
    payment is cited when it matches on exact signed cents AND its payment
    reference digits appear in the line's text (reference outranks amount);
    an amount-only singleton is cited as uncorroborated.  Ambiguous
    amount-only sets are never guessed."""
    if not edison_idx:
        return ""
    cands = edison_idx.get(bsl.amount_cents, [])
    if not cands:
        return ""
    line_digits = {d for d in
                   (bsl.ref_digits | digit_runs(bsl.additional_info)
                    | digit_runs(bsl.account_servicer_reference))
                   if len(d) >= 6}
    def _tied(ref):
        z = znorm(ref)
        return bool(z) and len(z) >= 6 and any(
            z in d or d in z for d in line_digits)
    tied = sorted((p for p in cands if _tied(p[0])))
    chosen, qual = None, ""
    if tied:
        chosen = tied[0]
    elif len(cands) == 1:
        chosen = cands[0]
        qual = " (amount-only; uncorroborated)"
    if chosen is None:
        return ""
    ref, inv, dt = chosen
    note = f" Edison: State of TN payment {ref}"
    if inv:
        note += f" for invoice {inv!r}"
        st_v = inv_info.get(inv)
        if st_v and st_v[0]:
            note += f" ({st_v[0]}" + (f", voucher {st_v[1]}" if st_v[1] else "") + ")"
    if dt:
        note += f" paid {dt}"
    return note + f" matches this line exactly{qual}."


def _coa_tag(e, coa):
    """" (dept <DEP_DESC>, entity <ENT_DESC>)" for a named ST that carries a
    CoA-decodable asset combo, else "".  ADVISORY text only (rule 8c); returns
    "" when the CoA is absent, so Review output is byte-identical without it."""
    seg = getattr(e, "asset_segments", "")
    if not coa or not seg:
        return ""
    d = coa_decode(seg, coa)
    if not d:
        return ""
    bits = []
    if d.get("dep_desc"):
        bits.append(f"dept {d['dep_desc']}")
    if d.get("ent_desc"):
        bits.append(f"entity {d['ent_desc']}")
    return f" ({', '.join(bits)})" if bits else ""


def _p10_review_cause(bsl, pool, feed_errors, deposit_index=None, coa=None,
                      exact_entries=None):
    """Name the dominant Review cause (Section 9 P10), distinguishing open
    counterparts from already-reconciled ones and testing the ties it names.
    When a Chart of Accounts is loaded, each named ST that carries an asset
    combo is annotated with its decoded department/entity (advisory only).
    `exact_entries` is the caller's precomputed exact-amount slice."""
    exact = (exact_entries if exact_entries is not None
             else [e for e in pool if e.amount_cents == bsl.amount_cents])
    open_exact = [e for e in exact if e.available]
    closed_exact = [e for e in exact if not e.available]

    def _named(entries):
        return ", ".join(f"{e.id}{_coa_tag(e, coa)}" for e in entries)

    def _tied(entries):
        return [e for e in entries
                if cross_reference_tie(bsl, e)
                or reference_tie(bsl.recon_reference, e.reference)
                or (e.payer_tokens & bsl.payer_tokens)]

    codes = []
    if not exact:
        # Orphan doctrine signature #6 (owner, 2026-07-19): a stranded line
        # with NO counterpart at exact cents, but an ALL-closed MET deposit
        # summing exactly to it — the deposit's receipts were cherry-picked
        # onto other lines by earlier (auto) reconciliations.  Stop searching
        # the open pool; the consuming groups need an unwind first.
        def _reconciled_closed(e):
            # A genuinely consumed-by-prior-reconciliation member: unavailable
            # BECAUSE its status is closed (REC/cleared), not because it is a
            # foreign-account shadow entry or a dual-fire-suppressed twin
            # (both available=False for unrelated reasons — counting them
            # would falsely accuse a deposit of being cherry-picked).
            return (not e.available and not e.foreign_account
                    and _norm_header(e.status) in CLOSED_RECEIPT_STATUSES)
        for dep in sorted(deposit_index or {}):
            members = deposit_index[dep]
            if members and all(_reconciled_closed(e) for e in members) and \
                    sum(e.amount_cents for e in members) == bsl.amount_cents:
                codes.append("POSSIBLE_AUTO_REC_SPLIT")
                expl = (f"No open counterpart, but CLOSED ORT deposit d:{dep} "
                        f"({len(members)} already-reconciled receipt(s): "
                        + _named(_sorted(members)[:8])
                        + ") sums exactly to this line — its receipts were "
                        "cherry-picked onto other lines. Run Unreconcile2; "
                        "the line closes only after those groups release.")
                return list(dict.fromkeys(codes)), expl
        codes.append("NO_MATCH_FOUND")
        expl = "No exact-amount counterpart in the pool."
    elif bsl.lane == LANE_MERCHANT and open_exact:
        codes.append("MID_GUARDRAIL")
        expl = "Merchant line with exact amount but no in-window card group."
    elif open_exact:
        tied = _tied(open_exact)
        if tied:
            codes.append("DATE_CONFLICT")
            expl = "Exact amount + reference/payer tie exists but date is out of band."
        else:
            codes.append("MISSING_REFERENCE")
            expl = ("Exact-amount open entry exists but shares no reference "
                    "or payer tie.")
    else:
        tied = _tied(closed_exact)
        codes.append("ALREADY_RECONCILED_COUNTERPART")
        if tied:
            expl = ("Only already-reconciled counterpart(s) at this amount — "
                    + ("payer/reference tie " if tied else "")
                    + "suggests an auto-rec error: "
                    + _named(sorted(tied, key=lambda e: e.id))
                    + ". Run Unreconcile2.")
        else:
            expl = ("Only already-reconciled (closed) counterpart(s) at this "
                    "amount; no open entry. Run Unreconcile2 if misdirected.")
    if not bsl.recon_reference:
        codes.append("MISSING_REFERENCE")
    return list(dict.fromkeys(codes)), expl


def recommend_gl_string(bsl, loaded, exact_entries=None):
    """Best-effort GL account string for a manual ECT (Section 6 sources)."""
    # MID master only (merchant lane); MISC Receipts is an ALL_DATA sheet,
    # never loaded here — returns "" when there is no MID hit.
    mid_master = loaded.get("MID_MASTER")
    if bsl.mid and mid_master:
        gl = mid_master.get("mid_gl", {}).get(bsl.mid)
        if gl:
            return gl
    # CoA fallback (owner, 2026-07-19; hardened 2026-07-19 adversarial
    # review): after a MID miss, an exact-amount MET counterpart's OFFSET
    # combo — the ECT posting side (§6) — yields the recommendation.  The
    # cash-side asset combo is NEVER recommended (its account segment is the
    # bank's own cash GL), and foreign-account shadow entries are excluded
    # (rule 8g: another depository's posting must not steer this account's
    # ECT).  `exact_entries` is the caller's precomputed exact-amount slice
    # (P10 already scanned the pool once — no second full scan).  Advisory
    # only; returns "" when the CoA is absent so output is byte-identical.
    coa = loaded.get("CHART_OF_ACCOUNTS")
    if coa and exact_entries:
        for e in _sorted([p for p in exact_entries
                          if not p.foreign_account
                          and getattr(p, "offset_segments", "")]):
            d = coa_decode(e.offset_segments, coa)
            if d:
                label = " / ".join(x for x in
                                   (d.get("ent_desc"), d.get("dep_desc"),
                                    d.get("act_desc")) if x)
                return (f"{e.offset_segments} = {label}" if label
                        else e.offset_segments)
    return ""


def _assert_conservation(bsls, placements):
    """Every BSL placed exactly once; no ST id consumed twice across Match+
    Candidate (Section 9 trailer)."""
    assert len(placements) == len(bsls), (
        f"conservation: {len(placements)} placements != {len(bsls)} BSLs")
    seen_lines = set()
    consumed = {}
    for p in placements:
        assert p.bsl.line_key not in seen_lines, f"BSL {p.bsl.line_key} placed twice"
        seen_lines.add(p.bsl.line_key)
        # MISDIRECTED included (hard guardrail, 2026-07-19): one foreign
        # entry must never be cited as the counterpart of two bank lines —
        # the audit's C3 already spans the Misdirected tab; the engine's own
        # invariant now matches it.
        if p.kind in (MATCH, CANDIDATE, MISDIRECTED):
            for e in p.st_entries:
                assert e.id not in consumed, (
                    f"ST {e.id} consumed twice ({consumed.get(e.id)} and {p.pass_name})")
                consumed[e.id] = p.pass_name

# ======================================================================
# Section 13 — Workbook writer (the §10.6 unwind writer is Unreconcile2's)
# ======================================================================

FONT_NAME = "Carlito"
FONT_SIZE = 11
NAVY = "FF1F4E78"
WHITE = "FFFFFFFF"


def _usd(c):
    if c is None:
        return ""
    neg = c < 0
    dollars = abs(c) / 100
    s = f"{dollars:,.2f}"
    return f"({s})" if neg else s


def _fmt_date(d):
    if d is None:
        return ""
    return d.strftime("%Y-%m-%d")


def _safe_cell(v):
    """Prepend a space to any value starting with '=' (Section 3/13)."""
    s = "" if v is None else str(v)
    if s.startswith("="):
        return " " + s
    return s


def _join_multi(values):
    """Multi-value cell: comma-space, or semicolon-space if any value embeds a
    thousands-separator comma (Section 13)."""
    strs = [str(v) for v in values]
    sep = "; " if any("," in s for s in strs) else ", "
    return sep.join(strs)


def _write_workbook(path, title, tabs, header_fill_hex):
    """Generic writer applying the Section 13 formatting standard."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    base_font = Font(name=FONT_NAME, size=FONT_SIZE)
    header_font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=WHITE)
    fill = PatternFill(start_color=header_fill_hex, end_color=header_fill_hex, fill_type="solid")

    for tab_title, columns, rows in tabs:
        ws = wb.create_sheet(title=tab_title[:31])
        # Row 1 title, row 2 blank, row 3 headers, row 4+ data (Section 13).
        ws.cell(row=1, column=1, value=title).font = header_font
        for ci, col in enumerate(columns, start=1):
            c = ws.cell(row=3, column=ci, value=col)
            c.font = header_font
            c.fill = fill
            c.alignment = Alignment(vertical="center")
        for ri, row in enumerate(rows, start=4):
            for ci, val in enumerate(row, start=1):
                c = ws.cell(row=ri, column=ci, value=_safe_cell(val))
                c.font = base_font
        ws.freeze_panes = "A4"
        # Column widths (cosmetic, deterministic).
        for ci, col in enumerate(columns, start=1):
            width = max(12, min(60, len(col) + 4))
            ws.column_dimensions[get_column_letter(ci)].width = width
    wb.save(path)


# HARD GUARDRAIL (owner, 2026-07-11): the output carries ALL BSL identifier
# fields and ALL ST detail fields — they are essential for reconciliation.
# The ST Number(s) cell lists EVERY cited ST, never truncated; the audit's
# C9 pins this exact 19-column layout and C10 rejects extra columns.
RECON_COLUMNS = [
    "BSL Date", "BSL Line Info", "BSL Amount",
    "BSL Reference", "BSL Additional Information", "BSL Customer Reference",
    "BSL Account Servicer Reference", "BSL Transaction Type",
    "ST Date(s)", "ST Number(s)", "ST Amount(s)", "ST Reference(s)",
    "ST Structured Payment Reference(s)", "ST Counterparty(ies)",
    "ST Source(s)",
    "Confidence", "ORT d:", "ORT r:", "Explanation",
]


def _placement_row(p: Placement):
    ents = p.st_entries
    st_dates = _join_multi([_fmt_date(e.date) for e in ents]) if ents else ""
    st_numbers = _join_multi([e.id for e in ents]) if ents else ""
    st_amounts = _join_multi([_usd(e.amount_cents) for e in ents]) if ents else ""
    st_refs = _join_multi([e.reference or "" for e in ents]) if ents else ""
    st_sprs = _join_multi([e.spr or "" for e in ents]) if ents else ""
    st_cps = _join_multi([e.counterparty or "" for e in ents]) if ents else ""
    st_srcs = _join_multi([e.source or "" for e in ents]) if ents else ""
    dep = _join_multi(sorted(set(p.deposit_ids))) if p.deposit_ids else ""
    rec = _join_multi(sorted(set(p.receipt_ids))) if p.receipt_ids else ""
    codes = (" [" + ", ".join(p.codes) + "]") if p.codes else ""
    return [
        _fmt_date(p.bsl.date),
        p.bsl.line_info,
        _usd(p.bsl.amount_cents),
        p.bsl.reference_raw,
        p.bsl.additional_info,
        p.bsl.customer_reference,
        p.bsl.account_servicer_reference,
        p.bsl.transaction_type,
        st_dates,
        st_numbers,
        st_amounts,
        st_refs,
        st_sprs,
        st_cps,
        st_srcs,
        p.confidence,
        dep,
        rec,
        (p.explanation + codes).strip(),
    ]


def write_reconciliation_workbook(path, account, placements):
    """Matches / Candidate Matches / Misdirected / Review Notes (Section 13;
    Misdirected tab per owner hard guardrail 2026-07-18)."""
    matches = [p for p in placements if p.kind == MATCH]
    candidates = [p for p in placements if p.kind == CANDIDATE]
    misdirected = [p for p in placements if p.kind == MISDIRECTED]
    reviews = [p for p in placements if p.kind == REVIEW]
    title = f"UT Reconciliation — {account}"
    tabs = [
        ("Matches", RECON_COLUMNS, [_placement_row(p) for p in matches]),
        ("Candidate Matches", RECON_COLUMNS, [_placement_row(p) for p in candidates]),
        ("Misdirected", RECON_COLUMNS, [_placement_row(p) for p in misdirected]),
        ("Review Notes", RECON_COLUMNS, [_placement_row(p) for p in reviews]),
    ]
    _write_workbook(path, title, tabs, NAVY)
    return {"matches": len(matches), "candidates": len(candidates),
            "misdirected": len(misdirected), "reviews": len(reviews)}


# ======================================================================
# MID master loader (Section 4.2)
# ======================================================================
def load_mid_master(routed_file: RoutedFile):
    """Scan every sheet for MID tokens; build MID -> account-string map."""
    from openpyxl import load_workbook
    wb = load_workbook(routed_file.path, read_only=False, data_only=True)
    mid_gl = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            mids, gls = [], []
            for c in row:
                s = N(c)
                if is_mid(s):
                    mids.append(znorm(s))
                if pred_gl_string(s):
                    gls.append(s)
            if not mids or not gls:
                continue
            # Never let column position pick silently: a row naming two
            # distinct GL strings, or remapping a MID already mapped to a
            # different GL, is genuinely ambiguous — fail loud.
            if len(set(gls)) > 1:
                raise InvalidSourceData(
                    routed_file.filename, "MID_MASTER",
                    f"sheet {ws.title!r}: row maps MID(s) {sorted(set(mids))} "
                    f"to multiple distinct GL strings {sorted(set(gls))}")
            for mid in dict.fromkeys(mids):
                prev = mid_gl.setdefault(mid, gls[0])
                if prev != gls[0]:
                    raise InvalidSourceData(
                        routed_file.filename, "MID_MASTER",
                        f"MID {mid} maps to conflicting GL strings "
                        f"{prev!r} and {gls[0]!r}")
    wb.close()
    return {"mid_gl": mid_gl}


# ---- Chart of Accounts decode bundle (owner COA export, 2026-07-19) ----

def _coa_header_index(rows, marker):
    """First row carrying a header cell whose normalized form == marker.
    OTBI workbooks prepend title/run-date preamble rows, so the header is
    located, never assumed row 0 (matches the engine's header-scan doctrine)."""
    want = _norm_header(marker)
    for i, row in enumerate(rows):
        for c in row:
            if _norm_header(c) == want:
                return i
    return None


def _coa_colmap(header_row):
    """Normalized-header -> leftmost column index (duplicate headers bind the
    leftmost, per the binder's convention)."""
    idx = {}
    for j, h in enumerate(header_row):
        idx.setdefault(_norm_header(h), j)
    return idx


def _coa_read_acctcombos(rf, combo_decode):
    """Parse one AcctCombos shard into combo_decode[ACCOUNT_COMBO] -> labels.
    The *_DESC columns are pre-joined "<code>-<label>" pairs, so no per-segment
    dictionary lookup is needed on the common path."""
    rows, _ = read_rows(rf)
    hi = _coa_header_index(rows, "ACCOUNT_COMBO")
    if hi is None:
        return
    idx = _coa_colmap(rows[hi])
    ci = idx.get("ACCOUNTCOMBO")
    if ci is None:
        return
    c_ent = idx.get("ENTDESC")
    c_fnd = idx.get("FNDDESC")
    c_dep = idx.get("DEPDESC")
    c_act = idx.get("ACTDESC")
    c_itc = idx.get("ITCDESC")
    c_grp = idx.get("ACTGRPDESC")
    for r in rows[hi + 1:]:
        combo = N(_cell(r, ci))
        parts = segments_of(combo)
        if len(parts) < 4:
            continue
        key = "-".join(parts)
        combo_decode.setdefault(key, {
            "ent": parts[0], "ent_desc": _coa_label(_cell(r, c_ent)),
            "fund": parts[1], "fund_desc": _coa_label(_cell(r, c_fnd)),
            "dept": parts[2], "dep_desc": _coa_label(_cell(r, c_dep)),
            "account": parts[3], "act_desc": _coa_label(_cell(r, c_act)),
            "intercompany": parts[6] if len(parts) > 6 else "",
            "itc_desc": _coa_label(_cell(r, c_itc)),
            "act_grp_desc": _coa_label(_cell(r, c_grp)),
        })


def _coa_read_segments(rf, entity_desc):
    """Parse Segments.csv into entity_desc[VALUE] -> DESCRIPTION for
    SEGMENT_TYPE=Entity rows with ENABLED_FLAG='Y'.  This value set also decodes
    the Intercompany segment (it reuses the Entity codes).  Effective-date
    gating is intentionally omitted: the engine carries no clock (rule 9)."""
    rows, _ = read_rows(rf)
    hi = _coa_header_index(rows, "SEGMENT_TYPE")
    if hi is None:
        return
    idx = _coa_colmap(rows[hi])
    c_type = idx.get("SEGMENTTYPE")
    c_val = idx.get("VALUE")
    c_desc = idx.get("DESCRIPTION")
    c_en = idx.get("ENABLEDFLAG")
    if c_type is None or c_val is None or c_desc is None:
        return
    for r in rows[hi + 1:]:
        if _norm_header(_cell(r, c_type)) != "ENTITY":
            continue
        if c_en is not None and _norm_header(_cell(r, c_en)) not in ("Y", ""):
            continue
        val = N(_cell(r, c_val))
        if val:
            entity_desc.setdefault(val, N(_cell(r, c_desc)))


def _coa_read_combosets(rf, postable_efdp):
    """Parse ComboSets/CombosTech into the postable E-F-D-P whitelist.  The
    "Combination Set" column already carries the 4-segment key
    (Entity-Fund-Department-Program); the "1 of 1" footer and any non-4-part
    cell are dropped, and non-breaking spaces are stripped."""
    rows, _ = read_rows(rf)
    hi = _coa_header_index(rows, "Combination Set")
    if hi is None:
        return
    ci = _coa_colmap(rows[hi]).get("COMBINATIONSET")
    if ci is None:
        return
    for r in rows[hi + 1:]:
        key = N(_cell(r, ci)).replace("\xa0", "").strip()
        parts = [p.strip() for p in key.split("-")]
        if len(parts) == 4 and all(parts):
            postable_efdp.add("-".join(parts))


def load_chart_of_accounts(files):
    """Load the Chart of Accounts reference bundle (owner COA export,
    2026-07-19): the combination universe that DECODES a GL combo into human
    department/entity labels for Review and recommended-GL text.  ADVISORY
    ONLY — campus/entity "consistency" confers no matching evidence (rule 8c),
    so nothing here ever gates a placement.

    Accepts the routed CoA file set (the seven AcctCombos shards, Segments.csv,
    ComboSets/CombosTech); dispatches each by filename token and skips anything
    else (RelatedValueSets and the huge ORT_Activity_GL_Departments routing
    table are out of scope here).  Returns {combo_decode, entity_desc,
    postable_efdp} or None when nothing usable loaded (every consumer no-ops)."""
    combo_decode, entity_desc, postable_efdp = {}, {}, set()
    for rf in files:
        low = rf.filename.lower()
        if "acctcombos" in low:
            _coa_read_acctcombos(rf, combo_decode)
        elif "combosets" in low or "combostech" in low:
            _coa_read_combosets(rf, postable_efdp)
        elif "segments" in low:
            _coa_read_segments(rf, entity_desc)
    if not combo_decode and not entity_desc:
        return None
    return {"combo_decode": combo_decode, "entity_desc": entity_desc,
            "postable_efdp": postable_efdp}


# ---- CM Configuration exports (owner, 2026-07-19) --------------------
# The five Cash Management configuration reports (Transaction Creation
# Rules, Parse Rules, Matching Rules, Tolerance Rules, Recon Rulesets).
# Loaded for the ADVISORY config audit only (orphan-doctrine R5 activation:
# "CFG_TCR coverage activates only when those exports are present") — they
# never gate a placement.  Paginated exports: title + run-date preamble,
# header row located by marker (never assumed), column A blank padding,
# repeated headers / "N of M" footer rows dropped.

def _cfg_read_table(rf, marker, fields, keep_keys=None):
    """Read one CM_Configurations export into a list of dicts.  `fields` maps
    output key -> header alias; the header row is located by `marker`.  A row
    is kept when ANY of `keep_keys` (default: the first field) is non-blank
    and is not a repeated page header — the TCR export carries ORPHAN rules
    with a BLANK bank-account cell (rules detached from any account) that a
    single-column keep-rule would silently drop (doctrine rule 3)."""
    rows, _ = read_rows(rf)
    hi = _coa_header_index(rows, marker)
    if hi is None:
        raise InvalidSourceData(rf.filename, rf.role,
                                f"no header row carrying {marker!r} found")
    idx = _coa_colmap(rows[hi])
    cols = {k: idx.get(_norm_header(alias)) for k, alias in fields.items()}
    key0 = next(iter(fields))
    if cols[key0] is None:
        raise InvalidSourceData(rf.filename, rf.role,
                                f"marker column {fields[key0]!r} did not bind")
    keys = list(keep_keys or (key0,))
    header_norms = {_norm_header(alias) for alias in fields.values()}
    out = []
    for r in rows[hi + 1:]:
        vals = {k: (N(_cell(r, ci)) if ci is not None else "")
                for k, ci in cols.items()}
        keep = next((vals[k] for k in keys if vals.get(k)), "")
        if not keep or _norm_header(keep) in header_norms:
            continue  # blank spacer / repeated page header
        out.append(vals)
    return out


def load_cm_config(role, rf):
    """Dispatch one CFG_* routed file to its table shape."""
    if role == "CFG_TCR":
        return _cfg_read_table(rf, "BNK ACCNTNME", {
            "bank_account": "BNK ACCNTNME", "enabled": "ENBLD",
            "seq": "SQC NUM", "name": "RLE NME", "descr": "DSCRPT",
            "trx_code": "TRX CDE", "trx_type": "TRX TPE",
            "search_field": "SRCH FLD", "search_string": "SRCH STRNG",
            "cash": "CASH", "offset": "OFFSET",
            "last_update": "LST UPDTE", "updated_by": "UPDTEBY"},
            # Orphan rules (blank bank account, rule name present) must load —
            # 23 live on the real export, 3 of them still ENABLED.
            keep_keys=("bank_account", "name"))
    if role == "CFG_PARSE":
        return _cfg_read_table(rf, "PRS RLE SET", {
            "rule_set": "PRS RLE SET", "descr": "DSCRPT",
            "set_enabled": "RLE SET ENBLD", "seq": "SQNCENUM",
            "enabled": "ENBLD", "trx_code": "TRX CDE", "trx_type": "TRX TPE",
            "source_field": "SRCE FLD", "target_field": "TRGT FLD",
            "pattern": "PRSE RLE", "overwrite": "OVRWRTE",
            "last_update": "LST UPDTE", "updated_by": "UPDTE BY"})
    if role == "CFG_MATCHING":
        return _cfg_read_table(rf, "MTCH RLE NME", {
            "name": "MTCH RLE NME", "descr": "DSCRPT", "enabled": "ENBLD",
            "source": "TRX SRCE", "match_type": "MTCH TPE",
            "amount_match": "AMNT MTCH", "date_match": "DTE MTCH",
            "ref_match": "REF ID MTCH", "type_match": "TPE MTCH",
            "stmt_groupby": "STMT GRPBY", "trx_groupby": "TRX GRPBY",
            "criteria": "ADV CRTRA",
            "last_update": "LST UPDTE", "updated_by": "UPDTEBY"})
    if role == "CFG_TOLERANCE":
        return _cfg_read_table(rf, "TOL RLE NME", {
            "name": "TOL RLE NME", "descr": "DSCRPT",
            "date_enabled": "DTE ENBLD", "days_before": "DAYS BFR",
            "days_after": "DAYS AFTR", "amount_enabled": "AMNT ENBLD",
            "amount_below": "AMNT BELOW", "amount_above": "AMNT ABOVE",
            "percent_enabled": "PRCNT ENBLD",
            "last_update": "LST UPDTE", "updated_by": "LST UPDTE BY"})
    if role == "CFG_RULESETS":
        return _cfg_read_table(rf, "RLST NME", {
            "ruleset": "RLST NME", "descr": "RLST DSCRPT",
            "last_update": "RLST LST UPDTE", "updated_by": "RLST UPDTE BY",
            "seq": "SQNCE NUM", "matching_rule": "MTCH RLE NME",
            "tolerance_rule": "TOL RLE NME"})
    return None


CM_CONFIG_ROLES = ("CFG_TCR", "CFG_PARSE", "CFG_MATCHING",
                   "CFG_TOLERANCE", "CFG_RULESETS")


@functools.lru_cache(maxsize=8192)
def _like_regex(pattern, ci):
    """Compile an Oracle LIKE pattern (% = any run, _ = any one char) to a
    regex.  Used to SIMULATE Transaction Creation Rule firing against
    bank-line addenda — advisory config audit only."""
    out = []
    for ch in pattern:
        if ch == "%":
            out.append(".*")
        elif ch == "_":
            out.append(".")
        else:
            out.append(re.escape(ch))
    flags = re.DOTALL | (re.IGNORECASE if ci else 0)
    return re.compile("".join(out), flags)


def like_match(pattern, text, case_sensitive=True):
    """Oracle LIKE semantics, full-match.  Oracle applies LIKE to the whole
    column value (TCR search strings carry their own leading/trailing % when
    substring semantics are intended) and is CASE-SENSITIVE — the default
    here.  A pattern that matches case-insensitively but NOT case-sensitively
    is the config audit's case-bug signature ('%CANTALOUPE%' never firing on
    'Cantaloupe, Inc.')."""
    p = N(pattern)
    if not p:
        return False
    return _like_regex(p, not case_sensitive).fullmatch(N(text)) is not None


# ---- Configuration audit (owner, 2026-07-19) --------------------------
# ADVISORY ONLY: replays the Transaction Creation Rules against this run's
# bank lines and inspects parse-rule coverage, producing config-change
# recommendations (disable / fix / create) in the runlog and a markdown
# artifact.  Never touches placements, the workbook, or the audit.
# Real-data validation (FHB UTHSC FY27): TCR coverage IS reconciliation —
# 631 TCR-claimed lines were 100% reconciled while ALL 44 unreconciled
# lines were TCR-unclaimed, so coverage gaps are the cheapest UNR fix.

def _tcr_search_text(bsl, rule):
    """The BSL field a TCR's SRCH FLD points at."""
    fld = _norm_header(rule.get("search_field"))
    if not fld:
        return None                     # blank field: rule fires on code alone
    if "ADDENDA" in fld:
        return bsl.additional_info
    if "SERV" in fld:
        return bsl.account_servicer_reference
    if "CUSTOMER" in fld:
        return bsl.customer_reference
    if "RECON" in fld:
        return bsl.recon_reference
    return bsl.additional_info


def _tcr_claims(bsl, rule, case_sensitive=True):
    """Would this TCR fire on this bank line?  Code must match; a blank
    search string claims every line of the code (Oracle behavior)."""
    if N(rule.get("trx_code")) != N(bsl.transaction_code):
        return False
    pat = N(rule.get("search_string"))
    if not pat:
        return True
    text = _tcr_search_text(bsl, rule)
    if text is None:
        return True
    if "%" not in pat and "_" not in pat:
        pat = f"%{pat}%"                # bare strings behave as contains
    return like_match(pat, text, case_sensitive=case_sensitive)


def _uncovered_signature(bsl):
    """Grouping signature for TCR-uncovered lines: the leading description
    text with digit runs collapsed, so recurring flows cluster."""
    s = re.sub(r"\d+", "#", N(bsl.additional_info))[:40].strip()
    return (N(bsl.transaction_code), s)


def config_audit(loaded, account, placements, pool, output_dir, runlog):
    """Replay the CM configuration against this run's outcomes (orphan
    doctrine R5 — activates only when CFG exports are present).  Writes
    runlog['config_audit'] plus <account>_config_recommendations.md/.json.
    ALL fire counts are SIMULATED against the BSL export's (truncated)
    addenda — an upper bound on coverage gaps, never placement evidence."""
    tcr_cfg = loaded.get("CFG_TCR")
    parse_cfg = loaded.get("CFG_PARSE")
    if not tcr_cfg and not parse_cfg:
        return None
    report = {}
    lines_md = [f"# Configuration recommendations — {account}",
                "", "All rule-fire counts are SIMULATED against the exported "
                "bank lines (Oracle evaluates its own untruncated addenda); "
                "treat them as leads, not proof.", ""]

    # ---- TCR replay -------------------------------------------------
    if tcr_cfg:
        mine = [r for r in tcr_cfg["rules"]
                if account_of_bank_name(r.get("bank_account")) == account]
        enabled = [r for r in mine if N(r.get("enabled")).upper() == "Y"]
        disabled = [r for r in mine if N(r.get("enabled")).upper() != "Y"]
        # Static config defects (need no bank lines) -------------------
        orphans = [r for r in tcr_cfg["rules"] if not N(r.get("bank_account"))]
        null_code = [r for r in enabled if not N(r.get("trx_code"))]
        no_cash = [r for r in mine if not N(r.get("cash"))]
        wrong_gl = [r for r in enabled
                    if N(r.get("cash"))
                    and account_of_gl_segments(r.get("cash")) not in (None, account)]
        dup_index = {}
        for r in enabled:
            key = (N(r.get("trx_code")), N(r.get("search_string")).upper())
            if key[1]:
                dup_index.setdefault(key, []).append(N(r.get("name")))
        dup_pairs = sorted((k, v) for k, v in dup_index.items() if len(v) > 1)
        # Amount-existence set for the creation-failure split: the WHOLE
        # pool including shadow/closed entries — absence at |cents| means
        # the TCR created nothing anywhere.
        pool_abs = {abs(e.amount_cents) for e in pool}
        claimed_rules = set()
        creation_failures = []     # claimed + NO pool entry at the amount
        stranded_claimed = []      # claimed + counterpart exists, unmatched
        case_bug_hits = []         # enabled TCR fires only case-insensitively
        disabled_hits = []         # Review lines only a DISABLED TCR claims
        uncovered = {}             # signature -> [placement, ...]
        for p in placements:
            bsl = p.bsl
            claimers = [r for r in enabled if _tcr_claims(bsl, r)]
            for r in claimers:
                claimed_rules.add(r.get("name"))
            if p.kind != REVIEW:
                continue
            if claimers:
                if abs(bsl.amount_cents) not in pool_abs:
                    creation_failures.append((bsl, claimers[0]))
                else:
                    stranded_claimed.append((bsl, claimers[0]))
                continue
            ci_only = [r for r in enabled
                       if _tcr_claims(bsl, r, case_sensitive=False)]
            if ci_only:
                case_bug_hits.append((bsl, ci_only[0]))
                continue
            dis = [r for r in disabled if _tcr_claims(bsl, r)]
            if dis:
                disabled_hits.append((bsl, dis[0]))
                continue
            uncovered.setdefault(_uncovered_signature(bsl), []).append(p)
        idle = sorted(N(r.get("name")) for r in enabled
                      if N(r.get("name")) not in claimed_rules)
        report["tcr"] = {
            "rules_for_account": len(mine),
            "enabled": len(enabled),
            "orphan_rules_no_account": len(orphans),
            "enabled_null_trx_code": len(null_code),
            "cash_combo_missing": len(no_cash),
            "cash_gl_foreign_account": len(wrong_gl),
            "duplicate_enabled_pairs": len(dup_pairs),
            "creation_failures": len(creation_failures),
            "stranded_but_claimed": len(stranded_claimed),
            "case_bug_suspects": len(case_bug_hits),
            "claimed_by_disabled_rule": len(disabled_hits),
            "uncovered_review_lines": sum(len(v) for v in uncovered.values()),
            "idle_rules_this_window": len(idle),
        }
        if null_code:
            lines_md += ["## Enabled rules with NO transaction code (FIX)",
                         "A null TRX CDE can never fire as exported.", ""]
            lines_md += [f"- `{N(r.get('name'))}` (updated "
                         f"{N(r.get('last_update'))} by {N(r.get('updated_by'))})"
                         for r in null_code]
            lines_md.append("")
        if wrong_gl:
            lines_md += ["## Rules whose CASH combo posts to ANOTHER "
                         "depository (FIX)", ""]
            for r in wrong_gl:
                lines_md.append(
                    f"- `{N(r.get('name'))}` CASH `{N(r.get('cash'))}` -> "
                    f"{account_of_gl_segments(r.get('cash'))} (this account is {account})")
            lines_md.append("")
        if no_cash:
            lines_md += ["## Rules with NO CASH combo (FIX before enabling)", ""]
            lines_md += [f"- `{N(r.get('name'))}` (enabled={N(r.get('enabled'))})"
                         for r in no_cash]
            lines_md.append("")
        if dup_pairs:
            lines_md += ["## Duplicate enabled rules — same code + search "
                         "string (DISABLE one of each)", ""]
            for (code, s), names in dup_pairs:
                lines_md.append(f"- code {code} `{s[:60]}`: " + ", ".join(
                    f"`{n}`" for n in sorted(names)))
            lines_md.append("")
        if orphans:
            lines_md += [f"## Orphan rules bound to NO bank account "
                         f"({len(orphans)}; CLEAN UP)",
                         "Rebind to the intended account (parseable from the "
                         "name) or delete.", ""]
            lines_md += [f"- `{N(r.get('name'))}` (enabled={N(r.get('enabled'))})"
                         for r in sorted(orphans, key=lambda r: N(r.get("name")))[:30]]
            lines_md.append("")
        if creation_failures:
            lines_md += ["## TCR creation FAILURES (INVESTIGATE the feed)",
                         "An enabled rule claims each line, yet NO system "
                         "transaction exists at that amount anywhere in the "
                         "pool — the rule fired (or should have) and created "
                         "nothing.", ""]
            for bsl, r in creation_failures[:25]:
                lines_md.append(
                    f"- {bsl.date} {_usd(bsl.amount_cents)} code {bsl.transaction_code}"
                    f" — rule `{r.get('name')}` | {N(bsl.additional_info)[:70]}")
            lines_md.append("")
        if stranded_claimed:
            lines_md += ["## TCR fired-but-stranded (FIX / INVESTIGATE)",
                         "An enabled Transaction Creation Rule claims each of "
                         "these lines and a counterpart amount exists in the "
                         "pool, yet the line is still unreconciled.", ""]
            for bsl, r in stranded_claimed[:25]:
                lines_md.append(
                    f"- {bsl.date} {_usd(bsl.amount_cents)} code {bsl.transaction_code}"
                    f" — rule `{r.get('name')}` | {N(bsl.additional_info)[:70]}")
            lines_md.append("")
        if case_bug_hits:
            lines_md += ["## Case-sensitivity bugs (FIX the search string)",
                         "The rule matches this line case-INSENSITIVELY only; "
                         "Oracle LIKE is case-sensitive, so the rule never "
                         "fires ('%CANTALOUPE%' vs 'Cantaloupe, Inc.').", ""]
            for bsl, r in case_bug_hits[:25]:
                lines_md.append(
                    f"- rule `{r.get('name')}` (search `{r.get('search_string')}`)"
                    f" vs line {bsl.date} {_usd(bsl.amount_cents)}:"
                    f" {N(_tcr_search_text(bsl, r))[:70]}")
            lines_md.append("")
        if disabled_hits:
            lines_md += ["## Claimed only by a DISABLED rule (ENABLE or replace)", ""]
            for bsl, r in disabled_hits[:25]:
                lines_md.append(
                    f"- {bsl.date} {_usd(bsl.amount_cents)} — disabled rule "
                    f"`{r.get('name')}`")
            lines_md.append("")
        recurring = sorted(((sig, ps) for sig, ps in uncovered.items()
                            if len(ps) >= 2),
                           key=lambda kv: (-len(kv[1]), kv[0]))
        if recurring:
            lines_md += ["## Uncovered recurring flows (CREATE a TCR)",
                         "No enabled TCR claims these unreconciled lines; "
                         "each signature recurs in this run.  Proposed search "
                         "string uses the line's merchant number when one is "
                         "present.", ""]
            for (code, sig), ps in recurring[:15]:
                mids = sorted({p.bsl.mid for p in ps if p.bsl.mid})
                prop = (f"%MERCHANT%DEPOSIT%{mids[0]}%" if mids
                        else f"%{sig.replace('#', '%').strip('% ')}%")
                total = sum(p.bsl.amount_cents for p in ps)
                lines_md.append(
                    f"- code {code} × {len(ps)} lines ({_usd(total)}): sig "
                    f"`{sig}` -> proposed SRCH STRNG `{prop}`")
            lines_md.append("")
        report["tcr"]["uncovered_recurring_signatures"] = len(recurring)
        if idle:
            lines_md += [f"## Idle rules this window ({len(idle)} of "
                         f"{len(enabled)} enabled claimed nothing)",
                         "Window-level only — check a full-year replay before "
                         "pruning.", ""]
            lines_md += [f"- `{n}`" for n in idle[:30]]
            lines_md.append("")

    # ---- Parse-rule coverage ---------------------------------------
    if parse_cfg:
        bank = account.split("_", 1)[0]      # FHB / REGIONS
        bank_rules = [r for r in parse_cfg["rules"]
                      if bank.lower() in N(r.get("rule_set")).lower()
                      and N(r.get("enabled")).upper() == "Y"]
        covered_codes = {N(r.get("trx_code")) for r in bank_rules}
        code_stats = {}
        for p in placements:
            bsl = p.bsl
            code = N(bsl.transaction_code)
            st = code_stats.setdefault(code, {"lines": 0, "no_ref": 0})
            st["lines"] += 1
            if not bsl.recon_reference:
                st["no_ref"] += 1
        gaps = {c: st for c, st in sorted(code_stats.items())
                if c not in covered_codes and st["no_ref"]}
        report["parse"] = {
            "bank": bank, "rules": len(bank_rules),
            "covered_codes": sorted(covered_codes),
            "uncovered_codes_with_blank_refs": {
                c: st["no_ref"] for c, st in gaps.items()},
        }
        if gaps:
            lines_md += ["## Parse-rule gaps (CREATE parse rules)",
                         f"Bank rule set covers codes "
                         f"{', '.join(sorted(covered_codes))}; these codes in "
                         "this run carry NO parse rule, leaving the "
                         "reconciliation reference blank:", ""]
            for c, st in gaps.items():
                lines_md.append(
                    f"- code {c}: {st['no_ref']} of {st['lines']} lines have a "
                    "blank reference -> add ACCNT_SERVICER_REF -> "
                    "RECON_REFERENCE `(X~)`")
            lines_md += ["",
                         "Also recommended (raw-BAI2 evidence): extract "
                         "`Customer ID:` / `Trace Number:` / `TRN1` and check "
                         "numbers from ADDENDA — the bank's servicer-ref "
                         "field letter-strips alphanumeric ACH ids, and no "
                         "current rule captures these keys.", ""]

    # ---- Tolerance note ---------------------------------------------
    tol = loaded.get("CFG_TOLERANCE")
    if tol and tol["rules"]:
        t = tol["rules"][0]
        report["tolerance"] = {"name": t.get("name"),
                               "days_before": t.get("days_before"),
                               "days_after": t.get("days_after"),
                               "amount_enabled": t.get("amount_enabled")}

    path = os.path.join(output_dir, f"{account}_config_recommendations.md")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines_md) + "\n")
    jpath = os.path.join(output_dir, f"{account}_config_recommendations.json")
    with open(jpath, "w", encoding="utf-8") as fh:
        json.dump(report, fh, indent=2, sort_keys=True, default=str)
    report["report_path"] = path
    report["json_path"] = jpath
    runlog["config_audit"] = report
    return report


# ======================================================================
# Section 9 P0 — Load & validate; build BSL list
# ======================================================================

def load_and_bind(by_role, runlog):
    """Read + bind every routed file to its required roles.  Returns `loaded`
    dict.  Raises on any required role that cannot be bound."""
    loaded = {}
    role_specs = {
        "BSL": BSL_ROLES,
        "ST": ST_ROLES,
        "RECEIPTS": RECEIPTS_ROLES,
        "PAYMENTS": PAYMENTS_ROLES,
        "ALL_BSL": ALL_BSL_ROLES,
        "MET": MET_ROLES,
        "BAI2": BAI2_ROLES,
        "DEPT_INFO": DEPT_INFO_ROLES,
        "EDISON_PAY": EDISON_PAY_ROLES,
        "EDISON_INV": EDISON_INV_ROLES,
        "APPLIED_UNAPPLIED": APPLIED_UNAPPLIED_ROLES,
        "AR_MATCHED": AR_MATCHED_ROLES,
        "ENRICHED": ENRICHED_ROLES,
        "ORT_AR": {"trx_id": _rs(False, ["Transaction Number"], pred_reference),
                   "parked_receipt_id": _rs(False, ["Parked Receipt ID"], pred_number),
                   "bank_name": _rs(False, ["BANK_NAME", "Bank Account Name"], pred_any)},
        "ORT_MISC": {"trx_id": _rs(False, ["Transaction Number"], pred_reference),
                     "parked_receipt_id": _rs(False, ["Parked Receipt ID"], pred_number),
                     "bank_name": _rs(False, ["BANK_NAME", "Bank Account Name"], pred_any)},
    }
    bound_report = {}
    for role, files in by_role.items():
        if role in ("ALL_DATA", "RECONCILED"):
            continue  # recognized but never loaded (Unreconcile2 / offline
            #           config-audit fuel — never the forward engine's input)
        if role == "CHART_OF_ACCOUNTS":
            # Multi-file reference bundle: pass the WHOLE routed set to the
            # loader (never _pick_newest — the shards are complementary, not
            # date-competing).  Optional at every consumer, so None is fine.
            coa = load_chart_of_accounts(files)
            if coa is not None:
                loaded["CHART_OF_ACCOUNTS"] = coa
                bound_report[role] = {"files": [f.filename for f in files],
                                      "combos": len(coa["combo_decode"]),
                                      "entities": len(coa["entity_desc"]),
                                      "postable_efdp": len(coa["postable_efdp"])}
            continue
        if role in CM_CONFIG_ROLES:
            # CM configuration reports (orphan-doctrine R5 activation):
            # loaded for the advisory config audit only, never a gate.
            rf = _pick_newest(files, role)
            rules = load_cm_config(role, rf)
            if rules:
                loaded[role] = {"rules": rules, "file": rf.filename}
                bound_report[role] = {"file": rf.filename, "rows": len(rules)}
            continue
        specs = role_specs.get(role)
        if specs is None and role != "MID_MASTER":
            continue  # recognized; no binding required at this layer
        # Multi-BAI2 union (owner review, 2026-07-19): a run's open lines
        # can span months, but newest-wins read only ONE BAI2 file — a July
        # transmission could never enrich June's lines.  Load EVERY BAI2
        # file (each with its OWN binding; raw-.txt and spreadsheet shapes
        # legitimately differ) and union at the INDEX level in _bai2_index.
        # BAI2 is enrichment-only — never a pool entry — so this changes
        # only which addenda are available, never conservation.
        if role == "BAI2" and len(files) > 1:
            parsed = []
            for rf in files:                  # newest-first route order
                rows, _ = read_rows(rf)
                mapping, hi = bind_columns(rows, specs, filename=rf.filename)
                parsed.append({"rows": rows, "map": mapping,
                               "header_index": hi, "file": rf.filename})
            loaded[role] = {"files": parsed,
                            "file": " + ".join(p["file"] for p in parsed)}
            bound_report[role] = {"files": [p["file"] for p in parsed],
                                  "union": "index-level"}
            continue
        # Pagination union (owner review, 2026-07-19; generalized beyond MET
        # 2026-07-19): real Oracle exports arrive as same-date "_2/_3" page
        # shards.  Same-date files of a paginatable role are PAGES of one
        # export — union them (identical column binding required, duplicate
        # rows barred).  Different-date files stay newest-wins (refreshes).
        if role in PAGINATABLE_ROLES and len(files) > 1 and \
                len({_leading_date_key(f.filename) for f in files}) == 1:
            entry = _union_same_date_shards(files, specs, role)
            loaded[role] = {k: entry[k] for k in
                            ("rows", "map", "header_index", "file")}
            bound_report[role] = {"file": entry["file"],
                                  "columns": entry["map"],
                                  "header_index": entry["header_index"],
                                  "union_of": entry["union_of"]}
            continue
        rf = _pick_newest(files, role)  # newest YYYYMMDD stamp wins
        if role == "MID_MASTER":
            loaded["MID_MASTER"] = load_mid_master(rf)
            continue
        try:
            rows, _title = read_rows(rf)
        except InvalidSourceData:
            raise
        mapping, hi = bind_columns(rows, specs, filename=rf.filename)
        loaded[role] = {"rows": rows, "map": mapping, "header_index": hi, "file": rf.filename}
        bound_report[role] = {"file": rf.filename, "columns": mapping, "header_index": hi}
    runlog["roles_bound"] = bound_report
    return loaded


# Roles whose real exports paginate into same-date "_2/_3" shards.  Every
# other role keeps single-file semantics (a same-date second file is a
# conflict, not a page): MID_MASTER / DEPT_INFO / CFG_* / EDISON_* /
# ENRICHED stay newest-wins-or-fail-loud; CHART_OF_ACCOUNTS has its own
# multi-file loader; BAI2 has its own index-level union.
PAGINATABLE_ROLES = {"BSL", "ST", "RECEIPTS", "PAYMENTS", "ALL_BSL", "MET"}


def _union_same_date_shards(files, specs, role):
    """Union same-date page shards of one paginated export.  Requires every
    shard to bind the IDENTICAL column mapping (else fail loud naming both
    files), and bars duplicate data rows across shards — a re-uploaded copy
    of the same export is a DUPLICATE, not a page, and unioning it would
    double-count money (fatal for BSL conservation)."""
    shards = sorted(files, key=lambda f: f.filename)
    rows, _ = read_rows(shards[0])
    mapping, hi = bind_columns(rows, specs, filename=shards[0].filename)
    rows = list(rows)
    seen_rows = {tuple(N(c) for c in r) for r in rows[hi + 1:]
                 if any(N(c) for c in r)}
    for extra in shards[1:]:
        erows, _ = read_rows(extra)
        emap, ehi = bind_columns(erows, specs, filename=extra.filename)
        if emap != mapping:
            raise InvalidSourceData(
                extra.filename, role,
                f"page shard binds different columns than "
                f"{shards[0].filename} ({emap} vs {mapping}) — "
                "not pages of one export")
        for r in erows[ehi + 1:]:
            key = tuple(N(c) for c in r)
            if any(key):
                if key in seen_rows:
                    raise InvalidSourceData(
                        extra.filename, role,
                        f"data row duplicated across same-date shards "
                        f"({shards[0].filename} + {extra.filename}) — a "
                        "duplicate upload, not pagination; remove one file "
                        "(unioning it would double-count)")
                seen_rows.add(key)
            rows.append(r)
    # BSL pages carry globally unique statement-line keys; a repeat across
    # shards means the same bank line twice — conservation poison.
    if role == "BSL":
        lk = mapping.get("line_key")
        if lk is not None:
            seen_lk = set()
            for r in rows[hi + 1:]:
                v = N(_cell(r, lk))
                if not v:
                    continue
                if v in seen_lk:
                    raise InvalidSourceData(
                        shards[0].filename, role,
                        f"statement line {v!r} appears in more than one "
                        "same-date BSL shard — duplicate upload, not "
                        "pagination")
                seen_lk.add(v)
    fname = " + ".join(f.filename for f in shards)
    return {"rows": rows, "map": mapping, "header_index": hi, "file": fname,
            "union_of": [f.filename for f in shards]}


def _pick_newest(files, role):
    """Newest-wins file selection for a role (spec §4.1).  Two files whose
    date keys tie would mean one silently ignored — fail loud instead and
    ask for a disambiguating YYYYMMDD stamp."""
    if len(files) > 1 and _leading_date_key(files[0].filename) == _leading_date_key(files[1].filename):
        raise InvalidSourceData(
            files[0].filename, role,
            f"multiple files tie for role {role}: "
            f"{[f.filename for f in files]} — add a YYYYMMDD date stamp so "
            "the newest can be chosen (the rest are never read silently)")
    return files[0]


def _bai2_index(loaded):
    """Index BAI2 rows by (date, signed cents).  DETAIL1..DETAILn columns are
    located by exact header name PER FILE (deterministic; they carry the full
    addenda the Oracle BSL feed truncates at ~1000 chars).  Accepts both the
    multi-file shape ({"files": [...]}) and the legacy single-file dict;
    candidate lists union across files in newest-first order — _bai2_enrich's
    decline-on-ambiguity already arbitrates merged lists conservatively."""
    bai = loaded.get("BAI2")
    if not bai:
        return None
    parts = bai["files"] if "files" in bai else [bai]
    idx = {}
    for part in parts:
        rows, m, hi = part["rows"], part["map"], part["header_index"]
        header = rows[hi]
        detail_cols = [i for i, h in enumerate(header)
                       if re.fullmatch(r"DETAIL\d+", _norm_header(h) or "")]
        for r in rows[hi + 1:]:
            dt = parse_date(_cell(r, m.get("post_date")))
            amt = cents(_cell(r, m.get("amount")))
            if dt is None or amt is None:
                continue
            details = "".join(N(_cell(r, c)) for c in detail_cols)
            cand = {
                "details": details,
                "description": N(_cell(r, m.get("description"))),
                "bank_reference": clean_ref(_cell(r, m.get("bank_reference"))),
                "customer_reference": clean_ref(_cell(r, m.get("customer_reference"))),
            }
            bucket = idx.setdefault((dt, amt), [])
            # Same-transaction dedup across overlapping BAI2 windows (the
            # 0715 csv and 0718 txt both carry July 1-15): the bank
            # reference identifies the transaction — verified identical
            # across file formats on real UTC data (326/350 overlapping
            # keys).  Keep ONE candidate per bank reference, preferring the
            # richer addenda (the raw transmission); without it, treating
            # the duplicate as a second candidate made _bai2_enrich decline
            # 31 previously-clean joins.  Blank-reference twins dedup on a
            # space-normalized details prefix; genuinely distinct same-day
            # same-amount transactions keep both (the decline is then
            # correct).
            dup = None
            for prev in bucket:
                if cand["bank_reference"] and \
                        prev["bank_reference"] == cand["bank_reference"]:
                    dup = prev
                    break
                pa = prev["details"].replace(" ", "")
                ca = cand["details"].replace(" ", "")
                if pa and ca and (pa.startswith(ca) or ca.startswith(pa)):
                    dup = prev
                    break
            if dup is None:
                bucket.append(cand)
            elif len(cand["details"]) > len(dup["details"]):
                dup.update(cand)
    return idx


def _bai2_enrich(info, dt, amt, bai2_idx, counters):
    """Resolve the BSL line's BAI2 row by (date, cents), tiebreaking on the
    space-normalized addenda prefix; never guess on a residual tie.  Returns
    the enriched additional-information text."""
    cands = bai2_idx.get((dt, amt), [])
    if not cands:
        counters["no_hit"] += 1
        return info
    chosen = None
    if len(cands) == 1:
        chosen = cands[0]
    else:
        key = N(info)[:40].replace(" ", "")
        pref = [c for c in cands
                if key and c["details"].replace(" ", "").startswith(key)]
        if len(pref) == 1:
            chosen = pref[0]
    if chosen is None:
        counters["ambiguous"] += 1
        return info
    counters["joined"] += 1
    extra = " ".join(x for x in (chosen["description"], chosen["details"],
                                 chosen["bank_reference"], chosen["customer_reference"]) if x)
    return (N(info) + " | BAI2: " + extra).strip(" |")


def _enriched_index(loaded):
    """Index the pre-parsed enriched-BSL workbook by Statement line key and by
    (date, signed cents)."""
    en = loaded.get("ENRICHED")
    if not en:
        return None
    rows, m, hi = en["rows"], en["map"], en["header_index"]
    by_line, by_da = {}, {}
    for r in rows[hi + 1:]:
        dt = parse_date(_cell(r, m.get("date")))
        amt = cents(_cell(r, m.get("amount")))
        if amt is None:
            continue
        rec = {
            "info": N(_cell(r, m.get("additional_info"))),
            "trace": clean_ref(_cell(r, m.get("parsed_trace"))),
            "mid": znorm(clean_ref(_cell(r, m.get("parsed_mid")))),
            "key": N(_cell(r, m.get("recon_key"))),
            "category": N(_cell(r, m.get("category"))),
        }
        lk = znorm(_cell(r, m.get("line_key")))
        if lk:
            by_line.setdefault(lk, []).append(rec)
        if dt is not None:
            by_da.setdefault((dt, amt), []).append(rec)
    return {"by_line": by_line, "by_da": by_da}


def _enriched_enrich(info, line_key, dt, amt, en_idx, counters):
    """Join one BSL to its enriched-workbook row: exact Statement line key
    first, unique (date, cents) as fallback; never guess a residual tie.
    Appends the parsed trace/MID/recon-key text to the addenda."""
    cands = en_idx["by_line"].get(znorm(line_key), [])
    if len(cands) != 1:
        da = en_idx["by_da"].get((dt, amt), [])
        if len(da) == 1:
            cands = da
        elif cands or da:
            counters["ambiguous"] += 1
            return info
        else:
            counters["no_hit"] += 1
            return info
    chosen = cands[0]
    counters["joined"] += 1
    extra = " ".join(x for x in (
        chosen["category"],
        ("trace:" + chosen["trace"]) if chosen["trace"] else "",
        ("mid:" + chosen["mid"]) if chosen["mid"] else "",
        chosen["key"],
    ) if x)
    base = N(info)
    # Prefer the LONGER additional-information text (the enriched workbook
    # carries the untruncated addenda when no BAI2 export is present).
    if len(chosen["info"]) > len(base):
        base = chosen["info"]
    if not extra:
        return base
    return (base + " | ENR: " + extra).strip(" |")


def build_bsls(loaded, by_role, account, runlog=None):
    """Build the open BSL list from the BSL file, enriched with the matching
    BAI2 row's full addenda + raw bank references when a BAI2 export is
    present (the chain: BSL -> BAI2 -> ORT/MET text -> amounts -> ST)."""
    bsls = []
    bai2_idx = _bai2_index(loaded)
    en_idx = _enriched_index(loaded)
    counters = {"joined": 0, "ambiguous": 0, "no_hit": 0}
    en_counters = {"joined": 0, "ambiguous": 0, "no_hit": 0}
    if "BSL" in loaded:
        b = loaded["BSL"]
        rows, m, hi = b["rows"], b["map"], b["header_index"]
        for r in rows[hi + 1:]:
            amt = cents(_cell(r, m.get("amount")))
            dt = parse_date(_cell(r, m.get("date")))
            if amt is None:
                continue
            info = N(_cell(r, m.get("additional_info")))
            if bai2_idx is not None:
                info = _bai2_enrich(info, dt, amt, bai2_idx, counters)
            if en_idx is not None:
                info = _enriched_enrich(
                    info, _cell(r, m.get("line_key")) or "", dt, amt,
                    en_idx, en_counters)
            bsl = make_bsl(
                line_key=_cell(r, m.get("line_key")) or f"L{len(bsls)+1}",
                dt=dt,
                amount_cents=amt,
                reference=_cell(r, m.get("reference")),
                account_servicer_reference=_cell(r, m.get("account_servicer_reference")),
                additional_info=info,
                transaction_type=_cell(r, m.get("transaction_type")),
                transaction_code=_cell(r, m.get("transaction_code")),
                customer_reference=_cell(r, m.get("customer_reference")),
            )
            bsls.append(bsl)
    if runlog is not None and bai2_idx is not None:
        runlog["bai2_enrichment"] = counters
    if runlog is not None and en_idx is not None:
        runlog["enriched_bsl_enrichment"] = en_counters
    # Deterministic order; also disambiguate any duplicate line keys.
    bsls.sort(key=lambda b: (b.date.toordinal() if b.date else 0, b.amount_cents, b.line_key))
    seen = {}
    for b in bsls:
        if b.line_key in seen:
            seen[b.line_key] += 1
            b.line_key = f"{b.line_key}#{seen[b.line_key]}"
        else:
            seen[b.line_key] = 0
    return bsls


# ======================================================================
# Section 15 — run() orchestrator + JSON run log
# ======================================================================

def run(account_input_dir, output_dir="./outputs", present=True):
    """Single entry point (Section 15.1): route -> bind -> validate -> build
    pool -> forward P0-P10 -> write workbook -> audit -> present."""
    runlog = {"input_dir": account_input_dir, "stages": []}

    # P0 route + bind + validate
    skipped = []
    by_role = route_folder(account_input_dir, skipped)
    runlog["files_routed"] = {role: [f.filename for f in files]
                              for role, files in by_role.items()}
    # Doctrine rule 3 (never silently drop): route_folder announced every
    # skip on stderr the moment it happened; the runlog carries the record
    # (file, reason, suggested fix) for the manifest/report trail.
    if skipped:
        runlog["files_ignored_by_name"] = skipped
    not_loaded = sorted(r for r in by_role if r in _RECOGNIZED_NOT_LOADED)
    if not_loaded:
        runlog["roles_recognized_not_loaded"] = {
            r: _RECOGNIZED_NOT_LOADED[r] for r in not_loaded}
    account = _resolve_account(by_role)
    runlog["account"] = account
    runlog["stages"].append("route")

    # ALL_DATA feeds nothing forward — it is the backward engine's fuel,
    # which now lives in the Unreconcile2 project.  Recognized, noted, and
    # deliberately not loaded (this is the forward engine's speed win).
    if "ALL_DATA" in by_role:
        runlog["all_data"] = ("present: not loaded (forward-only engine; "
                              "run Unreconcile2 for the backward re-audit)")
    if "RECONCILED" in by_role:
        runlog["reconciled_exports"] = (
            "present: not loaded (forensic reconciled exports — offline "
            "config-audit / Unreconcile2 fuel, never forward-engine input)")

    loaded = load_and_bind(by_role, runlog)
    runlog["stages"].append("bind")

    pool = build_pool(loaded, account, runlog)
    runlog["stages"].append("pool")

    bsls = build_bsls(loaded, by_role, account, runlog)
    runlog["bsl_count"] = len(bsls)
    runlog["bsl_by_lane"] = _count_lanes(bsls)

    # Forward P0-P10
    placements = forward_reconcile(bsls, pool, loaded, account, runlog)
    runlog["stages"].append("forward")

    # Writer
    os.makedirs(output_dir, exist_ok=True)
    recon_path = os.path.join(output_dir, f"{account}_reconciliation.xlsx")
    recon_summary = write_reconciliation_workbook(recon_path, account, placements)
    runlog["recon_workbook"] = recon_path
    runlog["recon_summary"] = recon_summary
    runlog["stages"].append("write")

    # Conservation ledger (Section 15.3): input == Matches + Candidates +
    # Misdirected + Review.
    total = (recon_summary["matches"] + recon_summary["candidates"]
             + recon_summary["misdirected"] + recon_summary["reviews"])
    assert total == len(bsls), f"conservation: {total} placed != {len(bsls)} BSLs"

    # Audit (Section 14) — imported here so the engine stays independently
    # importable; the audit module imports nothing from the engine.
    audit_result = None
    try:
        import recon_audit
        audit_result = recon_audit.audit(account_input_dir, recon_path, account)
    except ImportError:
        audit_result = {"status": "SKIPPED", "reason": "recon_audit not importable"}
    runlog["audit"] = audit_result
    runlog["stages"].append("audit")

    # Configuration audit (owner, 2026-07-19; orphan-doctrine R5): advisory
    # replay of the CM configuration against this run's outcomes.  Optional —
    # no CFG exports, no-op; never affects placements, workbook, or audit.
    if config_audit(loaded, account, placements, pool, output_dir, runlog):
        runlog["stages"].append("config_audit")

    # Write JSON run log (Section 15.7).
    log_path = os.path.join(output_dir, f"{account}_runlog.json")
    with open(log_path, "w", encoding="utf-8") as fh:
        json.dump(runlog, fh, indent=2, default=_json_default)
    runlog["runlog_path"] = log_path

    if present and audit_result and audit_result.get("status") not in ("PASS", "SKIPPED"):
        raise ReconError(f"Audit failed; workbooks withheld: {audit_result}")
    return runlog


def _resolve_account(by_role):
    for role in ("BSL", "ALL_DATA"):
        for f in by_role.get(role, []):
            acc = infer_account(f.filename)
            if acc:
                return acc
    return "UNKNOWN"


def _count_lanes(bsls):
    out = {}
    for b in bsls:
        out[b.lane] = out.get(b.lane, 0) + 1
    return out


def _json_default(o):
    if isinstance(o, (date, datetime)):
        return o.isoformat()
    if isinstance(o, set):
        return sorted(o)
    return str(o)


# ======================================================================
# CLI
# ======================================================================

def main(argv=None):
    ap = argparse.ArgumentParser(description="UT Cash Management Reconciliation Engine")
    ap.add_argument("input_dir", help="Folder of source files for one account")
    ap.add_argument("-o", "--output", default="./outputs", help="Output folder")
    ap.add_argument("--no-present-gate", action="store_true",
                    help="Do not raise if the audit fails (write anyway)")
    args = ap.parse_args(argv)
    runlog = run(args.input_dir, args.output, present=not args.no_present_gate)
    print(json.dumps({
        "account": runlog.get("account"),
        "bsl_count": runlog.get("bsl_count"),
        "recon_summary": runlog.get("recon_summary"),
        "audit": (runlog.get("audit") or {}).get("status"),
        "runlog": runlog.get("runlog_path"),
    }, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
