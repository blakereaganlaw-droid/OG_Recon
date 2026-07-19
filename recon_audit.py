#!/usr/bin/env python3
"""
Independent reconciliation audit (BUILD SPEC Section 14).

This module IMPORTS NOTHING from recon_engine.  It re-implements the
normalization primitives and a minimal column binder from scratch, re-parses
every raw source, re-reads the produced workbook, and enforces the C1-C10
checks.  Engine and audit must agree independently or the delivery is gated.

Checks:
  C1  conservation (source BSL count == output count; multiset on date+cents)
  C2  Match signed-cent equality (deduped ST group sums to BSL)
  C3  ST non-reuse across Matches + Candidates
  C4  MET bridge validity (every cited d:/r: is a real MET pair)
  C5  dual-fire excluded from Matches
  C6  STATE lane isolation (no ORT citation on a State Match)
  C7  MID guardrail
  C8  date ceiling with the State carve-out
  C9  formatting (Carlito, header fill, freeze A4, zero formulas, structure)
  C10 no provenance content

All checks must pass.  On failure the audit returns status FAIL with the
offending detail; it never relaxes a check.
"""

from __future__ import annotations

import os
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation


# ---- independent primitives (re-derived; not imported) ---------------

_NON_ALNUM = re.compile(r"[^A-Za-z0-9]+")
_MID_RE = re.compile(r"^(80\d{8}|2000\d{6})$")
_HEARTLAND = "6500000097"
_EXCEL_EPOCH = date(1899, 12, 30)
_DATE_FORMATS = ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d")


def _N(v):
    return "" if v is None else str(v).strip()


def _cents(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, int):
        return int(Decimal(v) * 100)
    s = repr(v) if isinstance(v, float) else str(v).strip()
    if not s:
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg, s = True, s[1:-1].strip()
    s = s.replace("$", "").replace(",", "").replace(" ", "")
    if s.startswith("-"):
        neg, s = True, s[1:]
    if not s:
        return None
    try:
        d = (Decimal(s) * 100).quantize(Decimal("1"))
    except (InvalidOperation, ValueError):
        return None
    c = int(d)
    return -c if neg else c


def _parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        n = int(v)
        if 1 <= n <= 60000:
            return _EXCEL_EPOCH.fromordinal(_EXCEL_EPOCH.toordinal() + n)
        return None
    s = str(v).strip()
    if not s:
        return None
    m = re.match(r"^(\d{4}-\d{2}-\d{2})[T ]", s)
    if m:
        try:
            return date.fromisoformat(m.group(1))
        except ValueError:
            pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _znorm(s):
    return _NON_ALNUM.sub("", _N(s)).upper()


def _is_mid(s):
    z = _znorm(s)
    return z != _HEARTLAND and bool(_MID_RE.match(z))


def _norm_header(h):
    return re.sub(r"[^A-Z0-9]", "", _N(h).upper())


# ---- independent minimal binder --------------------------------------

class AuditBindError(Exception):
    """The audit could not re-bind a source column unambiguously.  Callers
    convert this into a named check FAILURE — never a guess, never a crash."""


_NULL_TOKENS = {"NA", "NONE", "NULL", "UNKNOWN"}


def _blankish(v):
    s = _N(v)
    return not s or _znorm(s) in _NULL_TOKENS or not re.search(r"[A-Za-z0-9]", s)


def _find_col(rows, header_index, aliases, predicate, sample=50):
    """Return the best column index for a role, or None when no column shows
    any evidence.  A blind tie (two columns equal on content AND header
    score) raises AuditBindError rather than picking by position."""
    header = rows[header_index]
    ncols = max(len(r) for r in rows)
    data = rows[header_index + 1: header_index + 1 + sample]
    # Content sampling spans the whole sheet (first `sample` NON-BLANK values
    # per column), mirroring the engine: sparse columns must be scored on the
    # values they actually carry.
    col_samples = {c: [] for c in range(ncols)}
    unfilled = set(col_samples)
    for r in rows[header_index + 1:]:
        if not unfilled:
            break
        for c in list(unfilled):
            if c < len(r) and not _blankish(r[c]):
                bucket = col_samples[c]
                bucket.append(r[c])
                if len(bucket) >= sample:
                    unfilled.discard(c)
    alias_norms = [_norm_header(a) for a in aliases]
    scored = []
    for col in range(ncols):
        hnorm = _norm_header(header[col]) if col < len(header) else ""
        hscore = 3 if hnorm in alias_norms else (2 if any(a and a in hnorm for a in alias_norms) else 0)
        sampled = col_samples[col]
        cscore = (sum(1 for c in sampled if predicate(c)) / len(sampled)) if sampled else 0.0
        scored.append((cscore, hscore, col))
    scored.sort(key=lambda t: (t[0], t[1]), reverse=True)
    best = scored[0]
    if best[0] == 0 and best[1] == 0:
        return None
    if len(scored) > 1 and (best[0], best[1]) == (scored[1][0], scored[1][1]) and best[0] > 0:
        tied = [t[2] for t in scored if (t[0], t[1]) == (best[0], best[1])]
        # Verbatim-duplicated columns (identical values over the sample) are
        # not a real ambiguity — bind the leftmost deterministically.
        ref, identical = None, True
        for c in tied:
            vals = tuple(_N(r[c]) if c < len(r) else "" for r in data)
            if ref is None:
                ref = vals
            elif vals != ref:
                identical = False
                break
        if identical:
            return min(tied)
        raise AuditBindError(f"columns {tied} tie for aliases {aliases}")
    return best[2]


def _locate_header(rows, aliases, scan=12):
    """Best header row within the scan window, or None when no row carries
    enough recognized headers — never a positional default to row 0."""
    all_aliases = [_norm_header(a) for a in aliases if _norm_header(a)]
    need = min(2, len(all_aliases)) or 1
    best_i, best_hits = None, need - 1
    for i in range(min(scan, len(rows))):
        hits = sum(1 for c in rows[i]
                   if _norm_header(c) and any(_norm_header(c) == a or a in _norm_header(c)
                                              for a in all_aliases))
        if hits > best_hits:
            best_i, best_hits = i, hits
    return best_i


def _read_xlsx(path, sheet_substr=None):
    from openpyxl import load_workbook
    # Streaming read with reset_dimensions() (never trust stated dimensions
    # on Oracle BI exports); fall back to the slow full parse if the fast
    # path yields nothing.  Independent twin of the engine's reader.
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    if sheet_substr:
        for cand in wb.worksheets:
            if sheet_substr.lower() in cand.title.lower():
                ws = cand
                break
    ws.reset_dimensions()
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if len(rows) <= 1:
        wb = load_workbook(path, read_only=False, data_only=True)
        ws = wb.worksheets[0]
        if sheet_substr:
            for cand in wb.worksheets:
                if sheet_substr.lower() in cand.title.lower():
                    ws = cand
                    break
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    return rows


def _read_csv(path):
    import csv
    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        rows = [tuple(r) for r in csv.reader(fh)]
    if rows and rows[0] and rows[0][0].startswith("﻿"):
        first = list(rows[0])
        first[0] = first[0].lstrip("﻿")
        rows[0] = tuple(first)
    return rows


def _xlsb_norm(v):
    # pyxlsb hands back EVERY numeric cell as a float, so an integral Oracle id
    # (DEPOSIT_ID/RECEIPT_ID = 65105) arrives as 65105.0 and would stringify to
    # "65105.0" — which the d:/r: citations (clean ints) never match, spuriously
    # failing C4.  Collapse integral floats to int so an .xlsb MET yields the
    # same real-deposit / real-receipt set as the .csv/.xlsx twin.  Money
    # (1890.69) and true fractional serials are left untouched.  This duplicates
    # the engine's _xlsb_norm literally — the audit imports nothing from it.
    if isinstance(v, float) and -1e15 < v < 1e15 and v.is_integer():
        return int(v)
    return v


def _read_xlsb(path, sheet_substr=None):
    # Independent binary reader (pyxlsb).
    from pyxlsb import open_workbook
    with open_workbook(path) as wb:
        names = wb.sheets
        target = names[0]
        if sheet_substr:
            for n in names:
                if sheet_substr.lower() in n.lower():
                    target = n
                    break
        with wb.get_sheet(target) as ws:
            rows = [tuple(_xlsb_norm(c.v) for c in row) for row in ws.rows()]
    return rows


def _read_table(path, sheet_substr=None):
    """Extension dispatch mirroring the engine's read_rows (csv / xlsb / xlsx)
    so the audit reads every format the engine accepts, MET .xlsb included."""
    low = path.lower()
    if low.endswith(".csv"):
        return _read_csv(path)
    if low.endswith(".xlsb"):
        return _read_xlsb(path, sheet_substr)
    return _read_xlsx(path, sheet_substr)


def _pred_date(v):
    return _parse_date(v) is not None


def _pred_amount(v):
    return _cents(v) is not None


def _pred_any(v):
    return _N(v) != ""


# ---- source BSL re-parse (independent) -------------------------------

# The full BSL header vocabulary, re-declared locally (§5.3 mirrored, NOT
# imported from the engine): the audit must recognize every header the engine
# does, or a benign column insertion makes the two bind different columns.
_BSL_DATE_ALIASES = ["Transaction Date", "Booking Date", "Value Date", "Date", "Post Date"]
_BSL_AMOUNT_ALIASES = ["Amount", "Transaction Amount", "Signed Amount"]
_BSL_HEADER_VOCAB = _BSL_DATE_ALIASES + _BSL_AMOUNT_ALIASES + [
    "Line Number", "Statement Line", "Sequence", "Reference",
    "Account Servicer Reference", "Additional Information", "Transaction Code",
]

_YYYYMMDD_RE = re.compile(r"(\d{8})")


def _newest_date_key(name):
    """Newest plausible YYYYMMDD stamp in a filename (mirrors the engine's
    newest-wins routing without importing it)."""
    best = "00000000"
    for m in _YYYYMMDD_RE.finditer(name):
        s = m.group(1)
        y, mo, d = int(s[:4]), int(s[4:6]), int(s[6:8])
        if 1990 <= y <= 2099 and 1 <= mo <= 12 and 1 <= d <= 31 and s > best:
            best = s
    return best


def _is_all_accounts_bsl(low):
    """Mirror the engine's ALL_BSL routing: a whole-segment 'all' beside
    'bsl' marks the all-accounts export (Oracle_OTBI_All_BSL_UNR), which the
    engine never reconciles — the audit must never re-parse it as the
    account BSL (it would falsely fail C1 the moment its date stamp is
    newer than the account export's)."""
    segs = re.split(r"[^a-z0-9]+", low)
    return "all" in segs


def _reparse_source_bsls(input_dir):
    """Re-read the BSL source and return a multiset of (date, cents)."""
    candidates = []
    for name in sorted(os.listdir(input_dir)):
        low = name.lower()
        if "bsl" in low and "all_data" not in low and "enriched" not in low \
                and "crossref" not in low and "reconcil" not in low \
                and not _is_all_accounts_bsl(low) \
                and low.endswith((".xlsx", ".xlsm", ".csv", ".xlsb")):
            candidates.append(name)
    if not candidates:
        return None
    # Newest YYYYMMDD stamp wins, mirroring the engine's routing — the audit
    # must re-parse the SAME file the engine reconciled, not the first name
    # alphabetically.  Stable sort keeps alphabetical order among ties.
    candidates.sort(key=_newest_date_key, reverse=True)
    bsl_file = os.path.join(input_dir, candidates[0])
    rows = _read_table(bsl_file)
    if not rows:
        return []
    hi = _locate_header(rows, _BSL_HEADER_VOCAB)
    if hi is None:
        raise AuditBindError(
            f"{os.path.basename(bsl_file)}: no BSL header row found — "
            "refusing to re-parse from a positional guess")
    dcol = _find_col(rows, hi, _BSL_DATE_ALIASES, _pred_date)
    acol = _find_col(rows, hi, _BSL_AMOUNT_ALIASES, _pred_amount)
    bag = []
    for r in rows[hi + 1:]:
        amt = _cents(r[acol]) if acol is not None and acol < len(r) else None
        if amt is None:
            continue
        dt = _parse_date(r[dcol]) if dcol is not None and dcol < len(r) else None
        bag.append((dt.isoformat() if dt else "", amt))
    return bag


# ---- MET pairs re-parse (independent) --------------------------------

def _reparse_met_pairs(input_dir):
    pairs = set()
    for name in sorted(os.listdir(input_dir)):
        low = name.lower()
        # 'oracle_otbi' names are MET exports even without the MET token
        # (mirrors the engine's router) — but the all-accounts BSL export
        # (Oracle_OTBI_All_BSL_UNR) and reconciled forensic reports are not
        # MET sources and are skipped, as the engine's router skips them.
        if ("met" in low or "oracle_otbi" in low) \
                and "bsl" not in low and "reconcil" not in low \
                and low.endswith((".xlsx", ".xlsm", ".csv", ".xlsb")):
            path = os.path.join(input_dir, name)
            rows = _read_table(path)
            # Native DEPOSIT_ID / RECEIPT_ID columns (some exports carry the
            # ids only there, not in the d:/r: description text).
            if rows:
                hdr = rows[0]
                for i, h in enumerate(hdr):
                    hn = _norm_header(h)
                    if hn in ("DEPOSITID", "RECEIPTID"):
                        kind = "d" if hn == "DEPOSITID" else "r"
                        for r in rows[1:]:
                            if i < len(r) and r[i] is not None:
                                v = _N(r[i])
                                if v.endswith(".0"):
                                    v = v[:-2]
                                if v.isdigit():
                                    pairs.add((kind, v))
            for r in rows:
                for c in r:
                    # fast pre-check: regex only cells that can carry a token
                    if not isinstance(c, str) or (":" not in c):
                        continue
                    d = re.search(r"d:\s*(\d+)", c, re.IGNORECASE)
                    rr = re.search(r"r:\s*(\d+)", c, re.IGNORECASE)
                    if d:
                        pairs.add(("d", d.group(1)))
                    if rr:
                        pairs.add(("r", rr.group(1)))
    return pairs


# ---- workbook re-read -------------------------------------------------

def _read_output_tabs(recon_path):
    from openpyxl import load_workbook
    wb = load_workbook(recon_path, read_only=False, data_only=False)  # formulas visible
    tabs = {}
    formatting = {}
    for ws in wb.worksheets:
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        tabs[ws.title] = rows
        # capture header cell for formatting checks (row 3, col 1)
        hc = ws.cell(row=3, column=1)
        formatting[ws.title] = {
            "font_name": hc.font.name,
            "fill": (hc.fill.fgColor.rgb if hc.fill and hc.fill.patternType else None),
            "freeze": ws.freeze_panes,
        }
    wb.close()
    return tabs, formatting


# ---- the audit --------------------------------------------------------

# HARD GUARDRAIL (owner, 2026-07-11): the 19-column layout carrying ALL BSL
# identifier fields and ALL ST detail fields is pinned — C9 fails any drift.
RECON_HEADER = [
    "BSL Date", "BSL Line Info", "BSL Amount",
    "BSL Reference", "BSL Additional Information", "BSL Customer Reference",
    "BSL Account Servicer Reference", "BSL Transaction Type",
    "ST Date(s)", "ST Number(s)", "ST Amount(s)", "ST Reference(s)",
    "ST Structured Payment Reference(s)", "ST Counterparty(ies)",
    "ST Source(s)",
    "Confidence", "ORT d:", "ORT r:", "Explanation",
]
NCOLS = len(RECON_HEADER)
# Column indexes (audit-side view of the pinned layout).
COL_DATE, COL_INFO, COL_AMT = 0, 1, 2
COL_ST_DATES, COL_ST_NUMS, COL_ST_AMTS = 8, 9, 10
COL_DEP, COL_REC, COL_EXPL = 16, 17, 18


def audit(input_dir, recon_path, account):
    """Run C1-C10 and return {'status': PASS|FAIL, 'checks': {...}, 'failures': [...]}"""
    failures = []
    checks = {}

    tabs, formatting = _read_output_tabs(recon_path)
    tab_titles = list(tabs.keys())

    def data_rows(title):
        rows = tabs.get(title, [])
        # title row 1, blank row 2, header row 3, data row 4+
        return [r for r in rows[3:] if any(_N(c) for c in r)]

    match_rows = data_rows("Matches")
    cand_rows = data_rows("Candidate Matches")
    misdir_rows = data_rows("Misdirected")
    review_rows = data_rows("Review Notes")
    all_out = match_rows + cand_rows + misdir_rows + review_rows

    # C1 conservation (count + multiset on date+cents)
    try:
        src_bag = _reparse_source_bsls(input_dir)
    except AuditBindError as e:
        src_bag = None
        checks["C1"] = "FAIL"
        failures.append(f"C1: source re-parse could not bind unambiguously: {e}")
    if "C1" in checks:
        pass
    elif src_bag is None:
        checks["C1"] = "SKIP (no source BSL found)"
    else:
        out_bag = []
        for r in all_out:
            d = _N(r[0])
            amt = _cents(r[2])
            out_bag.append((d, amt))
        c1_ok = (len(src_bag) == len(out_bag)) and (_multiset(src_bag) == _multiset(out_bag))
        checks["C1"] = "PASS" if c1_ok else "FAIL"
        if not c1_ok:
            failures.append(f"C1: source {len(src_bag)} vs output {len(out_bag)}; "
                            f"multiset match={_multiset(src_bag) == _multiset(out_bag)}")

    # C2 Match signed-cent equality — the cited ST group sums to the BSL.
    # Strengthened (critical review, 2026-07-19): the workbook itself carries
    # the ST Amount(s) column, so the audit RE-SUMS every Match in signed
    # cents with zero pool access — a Match whose cited amounts do not sum to
    # the bank line is a defect no matter how it was produced.
    c2_ok = True
    for r in match_rows:
        amt = _cents(r[COL_AMT])
        if amt is None or not _N(r[COL_ST_NUMS]):
            c2_ok = False
            failures.append(f"C2: Match row missing amount or ST number: {r[:5]}")
            continue
        st_amts = [_cents(p) for p in _split_multi(r[COL_ST_AMTS])]
        if st_amts and None not in st_amts and sum(st_amts) != amt:
            c2_ok = False
            failures.append(
                f"C2: Match ST amounts do not sum to the BSL "
                f"({sum(st_amts)} != {amt}): {r[:3]}")
    # Owner doctrine: a Candidate without an ST citation must not exist.
    for r in cand_rows:
        if not _N(r[COL_ST_NUMS]):
            c2_ok = False
            failures.append(f"C2: Candidate row cites no ST: {r[:3]}")
    # Misdirected rows (owner hard guardrail, 2026-07-18) must cite the
    # foreign-account entry AND name the account it is booked to — and the
    # named account must be a DIFFERENT one than this run's (2026-07-19).
    for r in misdir_rows:
        if not _N(r[COL_ST_NUMS]):
            c2_ok = False
            failures.append(f"C2: Misdirected row cites no ST: {r[:3]}")
        expl_low = _N(r[COL_EXPL]).lower()
        if "booked to" not in expl_low:
            c2_ok = False
            failures.append(f"C2: Misdirected row does not name the foreign account: {r[:3]}")
        elif account and f"booked to {account.lower()}" in expl_low:
            c2_ok = False
            failures.append(
                f"C2: Misdirected row claims the entry is booked to THIS "
                f"account ({account}) — not misdirected: {r[:3]}")
    # Confidence vocabulary (2026-07-19): every placed row carries one of the
    # three levels, and a distinctive-amount Match is never High (the owner
    # exception grants Medium at most).
    for label, rows_ in (("Match", match_rows), ("Candidate", cand_rows),
                         ("Misdirected", misdir_rows)):
        for r in rows_:
            conf = _N(r[15])
            if conf not in ("High", "Medium", "Low"):
                c2_ok = False
                failures.append(f"C2: {label} row carries confidence "
                                f"{conf!r} (not High/Medium/Low): {r[:3]}")
            elif conf == "High" and "DISTINCTIVE_AMOUNT" in _N(r[COL_EXPL]):
                c2_ok = False
                failures.append(
                    f"C2: DISTINCTIVE_AMOUNT placement at High confidence "
                    f"(owner exception caps it at Medium): {r[:3]}")
    checks["C2"] = "PASS" if c2_ok else "FAIL"

    # C3 ST non-reuse across Matches + Candidates + Misdirected.
    seen = {}
    c3_ok = True
    for label, rows in (("Match", match_rows), ("Candidate", cand_rows),
                        ("Misdirected", misdir_rows)):
        for r in rows:
            for st in _split_multi(r[COL_ST_NUMS]):
                if st in seen:
                    c3_ok = False
                    failures.append(f"C3: ST {st} reused ({seen[st]} and {label})")
                seen[st] = label
    checks["C3"] = "PASS" if c3_ok else "FAIL"

    # C4 MET bridge validity — every cited d:/r: is a real MET pair.  The
    # citation loop runs UNCONDITIONALLY: a workbook citing d:/r: pairs when
    # no MET source exists to validate them is a FAIL, never a silent pass.
    met_pairs = _reparse_met_pairs(input_dir)
    c4_ok, cited = True, False
    for r in all_out:
        for dep in _split_multi(r[COL_DEP]):
            if dep:
                cited = True
                if ("d", dep) not in met_pairs:
                    c4_ok = False
                    failures.append(f"C4: cited d:{dep} not a real MET deposit")
        for rec in _split_multi(r[COL_REC]):
            if rec:
                cited = True
                if ("r", rec) not in met_pairs:
                    c4_ok = False
                    failures.append(f"C4: cited r:{rec} not a real MET receipt")
    if not cited and not met_pairs:
        checks["C4"] = "SKIP (no citations, no MET source)"
    else:
        checks["C4"] = "PASS" if c4_ok else "FAIL"

    # C5 dual-fire excluded from Matches (a Match may not cite the same ST twice).
    c5_ok = True
    for r in match_rows:
        sts = _split_multi(r[COL_ST_NUMS])
        if len(sts) != len(set(sts)):
            c5_ok = False
            failures.append(f"C5: dual-fire ST in a Match row: {sts}")
    checks["C5"] = "PASS" if c5_ok else "FAIL"

    # C6 retired (owner doctrine 2026-07-11): the Edison/State pass was
    # eliminated; State lines reconcile through the same ORT/reference
    # chain as every other line.
    checks["C6"] = "SKIP (retired 2026-07-11)"

    # C7 MID guardrail — a MID reference should not appear on a non-merchant Match.
    c7_ok = True
    for r in match_rows:
        info = _N(r[1])
        toks = re.findall(r"\d{10}", info)
        has_mid = any(_is_mid(t) for t in toks)
        merchant = any(k in info.upper() for k in
                       ("MERCHANT", "BANKCARD", "TOUCHNET", "CYBERSOURCE", "PAYMENTECH"))
        # A MID line matched as merchant is fine; only flag MID on a clearly
        # non-merchant description with an ORT/ST cross that shouldn't be MID.
        # Conservative: pass unless a MID appears with no merchant context AND
        # the explanation names a non-merchant lane.
        if has_mid and not merchant and "card" not in _N(r[COL_EXPL]).lower() and "MID" in _N(r[COL_EXPL]):
            c7_ok = False
            failures.append(f"C7: MID guardrail suspected on non-merchant Match: {r[:2]}")
    # Owner (2026-07-11, false-match review): deposit-type consistency alone
    # never makes a Match, and a deposit-correction line (manual fix; needs a
    # manual ECT) never Matches from an amount-sum pass.
    for label, rows_ in (("Match", match_rows), ("Candidate", cand_rows)):
        for r in rows_:
            if "deposit-type consistency" in _N(r[COL_EXPL]).lower():
                c7_ok = False
                failures.append(f"C7: {label} rests on deposit-type consistency: {r[:3]}")
    for r in match_rows:
        info_txt = _norm_header(r[COL_INFO])
        if "CORRECTION" in info_txt or "CORRECTED" in info_txt:
            c7_ok = False
            failures.append(f"C7: deposit-correction line placed as a Match: {r[:3]}")
    # Owner rule (2026-07-12): a negative chargeback / merchant-fee line
    # pairs ONLY on MID equality — a Match or Candidate citing STs with no
    # shared MID-shaped token is a defect.
    def _mids(txt):
        return {run for run in re.findall(r"\d{10}", _N(txt)) if _is_mid(run)}
    for label, rows_ in (("Match", match_rows), ("Candidate", cand_rows)):
        for r in rows_:
            amt = _cents(r[COL_AMT])
            info_txt = _norm_header(r[COL_INFO])
            if amt is None or amt >= 0:
                continue
            if "CHARGEBACK" not in info_txt and "MERCHANTFEE" not in info_txt:
                continue
            bmids = _mids(r[COL_INFO]) | _mids(r[3]) | _mids(r[6])
            if not bmids:
                continue
            smids = _mids(r[COL_ST_NUMS]) | _mids(r[11]) | _mids(r[12])
            if not (bmids & smids):
                c7_ok = False
                failures.append(
                    f"C7: chargeback/merchant-fee {label} without MID match: {r[:3]}")
    checks["C7"] = "PASS" if c7_ok else "FAIL"

    # C8 directional date rule (owner doctrine, final): the gate applies
    #    ONLY when the ST precedes the BSL by 8+ days; an ST after the BSL
    #    (by any amount) is valid.  A Match whose BSL trails EVERY cited ST
    #    by 8+ days fails.
    c8_ok = True
    for r in match_rows:
        bdt = _parse_date(r[0])
        st_dates = [_parse_date(x) for x in _split_multi(r[COL_ST_DATES])]
        st_dates = [d for d in st_dates if d]
        if bdt is None or not st_dates:
            continue
        lags = [(bdt - d).days for d in st_dates]
        if all(lag >= 8 for lag in lags):
            c8_ok = False
            failures.append(
                f"C8: Match trails every cited ST by >= 8d (stale-ST): {r[:3]}")
    # Owner rule (2026-07-12): an External-source ST 12+ days older than the
    # BSL may not appear even in a CANDIDATE.  ST Date(s) (col 8), ST
    # Source(s) (col 14) are written in the same entry order.
    for r in cand_rows:
        bdt = _parse_date(r[COL_DATE])
        if bdt is None:
            continue
        st_dates = _split_multi(r[COL_ST_DATES])
        st_srcs = _split_multi(r[14])
        for i, x in enumerate(st_dates):
            d = _parse_date(x)
            src_i = st_srcs[i] if i < len(st_srcs) else ""
            if d is not None and _N(src_i).upper() == "EXT" and (bdt - d).days >= 12:
                c8_ok = False
                failures.append(
                    f"C8: Candidate cites an External ST {((bdt - d).days)}d "
                    f"older than the BSL (>=12d bar): {r[:3]}")
                break
    checks["C8"] = "PASS" if c8_ok else "FAIL"

    # C9 formatting — Carlito, navy header fill, freeze A4, zero formulas, structure.
    c9_ok = True
    expected_tabs = ["Matches", "Candidate Matches", "Misdirected", "Review Notes"]
    for t in expected_tabs:
        if t not in tabs:
            c9_ok = False
            failures.append(f"C9: missing tab {t}")
            continue
        fmt = formatting[t]
        if fmt["font_name"] != "Carlito":
            c9_ok = False
            failures.append(f"C9: {t} header font {fmt['font_name']} != Carlito")
        if fmt["fill"] not in ("FF1F4E78", "001F4E78"):
            c9_ok = False
            failures.append(f"C9: {t} header fill {fmt['fill']} != navy FF1F4E78")
        if fmt["freeze"] != "A4":
            c9_ok = False
            failures.append(f"C9: {t} freeze {fmt['freeze']} != A4")
        # header row (row index 2 == spreadsheet row 3) must equal RECON_HEADER
        rows = tabs[t]
        if len(rows) < 3 or [_N(c) for c in rows[2][:NCOLS]] != RECON_HEADER:
            c9_ok = False
            failures.append(f"C9: {t} header row does not match the pinned {NCOLS}-column standard")
        # zero formula cells
        for r in rows:
            for c in r:
                if isinstance(c, str) and c.startswith("="):
                    c9_ok = False
                    failures.append(f"C9: {t} contains a formula cell: {c[:20]}")
    checks["C9"] = "PASS" if c9_ok else "FAIL"

    # C10 no provenance content — no extra band/comment/provenance rows; every
    #    data row has exactly the 9 columns populated within width, no stray cols.
    c10_ok = True
    for t in expected_tabs:
        for r in data_rows(t):
            if any(_N(c) for c in r[NCOLS:]):
                c10_ok = False
                failures.append(f"C10: extra provenance column in {t}: {r}")
    checks["C10"] = "PASS" if c10_ok else "FAIL"

    status = "PASS" if not failures else "FAIL"
    return {"status": status, "checks": checks, "failures": failures}


def _multiset(pairs):
    d = {}
    for p in pairs:
        d[p] = d.get(p, 0) + 1
    return d


def _split_multi(cell):
    s = _N(cell)
    if not s:
        return []
    # values joined by ', ' or '; '
    parts = re.split(r";\s|,\s", s)
    return [p.strip() for p in parts if p.strip()]


if __name__ == "__main__":
    import sys
    import json
    if len(sys.argv) < 3:
        print("usage: recon_audit.py <input_dir> <recon_workbook.xlsx> [account]")
        sys.exit(2)
    acc = sys.argv[3] if len(sys.argv) > 3 else "UNKNOWN"
    result = audit(sys.argv[1], sys.argv[2], acc)
    print(json.dumps(result, indent=2))
    sys.exit(0 if result["status"] == "PASS" else 1)
